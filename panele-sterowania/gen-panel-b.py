#!/usr/bin/env python3
# gen-panel-b.py — generuje Panel-B.xlsx (Grupa B: miasto, ekonomia, społeczeństwo, Wealth, budynki, surowce, tech)
# Uruchom z root projektu: python panele-sterowania/gen-panel-b.py
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
OUT = os.path.join(os.path.dirname(__file__), "Panel-B.xlsx")

HDR = ["ID", "Parametr", "Opis", "Wartość", "Zakres", "Jednostka", "Wpływ", "Plik"]
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SEC_FILL = PatternFill("solid", fgColor="D6E4F0")
SEC_FONT = Font(bold=True, size=11)

BAZA_KEYS = ["praca", "pieniadz", "zywnosc", "nauka", "kultura", "zadowolenie", "obrona", "mnoznik"]
BUILDINGS_COLS = (
    ["id", "nazwa", "kosztBudowy", "przyrostKosztu", "utrzymanie"]
    + [f"baza.{k}" for k in BAZA_KEYS]
    + [f"przyrost.{k}" for k in BAZA_KEYS]
    + ["techUnlock"]
)
RESOURCES_COLS = ["Surowiec", "Typ", "Źródło / budynek", "Uwagi"]
TECH_COLS = [
    "Technologia",
    "Epoka",
    "Poziom",
    "Dostęp do surowca.",
    "wymagany budynek",
    "Wymaga (prereq)",
    "Odblokowuje surowiec.",
    "Odblokowuje budynek",
    "Koszt nauki",
    "Uwagi",
    "Odblokowuje ulepszenie terenu",
]


def load_json(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)


def write_info(ws):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70
    rows = [
        ("Panel sterowania — GRUPA B (Miasto / Okolica / Społeczeństwo)", ""),
        ("", ""),
        ("Jak używać", ""),
        ("1.", "Edytuj kolumnę Wartość (parametry) lub komórki w arkuszach tabelarycznych."),
        ("2.", "Zapisz plik."),
        ("3.", "python panele-sterowania/export-b.py"),
        ("4.", "Playtest: Gra-podglad-OKOLICA-UX.html lub gra po integracji MASTER."),
        ("", ""),
        ("Nie edytuj tutaj", "FOOD / ulepszenia terenu → Panel-A (Grupa A); plony → Panel-A"),
        ("Power / Manpower", "Potega-P-A (9 współcz.), Potega-opcje, Manpower-epoki → export-b.py → gra/data/*.json"),
        ("Eksport Macieja", "W czacie: eksportuj panel — agent odpala export-b.py"),
        ("Kalkulator Excel", "docs/decyzje/POWER-kalkulator-Maciej.xlsx (python tools/build-power-kalkulator-xlsx.py)"),
        ("Lane", "EKONOMIA + UI (prototypy OKOLICA)"),
        ("Wygenerowano", "gen-panel-b.py z gra/data/*.json"),
    ]
    for i, (a, b) in enumerate(rows, 1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=14)


def write_param_sheet(wb, title, items, json_file):
    ws = wb.create_sheet(title)
    for c, h in enumerate(HDR, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    widths = [6, 32, 48, 12, 14, 14, 36, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for idx, it in enumerate(items, 2):
        ws.cell(idx, 1, idx - 1)
        ws.cell(idx, 2, it["key"])
        ws.cell(idx, 3, it.get("opis", ""))
        ws.cell(idx, 4, it["value"])
        ws.cell(idx, 5, it.get("zakres", ""))
        ws.cell(idx, 6, it.get("jednostka", ""))
        ws.cell(idx, 7, it.get("wplyw", ""))
        ws.cell(idx, 8, json_file)
        for c in (3, 7):
            ws.cell(idx, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


def write_table_sheet(wb, title, columns, rows):
    """Arkusz tabelaryczny (Budynki, Surowce, Technologie) — nagłówek w wierszu 1."""
    ws = wb.create_sheet(title)
    for c, h in enumerate(columns, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for i, row in enumerate(rows, 2):
        for j, col in enumerate(columns, 1):
            val = row.get(col) if isinstance(row, dict) else None
            ws.cell(i, j, val)
    ws.freeze_panes = "A2"
    return ws


def flatten_triple(section_dict, json_file, wplyw_prefix=""):
    out = []
    for key, obj in section_dict.items():
        if not isinstance(obj, dict):
            continue
        normal = obj.get("normal", obj.get("wartosc"))
        zakres = ""
        if "easy" in obj and "hard" in obj:
            zakres = f"easy={obj['easy']} … hard={obj['hard']}"
        out.append({
            "key": key,
            "value": normal,
            "opis": obj.get("opis", ""),
            "jednostka": obj.get("jednostka", ""),
            "zakres": zakres,
            "wplyw": wplyw_prefix,
        })
    return out


def flatten_building(b):
    row = {}
    row["id"] = b.get("id")
    row["nazwa"] = b.get("nazwa")
    row["kosztBudowy"] = b.get("kosztBudowy")
    row["przyrostKosztu"] = b.get("przyrostKosztu")
    row["utrzymanie"] = b.get("utrzymanie")
    row["techUnlock"] = b.get("techUnlock")
    for k in BAZA_KEYS:
        row[f"baza.{k}"] = b.get("baza", {}).get(k)
        row[f"przyrost.{k}"] = b.get("przyrost", {}).get(k)
    return row


def flatten_buildings(data):
    return [flatten_building(b) for b in data]


def flatten_resources(data):
    return [{col: r.get(col) for col in RESOURCES_COLS} for r in data]


def flatten_technologie(tech_data):
    techs = tech_data.get("technologie", tech_data) if isinstance(tech_data, dict) else tech_data
    return [{col: t.get(col) for col in TECH_COLS} for t in techs]


def flatten_miasto(data):
    out = []
    for key, obj in data.items():
        if not isinstance(obj, dict):
            continue
        out.append({
            "key": key,
            "value": obj.get("wartosc"),
            "opis": obj.get("opis", ""),
            "jednostka": obj.get("jednostka", ""),
            "zakres": "",
            "wplyw": "miasto-params → production, cities, okolica",
        })
    return out


def flatten_power_skladniki(data):
    """Współczynniki P-A z power-params.json → arkusz Potega-P-A."""
    out = []
    for key, obj in data.get("skladniki", {}).items():
        if not isinstance(obj, dict):
            continue
        out.append({
            "key": key,
            "value": obj.get("pkt"),
            "opis": obj.get("opis", ""),
            "jednostka": "pkt",
            "zakres": ">= 0",
            "wplyw": "power-objective.ts → computeObjectivePower (globalny kanon)",
        })
    kal = data.get("kalibracja", {})
    if kal.get("powerOczekiwany") is not None:
        out.append({
            "key": "_kalibracja_power_oczekiwany",
            "value": kal["powerOczekiwany"],
            "opis": "INFO: scenariusz testowy Macieja (nie eksportowane do silnika)",
            "jednostka": "pkt",
            "zakres": "INFO",
            "wplyw": "power-objective-test.cjs",
        })
    return out


def flatten_power_opcje(data):
    """Opcje przyszłe (P-C1..C3, P-B) — power-params.json opcje."""
    out = []
    for key, obj in data.get("opcje", {}).items():
        if not isinstance(obj, dict):
            continue
        typ = obj.get("typ", "string")
        jed = "bool" if typ == "bool" else "tekst"
        out.append({
            "key": key,
            "value": obj.get("wartosc"),
            "opis": obj.get("opis", ""),
            "jednostka": jed,
            "zakres": "tak/nie lub tekst",
            "wplyw": "power-objective.ts / HUD (gdy wdrozone)",
        })
    return out


EPOKA_MANPOWER_COLS = ["epoka", "ludekNaLudka", "manpowerNaLudka", "manpowerNaJednostke"]


def flatten_epoka_manpower(data):
    return [
        {col: row.get(col) for col in EPOKA_MANPOWER_COLS}
        for row in data.get("epoki", [])
        if isinstance(row, dict)
    ]


def write_eksporty(wb):
    ws = wb.create_sheet("_Eksporty")
    ws.cell(1, 1, "Po edycji Panel-B.xlsx").font = Font(bold=True, size=12)
    cmds = [
        ("Panel B (ten plik)", "python panele-sterowania/export-b.py"),
        ("Kalkulator Power", "python tools/build-power-kalkulator-xlsx.py"),
        ("Testy Power", "node gra/tools/power-objective-test.cjs"),
    ]
    for i, (label, cmd) in enumerate(cmds, 3):
        ws.cell(i, 1, label)
        ws.cell(i, 2, cmd)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 55


def main():
    miasto = load_json("miasto-params.json")
    econ = load_json("econ-params.json")
    society = load_json("society-params.json")
    buildings = load_json("buildings.json")
    resources = load_json("resources.json")
    tech = load_json("tech.json")
    power = load_json("power-params.json")
    epoka_mp = load_json("epoka-ludnosc-manpower.json")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_info(wb.create_sheet("_INFO", 0))

    write_param_sheet(wb, "Miasto", flatten_miasto(miasto), "miasto-params.json")
    write_param_sheet(
        wb, "Ekonomia", flatten_triple(econ["ekonomia_miasta"], "econ-params.json", "turn-economy, HUD suwaki"),
        "econ-params.json",
    )
    write_param_sheet(
        wb, "Wealth", flatten_triple(econ["wealth"], "econ-params.json", "wealth.ts, cityPanel"),
        "econ-params.json",
    )
    write_param_sheet(
        wb, "Globalne", flatten_triple(econ["globalne"], "econ-params.json", "economy, magazyny"),
        "econ-params.json",
    )
    write_param_sheet(
        wb, "Budynki-eco", flatten_triple(econ["budynki"], "econ-params.json", "economy.ts mnozniki"),
        "econ-params.json",
    )
    write_param_sheet(
        wb, "Teren-bonus", flatten_triple(econ["teren_mapa"], "econ-params.json", "okolica, heks bonus"),
        "econ-params.json",
    )

    for sheet_name, section_key, wplyw in [
        ("Zdrowie", "zdrowie", "society.ts zdrowie, wzrost pop"),
        ("Szczescie", "szczescie", "society.ts szczęście, Porządek"),
        ("Kultura", "kultura", "culture-religion, granice"),
        ("Religia", "religia", "culture-religion, szerzenie"),
    ]:
        write_param_sheet(
            wb, sheet_name, flatten_triple(society[section_key], "society-params.json", wplyw),
            "society-params.json",
        )

    por = flatten_triple(society["porzadek"], "society-params.json", "Porządek T1/T2")
    por += flatten_triple(society["prawo"], "society-params.json", "Prawo, garnizon")
    write_param_sheet(wb, "Porzadek", por, "society-params.json")

    write_param_sheet(
        wb, "Potega-P-A", flatten_power_skladniki(power), "power-params.json → skladniki.*.pkt",
    )
    write_param_sheet(
        wb, "Potega-opcje", flatten_power_opcje(power), "power-params.json → opcje.*",
    )
    write_table_sheet(wb, "Manpower-epoki", EPOKA_MANPOWER_COLS, flatten_epoka_manpower(epoka_mp))

    write_table_sheet(wb, "Budynki", BUILDINGS_COLS, flatten_buildings(buildings))
    write_table_sheet(wb, "Surowce", RESOURCES_COLS, flatten_resources(resources))
    write_table_sheet(wb, "Technologie", TECH_COLS, flatten_technologie(tech))

    write_eksporty(wb)
    wb.save(OUT)
    print(f"Zapisano: {OUT}")
    print(f"Arkusze: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
