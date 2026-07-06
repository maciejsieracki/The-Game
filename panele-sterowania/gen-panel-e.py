#!/usr/bin/env python3
# gen-panel-e.py — Panel-E.xlsx (Grupa E: start, meta, kreator, generator E2, zwycięstwo)
# Uruchom z root: python panele-sterowania/gen-panel-e.py
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
OUT = os.path.join(os.path.dirname(__file__), "Panel-E.xlsx")

HDR = ["ID", "Parametr", "Opis", "Wartość", "Zakres/dozwolone", "Jednostka", "Wpływ na grę", "Plik źródłowy"]
HDR_FILL = PatternFill("solid", fgColor="2D6A4F")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
VAL_FONT = Font(color="0000AA")


def load_json(name):
    p = os.path.join(DATA, name)
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def write_info(ws):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    rows = [
        ("Panel sterowania — GRUPA E (Start / Meta / Kreator / Zwycięstwo)", ""),
        ("", ""),
        ("Jak używać (Maciej — bez terminala)", ""),
        ("1.", "Edytuj kolumnę Wartość (niebieska)."),
        ("2.", "Zapisz plik."),
        ("3.", "W czacie Grupy E napisz: eksportuj panel"),
        ("4.", "Agent odpali export-e.py → ui-params.json + e-start-params.json"),
        ("", ""),
        ("Nie koliduje z", "FOOD-HODOWLA (Panel-A/B) · barbarzyńcy liczby → Panel-D / ai-params.json"),
        ("UI-parametry", "Zastąpione — seed: gra/data/ui-params.json (arkusze Nowa-gra, Menu)"),
        ("Generator-E2 vs A", "Ten panel = domyślne suwaki kreatora UI; balans generatora mapy = Panel-A Generator-E2"),
        ("Lane", "Grupa E (UI/meta) · wpięcie TS → handoff Integrator"),
        ("Wygenerowano", f"gen-panel-e.py · {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
    ]
    for i, (a, b) in enumerate(rows, 1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=14)


def write_sheet(wb, title, rows):
    ws = wb.create_sheet(title[:31])
    for c, h in enumerate(HDR, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = [14, 36, 52, 14, 22, 12, 40, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for idx, r in enumerate(rows, 2):
        for c, val in enumerate(r, 1):
            cell = ws.cell(idx, c, val)
            if c == 4:
                cell.font = VAL_FONT
            if c in (3, 7):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


def row(sheet_id, param, opis, wartosc, zakres, jedn, wplyw, plik):
    return [f"E-{sheet_id}-{param}", param, opis, wartosc, zakres, jedn, wplyw, plik]


def nowa_gra_rows(ui):
    out = []
    for s in ui.get("nowa_gra", {}).get("ustawienia", []):
        key = s["key"]
        dom = s.get("domyslny", 0)
        opts = ", ".join(s.get("opts", []))
        out.append(row(
            "NG", key,
            f"{s.get('label', key)} — indeks domyślny w opts",
            dom,
            f"0…{max(0, len(s.get('opts', [])) - 1)} · opts: {opts[:80]}{'…' if len(opts) > 80 else ''}",
            "idx",
            "Domyślny wybór w kreatorze krok 4",
            "gra/data/ui-params.json",
        ))
    return out


def menu_rows(ui):
    out = []
    out.append(row(
        "MENU", "menu_wersja",
        "Tekst wersji w menu głównym",
        ui.get("menu", {}).get("wersja", ""),
        "dowolny tekst",
        "—",
        "Wyświetlane pod tytułem menu",
        "gra/data/ui-params.json",
    ))
    for s in ui.get("menu", {}).get("ustawienia", []):
        key = s["key"]
        out.append(row(
            "MENU", key,
            f"{s.get('label', key)} — domyślny indeks",
            s.get("domyslny", 0),
            f"0…{max(0, len(s.get('opts', [])) - 1)}",
            "idx",
            "Domyślne ustawienie audio/grafika/UI w menu",
            "gra/data/ui-params.json",
        ))
    return out


def skala_mapy_rows(ep):
    out = []
    for label, obj in ep.get("skala_mapy", {}).items():
        for field, opis, jedn, wplyw in [
            ("rywale_ai", "Domyślna liczba rywali AI", "szt.", "Gęstość starć na mapie"),
            ("miasta_panstwa", "Domyślna liczba miast w klastrze gracza", "szt.", "Suwak ±1 (Sparta, Kapua…)"),
            ("typy_cywilizacji", "Domyślna liczba obcych typów/klastrów", "szt.", "Ta sama skala co miasta-państwa"),
            ("hex_w", "Szerokość mapy w heksach", "hex", "Informacyjnie"),
            ("hex_h", "Wysokość mapy w heksach", "hex", "Informacyjnie"),
        ]:
            out.append(row(
                "MAP", f"{label}.{field}",
                f"{label}: {opis}",
                obj.get(field, ""),
                "liczba dodatnia",
                jedn,
                wplyw,
                "gra/data/e-start-params.json",
            ))
    return out


def generator_rows(ep):
    g = ep.get("generator_e2", {})
    specs = [
        ("resource_mult_low", "Mnożnik surowców — Mało", 0.6, "0.1…2", "×", "Mniej złóż na mapie"),
        ("resource_mult_normal", "Mnożnik surowców — Normalnie", 1.0, "0.1…2", "×", "Baza E2-Q2"),
        ("resource_mult_high", "Mnożnik surowców — Dużo", 1.4, "0.1…2", "×", "Więcej złóż"),
        ("resource_baseline_rarity", "Boost bazy rarity (Normalnie bogatsze)", 1.35, "1…2", "×", "Gdy kreator przekazał preset E2"),
        ("river_base_low", "Baza max rzek — Mało", 2, "0…20", "szt.", "E2-Q3"),
        ("river_base_normal", "Baza max rzek — Normalnie", 5, "0…20", "szt.", "Skaluje z rozmiarem mapy"),
        ("river_base_high", "Baza max rzek — Dużo", 8, "0…20", "szt.", ""),
        ("river_scale_mala", "Skala rzek — bucket mała", 1.0, "0.5…3", "×", ""),
        ("river_scale_srednia", "Skala rzek — bucket średnia", 1.35, "0.5…3", "×", ""),
        ("river_scale_duza", "Skala rzek — bucket duża", 1.7, "0.5…3", "×", ""),
        ("river_scale_ogromna", "Skala rzek — bucket ogromna", 2.1, "0.5…3", "×", ""),
        ("desert_threshold_low", "Próg szumu pustyni — Mało", 0.68, "0.4…0.9", "0…1", "Wyżej = mniej pustyni"),
        ("desert_threshold_normal", "Próg pustyni — Normalnie", 0.63, "0.4…0.9", "0…1", ""),
        ("desert_threshold_high", "Próg pustyni — Dużo", 0.58, "0.4…0.9", "0…1", "Niżej = więcej pustyni"),
        ("forest_threshold_low", "Próg lasu — Mało", 0.65, "0.3…0.8", "0…1", "Wyżej = mniej lasu"),
        ("forest_threshold_normal", "Próg lasu — Normalnie", 0.58, "0.3…0.8", "0…1", ""),
        ("forest_threshold_high", "Próg lasu — Dużo", 0.5, "0.3…0.8", "0…1", "Niżej = więcej lasu"),
    ]
    out = []
    for key, opis, default, zakres, jedn, wplyw in specs:
        out.append(row("GEN", key, opis, g.get(key, default), zakres, jedn, wplyw, "gra/data/e-start-params.json"))
    return out


def tempo_rows(ep):
    t = ep.get("tempo_gry", {})
    out = []
    for key, opis, wplyw in [
        ("szybka", "Mnożnik kosztu badań — Szybka", "Tańsze tech"),
        ("standardowa", "Mnożnik — Standardowa", "Baza tech.json ×1"),
        ("dluga", "Mnożnik — Długa", "5× koszt badań"),
    ]:
        out.append(row("TEMPO", key, opis, t.get(key, 1.0), "0.05…10", "×", wplyw, "e-start-params.json → tech-tempo.ts"))
    return out


def zwyciestwo_rows(ep):
    z = ep.get("zwyciestwo", {})
    return [
        row("VIC", "ostatnia_epoka_v1", "Ostatnia epoka v0.1 (Żelazo=3)", z.get("ostatnia_epoka_v1", 3), "1…6", "era id", "Gate dominacji", "victory.ts"),
        row("VIC", "prog_dominacji_power", "Próg Power (10=A*)", z.get("prog_dominacji_power", 0.5), "0.3…0.8", "udział", ">50% = dominacja", "victory.ts"),
        row("VIC", "dominacja_wymaga_ostatniej_epoki", "Dominacja tylko ostatnia epoka", z.get("dominacja_wymaga_ostatniej_epoki", True), "tak/nie", "bool", "", "victory.ts"),
        row("VIC", "nauka_wymaga_rakiety", "Nauka = tech + rakieta", z.get("nauka_wymaga_rakiety", True), "tak/nie", "bool", "10=A*", "victory.ts"),
    ]


def defaulty_rows(ep):
    d = ep.get("defaulty", {})
    return [
        row("DEF", "player_civ_id", "Domyślna cywilizacja (ikonaId)", d.get("player_civ_id", "rzymianie"), "z civs.json", "id", "Preselect kreator", "newGameMapDefaults.ts"),
        row("DEF", "start_epoch_id", "Domyślna epoka startu", d.get("start_epoch_id", "kamien"), "kamien, brz", "id", "Kamień default", "newGameMapDefaults.ts"),
        row("DEF", "map_quality_default", "Jakość mapy (bundle GPU)", d.get("map_quality_default", "Średnia"), "Niska, Średnia, Wysoka", "etykieta", "E1-Q-BUNDLE", "kreator zaaw."),
    ]


def zaawansowane_rows(ep):
    k = ep.get("kreator_zaawansowane", {})
    return [
        row("ADV", "manual_seed_default", "Seed ręczny", k.get("manual_seed_default", 424242), "1…999999999", "int", "Powtarzalna mapa", "newGameFlow.ts"),
        row("ADV", "barbarians_enabled_default", "Barbarzyńcy wł.", k.get("barbarians_enabled_default", True), "tak/nie", "bool", "Zaawansowane", "newGameFlow.ts"),
        row("ADV", "battle_always_manual_default", "Bitwy zawsze ręczne", k.get("battle_always_manual_default", False), "tak/nie", "bool", "", "newGameFlow.ts"),
        row("ADV", "fog_debug_reveal_all_default", "Mgła debug", k.get("fog_debug_reveal_all_default", False), "tak/nie", "bool", "Tylko testy", "newGameFlow.ts"),
        row("ADV", "victory_power_and_dominance_default", "Warunki zwycięstwa w grze", k.get("victory_power_and_dominance_default", True), "tak/nie", "bool", "", "newGameFlow.ts"),
    ]


def decyzje_rows(ep):
    d = ep.get("decyzje_kanon", {})
    labels = {
        "e1_reset_nowa_gra": "E1-Q1=A: pełny reset przy Nowa gra",
        "e1_tech_kaskada_epok": "E1-Q2=B: tech wcześniejszych epok zbadane",
        "e1_ziemia_preset_staly": "E1-Q3=A: Ziemia = stały preset",
        "e1_zloza_tylko_gory": "E1-Q8=B: metale tylko Góry",
        "e1_zloza_ukryte_do_epoki": "E1-Q9=B: złoża ukryte do epoki",
        "e2_barbarzyncy_do_przed_sredniowiecza": "E2-Q11=C: barbarzyńcy do przed-Średniowiecza",
        "e2_buntownicy_od_sredniowiecza": "E2-Q11=C: buntownicy od Średniowiecza",
    }
    return [row("KANON", key, opis, d.get(key, True), "tak/nie", "bool", "Reguła produktowa", "docs/grupa-e/") for key, opis in labels.items()]


def write_eksporty(wb):
    ws = wb.create_sheet("_Eksporty")
    ws.cell(1, 1, "Po edycji Panel-E.xlsx").font = Font(bold=True, size=12)
    for i, (label, cmd) in enumerate([
        ("Panel E", "python panele-sterowania/export-e.py"),
        ("Regeneruj z JSON", "python panele-sterowania/gen-panel-e.py"),
    ], 3):
        ws.cell(i, 1, label)
        ws.cell(i, 2, cmd)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55


def main():
    ui = load_json("ui-params.json")
    ep = load_json("e-start-params.json") if os.path.isfile(os.path.join(DATA, "e-start-params.json")) else {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_info(wb.create_sheet("_INFO", 0))
    write_sheet(wb, "Defaulty-startu", defaulty_rows(ep))
    write_sheet(wb, "Nowa-gra", nowa_gra_rows(ui))
    write_sheet(wb, "Menu", menu_rows(ui))
    write_sheet(wb, "Skala-mapy", skala_mapy_rows(ep))
    write_sheet(wb, "Generator-E2", generator_rows(ep))
    write_sheet(wb, "Tempo-gry", tempo_rows(ep))
    write_sheet(wb, "Zwyciestwo", zwyciestwo_rows(ep))
    write_sheet(wb, "Kreator-zaaw", zaawansowane_rows(ep))
    write_sheet(wb, "Decyzje-kanon", decyzje_rows(ep))
    write_eksporty(wb)

    wb.save(OUT)
    print(f"OK: {OUT}")


if __name__ == "__main__":
    main()
