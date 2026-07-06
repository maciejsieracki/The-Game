#!/usr/bin/env python3
# export-a.py — Panel-A.xlsx → gra/data (Grupa A: terrain-improvements + map-gen-params)
# Uruchom: python panele-sterowania/export-a.py
# Maciej nie używa terminala — agent odpala po "eksportuj panel".
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "Panel-A.xlsx")

COL_PARAM = 2
COL_VALUE = 4

TERRAIN_SHEETS = ("Ulepszenia-FOOD", "Ulepszenia-inne", "Zasięgi-mgła")
MAP_GEN_SHEETS = (
    "Generator-E2", "Mapa-skala", "Zasięgi-mgła",
    "Generator-rozmiary", "Zloza-generator",
)

SKIP_PARAMS = {"nota.miasto_zasieg_terytorium"}


def _num_equal(a, b):
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return False


def coerce(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return old, False
    if isinstance(old, bool):
        if isinstance(val, bool):
            return val, val != old
        nv = str(val).strip().lower() in ("tak", "true", "1", "prawda")
        return nv, nv != old
    if isinstance(old, int):
        try:
            nv = int(round(float(val)))
            return nv, not _num_equal(nv, old)
        except (TypeError, ValueError):
            return old, False
    if isinstance(old, float):
        try:
            nv = float(val)
            return nv, not _num_equal(nv, old)
        except (TypeError, ValueError):
            return old, False
    if old is None:
        if isinstance(val, (int, float)):
            return val, True
        try:
            f = float(str(val).replace(",", "."))
            return (int(f) if f.is_integer() else f), True
        except ValueError:
            return val, True
    nv = str(val)
    return nv, nv != old


def read_sheet_params(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, COL_PARAM).value
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        if key in SKIP_PARAMS:
            continue
        out[key] = ws.cell(r, COL_VALUE).value
    return out


def overlay_terrain_param(root, param_key, val):
    parts = param_key.split(".")
    if len(parts) < 2:
        return 0
    imp_key = parts[0]
    if imp_key not in root or not isinstance(root[imp_key], dict):
        return 0
    node = root[imp_key]
    if parts[1] == "koszt_praca":
        nv, did = coerce(node.get("koszt_praca"), val)
        if did:
            node["koszt_praca"] = nv
        return 1 if did else 0
    if parts[1] == "bonus" and len(parts) == 3:
        bonus = node.setdefault("bonus", {})
        bk = parts[2]
        old = bonus.get(bk, 0)
        nv, did = coerce(old if isinstance(old, (int, float)) else 0, val)
        if did:
            bonus[bk] = nv
        return 1 if did else 0
    if parts[1] == "wycinka" and len(parts) == 3:
        wyc = node.setdefault("wycinka", {})
        wk = parts[2]
        old = wyc.get(wk, 0)
        nv, did = coerce(old if isinstance(old, (int, float)) else 0, val)
        if did:
            wyc[wk] = nv
        return 1 if did else 0
    if parts[1] in node:
        nv, did = coerce(node[parts[1]], val)
        if did:
            node[parts[1]] = nv
        return 1 if did else 0
    return 0


def overlay_terrain_from_sheets(wb, terrain):
    params = {}
    for sheet in TERRAIN_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        params.update(read_sheet_params(wb[sheet]))
    changed = 0
    skipped = []
    for param_key, val in params.items():
        if param_key.startswith("mgla.") or param_key.startswith("gestosc.") or param_key.startswith("mapa_skala."):
            continue
        if param_key.startswith("plony.") or param_key.startswith("ruch.") or param_key.startswith("zloza.") or param_key.startswith("generator."):
            continue
        parts = param_key.split(".")
        if len(parts) < 2:
            continue
        imp_key = parts[0]
        if imp_key not in terrain or not isinstance(terrain[imp_key], dict):
            skipped.append(param_key)
            continue
        changed += overlay_terrain_param(terrain, param_key, val)
    if skipped:
        print(f"  (pominięto {len(skipped)} wierszy — brak klucza w terrain-improvements.json)")
    return changed


def set_nested(obj, path_parts, val):
    cur = obj
    for p in path_parts[:-1]:
        if p not in cur or not isinstance(cur[p], dict):
            cur[p] = {}
        cur = cur[p]
    key = path_parts[-1]
    old = cur.get(key)
    nv, did = coerce(old, val)
    if did:
        cur[key] = nv
    return 1 if did else 0


def default_map_gen():
    return {
        "_meta": {
            "opis": "Panel-A export — generator E2 + mgła. Kod czyta po P3 / handoff Integratora.",
            "panel": "panele-sterowania/Panel-A.xlsx",
            "export": "panele-sterowania/export-a.py",
        },
        "mgla": {
            "default_sight_jednostki": {"wartosc": 3, "opis": "Domyślny promień wzroku jednostki"},
        },
        "gestosc": {
            "surowce_mult": {"low": 0.6, "medium": 1.0, "high": 1.4},
            "baseline_rarity_mult": 1.35,
            "rzeki_max_mala_mapa": {"low": 2, "medium": 5, "high": 8},
            "river_scale": {"mala": 1.0, "srednia": 1.35, "duza": 1.7, "ogromna": 2.1},
            "desert_noise_threshold": {"low": 0.68, "medium": 0.63, "high": 0.58},
            "forest_noise_threshold": {"low": 0.65, "medium": 0.58, "high": 0.50},
        },
        "mapa_skala": {
            "aktywne_typy": {"mala": 3, "srednia": 5, "duza": 7, "ogromna": 9},
            "domyslni_rywale": {"mala": 2, "srednia": 4, "duza": 6, "ogromna": 8},
        },
        "generator": {
            "default_width": 36,
            "default_height": 28,
            "rozmiar_dims": {
                "malenki": [38, 26], "maly": [54, 37], "standardowy": [84, 60],
                "duzy": [120, 84], "ogromny": [168, 119],
            },
        },
        "deposit_rules": {
            z: {"rarity": r} for z, r in {
                "miedz": 0.10, "zelazo": 0.08, "glina": 0.10, "konie": 0.10, "wegiel": 0.10,
                "owce": 0.08, "bydlo": 0.07, "lama": 0.06, "luksus": 0.06, "sol": 0.12,
            }.items()
        },
        "metal_deposit_min_era": {"miedz": 2, "zelazo": 3},
    }


def overlay_map_gen(wb, map_gen):
    params = {}
    for sheet in MAP_GEN_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        params.update(read_sheet_params(wb[sheet]))
    changed = 0
    for param_key, val in params.items():
        if param_key.startswith("nota."):
            continue
        parts = param_key.split(".")
        if param_key == "mgla.default_sight_jednostki":
            mg = map_gen.setdefault("mgla", {})
            ds = mg.setdefault("default_sight_jednostki", {"wartosc": 3, "opis": ""})
            nv, did = coerce(ds.get("wartosc", 3), val)
            if did:
                ds["wartosc"] = nv
                changed += 1
            continue
        if param_key == "gestosc.baseline_rarity_mult":
            nv, did = coerce(map_gen.setdefault("gestosc", {}).get("baseline_rarity_mult", 1.35), val)
            if did:
                map_gen["gestosc"]["baseline_rarity_mult"] = nv
                changed += 1
            continue
        if param_key == "generator.default_width":
            g = map_gen.setdefault("generator", {})
            nv, did = coerce(g.get("default_width", 36), val)
            if did:
                g["default_width"] = nv
                changed += 1
            continue
        if param_key == "generator.default_height":
            g = map_gen.setdefault("generator", {})
            nv, did = coerce(g.get("default_height", 28), val)
            if did:
                g["default_height"] = nv
                changed += 1
            continue
        if len(parts) == 4 and parts[0] == "generator" and parts[1] == "rozmiar_dims" and parts[3] in ("w", "h"):
            g = map_gen.setdefault("generator", {})
            dims = g.setdefault("rozmiar_dims", {})
            size_key = parts[2]
            arr = dims.get(size_key, [84, 60])
            if not isinstance(arr, list):
                arr = [84, 60]
            idx = 0 if parts[3] == "w" else 1
            nv, did = coerce(arr[idx] if idx < len(arr) else 0, val)
            if did:
                while len(arr) <= idx:
                    arr.append(0)
                arr[idx] = nv
                dims[size_key] = arr
                changed += 1
            continue
        if len(parts) == 3 and parts[0] == "zloza" and parts[2] == "rarity":
            deps = map_gen.setdefault("deposit_rules", {})
            node = deps.setdefault(parts[1], {})
            old = node.get("rarity", 0.1) if isinstance(node, dict) else 0.1
            nv, did = coerce(old, val)
            if did:
                deps[parts[1]] = {"rarity": nv}
                changed += 1
            continue
        if len(parts) == 3 and parts[0] == "zloza" and parts[2] == "min_epoka":
            metal = map_gen.setdefault("metal_deposit_min_era", {})
            nv, did = coerce(metal.get(parts[1], 1), val)
            if did:
                metal[parts[1]] = nv
                changed += 1
            continue
        if len(parts) >= 3:
            changed += set_nested(map_gen, parts, val)
    return changed


PLONY_TYPE_SLUG = {
    "laka": "Łąka", "rownina": "Równina", "wzgorza": "Wzgórza", "gory": "Góry",
    "wybrzeze": "Wybrzeże", "morze": "Morze", "pustynia": "Pustynia",
}
PLONY_MOD_SLUG = {"rzeka": "Rzeka", "las": "Las (nakładka)"}
PLONY_FIELD = {
    "zywnosc": "Żywność", "praca": "Praca", "handel": "Handel",
    "drewno": "Drewno", "kamien": "Kamień",
}
RUCH_COST_KEY = {
    "Laka": "Laka", "Rownina": "Rownina", "Pustynia": "Pustynia",
    "Wybrzeze": "Wybrzeze", "Wzgorza": "Wzgorza", "Gory": "Gory", "Morze": "Morze",
}


def overlay_terrain_yields(data, params):
    changed = 0
    for param_key, val in params.items():
        if not param_key.startswith("plony."):
            continue
        parts = param_key.split(".")
        if len(parts) != 4:
            continue
        _, kind, slug, fk = parts
        col = PLONY_FIELD.get(fk)
        if not col:
            continue
        if kind == "types":
            teren = PLONY_TYPE_SLUG.get(slug)
            if not teren:
                continue
            for row in data.get("terrain_types", []):
                if row.get("Teren") != teren:
                    continue
                nv, did = coerce(row.get(col), val)
                if did:
                    row[col] = nv
                    changed += 1
                break
        elif kind == "mod":
            mod = PLONY_MOD_SLUG.get(slug)
            if not mod:
                continue
            for row in data.get("terrain_modifiers", []):
                if row.get("Modyfikator") != mod:
                    continue
                nv, did = coerce(row.get(col), val)
                if did:
                    row[col] = nv
                    changed += 1
                break
    return changed


def overlay_terrain_movement(data, params):
    changed = 0
    for param_key, val in params.items():
        if param_key == "ruch.forestExtra":
            nv, did = coerce(data.get("forestExtra", 1), val)
            if did:
                data["forestExtra"] = nv
                changed += 1
            continue
        if not param_key.startswith("ruch.costs."):
            continue
        ck = param_key.split(".", 2)[2]
        costs = data.setdefault("costs", {})
        if ck not in costs and ck not in RUCH_COST_KEY:
            continue
        old = costs.get(ck, 1)
        nv, did = coerce(old, val)
        if did:
            costs[ck] = nv
            changed += 1
    return changed


def save_json(obj, path):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
        f.write("\n")
    json.load(open(tmp, encoding="utf-8"))
    os.replace(tmp, path)


def main():
    ap = argparse.ArgumentParser(description="Eksport Panel-A.xlsx → JSON Grupy A")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--data-dir", default=DATA)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"Brak pliku: {args.xlsx}")
        print("Uruchom najpierw: python panele-sterowania/gen-panel-a.py")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    total = 0
    all_params = {}
    for sn in wb.sheetnames:
        if sn.startswith("_"):
            continue
        all_params.update(read_sheet_params(wb[sn]))

    terrain_path = os.path.join(args.data_dir, "terrain-improvements.json")
    terrain = json.load(open(terrain_path, encoding="utf-8"))
    ch = overlay_terrain_from_sheets(wb, terrain)
    print(f"terrain-improvements.json: {ch} zmian")
    if ch and not args.dry_run:
        save_json(terrain, terrain_path)
    total += ch

    yields_path = os.path.join(args.data_dir, "terrain-yields.json")
    if os.path.isfile(yields_path):
        yields_data = json.load(open(yields_path, encoding="utf-8"))
        ch_y = overlay_terrain_yields(yields_data, all_params)
        print(f"terrain-yields.json: {ch_y} zmian")
        if ch_y and not args.dry_run:
            save_json(yields_data, yields_path)
        total += ch_y

    move_path = os.path.join(args.data_dir, "terrain-movement.json")
    if os.path.isfile(move_path):
        movement = json.load(open(move_path, encoding="utf-8"))
        ch_m = overlay_terrain_movement(movement, all_params)
        print(f"terrain-movement.json: {ch_m} zmian")
        if ch_m and not args.dry_run:
            save_json(movement, move_path)
        total += ch_m

    map_gen_path = os.path.join(args.data_dir, "map-gen-params.json")
    if os.path.isfile(map_gen_path):
        map_gen = json.load(open(map_gen_path, encoding="utf-8"))
    else:
        map_gen = default_map_gen()
    ch2 = overlay_map_gen(wb, map_gen)
    print(f"map-gen-params.json: {ch2} zmian")
    if ch2 and not args.dry_run:
        save_json(map_gen, map_gen_path)
    elif not os.path.isfile(map_gen_path) and not args.dry_run:
        save_json(map_gen, map_gen_path)
        print("map-gen-params.json: utworzono domyślny plik")
    total += ch2

    print(f"RAZEM: {total} zmian")
    if args.dry_run:
        print("(dry-run — JSON nie zapisany)")
    if total == 0:
        print("(brak zmian — Wartość pusta lub identyczna z JSON)")


if __name__ == "__main__":
    main()
