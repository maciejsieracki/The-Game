#!/usr/bin/env python3
"""
Eksport 26 jednostek PL0 — tabela zaakceptowana przez Macieja (2026-07-06).
Wyjście: panele-sterowania/Jednostki-PL0-MACIEJ-20260706.xlsx
"""
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    sys.exit("pip install openpyxl")

OUT = os.path.join(os.path.dirname(__file__), "Jednostki-PL0-MACIEJ-20260706.xlsx")

ROWS = [
    # nr, jednostka, typ, hp, atk, obraz, obr, dist, uwaga, stara_nazwa
    (42, "Soldurii", "Miecznik", 44, 7, 5, 2, 0, "", ""),
    (51, "Konnica lancowa asyryjska", "Konnica", 34, 9, 8, 5, 0, "", ""),
    (52, "Konnica łucznicza asyryjska", "Konnica łucznicza", 30, 4, 3, 2, 6, "", ""),
    (53, "Łucznik asyryjski", "Łucznik", 16, 2, 2, 3, 4, "", ""),
    (54, "Drużynnik", "Włócznik", 31, 7, 6, 4, 7, "3× krótki oszczep przed wręcz", ""),
    (55, "Jeździec z oszczepami", "Konnica", 28, 6, 5, 4, 7, "", "Jeździec z szczepnikami"),
    (56, "Strażnik bram Harappy", "Włócznik", 40, 4, 4, 9, 0, "", ""),
    (57, "Piechota induska", "Włócznik", 24, 6, 5, 5, 0, "", ""),
    (58, "Garnizon Harappy", "Włócznik", 26, 5, 4, 8, 0, "", ""),
    (59, "Rydwan Kapadokijski", "Rydwan", 36, 7, 7, 3, 0, "", ""),
    (60, "Piechota hetycka", "Włócznik", 32, 5, 8, 8, 0, "", ""),
    (61, "Gwardia hetycka", "Włócznik", 40, 7, 6, 7, 0, "", ""),
    (62, "Gwardia Ishtar", "Włócznik", 35, 8, 7, 7, 0, "", ""),
    (63, "Wojownik babiloński", "Miecznik", 24, 6, 5, 5, 0, "", ""),
    (64, "Piechota neobabilońska", "Włócznik", 26, 6, 5, 7, 0, "", ""),
    (65, "Tyrski miecznik", "Miecznik", 24, 7, 6, 3, 0, "", ""),
    (66, "Wojownik fenicki", "Włócznik", 20, 5, 4, 4, 0, "", ""),
    (67, "Gwardia Tyreńska", "Włócznik", 24, 7, 6, 6, 0, "", "Gwardia Tyr"),
    (68, "Thorakites", "Włócznik", 42, 5, 4, 8, 6, "2× oszczep przed wręcz (jak legion)", ""),
    (69, "Legion Rzymski", "Miecznik", 48, 7, 6, 7, 6, "2× pilum przed wręcz", "Legionarius"),
    (70, "Evocati", "Miecznik", 55, 8, 8, 8, 6, "2× pilum przed wręcz", ""),
    (71, "iButho z iklwa", "Włócznik", 29, 5, 4, 7, 0, "", ""),
    (72, "Gwardzista z champi", "Miecznik", 48, 7, 6, 5, 5, "2× oszczep przed wręcz", ""),
    (73, "Wojownik z żelaznym khopesh", "Miecznik", 23, 6, 6, 6, 0, "", ""),
    (74, "Mur tarcz (Sargonid)", "Włócznik", 34, 4, 3, 10, 0, "", ""),
    (75, "Miecznik galijski", "Miecznik", 38, 8, 6, 2, 0, "", ""),
]

HEADERS = [
    "#", "Jednostka", "Typ", "HP", "Atak", "Obraż", "Obrona", "Dystans",
    "Uwaga", "Stara nazwa (JSON)", "Akceptacja",
]

NOTE = (
    "Skala combat.ts: health/meleeAttack/weaponDamage/meleeDefence/missileAttack. "
    "Wartości finalne — bez dzielenia/mnożenia. "
    "Oszczep/pilum = kolumna Dystans; typ = Miecznik lub Włócznik (nie Oszczepnik)."
)


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PL0-staty"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "Jednostki PL0 — staty finalne (Maciej 2026-07-06)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:K1")
    ws["A2"] = NOTE
    ws["A2"].alignment = wrap
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 36

    start = 4
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(ROWS, start + 1):
        nr, name, typ, hp, atk, ob, obr, dist, uwaga, old = row
        vals = [nr, name, typ, hp, atk, ob, obr, dist, uwaga, old, ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = border
            cell.alignment = wrap
            if c >= 4 and c <= 8 and isinstance(v, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="top")

    widths = [5, 32, 18, 6, 6, 6, 8, 8, 28, 24, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = f"A{start + 1}"
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
