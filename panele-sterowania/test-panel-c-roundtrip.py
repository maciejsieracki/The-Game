#!/usr/bin/env python3
# test-panel-c-roundtrip.py — JSON↔Excel sync Panel-C (Koszty-jednostek F250 ×5 + Stale-walki)
import json
import os
import shutil
import subprocess
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    print("SKIP: brak openpyxl")
    sys.exit(0)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
PANEL = os.path.join(os.path.dirname(__file__), "Panel-C.xlsx")
EXPORT = os.path.join(os.path.dirname(__file__), "export-c.py")
DATA = os.path.join(ROOT, "gra", "data")

RESOURCE_FIELDS = (
    "Surowiec",
    "Surowiec (ilość)",
    "Utrzymanie surowiec",
    "Utrzymanie surowiec (ilość)",
)


def fail(msg):
    print(f"FAIL: {msg}")
    sys.exit(1)


def norm_val(val):
    if val is None:
        return None
    if isinstance(val, float) and val.is_integer():
        return int(val)
    if isinstance(val, str):
        s = val.strip()
        return s if s else None
    return val


def load_units(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def read_koszty_from_excel(wb):
    if "Koszty-jednostek" not in wb.sheetnames:
        fail("brak arkusza Koszty-jednostek")
    ws = wb["Koszty-jednostek"]
    out = {}
    for r in range(2, ws.max_row + 1):
        param = ws.cell(r, 2).value
        if not param or " · " not in str(param):
            continue
        uname, field = str(param).split(" · ", 1)
        field = field.strip()
        if field not in RESOURCE_FIELDS:
            continue
        out.setdefault(uname.strip(), {})[field] = norm_val(ws.cell(r, 4).value)
    return out


def unit_has_resource_fields(u):
    return any(f in u for f in RESOURCE_FIELDS)


def verify_json_excel_parity(units, excel_koszty):
    checked = 0
    for u in units:
        if not unit_has_resource_fields(u):
            continue
        name = u["Jednostka"]
        ex = excel_koszty.get(name)
        if not ex:
            fail(f"brak wierszy Koszty-jednostek dla {name!r}")
        for field in RESOURCE_FIELDS:
            if field not in u:
                continue
            expected = norm_val(u[field])
            actual = ex.get(field)
            if actual != expected:
                fail(f"{name}.{field}: JSON={expected!r}, Excel={actual!r}")
            checked += 1
    if checked == 0:
        fail("brak pól surowcowych w units.json")
    return checked


def verify_f250_x5(units):
    """Rekrutacja = utrzymanie ×5 (FALA 250 / R-DYST-DREWNO)."""
    n = 0
    for u in units:
        if "Surowiec (ilość)" not in u or "Utrzymanie surowiec (ilość)" not in u:
            continue
        rec = u["Surowiec (ilość)"]
        upk = u["Utrzymanie surowiec (ilość)"]
        if rec != upk * 5:
            fail(
                f"{u['Jednostka']}: F250 ×5 — rekrutacja {rec} != utrzymanie {upk}×5"
            )
        n += 1
    if n == 0:
        fail("brak jednostek z Surowiec (ilość) do weryfikacji F250")
    return n


def find_koszty_row(ws, unit_name, field):
    needle = f"{unit_name} · {field}"
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value == needle:
            return r
    return None


def test_koszty_json_to_excel(xlsx, units):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    excel_koszty = read_koszty_from_excel(wb)
    n_fields = verify_json_excel_parity(units, excel_koszty)
    n_units = verify_f250_x5(units)
    print(f"OK: JSON→Excel Koszty-jednostek ({n_fields} pól, {n_units} jednostek F250 ×5)")


def test_koszty_export_roundtrip(data_dir, xlsx):
    units_path = os.path.join(data_dir, "units.json")
    units = load_units(units_path)
    probe_unit = "Konnica"
    u = next((x for x in units if x.get("Jednostka") == probe_unit), None)
    if not u:
        fail(f"brak jednostki {probe_unit!r} w units.json (tmp)")

    before_rec = u["Surowiec (ilość)"]
    before_upk = u["Utrzymanie surowiec (ilość)"]
    probe_rec = before_rec + 5
    probe_upk = before_upk + 1
    probe_surowiec = "Żelazo" if u.get("Surowiec") != "Żelazo" else "Brąz"

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Koszty-jednostek"]
    for field, probe in (
        ("Surowiec", probe_surowiec),
        ("Surowiec (ilość)", probe_rec),
        ("Utrzymanie surowiec", probe_surowiec),
        ("Utrzymanie surowiec (ilość)", probe_upk),
    ):
        row = find_koszty_row(ws, probe_unit, field)
        if not row:
            fail(f"nie znaleziono wiersza {probe_unit} · {field}")
        ws.cell(row, 4, probe)
    wb.save(xlsx)

    cmd = [sys.executable, EXPORT, "--xlsx", xlsx, "--data-dir", data_dir]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        fail(r.stderr or r.stdout)

    after = next(x for x in load_units(units_path) if x.get("Jednostka") == probe_unit)
    checks = (
        ("Surowiec", probe_surowiec),
        ("Surowiec (ilość)", probe_rec),
        ("Utrzymanie surowiec", probe_surowiec),
        ("Utrzymanie surowiec (ilość)", probe_upk),
    )
    for field, expected in checks:
        if after.get(field) != expected:
            fail(f"export round-trip {probe_unit}.{field}: oczekiwano {expected!r}, jest {after.get(field)!r}")

    print(f"OK: Excel→export→JSON Koszty-jednostek ({probe_unit})")


def test_stale_walki_roundtrip(data_dir, xlsx):
    params_path = os.path.join(data_dir, "combat-params.json")
    before = json.load(open(params_path, encoding="utf-8"))["tw_v3"]["hit_base"]
    probe = before + 1

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Stale-walki"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "C-TW-HIT-BASE":
            ws.cell(r, 4, probe)
            break
    else:
        fail("nie znaleziono C-TW-HIT-BASE")
    wb.save(xlsx)

    cmd = [sys.executable, EXPORT, "--xlsx", xlsx, "--data-dir", data_dir]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        fail(r.stderr or r.stdout)

    after = json.load(open(params_path, encoding="utf-8"))["tw_v3"]["hit_base"]
    if after != probe:
        fail(f"Stale-walki round-trip: oczekiwano {probe}, jest {after}")

    print("OK: Panel-C round-trip (Stale-walki)")


def main():
    if not os.path.isfile(PANEL):
        fail("brak Panel-C.xlsx — uruchom gen-panel-c.py")

    units_live = load_units(os.path.join(DATA, "units.json"))
    test_koszty_json_to_excel(PANEL, units_live)

    with tempfile.TemporaryDirectory() as tmp:
        data_dir = os.path.join(tmp, "data")
        shutil.copytree(DATA, data_dir)
        xlsx = os.path.join(tmp, "Panel-C.xlsx")
        shutil.copy(PANEL, xlsx)

        test_koszty_export_roundtrip(data_dir, xlsx)

        shutil.copytree(DATA, data_dir, dirs_exist_ok=True)
        shutil.copy(PANEL, xlsx)
        test_stale_walki_roundtrip(data_dir, xlsx)

    live_after = load_units(os.path.join(DATA, "units.json"))
    if live_after != units_live:
        fail("żywy gra/data/units.json został zmodyfikowany — test musi używać tylko tmp")

    print("OK: panel-c round-trip (koszty F250 + stale-walki)")


if __name__ == "__main__":
    main()
