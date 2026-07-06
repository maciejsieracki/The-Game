#!/usr/bin/env python3
"""
gen-cyw-macierz.py — arkusz Cyw-macierz w Panel-D.xlsx (1 wiersz = 1 cywilizacja, ~120 kolumn liczb).

  python panele-sterowania/gen-cyw-macierz.py
  python panele-sterowania/gen-cyw-macierz.py --json-only

Eksport modulow: python panele-sterowania/export-cyw-macierz.py -> gra/data/civ-matrix.json
"""
from __future__ import annotations

import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, ".."))
sys.path.insert(0, HERE)
from cyw_macierz_schema import (
    COLUMN_DEFS, DEFAULTS, PARAM_META, PARAM_IDS, IDENT_COLS,
    DOMAIN_SHEETS, PARAMS_BY_DOMAIN, SHEET_INFO, SHEET_LEGACY, ALL_CYw_SHEETS,
    RESERVED_DOMAINS,
)

PANEL = os.path.join(HERE, "Panel-D.xlsx")
STANDALONE = os.path.join(HERE, "Cyw-macierz-REVIEW.xlsx")
CIVS = os.path.join(ROOT, "gra", "data", "civs.json")
DRAFT = os.path.join(ROOT, "Civ-CYWILIZACJE", "draft", "roster-6-REZERWA.json")
UNITS = os.path.join(ROOT, "gra", "data", "units.json")
CIV_AI = os.path.join(ROOT, "gra", "data", "civ-ai.json")
CIV_PARAMS = os.path.join(ROOT, "gra", "data", "civ-params.json")
DIP = os.path.join(ROOT, "gra", "data", "diplomacy.json")
OUT_JSON = os.path.join(ROOT, "gra", "data", "civ-matrix.json")

SHEET = SHEET_LEGACY  # kompat wsteczna

ARCHETYPE_AGG = {
    "grecy": 0.40, "rzymianie": 0.75, "chinczycy": 0.20, "inkowie": 0.45,
    "zulusi": 0.90, "egipt": 0.35, "babilon": 0.30, "sumer": 0.30,
    "celtowie": 0.60, "germanie": 0.65,
    "harappa": 0.25, "hetyci": 0.45, "slowianie": 0.55,
    "babilonia": 0.30, "asyria": 0.80, "fenicjanie": 0.25,
}
ARCHETYPE_TRADE = {
    "grecy": 0.75, "rzymianie": 0.50, "chinczycy": 0.85, "inkowie": 0.25,
    "zulusi": 0.20, "egipt": 0.60, "babilon": 0.65, "sumer": 0.65,
    "celtowie": 0.35, "germanie": 0.30,
    "harappa": 0.70, "hetyci": 0.50, "slowianie": 0.40,
    "babilonia": 0.65, "asyria": 0.25, "fenicjanie": 0.90,
}

# Jednostki specjalne nowych cyw (brak w units.json) -> szablon bazowy + mnoznik elity
SPEC_UNIT_TEMPLATE = {
    "Strażnik bram Harappy": ("Włócznik", 1.12),
    "Rydwan Kapadokijski": ("Rydwan konny", 1.15),
    "Drużynnik": ("Włócznik", 1.14),
    "Gwardia Ishtar": ("Wojownik z khopesh", 1.13),
    "Łucznik asyryjski": ("Łucznik", 1.16),
    "Tyrski miecznik": ("Wojownik z mieczem i tarczą", 1.12),
}

UNIT_FIELD_MAP = {
    "Atak": "Atak",
    "Obrazenia": "Obrażenia",
    "Obrona": "Obrona",
    "Uderzenie": "Uderzenie",
    "Pancerz": "Pancerz",
    "Przebicie": "Przebicie",
    "Health": "Health",
    "Atak_dystansowy": "Atak dystansowy",
    "Zasieg_hex": "Zasięg ataku (hex)",
    "Pociski": "Ilość pocisków",
    "Ruch_bitwa": "Ruch w bitwie (heksy)",
    "Ruch_mapa": "Ruch",
    "Widok": "Widok pola",
    "Dezercja_proc": "Próg dezercji (% health)",
    "Morale": "Morale bazowe",
    "Koszt_pieniadz": "Pieniądz (koszt)",
    "Utrzymanie": "Utrzymanie (Pieniądz/turę)",
    "Zywnosc_ture": "żywność/turę",
}

H_FILL = PatternFill("solid", fgColor="1B4332")
H2_FILL = PatternFill("solid", fgColor="2D6A4F")
ID_FILL = PatternFill("solid", fgColor="3D5A80")
H_FONT = Font(bold=True, color="FFFFFF", size=9)
H2_FONT = Font(bold=True, color="E8F5E9", size=8)
VAL_FONT = Font(color="0000FF", size=9)
THIN = Side(style="thin", color="CCCCCC")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def norm(s):
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def find_unit(units, name):
    if not name:
        return None
    n = norm(name.split("(")[0].split("/")[0])
    for u in units:
        j = norm(u.get("Jednostka", ""))
        if j == n or n in j or j in n:
            return u
    return None


def num_field(u, key):
    v = u.get(key)
    if v is None or v == "—" or v == "-":
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return 0


def elite_unit_stats(units, spec_name):
    u = find_unit(units, spec_name)
    if u:
        return u
    tpl = SPEC_UNIT_TEMPLATE.get(spec_name.split("(")[0].strip())
    if not tpl:
        for k, v in SPEC_UNIT_TEMPLATE.items():
            if k.lower() in spec_name.lower() or spec_name.lower() in k.lower():
                tpl = v
                break
    if not tpl:
        return None
    base_name, mult = tpl
    base = find_unit(units, base_name)
    if not base:
        return None
    out = dict(base)
    for k in ("Atak", "Obrażenia", "Obrona", "Uderzenie", "Pancerz", "Przebicie", "Health", "Atak dystansowy"):
        if k in out and isinstance(out[k], (int, float)):
            out[k] = round(out[k] * mult, 2)
    return out


def blank_row():
    return dict(DEFAULTS)


def set_epoki(row, epoki):
    row["meta_epoka_kamien"] = 1 if "kamien" in epoki else 0
    row["meta_epoka_braz"] = 1 if "braz" in epoki else 0
    row["meta_epoka_zelazo"] = 1 if "zelazo" in epoki else 0


def apply_bonus(row, b):
    typ = b.get("typ", "")
    cel = b.get("cel", "wszystkie")
    if cel == "dystans":
        cel = "lukownicy"
    if cel == "rekruci":
        pass
    val = b.get("wartosc", 0)
    opis = (b.get("opis") or "").lower()

    if typ == "bonus_walka":
        if "uderzen" in opis or ("szarz" in opis and cel == "kawaleria"):
            key = f"walka_uderzenie_{cel}"
        elif "dystans" in opis or cel == "lukownicy":
            key = f"walka_dystans_{cel if cel in ('lukownicy', 'rydwany') else 'lukownicy'}"
        elif "hp" in opis and "obron" in opis:
            key = f"walka_hp_{cel if cel in ('piechota', 'kawaleria', 'rydwany') else 'piechota'}"
        elif "pancerz" in opis and "atak" in opis:
            row[f"walka_atak_{cel}"] = row.get(f"walka_atak_{cel}", 0) + (val if isinstance(val, (int, float)) else 0)
            row[f"walka_pancerz_{cel}"] = row.get(f"walka_pancerz_{cel}", 0) + (val if isinstance(val, (int, float)) else 0)
            return
        elif "rekrutac" in opis or "koszt rekrutacji" in opis:
            if isinstance(val, (int, float)):
                row["walka_koszt_rekrutacji_proc"] = row.get("walka_koszt_rekrutacji_proc", 0) + val
                row["prod_koszt_jednostki_proc"] = row.get("prod_koszt_jednostki_proc", 0) + val
            return
        elif "ruch" in opis or "predkosc" in opis or "prędko" in opis:
            if isinstance(val, (int, float)):
                row["walka_ruch_bitwa_proc"] = row.get("walka_ruch_bitwa_proc", 0) + val
            return
        elif "zasi" in opis:
            if isinstance(val, (int, float)):
                row["walka_zasieg_proc"] = row.get("walka_zasieg_proc", 0) + val
            return
        elif "les" in opis or "dzungl" in opis or "gor" in opis:
            if isinstance(val, (int, float)):
                if "pierwsz" in opis or "szarz" in opis or "zasadz" in opis:
                    row["walka_atak_piechota_runda_szarzy"] = row.get("walka_atak_piechota_runda_szarzy", 0) + val
                row["walka_atak_piechota_teren_las"] = row.get("walka_atak_piechota_teren_las", 0) + val
            return
        elif "pierwsz" in opis or "szarz" in opis or "starciu" in opis or "uderzen" in opis:
            key = "walka_atak_piechota_runda_szarzy"
        else:
            key = f"walka_atak_{cel}"
        if isinstance(val, (int, float)):
            row[key] = row.get(key, 0) + val

    elif typ == "bonus_obrona":
        if "terytor" in opis or "wlasn" in opis:
            key = "walka_obrona_piechota_terytorium_wlasne"
        elif "mur" in opis or "fort" in opis:
            key = "walka_obrona_piechota_w_murze"
        else:
            key = f"walka_obrona_{cel}"
        if isinstance(val, (int, float)):
            row[key] = row.get(key, 0) + val
            if "hp" in opis:
                row[f"walka_hp_{cel}"] = row.get(f"walka_hp_{cel}", 0) + val

    elif typ == "bonus_zloto":
        if isinstance(val, (int, float)):
            if "port" in opis or "morsk" in opis:
                row["eko_pieniadz_port_proc"] = row.get("eko_pieniadz_port_proc", 0) + val
            row["eko_handel_brutto_proc"] = row.get("eko_handel_brutto_proc", 0) + val
            row["eko_pieniadz_proc"] = row.get("eko_pieniadz_proc", 0) + val

    elif typ == "bonus_nauka":
        if isinstance(val, (int, float)):
            row["eko_nauka_proc"] = row.get("eko_nauka_proc", 0) + val

    elif typ == "bonus_pobor_regen":
        if isinstance(val, (int, float)):
            row["mp_regen_proc"] = row.get("mp_regen_proc", 0) + val

    elif typ == "koszt_redukcja" and cel == "budynki":
        if isinstance(val, (int, float)):
            row["prod_koszt_budynku_proc"] = row.get("prod_koszt_budynku_proc", 0) + val


def suggest_ai_defaults(name, typ):
    """Obiektywne domyslne AI gdy brak w civ-ai.json."""
    presets = {
        "Harappa": dict(agresywnosc=2, priorytetMilitarny=4, priorytetEkonomia=7, priorytetNauka=5, sklonnoscDoPodboju=1),
        "Hetyci": dict(agresywnosc=5, priorytetMilitarny=6, priorytetEkonomia=5, priorytetNauka=4, sklonnoscDoPodboju=3),
        "Słowianie": dict(agresywnosc=6, priorytetMilitarny=6, priorytetEkonomia=5, priorytetNauka=4, sklonnoscDoPodboju=3),
        "Babilonia": dict(agresywnosc=3, priorytetMilitarny=4, priorytetEkonomia=6, priorytetNauka=8, sklonnoscDoPodboju=2),
        "Asyria": dict(agresywnosc=8, priorytetMilitarny=8, priorytetEkonomia=4, priorytetNauka=5, sklonnoscDoPodboju=5),
        "Fenicjanie": dict(agresywnosc=3, priorytetMilitarny=3, priorytetEkonomia=8, priorytetNauka=5, sklonnoscDoPodboju=1),
    }
    p = presets.get(name, {})
    return {
        "ai_agresywnosc": p.get("agresywnosc", 5),
        "ai_ekspansywnosc": 4 if typ in ("slowianie", "asyria") else 3,
        "ai_priorytet_militarny": p.get("priorytetMilitarny", 5),
        "ai_priorytet_ekonomia": p.get("priorytetEkonomia", 5),
        "ai_priorytet_nauka": p.get("priorytetNauka", 5),
        "ai_tolerancja_ryzyka": 4,
        "ai_sklonnosc_podboju": p.get("sklonnoscDoPodboju", 3),
        "ai_profil_obronna": 1 if name in ("Harappa", "Hetyci", "Inkowie") else 0,
    }


def suggest_dip_defaults(name, per):
    if per:
        return {
            "dip_sklonnosc_sojusze": per.get("sklonnoscSojusze", 5),
            "dip_lojalnosc": per.get("lojalnosc", 5),
            "dip_prog_wojny": per.get("progWojny", 5),
            "dip_pamietliwosc": per.get("pamietliwosc", 5),
            "dip_otwartosc_handel": per.get("otwartoscHandel", 5),
            "dip_nastawienie_bazowe": per.get("nastawienieBazowe", 50),
        }
    neutral = {
        "Harappa": dict(sklonnoscSojusze=7, lojalnosc=6, progWojny=2, pamietliwosc=5, otwartoscHandel=8, nastawienieBazowe=58),
        "Hetyci": dict(sklonnoscSojusze=5, lojalnosc=6, progWojny=5, pamietliwosc=6, otwartoscHandel=5, nastawienieBazowe=52),
        "Słowianie": dict(sklonnoscSojusze=4, lojalnosc=5, progWojny=6, pamietliwosc=6, otwartoscHandel=4, nastawienieBazowe=48),
        "Babilonia": dict(sklonnoscSojusze=6, lojalnosc=5, progWojny=4, pamietliwosc=5, otwartoscHandel=6, nastawienieBazowe=55),
        "Asyria": dict(sklonnoscSojusze=2, lojalnosc=4, progWojny=9, pamietliwosc=8, otwartoscHandel=3, nastawienieBazowe=38),
        "Fenicjanie": dict(sklonnoscSojusze=5, lojalnosc=4, progWojny=3, pamietliwosc=4, otwartoscHandel=9, nastawienieBazowe=62),
    }
    d = neutral.get(name, {})
    return {
        "dip_sklonnosc_sojusze": d.get("sklonnoscSojusze", 5),
        "dip_lojalnosc": d.get("lojalnosc", 5),
        "dip_prog_wojny": d.get("progWojny", 5),
        "dip_pamietliwosc": d.get("pamietliwosc", 5),
        "dip_otwartosc_handel": d.get("otwartoscHandel", 5),
        "dip_nastawienie_bazowe": d.get("nastawienieBazowe", 50),
    }


def build_civ_list():
    civs = load_json(CIVS).get("cywilizacje", [])
    draft = load_json(DRAFT).get("cywilizacje", [])
    by_name = {c["Cywilizacja"]: c for c in civs}
    for d in draft:
        if d["Cywilizacja"] not in by_name:
            by_name[d["Cywilizacja"]] = d
    order = [
        "Grecy", "Rzymianie", "Chińczycy", "Inkowie", "Zulusi", "Egipt", "Sumerowie",
        "Celtowie", "Germanie", "Harappa", "Hetyci", "Słowianie", "Babilonia", "Asyria", "Fenicjanie",
    ]
    out = []
    for name in order:
        if name in by_name:
            out.append(by_name[name])
    return out


def build_matrix_rows():
    units = load_json(UNITS)
    civ_ai = {r["Cywilizacja"]: r for r in load_json(CIV_AI).get("cywilizacje", [])}
    civ_params = {r["Cywilizacja"]: r for r in load_json(CIV_PARAMS).get("cywilizacje", [])}
    dip_rows = {r["Cywilizacja"]: r for r in load_json(DIP).get("perNacja", [])}

    rows_out = []
    for cyw in build_civ_list():
        name = cyw["Cywilizacja"]
        typ = cyw.get("typCywilizacji", "")
        if name == "Sumerowie":
            typ = "sumer"
        ikona = cyw.get("ikonaId", typ)
        if name == "Sumerowie":
            ikona = "sumer"
        tier = cyw.get("tier", 1)

        row = blank_row()
        row["_Cywilizacja"] = name
        row["_typCywilizacji"] = typ
        row["_ikonaId"] = ikona
        row["_tier"] = tier

        set_epoki(row, cyw.get("epokiStartowe") or [])
        row["meta_mnoznik_waluta"] = float(cyw.get("mnoznikHandelPieniadz", 2.0))
        row["meta_tier_roster"] = tier

        for b in cyw.get("bonusy") or []:
            apply_bonus(row, b)

        # modWzrostu / modEkonomii -> lud_wzrost / eko_praca
        cp = civ_params.get(name)
        if cp:
            mw = float(cp.get("modWzrostu", 1.0))
            me = float(cp.get("modEkonomii", 1.0))
            row["lud_wzrost_proc"] = round(mw - 1.0, 3)
            row["eko_praca_proc"] = round(me - 1.0, 3)
        elif name in ("Harappa", "Fenicjanie"):
            row["lud_wzrost_proc"] = 0.0
            row["eko_praca_proc"] = 0.04 if name == "Harappa" else 0.05
        elif name == "Słowianie":
            row["lud_wzrost_proc"] = 0.03
            row["eko_praca_proc"] = 0.0
        elif name == "Asyria":
            row["lud_wzrost_proc"] = 0.0
            row["eko_praca_proc"] = -0.02

        # AI
        ai = civ_ai.get(name) or cyw.get("civAi")
        if ai:
            row["ai_agresywnosc"] = ai.get("agresywnosc", 5)
            row["ai_ekspansywnosc"] = ai.get("ekspansywnosc", 3)
            row["ai_priorytet_militarny"] = ai.get("priorytetMilitarny", 5)
            row["ai_priorytet_ekonomia"] = ai.get("priorytetEkonomia", 5)
            row["ai_priorytet_nauka"] = ai.get("priorytetNauka", 5)
            row["ai_tolerancja_ryzyka"] = ai.get("tolerancjaRyzyka", 4)
            row["ai_sklonnosc_podboju"] = ai.get("sklonnoscDoPodboju", 3)
            prof = ai.get("profilMapy", "")
            row["ai_profil_obronna"] = 1 if prof == "kopia_typu_obronna" else 0
        else:
            row.update(suggest_ai_defaults(name, typ))

        # Dyplomacja
        per = dip_rows.get(name) or cyw.get("perNacja")
        row.update(suggest_dip_defaults(name, per))
        tkey = typ if typ != "babilon" else "sumer" if name == "Sumerowie" else typ
        row["dip_agresja_archetyp"] = ARCHETYPE_AGG.get(tkey, ARCHETYPE_AGG.get(typ, 0.5))
        row["dip_handlowosc_archetyp"] = ARCHETYPE_TRADE.get(tkey, ARCHETYPE_TRADE.get(typ, 0.5))

        # Typowe minusy (obiektywne rezerwy balansu)
        if name == "Chińczycy":
            row["walka_obrona_piechota"] = row.get("walka_obrona_piechota", 0) - 0.05
        if name == "Zulusi":
            row["walka_dystans_lukownicy"] = -0.10
        if name == "Inkowie":
            row["walka_atak_kawaleria"] = -0.15
            row["walka_atak_rydwany"] = -0.15
        if name == "Fenicjanie":
            row["walka_obrona_piechota"] = -0.05
        if name == "Harappa":
            row["walka_atak_kawaleria"] = -0.08
        if name == "Hetyci":
            row["eko_pieniadz_port_proc"] = -0.05
        if name == "Słowianie":
            row["eko_nauka_proc"] = row.get("eko_nauka_proc", 0) - 0.05
        if name == "Asyria":
            row["dip_nastawienie_bazowe"] = min(row.get("dip_nastawienie_bazowe", 50), 40)

        # Jednostka spec — staty absolutne
        spec_name = cyw.get("Jednostka specjalna", "")
        for b in cyw.get("bonusy") or []:
            if b.get("typ") == "jednostka_specjalna":
                spec_name = str(b.get("wartosc", spec_name))
        u = elite_unit_stats(units, spec_name)
        if u:
            for pid, ukey in UNIT_FIELD_MAP.items():
                row[f"spec_{pid}"] = num_field(u, ukey)

        rows_out.append(row)
    return rows_out


def write_domain_sheet(ws, rows, param_ids):
    headers = list(IDENT_COLS) + param_ids
    unit_row = ["", "", "", ""] + [PARAM_META[p]["jednostka"] for p in param_ids]
    form_row = ["", "", "", ""] + [PARAM_META[p]["formula"] for p in param_ids]
    mod_row = ["", "", "", ""] + [PARAM_META[p]["modul"] for p in param_ids]

    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.fill = ID_FILL if j <= 4 else H_FILL
        c.font = H_FONT
        c.border = BD
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for r_idx, extra in enumerate([unit_row, form_row, mod_row], 2):
        for j, v in enumerate(extra, 1):
            c = ws.cell(r_idx, j, v)
            c.fill = H2_FILL
            c.font = H2_FONT
            c.border = BD

    data_start = 5
    for i, row in enumerate(rows):
        ri = data_start + i
        ws.cell(ri, 1, row["_Cywilizacja"]).border = BD
        ws.cell(ri, 2, row["_typCywilizacji"]).border = BD
        ws.cell(ri, 3, row["_ikonaId"]).border = BD
        ws.cell(ri, 4, row["_tier"]).border = BD
        for j, pid in enumerate(param_ids, 5):
            val = row.get(pid, DEFAULTS.get(pid, 0))
            c = ws.cell(ri, j, val)
            c.border = BD
            c.font = VAL_FONT
            if isinstance(val, float):
                c.number_format = "0.000"

    ws.freeze_panes = "E5"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 6
    for j in range(5, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 14


def write_info_sheet(ws):
    lines = [
        ["Cyw-macierz — 11 zakladek aktywnych (+ INFO)", ""],
        ["", ""],
        ["Potega / Power", "CZEKA modul Power — wagi tylko globalnie Panel-B (Potega-P-A). Cyw-12 per-cyw dopiero po wzorze od Macieja."],
        ["Zasada", "0 = brak efektu; ulamek 0.15 = +15% (mul_proc: baza * (1+val))"],
        ["Edycja", "Zmien niebieskie liczby w wierszach 5+"],
        ["Eksport", "python panele-sterowania/export-cyw-macierz.py"],
        ["JSON", "gra/data/civ-matrix.json"],
        ["API", "gra/src/game/civ-matrix.ts"],
        ["", ""],
        ["Nr", "Arkusz", "Domena", "Kolumn param", "Modul glowny"],
    ]
    for i, line in enumerate(lines, 1):
        for j, v in enumerate(line, 1):
            ws.cell(i, j, v)

    r = len(lines) + 1
    mod_prim = {
        "meta": "start-preview.ts",
        "walka": "civ-bonuses.ts",
        "jednostka_spec": "units.json",
        "ekonomia": "economy.ts",
        "produkcja": "production.ts",
        "ludnosc": "economy.ts",
        "manpower": "manpower.ts",
        "spoleczenstwo": "wealth.ts / order.ts",
        "obleczenie": "siege.ts",
        "dyplomacja": "diplomacy.ts",
        "ai": "civ-ai-data.ts",
    }
    for nr, (dom, sheet) in enumerate(DOMAIN_SHEETS, 1):
        ws.cell(r, 1, nr)
        ws.cell(r, 2, sheet)
        ws.cell(r, 3, dom)
        ws.cell(r, 4, len(PARAMS_BY_DOMAIN[dom]))
        note = mod_prim.get(dom, "")
        if dom in RESERVED_DOMAINS:
            note += " [REZERWA — czeka Power]"
        ws.cell(r, 5, note)
        r += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 28


def write_excel(rows, target_path):
    # Nowy plik review = tylko Cyw-* ; Panel-D = dolacz arkusze
    is_standalone = os.path.normpath(target_path) == os.path.normpath(STANDALONE)
    if is_standalone or not os.path.isfile(target_path):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(target_path)

    for old in list(wb.sheetnames):
        if old == SHEET_LEGACY or old.startswith("Cyw-"):
            del wb[old]

    ws_info = wb.create_sheet(SHEET_INFO, 0)
    write_info_sheet(ws_info)

    total_cols = 4
    for idx, (dom, sheet_name) in enumerate(DOMAIN_SHEETS, 1):
        pids = PARAMS_BY_DOMAIN[dom]
        ws = wb.create_sheet(sheet_name, idx)
        write_domain_sheet(ws, rows, pids)
        total_cols = max(total_cols, 4 + len(pids))

    wb.save(target_path)
    return total_cols, len(rows)


def write_json(rows):
    payload = {
        "_meta": {
            "version": "1.0",
            "source": "Cyw-macierz (11 arkuszy Cyw-01..11)",
            "formula_mul_proc": "wynik = baza * (1 + wartosc)",
            "formula_mul_abs": "wynik = baza * wartosc",
            "formula_add": "wynik = baza + wartosc",
            "formula_flag": "wartosc 1 = aktywny warunek",
            "kolumny": len(PARAM_IDS),
            "cywilizacje": len(rows),
        },
        "paramDefs": PARAM_META,
        "defaults": DEFAULTS,
        "cywilizacje": [],
    }
    for row in rows:
        entry = {
            "Cywilizacja": row["_Cywilizacja"],
            "typCywilizacji": row["_typCywilizacji"],
            "ikonaId": row["_ikonaId"],
            "tier": row["_tier"],
            "params": {pid: row.get(pid, DEFAULTS.get(pid, 0)) for pid in PARAM_IDS},
        }
        payload["cywilizacje"].append(entry)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main():
    json_only = "--json-only" in sys.argv
    rows = build_matrix_rows()
    write_json(rows)
    if not json_only:
        targets = []
        for path in (STANDALONE, PANEL):
            try:
                ncol, nrow = write_excel(rows, path)
                targets.append(path)
                print(f"OK Excel: {path} [12 arkuszy Cyw-00..11] {nrow} cyw")
            except PermissionError:
                print(f"SKIP (zablokowany): {path} — zamknij Excel i uruchom ponownie")
        if not targets:
            sys.exit(1)
    print(f"OK JSON: {OUT_JSON} ({len(PARAM_IDS)} params x {len(rows)} cyw)")


if __name__ == "__main__":
    main()
