#!/usr/bin/env python3
# gen-tereny-maciej-xlsx.py — terrain-yields.json + terrain-improvements.json → tabela do edycji
# Uruchom: python panele-sterowania/gen-tereny-maciej-xlsx.py
import json
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
OUT = os.path.join(os.path.dirname(__file__), "Tereny-i-ulepszenia-MACIEJ.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
INFO_FONT = Font(bold=True, size=11, color="1F4E79")

YIELD_COLS = ("Żywność", "Praca", "Podatek", "Drewno", "Kamień")
BONUS_COLS = (
    ("zywnosc", "Żywność (bonus)"),
    ("praca", "Praca (bonus)"),
    ("handel", "Podatek (bonus)"),
    ("pieniadz", "Pieniądz (bonus)"),
    ("kamien", "Kamień (bonus)"),
    ("drewno", "Drewno (bonus)"),
    ("glina", "Glina (bonus)"),
)


def load_json(name):
    path = os.path.join(DATA, name)
    if not os.path.isfile(path):
        print(f"Brak pliku: {path}")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def style_header(ws, headers, row=2):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row + 1, 1).coordinate


def write_info(wb):
    ws = wb.active
    ws.title = "_INFO"
    lines = [
        "Tereny i ulepszenia — edycja balansu (Maciej)",
        "",
        f"Wygenerowano z JSON: {date.today().isoformat()}",
        "",
        "Arkusze do edycji:",
        "  1) Plony-bazowe — typ terenu (Łąka, Równina, …)",
        "  2) Nakladki — Rzeka, Las (dodawane do terenu bazowego)",
        "  3) Ulepszenia — bonusy pól + surowiec/turę + koszt",
        "",
        "Po zapisaniu pliku napisz w czacie: eksportuj tereny maciej",
        "(agent zaktualizuje terrain-yields.json i terrain-improvements.json)",
        "",
        "Źródło prawdy w grze: gra/data/*.json — ten Excel to panel sterowania.",
    ]
    for i, line in enumerate(lines, 1):
        cell = ws.cell(i, 1, line)
        if i == 1:
            cell.font = INFO_FONT
    ws.column_dimensions["A"].width = 72


def write_yields_sheet(wb, title, key_field, rows):
    ws = wb.create_sheet(title)
    headers = [key_field, *YIELD_COLS, "Suma", "Uwagi"]
    style_header(ws, headers)
    widths = [18, 10, 10, 10, 10, 10, 8, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for idx, row in enumerate(rows, 3):
        ws.cell(idx, 1, row.get(key_field, ""))
        for j, col in enumerate(YIELD_COLS, 2):
            val = row.get(col, row.get("Handel", 0) if col == "Podatek" else 0)
            ws.cell(idx, j, val)
        ws.cell(idx, 7, row.get("Suma", 0))
        uwagi = row.get("Uwagi")
        ws.cell(idx, 8, uwagi if uwagi is not None else "")
        ws.cell(idx, 8).alignment = Alignment(wrap_text=True, vertical="top")


def write_improvements_sheet(wb, terrain):
    ws = wb.create_sheet("Ulepszenia")
    headers = [
        "Klucz (nie zmieniaj)",
        "Nazwa",
        "Epoka",
        *[label for _, label in BONUS_COLS],
        "Surowiec (ASCII)",
        "Ilość surowca/turę",
        "Koszt (Praca)",
        "Technologia",
        "Teren / gdzie",
        "Warunek",
    ]
    style_header(ws, headers)
    widths = [20, 22, 8, 12, 12, 12, 12, 12, 12, 12, 14, 14, 12, 28, 36]
    for i, w in enumerate(widths, 1):
        col = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws.column_dimensions[col].width = w

    skip = {"_meta", "_miasto_zasieg_ref"}
    row_idx = 3
    for imp_key, node in sorted(terrain.items()):
        if imp_key in skip or not isinstance(node, dict):
            continue
        bonus = node.get("bonus") or {}
        ws.cell(row_idx, 1, imp_key)
        ws.cell(row_idx, 2, node.get("nazwa", imp_key))
        ws.cell(row_idx, 3, node.get("epoka", ""))
        for j, (bf, _) in enumerate(BONUS_COLS, 4):
            val = bonus.get(bf, "")
            ws.cell(row_idx, j, val if val != "" else "")
        base_col = 4 + len(BONUS_COLS)
        ws.cell(row_idx, base_col, node.get("surowiecOdblokowany") or "")
        ws.cell(row_idx, base_col + 1, node.get("surowiec_ilosc_tura", ""))
        ws.cell(row_idx, base_col + 2, node.get("koszt_praca", ""))
        ws.cell(row_idx, base_col + 3, node.get("tech", ""))
        ws.cell(row_idx, base_col + 4, node.get("teren", ""))
        ws.cell(row_idx, base_col + 5, node.get("warunek", ""))
        ws.cell(row_idx, base_col + 5).alignment = Alignment(wrap_text=True, vertical="top")
        row_idx += 1


def main():
    yields_data = load_json("terrain-yields.json")
    terrain = load_json("terrain-improvements.json")

    wb = openpyxl.Workbook()
    write_info(wb)
    write_yields_sheet(wb, "Plony-bazowe", "Teren", yields_data.get("terrain_types", []))
    write_yields_sheet(wb, "Nakladki", "Modyfikator", yields_data.get("terrain_modifiers", []))
    write_improvements_sheet(wb, terrain)

    wb.properties.title = "Tereny i ulepszenia — MACIEJ"
    wb.properties.subject = f"gen {date.today().isoformat()}"
    wb.save(OUT)
    print(f"Zapisano: {OUT}")
    print(f"Arkusze: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
