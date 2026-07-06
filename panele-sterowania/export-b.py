#!/usr/bin/env python3

# export-b.py — Panel-B.xlsx → miasto/econ/society/buildings/resources/tech JSON

# Kolumna Wartość (4) → normal (econ/society) lub wartosc (miasto). Puste = brak zmiany.

# Arkusze tabelaryczne (Budynki, Surowce, Technologie): puste komórki = brak zmiany.

# Uruchom z root: python panele-sterowania/export-b.py

import argparse

import json

import os

import sys

try:

    import openpyxl

except ImportError:

    print("pip install openpyxl")

    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))

DATA = os.path.join(ROOT, "gra", "data")

DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "Panel-B.xlsx")

COL_PARAM = 2

COL_VALUE = 4

SHEET_MIASTO = {"Miasto": "miasto-params.json"}

SHEET_ECON = {

    "Ekonomia": "ekonomia_miasta",

    "Wealth": "wealth",

    "Globalne": "globalne",

    "Budynki-eco": "budynki",

    "Teren-bonus": "teren_mapa",

}

SHEET_SOCIETY = {

    "Zdrowie": "zdrowie",

    "Szczescie": "szczescie",

    "Kultura": "kultura",

    "Religia": "religia",

}

TECH_JSON_FIELDS = [

    "Technologia",

    "Epoka",

    "Poziom",

    "Dostęp do surowca.",

    "wymagany budynek",

    "Wymaga (prereq)",

    "Odblokowuje surowiec.",

    "Odblokowuje budynek",

    "Koszt nauki",

    "Uwagi",

    "Odblokowuje ulepszenie terenu",

]

def coerce(old, val):

    if val is None or (isinstance(val, str) and val.strip() == ""):

        return old, False

    if isinstance(old, bool):

        if isinstance(val, bool):

            return val, val != old

        nv = str(val).strip().lower() in ("tak", "true", "1", "prawda")

        return nv, nv != old

    if isinstance(old, int):

        try:

            nv = int(round(float(val)))

            return nv, nv != old

        except (TypeError, ValueError):

            return old, False

    if isinstance(old, float):

        try:

            nv = float(val)

            return nv, nv != old

        except (TypeError, ValueError):

            return old, False

    if old is None:

        return val, True

    nv = str(val)

    return nv, nv != old

def read_sheet_params(ws):

    out = {}

    for r in range(2, ws.max_row + 1):

        key = ws.cell(r, COL_PARAM).value

        if key is None or str(key).strip() == "":

            continue

        key = str(key).strip()

        val = ws.cell(r, COL_VALUE).value

        out[key] = val

    return out

def overlay_normal(section, params):

    changed = 0

    for key, val in params.items():

        if key not in section or not isinstance(section[key], dict):

            continue

        if "normal" not in section[key]:

            continue

        nv, did = coerce(section[key]["normal"], val)

        if did:

            section[key]["normal"] = nv

            changed += 1

    return changed

def overlay_wartosc(root, params):

    changed = 0

    for key, val in params.items():

        if key not in root or not isinstance(root[key], dict):

            continue

        if "wartosc" not in root[key]:

            continue

        nv, did = coerce(root[key]["wartosc"], val)

        if did:

            root[key]["wartosc"] = nv

            changed += 1

    return changed

def overlay_porzadek(society, ws):

    params = read_sheet_params(ws)

    changed = 0

    for sec_name in ("porzadek", "prawo"):

        if sec_name not in society:

            continue

        changed += overlay_normal(society[sec_name], params)

    return changed

def overlay_buildings(ws, orig):

    """Nakładka wartości z arkusza Budynki na buildings.json (płaskie kolumny z kropką)."""

    header = [c.value for c in ws[1]]

    by_id = {b["id"]: b for b in orig}

    changed = 0

    for r in range(2, ws.max_row + 1):

        row = {header[j]: ws.cell(r, j + 1).value for j in range(len(header)) if header[j]}

        rid = row.get("id")

        if rid is None or rid not in by_id:

            continue

        b = by_id[rid]

        for col, val in row.items():

            if not col or col in ("id",):

                continue

            if val is None or (isinstance(val, str) and val.strip() == ""):

                continue

            if "." in col:

                p, c = col.split(".", 1)

                if isinstance(b.get(p), dict) and c in b[p]:

                    nv, did = coerce(b[p][c], val)

                    if did:

                        b[p][c] = nv

                        changed += 1

            elif col in b:

                nv, did = coerce(b[col], val)

                if did:

                    b[col] = nv

                    changed += 1

    return changed

def overlay_resources(ws, orig):

    """Nakładka wartości z arkusza Surowce — klucz = pole Surowiec."""

    header = [c.value for c in ws[1]]

    by_name = {r["Surowiec"]: r for r in orig}

    changed = 0

    for r in range(2, ws.max_row + 1):

        row = {header[j]: ws.cell(r, j + 1).value for j in range(len(header)) if header[j]}

        name = row.get("Surowiec")

        if name is None or name not in by_name:

            continue

        rec = by_name[name]

        for col, val in row.items():

            if not col or col == "Surowiec":

                continue

            if val is None or (isinstance(val, str) and val.strip() == ""):

                continue

            if col in rec:

                nv, did = coerce(rec[col], val)

                if did:

                    rec[col] = nv

                    changed += 1

    return changed

def read_technologie_sheet(ws):

    """Czyta arkusz Technologie → lista słowników (format tech.json)."""

    col_map = {}

    for c in range(1, ws.max_column + 1):

        hdr = ws.cell(1, c).value

        if hdr is not None:

            col_map[str(hdr).strip()] = c

    missing = [f for f in TECH_JSON_FIELDS if f not in col_map]

    if missing:

        print(f"  (Technologie: brak kolumn {missing} — pominięto eksport tech)")

        return None

    techs = []

    for r in range(2, ws.max_row + 1):

        first = ws.cell(r, 1).value

        if first is None or not isinstance(first, str) or not first.strip():

            continue

        if len(first.strip()) > 60 and "=" in first:

            continue

        obj = {}

        for field in TECH_JSON_FIELDS:

            c = col_map[field]

            val = ws.cell(r, c).value

            if field == "Poziom":

                if val is not None:

                    try:

                        val = int(val)

                    except (TypeError, ValueError):

                        pass

            elif field == "Koszt nauki":

                if val is not None:

                    try:

                        f = float(val)

                        val = int(f) if f.is_integer() else f

                    except (TypeError, ValueError):

                        pass

            elif isinstance(val, str):

                val = val.strip() if val.strip() else None

            obj[field] = val

        techs.append(obj)

    return techs

def tech_lists_equal(a, b):

    if len(a) != len(b):

        return False

    for ta, tb in zip(a, b):

        if ta.get("Technologia") != tb.get("Technologia"):

            return False

        for field in TECH_JSON_FIELDS:

            if ta.get(field) != tb.get(field):

                return False

    return True

def save_json(obj, path):

    tmp = path + ".tmp"

    with open(tmp, "w", encoding="utf-8") as f:

        json.dump(obj, f, ensure_ascii=False, indent=2)

        f.write("\n")

    json.load(open(tmp, encoding="utf-8"))

    os.replace(tmp, path)

EPOKA_MANPOWER_COLS = ["epoka", "ludekNaLudka", "manpowerNaLudka", "manpowerNaJednostke"]

POWER_INFO_KEYS = {"_kalibracja_power_oczekiwany"}


def overlay_power_skladniki(power, params):

    changed = 0

    skladniki = power.setdefault("skladniki", {})

    for key, val in params.items():

        if key in POWER_INFO_KEYS:

            continue

        if key not in skladniki or not isinstance(skladniki[key], dict):

            continue

        nv, did = coerce(skladniki[key].get("pkt", 0), val)

        if did:

            skladniki[key]["pkt"] = nv

            changed += 1

    return changed


def overlay_power_opcje(power, params):

    changed = 0

    opcje = power.setdefault("opcje", {})

    for key, val in params.items():

        if key not in opcje or not isinstance(opcje[key], dict):

            continue

        old = opcje[key].get("wartosc")

        typ = opcje[key].get("typ", "string")

        if typ == "bool":

            nv, did = coerce(bool(old) if old is not None else False, val)

        else:

            nv, did = coerce(old if old is not None else "", val)

        if did:

            opcje[key]["wartosc"] = nv

            changed += 1

    return changed


def overlay_epoka_manpower(ws, orig):

    header = [c.value for c in ws[1]]

    by_epoka = {row["epoka"]: row for row in orig.get("epoki", []) if isinstance(row, dict)}

    changed = 0

    for r in range(2, ws.max_row + 1):

        row = {header[j]: ws.cell(r, j + 1).value for j in range(len(header)) if header[j]}

        ep = row.get("epoka")

        if ep is None:

            continue

        try:

            ep = int(ep)

        except (TypeError, ValueError):

            continue

        if ep not in by_epoka:

            continue

        rec = by_epoka[ep]

        for col in EPOKA_MANPOWER_COLS:

            if col == "epoka" or col not in rec:

                continue

            val = row.get(col)

            if val is None or (isinstance(val, str) and val.strip() == ""):

                continue

            nv, did = coerce(rec[col], val)

            if did:

                rec[col] = nv

                changed += 1

    return changed

def main():

    ap = argparse.ArgumentParser(description="Eksport Panel-B.xlsx → JSON Grupy B")

    ap.add_argument("--xlsx", default=DEFAULT_XLSX)

    ap.add_argument("--data-dir", default=DATA)

    ap.add_argument("--dry-run", action="store_true", help="pokaż zmiany bez zapisu JSON")

    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):

        print(f"Brak pliku: {args.xlsx}")

        print("Uruchom najpierw: python panele-sterowania/gen-panel-b.py")

        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    total = 0

    miasto_path = os.path.join(args.data_dir, "miasto-params.json")

    miasto = json.load(open(miasto_path, encoding="utf-8"))

    if "Miasto" in wb.sheetnames:

        ch = overlay_wartosc(miasto, read_sheet_params(wb["Miasto"]))

        if ch and not args.dry_run:

            save_json(miasto, miasto_path)

        print(f"miasto-params.json: {ch} zmian")

        total += ch

    econ_path = os.path.join(args.data_dir, "econ-params.json")

    econ = json.load(open(econ_path, encoding="utf-8"))

    econ_total = 0

    for sheet, section_key in SHEET_ECON.items():

        if sheet not in wb.sheetnames:

            continue

        ch = overlay_normal(econ[section_key], read_sheet_params(wb[sheet]))

        print(f"econ-params/{section_key} ({sheet}): {ch} zmian")

        econ_total += ch

    if econ_total and not args.dry_run:

        save_json(econ, econ_path)

    total += econ_total

    society_path = os.path.join(args.data_dir, "society-params.json")

    society = json.load(open(society_path, encoding="utf-8"))

    soc_total = 0

    for sheet, section_key in SHEET_SOCIETY.items():

        if sheet not in wb.sheetnames:

            continue

        ch = overlay_normal(society[section_key], read_sheet_params(wb[sheet]))

        print(f"society-params/{section_key} ({sheet}): {ch} zmian")

        soc_total += ch

    if "Porzadek" in wb.sheetnames:

        ch = overlay_porzadek(society, wb["Porzadek"])

        print(f"society-params/porzadek+prawo (Porzadek): {ch} zmian")

        soc_total += ch

    if soc_total and not args.dry_run:

        save_json(society, society_path)

        total += soc_total

    if "Potega-P-A" in wb.sheetnames or "Potega-opcje" in wb.sheetnames:

        power_path = os.path.join(args.data_dir, "power-params.json")

        power = json.load(open(power_path, encoding="utf-8"))

        power_ch = 0

        if "Potega-P-A" in wb.sheetnames:

            ch = overlay_power_skladniki(power, read_sheet_params(wb["Potega-P-A"]))

            print(f"power-params.json/skladniki (Potega-P-A): {ch} zmian")

            power_ch += ch

        if "Potega-opcje" in wb.sheetnames:

            ch = overlay_power_opcje(power, read_sheet_params(wb["Potega-opcje"]))

            print(f"power-params.json/opcje (Potega-opcje): {ch} zmian")

            power_ch += ch

        if power_ch and not args.dry_run:

            save_json(power, power_path)

        total += power_ch

    if "Manpower-epoki" in wb.sheetnames:

        ep_path = os.path.join(args.data_dir, "epoka-ludnosc-manpower.json")

        epoka_mp = json.load(open(ep_path, encoding="utf-8"))

        ch = overlay_epoka_manpower(wb["Manpower-epoki"], epoka_mp)

        if ch and not args.dry_run:

            save_json(epoka_mp, ep_path)

        print(f"epoka-ludnosc-manpower.json (Manpower-epoki): {ch} zmian")

        total += ch

    if "Budynki" in wb.sheetnames:

        bld_path = os.path.join(args.data_dir, "buildings.json")

        buildings = json.load(open(bld_path, encoding="utf-8"))

        ch = overlay_buildings(wb["Budynki"], buildings)

        if ch and not args.dry_run:

            save_json(buildings, bld_path)

        print(f"buildings.json (Budynki): {ch} zmian")

        total += ch

    if "Surowce" in wb.sheetnames:

        res_path = os.path.join(args.data_dir, "resources.json")

        resources = json.load(open(res_path, encoding="utf-8"))

        ch = overlay_resources(wb["Surowce"], resources)

        if ch and not args.dry_run:

            save_json(resources, res_path)

        print(f"resources.json (Surowce): {ch} zmian")

        total += ch

    if "Technologie" in wb.sheetnames:

        tech_path = os.path.join(args.data_dir, "tech.json")

        tech_data = json.load(open(tech_path, encoding="utf-8"))

        techs = read_technologie_sheet(wb["Technologie"])

        if techs is not None:

            existing = tech_data.get("technologie", [])

            if not tech_lists_equal(techs, existing):

                if not args.dry_run:

                    tempo = tech_data.get("tempo_gry")

                    tech_data["technologie"] = techs

                    if tempo is not None:

                        tech_data["tempo_gry"] = tempo

                    save_json(tech_data, tech_path)

                print(f"tech.json (Technologie): zaktualizowano {len(techs)} technologii")

                total += 1

            else:

                print(f"tech.json (Technologie): 0 zmian ({len(techs)} wierszy)")

    print(f"RAZEM: {total} zmian")

    if args.dry_run:

        print("(dry-run — JSON nie zapisany)")

    if total == 0:

        print("(brak zmian — kolumna Wartość pusta lub identyczna z JSON)")

if __name__ == "__main__":

    main()

