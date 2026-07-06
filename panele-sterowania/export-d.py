#!/usr/bin/env python3
"""
export-d.py — Panel Grupy D → JSON (targeted, NIGDY export-data.py).

Domyślnie (bez --full) eksportuje wszystkie arkusze Panel-D.xlsx:
  Dyplomacja, AI-*, Barbarzyńcy, Bonusy-cywilizacji, Cywilizacje-roster,
  Parametry-cyw, AI-per-nacja, Dyplomacja-akcje, Dyplomacja-per-nacja.

Użycie (z root projektu):
  python panele-sterowania/export-d.py
  python panele-sterowania/export-d.py --dry-run
  python panele-sterowania/export-d.py --full   # DEPRECATED — fallback legacy gdy brak arkusza
"""
import argparse
import json
import os
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, ".."))
PANEL = os.path.join(HERE, "Panel-D.xlsx")
DATA = os.path.join(ROOT, "gra", "data")
TOOLS = os.path.join(ROOT, "gra", "tools")

COL_PARAM = 2
COL_VALUE = 4

SHEET_BONUSY = "Bonusy-cywilizacji"
SHEET_ROSTER = "Cywilizacje-roster"
SHEET_PARAMS_CYW = "Parametry-cyw"
SHEET_AI_NACJA = "AI-per-nacja"
SHEET_DIP_AKCJE = "Dyplomacja-akcje"
SHEET_DIP_PER = "Dyplomacja-per-nacja"

LEGACY_FALLBACK = {
    SHEET_BONUSY: "export-bonusy-cyw.py",
    SHEET_ROSTER: "export-civs.py",
    SHEET_PARAMS_CYW: "export-civ-params.py",
    SHEET_AI_NACJA: "export-civ-ai.py",
    SHEET_DIP_PER: "export-civ-dyplomacy-nations.py",
    SHEET_DIP_AKCJE: None,  # brak osobnego legacy — tylko panel
}


def num(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return int(v) if isinstance(v, float) and v.is_integer() else v
    s = str(v).strip()
    if s.upper() in ("TAK", "TRUE", "YES"):
        return True
    if s.upper() in ("NIE", "FALSE", "NO"):
        return False
    s = s.replace(",", ".")
    f = float(s)
    return int(f) if f.is_integer() else f


def num_or_default(v, default=None):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return default


def str_val(v, default=""):
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def read_sheet_params(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, COL_PARAM).value
        val = ws.cell(r, COL_VALUE).value
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        if val is None or str(val).strip() == "":
            continue
        try:
            out[key] = num(val)
        except (TypeError, ValueError):
            out[key] = str(val).strip()
    return out


def read_table(wb, sheet_name):
    """Nagłówek w wierszu 1, dane od wiersza 2."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        empty = True
        for j, col in enumerate(header):
            if not col:
                continue
            val = ws.cell(r, j + 1).value
            if val is not None and str(val).strip() != "":
                empty = False
            rec[str(col).strip()] = val
        if not empty:
            rows.append(rec)
    return rows


def coerce_wartosc(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return old
    if isinstance(old, (int, float)) and not isinstance(old, bool):
        try:
            return float(val) if isinstance(old, float) else int(round(float(val)))
        except (TypeError, ValueError):
            return old
    if isinstance(old, str):
        return str(val).strip()
    if old is None:
        return val
    return val


def export_bonusy(wb, data_dir, dry_run=False):
    rows = read_table(wb, SHEET_BONUSY)
    if not rows:
        return 0, "brak arkusza/wierszy"
    by_nation = {}
    for rec in rows:
        nacja = str_val(rec.get("Nacja"))
        if not nacja:
            continue
        typ = rec.get("Typ efektu")
        cel = rec.get("Cel")
        wartosc = rec.get("Wartosc")
        if typ is None and cel is None and wartosc is None:
            continue
        entry = {
            "typ": str_val(typ),
            "cel": str_val(cel),
            "wartosc": wartosc if wartosc is not None else "",
            "opis": str_val(rec.get("Opis")),
            "realizuje": str_val(rec.get("Realizuje")),
        }
        by_nation.setdefault(nacja, []).append(entry)

    path = os.path.join(data_dir, "civs.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    changed = 0
    for row in data.get("cywilizacje", []):
        name = row.get("Cywilizacja")
        if not name or name not in by_nation:
            continue
        new_bonusy = []
        old_bonusy = row.get("bonusy") or []
        for i, nb in enumerate(by_nation[name]):
            ob = old_bonusy[i] if i < len(old_bonusy) else {}
            merged = {
                "typ": nb["typ"] or ob.get("typ", ""),
                "cel": nb["cel"] or ob.get("cel", ""),
                "wartosc": coerce_wartosc(ob.get("wartosc"), nb["wartosc"]),
                "opis": nb["opis"] or ob.get("opis", ""),
                "realizuje": nb["realizuje"] or ob.get("realizuje", ""),
            }
            if merged != (ob if i < len(old_bonusy) else {}):
                changed += 1
            new_bonusy.append(merged)
        row["bonusy"] = new_bonusy

    if dry_run:
        return len(rows), f"bonusy: {len(by_nation)} nacji, {len(rows)} wierszy"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return len(rows), f"bonusy: {len(by_nation)} nacji, zmian={changed}"


def export_roster(wb, data_dir, dry_run=False):
    rows = read_table(wb, SHEET_ROSTER)
    if not rows:
        return 0, "brak arkusza/wierszy"
    xlsx_data = {}
    for rec in rows:
        cyw = str_val(rec.get("Cywilizacja"))
        if not cyw:
            continue
        nazwy = []
        for i in range(1, 11):
            v = rec.get(f"Klaster_{i:02d}")
            if v is None or str(v).strip() == "":
                return 0, f"brak Klaster_{i:02d} dla {cyw}"
            nazwy.append(str(v).strip())
        mnoznik_raw = rec.get("mnoznikHandelPieniadz")
        if mnoznik_raw is None:
            return 0, f"brak mnoznikHandelPieniadz dla {cyw}"
        ikona = str_val(rec.get("ikonaId"))
        if not ikona:
            return 0, f"brak ikonaId dla {cyw}"
        xlsx_data[cyw] = {
            "klastry": nazwy,
            "mnoznikHandelPieniadz": float(mnoznik_raw),
            "ikonaId": ikona,
        }

    path = os.path.join(data_dir, "civs.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    matched = 0
    for cyw_obj in data.get("cywilizacje", []):
        cyw_key = str_val(cyw_obj.get("Cywilizacja"))
        if cyw_key in xlsx_data:
            d = xlsx_data[cyw_key]
            cyw_obj["nazwyKlastra"] = d["klastry"]
            cyw_obj["mnoznikHandelPieniadz"] = d["mnoznikHandelPieniadz"]
            cyw_obj["ikonaId"] = d["ikonaId"]
            matched += 1

    if dry_run:
        return len(rows), f"roster: {matched}/{len(xlsx_data)} cywilizacji"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return len(rows), f"roster: {matched} cywilizacji"


def export_civ_params(wb, data_dir, dry_run=False):
    rows = read_table(wb, SHEET_PARAMS_CYW)
    if not rows:
        return 0, "brak arkusza/wierszy"
    out_rows = []
    for rec in rows:
        name = str_val(rec.get("Cywilizacja"))
        if not name:
            continue
        out_rows.append({
            "Cywilizacja": name,
            "preferowaneBudynki": str_val(rec.get("preferowaneBudynki")),
            "preferowaneJednostki": str_val(rec.get("preferowaneJednostki")),
            "modWzrostu": num_or_default(rec.get("modWzrostu"), 1.0),
            "modEkonomii": num_or_default(rec.get("modEkonomii"), 1.0),
            "uwagi": str_val(rec.get("uwagi")),
        })
    out = {"cywilizacje": out_rows, "_meta": {"zrodlo": "Panel-D.xlsx/Parametry-cyw"}}
    if dry_run:
        return len(out_rows), f"civ-params: {len(out_rows)} nacji"
    path = os.path.join(data_dir, "civ-params.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return len(out_rows), f"civ-params: {len(out_rows)} nacji"


def export_civ_ai(wb, data_dir, dry_run=False):
    rows = read_table(wb, SHEET_AI_NACJA)
    if not rows:
        return 0, "brak arkusza/wierszy"
    out_rows = []
    for rec in rows:
        name = str_val(rec.get("Cywilizacja"))
        if not name:
            continue
        entry = {"Cywilizacja": name}
        for key in (
            "agresywnosc", "ekspansywnosc", "priorytetMilitarny",
            "priorytetEkonomia", "priorytetNauka", "tolerancjaRyzyka",
            "sklonnoscDoPodboju",
        ):
            entry[key] = num_or_default(rec.get(key), 0)
        entry["profilMapy"] = str_val(rec.get("profilMapy"))
        entry["uwagi"] = str_val(rec.get("uwagi"))
        out_rows.append(entry)
    out = {"cywilizacje": out_rows, "_meta": {"zrodlo": "Panel-D.xlsx/AI-per-nacja"}}
    if dry_run:
        return len(out_rows), f"civ-ai: {len(out_rows)} nacji"
    path = os.path.join(data_dir, "civ-ai.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return len(out_rows), f"civ-ai: {len(out_rows)} nacji"


def export_dip_akcje(wb, data_dir, dry_run=False):
    rows = read_table(wb, SHEET_DIP_AKCJE)
    if not rows:
        return 0, "brak arkusza/wierszy"
    akcje = []
    for rec in rows:
        akcja = str_val(rec.get("Akcja"))
        if not akcja:
            continue
        akcje.append({k: (str_val(v) if v is not None else "") for k, v in rec.items()})
    path = os.path.join(data_dir, "diplomacy.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    data["akcje_dyplomatyczne"] = akcje
    if dry_run:
        return len(akcje), f"akcje_dyplomatyczne: {len(akcje)}"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return len(akcje), f"akcje_dyplomatyczne: {len(akcje)}"


def export_dip_per_nacja(wb, data_dir, dry_run=False):
    rows = read_table(wb, SHEET_DIP_PER)
    if not rows:
        return 0, "brak arkusza/wierszy"
    per = []
    for rec in rows:
        name = str_val(rec.get("Cywilizacja"))
        if not name:
            continue
        per.append({
            "Cywilizacja": name,
            "sklonnoscSojusze": num_or_default(rec.get("sklonnoscSojusze"), 0),
            "lojalnosc": num_or_default(rec.get("lojalnosc"), 0),
            "progWojny": num_or_default(rec.get("progWojny"), 0),
            "pamietliwosc": num_or_default(rec.get("pamietliwosc"), 0),
            "otwartoscHandel": num_or_default(rec.get("otwartoscHandel"), 0),
            "nastawienieBazowe": num_or_default(rec.get("nastawienieBazowe"), 50),
            "uwagi": str_val(rec.get("uwagi")),
        })
    path = os.path.join(data_dir, "diplomacy.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    data["perNacja"] = per
    if dry_run:
        return len(per), f"perNacja: {len(per)}"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return len(per), f"perNacja: {len(per)}"


def merge_json(path, updates, nested_key=None):
    with open(path, encoding="utf-8") as f:
        doc = json.load(f)
    if nested_key:
        base = doc.get(nested_key, {})
        base.update(updates)
        doc[nested_key] = base
    else:
        for k, v in updates.items():
            if k in doc and isinstance(doc[k], dict) and "wartosc" in doc[k]:
                doc[k]["wartosc"] = v
            else:
                doc[k] = {
                    "wartosc": v,
                    "opis": doc.get(k, {}).get("opis", k) if isinstance(doc.get(k), dict) else k,
                }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
        f.write("\n")


def export_hub(wb, data_dir, dry_run=False):
    dip = read_sheet_params(wb, "Dyplomacja")
    bar = read_sheet_params(wb, "Barbarzyńcy")
    ai_t = read_sheet_params(wb, "AI-trudnosc")
    ai_a = read_sheet_params(wb, "AI-archetyp")
    ai_z = read_sheet_params(wb, "AI-zachowanie")
    bar.pop("epoka_cutoff_barbarzy", None)
    ai_updates = {**ai_t, **ai_a, **ai_z, **{k: v for k, v in bar.items() if k.startswith("barbarzyncy_")}}

    if dry_run:
        return len(dip) + len(ai_updates), f"hub: dip={len(dip)} ai/bar={len(ai_updates)}"

    dip_path = os.path.join(data_dir, "diplomacy.json")
    merge_json(dip_path, dip, "params")
    ai_path = os.path.join(data_dir, "ai-params.json")
    merge_json(ai_path, ai_updates)
    return len(dip) + len(ai_updates), f"hub: dip={len(dip)} ai/bar={len(ai_updates)}"


def run_tool(script):
    p = os.path.join(TOOLS, script)
    if not os.path.isfile(p):
        print(f"SKIP (brak): {script}")
        return
    print(f"--- legacy fallback: {script} ---")
    subprocess.check_call([sys.executable, p], cwd=os.path.join(ROOT, "gra"))


def export_sheet_or_fallback(wb, sheet, export_fn, data_dir, dry_run, full):
    if sheet in wb.sheetnames:
        n, msg = export_fn(wb, data_dir, dry_run)
        print(f"OK {sheet}: {msg}")
        return n
    if full and LEGACY_FALLBACK.get(sheet):
        if not dry_run:
            run_tool(LEGACY_FALLBACK[sheet])
        print(f"FALLBACK {sheet}: {LEGACY_FALLBACK[sheet]}")
        return 0
    print(f"SKIP {sheet}: brak arkusza")
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--full",
        action="store_true",
        help="DEPRECATED — fallback do gra/tools/export-*.py gdy brak arkusza w panelu",
    )
    ap.add_argument("--xlsx", default=PANEL, help="Ścieżka do Panel-D.xlsx")
    ap.add_argument("--data-dir", default=DATA, help="Katalog gra/data (domyślnie kanon)")
    args = ap.parse_args()

    if args.full:
        print("UWAGA: --full jest DEPRECATED; domyślny eksport czyta wyłącznie Panel-D.xlsx.")

    if not os.path.isfile(args.xlsx):
        sys.exit(f"Brak {args.xlsx} — najpierw: python panele-sterowania/gen-panel-d.py")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    stats = {}

    _, hub_msg = export_hub(wb, args.data_dir, args.dry_run)
    print(f"OK hub: {hub_msg}")

    stats["bonusy"] = export_sheet_or_fallback(
        wb, SHEET_BONUSY, export_bonusy, args.data_dir, args.dry_run, args.full)
    stats["roster"] = export_sheet_or_fallback(
        wb, SHEET_ROSTER, export_roster, args.data_dir, args.dry_run, args.full)
    stats["civ_params"] = export_sheet_or_fallback(
        wb, SHEET_PARAMS_CYW, export_civ_params, args.data_dir, args.dry_run, args.full)
    stats["civ_ai"] = export_sheet_or_fallback(
        wb, SHEET_AI_NACJA, export_civ_ai, args.data_dir, args.dry_run, args.full)
    stats["dip_akcje"] = export_sheet_or_fallback(
        wb, SHEET_DIP_AKCJE, export_dip_akcje, args.data_dir, args.dry_run, args.full)
    stats["dip_per"] = export_sheet_or_fallback(
        wb, SHEET_DIP_PER, export_dip_per_nacja, args.data_dir, args.dry_run, args.full)

    if args.dry_run:
        print(f"[dry-run] Panel-D: bonusy={stats['bonusy']} roster={stats['roster']} "
              f"params={stats['civ_params']} ai={stats['civ_ai']} akcje={stats['dip_akcje']} per={stats['dip_per']}")
    else:
        print("Eksport Panel-D zakonczony.")


if __name__ == "__main__":
    main()
