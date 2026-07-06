#!/usr/bin/env python3
# export-c.py — Panel-C.xlsx → gra/data (Grupa C only, bez export-data.py / legacy pipeline)
# Uruchom: python panele-sterowania/export-c.py
import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    raise SystemExit("Brak openpyxl: pip install openpyxl")

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_DATA = os.path.join(ROOT, "gra", "data")
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "Panel-C.xlsx")

STAT_FIELDS = {
    "MELEEATTACK": "meleeAttack",
    "MELEEDEFENCE": "meleeDefence",
    "MELEEDEFENSE": "meleeDefence",
    "WEAPONDAMAGE": "weaponDamage",
    "PIERCING": "piercing",
    "ARMOR": "armor",
    "CHARGEBONUS": "chargeBonus",
    "HEALTH": "health",
    "MISSILEATTACK": "missileAttack",
    "WALLATTACK": "wallAttack",
    "MORALE-BAZOWE": "Morale bazowe",
    # legacy IDs (stary Panel-C)
    "ATAK": "meleeAttack", "OBRONA": "meleeDefence",
    "OBRAENIA": "weaponDamage", "OBRAZENIA": "weaponDamage", "OBRAŻENIA": "weaponDamage",
    "PRZEBICIE": "piercing", "PANCERZ": "armor", "UDERZENIE": "chargeBonus",
    "ATAK-DYSTANSOWY": "missileAttack",
}

PL_FIELD_TO_JSON = {
    "Atak": "meleeAttack",
    "Obrona": "meleeDefence",
    "Obrażenia": "weaponDamage",
    "Przebicie": "piercing",
    "Pancerz": "armor",
    "Szarża": "chargeBonus",
    "Uderzenie": "chargeBonus",
    "Zdrowie": "health",
    "Health": "health",
    "Atak dystansowy": "missileAttack",
    "Atak vs Mur": "wallAttack",
    "Morale bazowe": "Morale bazowe",
}
COST_FIELDS = {
    "PIENIADZ-KOSZT": "Pieniądz (koszt)", "PIENIĄDZ-KOSZT": "Pieniądz (koszt)",
    "LUDNOSC": "Ludność", "LUDNOŚĆ": "Ludność",
    "UTRZYMANIE-PIENIADZ-TURE": "Utrzymanie (Pieniądz/turę)",
    "UTRZYMANIE-PIENIĄDZ-TURĘ": "Utrzymanie (Pieniądz/turę)",
    "ZYWNOŚĆ-TURĘ": "żywność/turę", "ZYWNOŚC-TURE": "żywność/turę",
}

STALE_MAP = {
    "C-TW-HIT-BASE": ("tw_v3", "hit_base"),
    "C-TW-HIT-MIN": ("tw_v3", "hit_min"),
    "C-TW-HIT-MAX": ("tw_v3", "hit_max"),
    "C-TW-MAX-RND": ("tw_v3", "max_rounds"),
    "C-COUNTER-MULT": (None, "counter_multiplier"),
    "C-RIVER-ATK": (None, "river_attack_mult"),
}

MOC_MAP = {
    "C-MOC-CHARGE-DIV": "charge_divisor",
    "C-MOC-MS-DIV": "missile_divisor",
    "C-MOC-HP-FIELD": "hp_field_divisor",
    "C-MOC-HP-SIEGE": "hp_siege_divisor",
}

OBL_MAP = {
    "C-SIEGE-WALL-BASE": "wall_base_obrona",
    "C-SIEGE-WALL-LVL": "wall_per_level_obrona",
    "C-SIEGE-WALL-MAX": "wall_max_level",
    "C-SIEGE-WALL-PANC": "wall_pancerz_fraction",
    "C-SIEGE-HILL": "hill_defense_mult",
    "C-SIEGE-MTN": "mountain_defense_mult",
    "C-SIEGE-FORT": "fortify_obrona_bonus",
    "C-SIEGE-MIL-POP": "militia_pop_fraction",
    "C-SIEGE-MIL-STR": "militia_strength_fraction",
    "C-SIEGE-MAX-RND": "siege_max_rounds",
}

SAI_MAP = {
    "C-SAI-T1": "t1_assault_ratio",
    "C-SAI-T2": "t2_build_min_ratio",
    "C-SAI-T3": "t3_starve_min_ratio",
    "C-SAI-WAIT": "t2_max_wait_turns",
    "C-SAI-MACH": "units_per_extra_machine",
}

# Auto-walka na mocy M (v2b) → gra/data/auto-battle-params.json
# Maciej kręci tylko COEF-WIN i COEF-LOSE w Panel-C (arkusz Auto-walka).
AUTO_BATTLE_MAP = {
    "C-AUTO-COEF-WIN": "coef_zwyciezca",
    "C-AUTO-COEF-LOSE": "coef_przegrany",
    "C-AUTO-L-MAX": "L_MAX",
    "C-AUTO-P-ATK": "p_atk",
    "C-AUTO-P-DEF": "p_def",
    "C-AUTO-P-EXP": "p",
    "C-AUTO-REMIS": "remis_pct",
}


def coerce_num(val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return int(val) if float(val).is_integer() else float(val)
    try:
        f = float(str(val).replace(",", "."))
        return int(f) if f.is_integer() else f
    except ValueError:
        return val


def parse_unit_field_from_id(pid):
    if not pid or not str(pid).startswith("C-U-"):
        return None, None
    parts = str(pid).split("-")
    if len(parts) < 4:
        return None, None
    field_key = parts[-1]
    unit_slug = "-".join(parts[2:-1])
    field = STAT_FIELDS.get(field_key.upper())
    return unit_slug, field


def parse_unit_field_from_param(param):
    if not param or " · " not in str(param):
        return None, None
    unit, field = str(param).split(" · ", 1)
    field = field.strip()
    json_field = PL_FIELD_TO_JSON.get(field, field)
    return unit.strip(), json_field


def slug(name):
    return re.sub(r"[^a-zA-Z0-9]+", "-", name.upper()).strip("-")[:40]


def export_stale(ws, params):
    changed = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        val = coerce_num(ws.cell(r, 4).value)
        if pid not in STALE_MAP or val is None:
            continue
        sec, key = STALE_MAP[pid]
        if sec:
            if params[sec][key] != val:
                params[sec][key] = val
                changed += 1
        else:
            if params[key] != val:
                params[key] = val
                changed += 1
    return changed


def export_stale_moc(ws, params):
    changed = 0
    if "unit_power" not in params:
        params["unit_power"] = {}
    up = params["unit_power"]
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        val = coerce_num(ws.cell(r, 4).value)
        if pid not in MOC_MAP or val is None:
            continue
        key = MOC_MAP[pid]
        if up.get(key) != val:
            up[key] = val
            changed += 1
    return changed


def export_auto_walka(ws, out_path):
    """Panel-C arkusz Auto-walka → auto-battle-params.json (sekcja straty)."""
    if os.path.isfile(out_path):
        with open(out_path, encoding="utf-8") as f:
            doc = json.load(f)
    else:
        doc = {
            "_opis": "Auto-walka v2b — eksport z Panel-C.xlsx (arkusz Auto-walka).",
            "kalibracja": "v2b-2026-06-30",
            "straty": {},
        }
    straty = doc.setdefault("straty", {})
    changed = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        val = coerce_num(ws.cell(r, 4).value)
        if pid not in AUTO_BATTLE_MAP or val is None:
            continue
        key = AUTO_BATTLE_MAP[pid]
        if straty.get(key) != val:
            straty[key] = val
            changed += 1
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return changed


def sync_units_power_cache(units):
    import sys
    root = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
    tools = os.path.join(root, "gra", "tools")
    if tools not in sys.path:
        sys.path.insert(0, tools)
    from unit_power import apply_power_cache
    return apply_power_cache(units)


def export_oblezenie(ws, params):
    changed = 0
    obl = params["oblężenie"]
    if "siege_ai" not in params:
        params["siege_ai"] = {}
    sai = params["siege_ai"]
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        val = coerce_num(ws.cell(r, 4).value)
        if val is None:
            continue
        if pid in OBL_MAP:
            key = OBL_MAP[pid]
            if obl[key] != val:
                obl[key] = val
                changed += 1
        elif pid in SAI_MAP:
            key = SAI_MAP[pid]
            if sai.get(key) != val:
                sai[key] = val
                changed += 1
    return changed


def export_jednostki_stats(ws, units):
    by_name = {u["Jednostka"]: u for u in units}
    by_slug = {slug(u["Jednostka"]): u for u in units}
    changed = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        param = ws.cell(r, 2).value
        val = coerce_num(ws.cell(r, 4).value)
        if val is None:
            continue
        unit_slug, field = parse_unit_field_from_id(pid)
        if not field:
            uname, field = parse_unit_field_from_param(param)
            u = by_name.get(uname) if uname else None
        else:
            u = by_slug.get(unit_slug)
        if not u or not field:
            continue
        if u.get(field) != val:
            u[field] = val
            changed += 1
    return changed


def export_macierz(ws, units):
    """Macierz-jednostek — wszystkie jednostki, pola meta (Widok pola, Ruch…)."""
    by_name = {u["Jednostka"]: u for u in units}
    changed = 0
    for r in range(2, ws.max_row + 1):
        param = ws.cell(r, 2).value
        val = ws.cell(r, 4).value
        if val is None or (isinstance(val, str) and str(val).strip() == ""):
            continue
        uname, field = parse_unit_field_from_param(param)
        u = by_name.get(uname) if uname else None
        if not u or field not in u:
            continue
        if field in ("Epoka", "Tech", "Rola (linia)"):
            nv = str(val).strip()
        else:
            nv = coerce_num(val)
        if u.get(field) != nv:
            u[field] = nv
            changed += 1
    return changed


def export_koszty(ws, units):
    by_name = {u["Jednostka"]: u for u in units}
    changed = 0
    for r in range(2, ws.max_row + 1):
        param = ws.cell(r, 2).value
        val = coerce_num(ws.cell(r, 4).value)
        if val is None or not param:
            continue
        uname, field = parse_unit_field_from_param(param)
        u = by_name.get(uname)
        if not u or field not in u:
            continue
        if u[field] != val:
            u[field] = val
            changed += 1
    return changed


def export_countery(ws, counters):
    changed = 0
    for r in range(2, ws.max_row + 1):
        idx = r - 2
        if idx >= len(counters):
            break
        val = ws.cell(r, 4).value
        if val is None:
            continue
        bonus = str(val).strip()
        if not bonus.startswith(("+", "-")):
            bonus = f"+{bonus}%" if "%" not in str(val) else str(val)
        elif "%" not in bonus:
            bonus = f"{bonus}%"
        if counters[idx].get("Bonus") != bonus:
            counters[idx]["Bonus"] = bonus
            changed += 1
    return changed


def export_teren(ws, terrain):
    changed = 0
    for r in range(2, ws.max_row + 1):
        idx = r - 2
        if idx >= len(terrain):
            break
        val = coerce_num(ws.cell(r, 4).value)
        if val is None:
            continue
        bonus = f"+{int(val)}%" if val >= 0 else f"{int(val)}%"
        if terrain[idx].get("Bonus Obrona") != bonus:
            terrain[idx]["Bonus Obrona"] = bonus
            changed += 1
    return changed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--data-dir", default=DEFAULT_DATA)
    args = ap.parse_args()
    if not os.path.isfile(args.xlsx):
        print("Brak Panel-C.xlsx — uruchom gen-panel-c.py", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    units_path = os.path.join(args.data_dir, "units.json")
    params_path = os.path.join(args.data_dir, "combat-params.json")
    counters_path = os.path.join(args.data_dir, "counters.json")
    terrain_path = os.path.join(args.data_dir, "terrain-combat.json")
    with open(units_path, encoding="utf-8") as f:
        units = json.load(f)
    with open(params_path, encoding="utf-8") as f:
        params = json.load(f)
    with open(counters_path, encoding="utf-8") as f:
        counters = json.load(f)
    with open(terrain_path, encoding="utf-8") as f:
        terrain = json.load(f)
    c1 = export_stale(wb["Stale-walki"], params) if "Stale-walki" in wb.sheetnames else 0
    c1b = export_stale_moc(wb["Stale-moc"], params) if "Stale-moc" in wb.sheetnames else 0
    c2 = export_oblezenie(wb["Oblezenie"], params) if "Oblezenie" in wb.sheetnames else 0
    c3 = export_jednostki_stats(wb["Jednostki-staty"], units) if "Jednostki-staty" in wb.sheetnames else 0
    c3b = export_macierz(wb["Macierz-jednostek"], units) if "Macierz-jednostek" in wb.sheetnames else 0
    c4 = export_koszty(wb["Koszty-jednostek"], units) if "Koszty-jednostek" in wb.sheetnames else 0
    c5 = export_countery(wb["Countery"], counters) if "Countery" in wb.sheetnames else 0
    c6 = export_teren(wb["Teren-walka"], terrain) if "Teren-walka" in wb.sheetnames else 0
    auto_path = os.path.join(args.data_dir, "auto-battle-params.json")
    c8 = (
        export_auto_walka(wb["Auto-walka"], auto_path)
        if "Auto-walka" in wb.sheetnames
        else 0
    )
    c7 = sync_units_power_cache(units)
    with open(units_path, "w", encoding="utf-8") as f:
        json.dump(units, f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(params_path, "w", encoding="utf-8") as f:
        json.dump(params, f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(counters_path, "w", encoding="utf-8") as f:
        json.dump(counters, f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(terrain_path, "w", encoding="utf-8") as f:
        json.dump(terrain, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"export-c OK: stale={c1} stale_moc={c1b} oblezenie={c2} staty={c3} macierz={c3b} koszty={c4} countery={c5} teren={c6} auto_walka={c8} moc_cache={c7}")


if __name__ == "__main__":
    main()
