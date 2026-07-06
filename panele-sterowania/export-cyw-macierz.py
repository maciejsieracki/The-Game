#!/usr/bin/env python3
"""
export-cyw-macierz.py — 11 arkuszy Cyw-01..11 -> gra/data/civ-matrix.json
(Potega/Power: CZEKA — nie w macierzy per-cyw)

Szuka arkuszy w Panel-D.xlsx lub Cyw-macierz-REVIEW.xlsx.

  python panele-sterowania/export-cyw-macierz.py
  python panele-sterowania/export-cyw-macierz.py --dry-run
"""
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, ".."))
PANEL = os.path.join(HERE, "Panel-D.xlsx")
REVIEW = os.path.join(HERE, "Cyw-macierz-REVIEW.xlsx")
OUT = os.path.join(ROOT, "gra", "data", "civ-matrix.json")

sys.path.insert(0, HERE)
from cyw_macierz_schema import (
    PARAM_IDS, PARAM_META, DEFAULTS, IDENT_COLS,
    DOMAIN_SHEETS, SHEET_TO_DOMAIN, SHEET_LEGACY, RESERVED_DOMAINS,
)


def num(v, default=0):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v) if isinstance(v, float) and v == int(v) else v
    try:
        f = float(str(v).replace(",", "."))
        return int(f) if f == int(f) else f
    except ValueError:
        return default


def pick_workbook():
    for path in (PANEL, REVIEW):
        if os.path.isfile(path):
            return path
    sys.exit("Brak Panel-D.xlsx ani Cyw-macierz-REVIEW.xlsx")


def read_sheet(ws, param_ids):
    """Wiersze 5+ = dane; wiersz 1 = naglowki param_id."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h else "" for h in headers]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    rows_by_name = {}
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, idx.get("Cywilizacja", 1)).value
        if not name:
            continue
        name = str(name).strip()
        entry = {
            "Cywilizacja": name,
            "typCywilizacji": str(ws.cell(r, idx.get("typCywilizacji", 2)).value or "").strip(),
            "ikonaId": str(ws.cell(r, idx.get("ikonaId", 3)).value or "").strip(),
            "tier": num(ws.cell(r, idx.get("tier", 4)).value, 1),
            "params": {},
        }
        for pid in param_ids:
            if pid in idx:
                entry["params"][pid] = num(ws.cell(r, idx[pid]).value, DEFAULTS.get(pid, 0))
        rows_by_name[name] = entry
    return rows_by_name


def merge_from_workbook(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s.startswith("Cyw-") and s not in ("Cyw-00-INFO", SHEET_LEGACY)]

    if not sheets:
        # fallback: stara jedna zakladka
        if SHEET_LEGACY in wb.sheetnames:
            ws = wb[SHEET_LEGACY]
            from cyw_macierz_schema import PARAM_IDS as ALL
            partial = read_sheet(ws, ALL)
            return list(partial.values())
        sys.exit(f"Brak arkuszy Cyw-01..12 w {path}")

    merged = {}
    for sheet_name in sorted(sheets):
        dom = SHEET_TO_DOMAIN.get(sheet_name)
        if not dom:
            continue
        if dom in RESERVED_DOMAINS:
            continue  # Potega: placeholder w Excelu, bez JSON do czasu Power
        from cyw_macierz_schema import PARAMS_BY_DOMAIN
        pids = PARAMS_BY_DOMAIN.get(dom, [])
        partial = read_sheet(wb[sheet_name], pids)
        for name, entry in partial.items():
            if name not in merged:
                merged[name] = {
                    "Cywilizacja": entry["Cywilizacja"],
                    "typCywilizacji": entry["typCywilizacji"],
                    "ikonaId": entry["ikonaId"],
                    "tier": entry["tier"],
                    "params": {pid: DEFAULTS.get(pid, 0) for pid in PARAM_IDS},
                }
            merged[name]["params"].update(entry["params"])
            # ident z pierwszego napotkanego arkusza (meta)
            if entry["typCywilizacji"]:
                merged[name]["typCywilizacji"] = entry["typCywilizacji"]
            if entry["ikonaId"]:
                merged[name]["ikonaId"] = entry["ikonaId"]

    order = [
        "Grecy", "Rzymianie", "Chińczycy", "Inkowie", "Zulusi", "Egipt", "Sumerowie",
        "Celtowie", "Germanie", "Harappa", "Hetyci", "Słowianie", "Babilonia", "Asyria", "Fenicjanie",
    ]
    out = []
    for name in order:
        if name in merged:
            out.append(merged[name])
    for name in sorted(merged.keys()):
        if name not in order:
            out.append(merged[name])
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--xlsx", default=None, help="sciezka do xlsx (domyslnie Panel-D lub REVIEW)")
    args = ap.parse_args()

    path = args.xlsx or pick_workbook()
    cywilizacje = merge_from_workbook(path)

    payload = {
        "_meta": {
            "version": "1.0",
            "source": f"{os.path.basename(path)} (12 arkuszy Cyw)",
            "formula_mul_proc": "wynik = baza * (1 + wartosc)",
            "formula_mul_abs": "wynik = baza * wartosc",
            "kolumny": len(PARAM_IDS),
            "cywilizacje": len(cywilizacje),
            "reservedDomains": list(RESERVED_DOMAINS),
            "uwaga_potega": "Cyw-12-POTEGA w Excelu; JSON/silnik po domknieciu modulu Power",
        },
        "paramDefs": PARAM_META,
        "defaults": DEFAULTS,
        "cywilizacje": cywilizacje,
    }

    if args.dry_run:
        print(f"DRY-RUN: {path} -> {len(cywilizacje)} cyw, {len(PARAM_IDS)} params")
        return

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"OK -> {OUT} (z {path})")


if __name__ == "__main__":
    main()
