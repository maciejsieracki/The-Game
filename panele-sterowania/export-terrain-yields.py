#!/usr/bin/env python3
# export-terrain-yields.py — Panel-A.xlsx (arkusze Teren-bazowy + Bonusy-nakladki) → gra/data/terrain-yields.json
# Uruchom: python panele-sterowania/export-terrain-yields.py
# Maciej nie używa terminala — agent odpala po "eksportuj plony terenu".
# SYNC-PANELI [17:05]: plony terenu scalone do Panel-A (dawniej osobny Panel-A-Plony-Terenu.xlsx → archiwum/).
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "Panel-A.xlsx")
YIELDS_PATH = os.path.join(DATA, "terrain-yields.json")

EDITABLE_FIELDS = ("Żywność", "Praca", "Podatek", "Uwagi")
COL_KEY = 1
COL_START = 2
COL_UWAGI = 5
DATA_ROW = 3


def _num_equal(a, b):
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return False


def coerce(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return old, False
    if isinstance(old, int):
        try:
            nv = int(round(float(val)))
            return nv, not _num_equal(nv, old)
        except (TypeError, ValueError):
            return old, False
    if old is None:
        if isinstance(val, (int, float)):
            return val, True
        s = str(val).strip()
        return (s if s else None), s != (old if old is not None else "")
    nv = str(val).strip()
    old_s = "" if old is None else str(old)
    return nv if nv else None, nv != old_s


def _podatek_val(row):
    return row.get("Podatek", row.get("Handel", 0))


def recalc_suma(row):
    if "Suma" not in row:
        return False
    total = sum(int(row.get(f, 0) or 0) for f in ("Żywność", "Praca")) + int(_podatek_val(row) or 0) + sum(
        int(row.get(f, 0) or 0) for f in ("Drewno", "Kamień")
    )
    if row["Suma"] == total:
        return False
    row["Suma"] = total
    return True


def read_sheet_rows(ws, key_header):
    rows = {}
    for r in range(DATA_ROW, ws.max_row + 1):
        key = ws.cell(r, COL_KEY).value
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        rows[key] = {
            "Żywność": ws.cell(r, COL_START).value,
            "Praca": ws.cell(r, COL_START + 1).value,
            "Podatek": ws.cell(r, COL_START + 2).value,
            "Uwagi": ws.cell(r, COL_UWAGI).value,
        }
    return rows


def overlay_section(json_rows, excel_rows, key_field):
    changed = 0
    for row in json_rows:
        key = row.get(key_field)
        if not key or key not in excel_rows:
            continue
        src = excel_rows[key]
        row_changed = False
        for field in EDITABLE_FIELDS:
            old = row.get(field, _podatek_val(row) if field == "Podatek" else None)
            src_field = field
            nv, did = coerce(old, src.get(src_field, src.get("Handel") if field == "Podatek" else None))
            if did:
                row[field] = nv
                row_changed = True
                changed += 1
        if row_changed and recalc_suma(row):
            changed += 1
    return changed


def save_json(obj, path):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
        f.write("\n")
    json.load(open(tmp, encoding="utf-8"))
    os.replace(tmp, path)


def main():
    ap = argparse.ArgumentParser(description="Eksport Panel-A-Plony-Terenu.xlsx → terrain-yields.json")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"Brak pliku: {args.xlsx}")
        print("Uruchom najpierw: python panele-sterowania/generate-terrain-yields-xlsx.py")
        sys.exit(1)
    if not os.path.isfile(YIELDS_PATH):
        print(f"Brak pliku: {YIELDS_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if "Teren-bazowy" not in wb.sheetnames or "Bonusy-nakladki" not in wb.sheetnames:
        print("Brak wymaganych arkuszy: Teren-bazowy, Bonusy-nakladki")
        sys.exit(1)

    data = json.load(open(YIELDS_PATH, encoding="utf-8"))
    base_rows = read_sheet_rows(wb["Teren-bazowy"], "Teren")
    mod_rows = read_sheet_rows(wb["Bonusy-nakladki"], "Modyfikator")

    ch_base = overlay_section(data.get("terrain_types", []), base_rows, "Teren")
    ch_mod = overlay_section(data.get("terrain_modifiers", []), mod_rows, "Modyfikator")
    total = ch_base + ch_mod

    print(f"terrain_types: {ch_base} zmian")
    print(f"terrain_modifiers: {ch_mod} zmian")
    print(f"RAZEM: {total} zmian")

    if total and not args.dry_run:
        save_json(data, YIELDS_PATH)
        print(f"Zapisano: {YIELDS_PATH}")
    elif args.dry_run:
        print("(dry-run — JSON nie zapisany)")
    elif total == 0:
        print("(brak zmian — puste lub identyczne z JSON)")


if __name__ == "__main__":
    main()
