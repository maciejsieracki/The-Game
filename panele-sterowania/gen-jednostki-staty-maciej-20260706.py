#!/usr/bin/env python3
"""
gen-jednostki-staty-maciej-20260706.py
Źródło listy: gra-robocza/JEDNOSTKI-DO-POPRAWY-staty.md
Wyjście: panele-sterowania/Jednostki-staty-MACIEJ-20260706.xlsx
"""
import json
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    sys.exit("pip install openpyxl")

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
UNITS_PATH = os.path.join(ROOT, "gra", "data", "units.json")
OUT = os.path.join(os.path.dirname(__file__), "Jednostki-staty-MACIEJ-20260706.xlsx")

TW_COLS = [
    "meleeAttack", "meleeDefence", "weaponDamage", "piercing", "armor",
    "chargeBonus", "health", "missileAttack", "wallAttack", "Morale bazowe",
]

# (ref_name, historia, propozycja TW dict, uzasadnienie)
# Propozycje: skala TW v3 jak Gaesatae/Hastati/Konnica; health ≈ round(Health_json/5)
SPECS = {
    "Soldurii": (
        "Gaesatae",
        "Elitarna gwardia wodza celtyckiego (II–I w. p.n.e.); miecz + tarcza + torc; identyczny profil jak Gaesatae (CELT-Q2).",
        {"meleeAttack": 7, "meleeDefence": 2, "weaponDamage": 5, "piercing": 1, "armor": 1, "chargeBonus": 5, "health": 11, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 100},
        "Kopia Gaesatae 1:1 — brak pól EN w JSON; uwagi jednostki.",
    ),
    "Konnica lancowa asyryjska": (
        "Konnica",
        "Konnica z kopją (VII–VI w. p.n.e.); elitarna kawaleria Nowoassyryjska; szarża frontalna, słabsza w zwarciu.",
        {"meleeAttack": 9, "meleeDefence": 4, "weaponDamage": 8, "piercing": 7, "armor": 3, "chargeBonus": 12, "health": 17, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 75},
        "Konnica +1 AP, +2 szarża; mniejsza OBR niż standard (lance charge, słabe zatrzymanie).",
    ),
    "Konnica łucznicza asyryjska": (
        "Konnica",
        "Łucznicy konni Asyrii; hit-and-run, osłabiacz przed szarżą lanc.",
        {"meleeAttack": 4, "meleeDefence": 2, "weaponDamage": 3, "piercing": 2, "armor": 2, "chargeBonus": 6, "health": 15, "missileAttack": 4, "wallAttack": 0, "Morale bazowe": 70},
        "Słabe zwarcie, AD jak lekki łucznik konny; HP jak konnica lekka.",
    ),
    "Łucznik asyryjski": (
        "Łucznik",
        "Kompozytowy łuk (Brąz); piechota dystansowa państwa; tarcza + krótki miecz w zwarciu.",
        {"meleeAttack": 2, "meleeDefence": 3, "weaponDamage": 2, "piercing": 1, "armor": 2, "chargeBonus": 0, "health": 8, "missileAttack": 4, "wallAttack": 0, "Morale bazowe": 55},
        "Między Procarzem a Łucznikiem; lepsza OBR niż proca (tarcza).",
    ),
    "Drużynnik": (
        "Gaesatae",
        "Słowiański drużynnik (VII–X w.); woj drużyny, tarcza + włócznia/miecz; bardziej defensywny niż najemnik celtycki.",
        {"meleeAttack": 7, "meleeDefence": 3, "weaponDamage": 5, "piercing": 2, "armor": 2, "chargeBonus": 4, "health": 11, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 80},
        "Jak Gaesatae +1 OBR (formacja drużyny), −1 szarża.",
    ),
    "Jeździec z szczepnikami": (
        "Konnica",
        "Lekka jazda słowiańska; oszczepy przed szarżą (AD niski), unik zwarci.",
        {"meleeAttack": 6, "meleeDefence": 3, "weaponDamage": 5, "piercing": 4, "armor": 2, "chargeBonus": 8, "health": 14, "missileAttack": 2, "wallAttack": 0, "Morale bazowe": 75},
        "Konnica obniżona; mały AD (oszczepy) z PL AD=2–5.",
    ),
    "Strażnik bram Harappy": (
        "Włócznik",
        "Garnizon bram Mohenjo-daro (Brąz); włócznia, tarcza, profil anty-konna.",
        {"meleeAttack": 4, "meleeDefence": 7, "weaponDamage": 4, "piercing": 3, "armor": 5, "chargeBonus": 2, "health": 14, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 70},
        "Defensywny włócznik; PL O=9 → wysoka OBR TW.",
    ),
    "Piechota induska": (
        "Wojownik z mieczem i tarczą",
        "Standardowa piechota Doliny Indus (Brąz); brak zbroi płytowej, tarcza.",
        {"meleeAttack": 6, "meleeDefence": 5, "weaponDamage": 5, "piercing": 2, "armor": 3, "chargeBonus": 3, "health": 12, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 60},
        "≈ Wojownik z mieczem; lekko niższe HP (mniej zbroi).",
    ),
    "Garnizon Harappy": (
        "Włócznik",
        "Elitarny garnizon miejski (Żelazo); doświadczeni obrońcy.",
        {"meleeAttack": 5, "meleeDefence": 8, "weaponDamage": 4, "piercing": 3, "armor": 5, "chargeBonus": 2, "health": 13, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 72},
        "Profil garnizonu: Włócznik +1 OBR, −1 AP.",
    ),
    "Rydwan Kapadokijski": (
        "Rydwan celtycki",
        "Rydwan hetycki/kapadocki (Brąz); 2 konie, łucznik + woźnica; szybki, kruchy.",
        {"meleeAttack": 7, "meleeDefence": 3, "weaponDamage": 6, "piercing": 3, "armor": 2, "chargeBonus": 10, "health": 18, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 85},
        "Jak Rydwan celtycki; więcej HP (2 woły w wersji standard).",
    ),
    "Piechota hetycka": (
        "Włócznik",
        "Piechota hetycka (Brąz); włócznie, tarcze, formacja górska.",
        {"meleeAttack": 5, "meleeDefence": 7, "weaponDamage": 4, "piercing": 3, "armor": 4, "chargeBonus": 3, "health": 13, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 65},
        "Defensywna piechota Brązu; PL 7/8.",
    ),
    "Gwardia hetycka": (
        "Hastati",
        "Gwardia królewska Hetytów (Żelazo); elitarna piechota pałacowa.",
        {"meleeAttack": 7, "meleeDefence": 7, "weaponDamage": 6, "piercing": 3, "armor": 6, "chargeBonus": 5, "health": 14, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 80},
        "Poniżej Hastati (brak pilum); elitarna OBR.",
    ),
    "Gwardia Ishtar": (
        "Evocati",
        "Gwardia świątynna Babilonu (Brąz); elitarna piechota sakralna.",
        {"meleeAttack": 8, "meleeDefence": 7, "weaponDamage": 7, "piercing": 3, "armor": 5, "chargeBonus": 4, "health": 15, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 85},
        "Między Evocati a Hastati; PL 9/8 ofensywno-defensywna.",
    ),
    "Wojownik babiloński": (
        "Wojownik z mieczem i tarczą",
        "Piechota babilońska (Brąz); tarcza, krótki miecz/kopis.",
        {"meleeAttack": 6, "meleeDefence": 5, "weaponDamage": 5, "piercing": 2, "armor": 3, "chargeBonus": 3, "health": 11, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 58},
        "≈ Wojownik z mieczem (standard Brąz).",
    ),
    "Piechota neobabilońska": (
        "Hastati",
        "Neobabilońska piechota (Żelazo); zbroja lamelkowa, tarcza.",
        {"meleeAttack": 6, "meleeDefence": 7, "weaponDamage": 5, "piercing": 3, "armor": 6, "chargeBonus": 4, "health": 13, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 70},
        "Defensywna piechota Żelaza; bez pilum.",
    ),
    "Tyrski miecznik": (
        "Gaesatae",
        "Elitarny miecznik Tyru (Żelazo); ofensywny, najemniczy charakter.",
        {"meleeAttack": 7, "meleeDefence": 3, "weaponDamage": 6, "piercing": 2, "armor": 2, "chargeBonus": 6, "health": 11, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 75},
        "Ofensywniejszy Gaesatae; fenicki handel = dobre morale.",
    ),
    "Wojownik fenicki": (
        "Wojownik",
        "Lekka piechota fenicka (Brąz); tarcza, włócznia; morska kolonia.",
        {"meleeAttack": 5, "meleeDefence": 4, "weaponDamage": 4, "piercing": 1, "armor": 2, "chargeBonus": 2, "health": 10, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 55},
        "Poniżej Wojownika z mieczem; tania piechota.",
    ),
    "Gwardia Tyr": (
        "Evocati",
        "Gwardia miejska Tyru (Żelazo); elitarna ochrona magistratu.",
        {"meleeAttack": 7, "meleeDefence": 6, "weaponDamage": 6, "piercing": 2, "armor": 4, "chargeBonus": 4, "health": 12, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 78},
        "Elitarna, ale lżejsza niż Gwardia Ishtar (miasto-handlowe).",
    ),
    "Thorakites": (
        "Falanga",
        "Grecki hoplita z tarczą (Thorax, V–III w. p.n.e.); profil defensywny.",
        {"meleeAttack": 5, "meleeDefence": 7, "weaponDamage": 5, "piercing": 2, "armor": 5, "chargeBonus": 3, "health": 13, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 65},
        "Korekta: PL O=9 → MD 7 (obecne EN 6/5 za słabe w obronie).",
    ),
    "Legionarius": (
        "Wojownik z mieczem i tarczą",
        "Wczesny legionariusz (Brąz/Republica); scutum + gladius przed reformą Mariusza.",
        {"meleeAttack": 6, "meleeDefence": 6, "weaponDamage": 6, "piercing": 2, "armor": 4, "chargeBonus": 4, "health": 12, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 70},
        "Zbalansowany; +1 MD vs obecne 5 (PL O=7).",
    ),
    "Evocati": (
        "Hastati",
        "Weterani wczesnego legionu (Brąz); doświadczeni, dobra zbroja.",
        {"meleeAttack": 8, "meleeDefence": 8, "weaponDamage": 8, "piercing": 3, "armor": 6, "chargeBonus": 5, "health": 15, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 90},
        "Obecne EN OK; lekka korekta WD=8 (spójność z AP).",
    ),
    "iButho z iklwa": (
        "Włócznik",
        "Zulu ibutho (Żelazo); tarcza + iklwa; formacja „głowy wołu”.",
        {"meleeAttack": 5, "meleeDefence": 7, "weaponDamage": 4, "piercing": 4, "armor": 3, "chargeBonus": 6, "health": 18, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 95},
        "Anty-konna formacja; +MD vs obecne 6; wysokie morale.",
    ),
    "Gwardzista z champi": (
        "Wojownik z mieczem i tarczą",
        "Gwardia Inków (Brąz); macana/champi + tarcza; elitarna piechota górska.",
        {"meleeAttack": 7, "meleeDefence": 5, "weaponDamage": 6, "piercing": 2, "armor": 3, "chargeBonus": 5, "health": 12, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 80},
        "Ofensywna elita; +AP vs obecne 6 (PL A=8).",
    ),
    "Wojownik z żelaznym khopesh": (
        "Wojownik z mieczem i tarczą",
        "Egipska piechota z khopesh (Żelazo); krótki miecz hakowy + tarcza.",
        {"meleeAttack": 6, "meleeDefence": 6, "weaponDamage": 6, "piercing": 3, "armor": 4, "chargeBonus": 4, "health": 12, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 65},
        "+1 MD vs obecne 5; khopesh = przebicie.",
    ),
    "Mur tarcz (Sargonid)": (
        "Falanga",
        "Formacja tarczowa Sargonidów (Żelazo); „tarcza na tarcz” — ekstremalna obrona.",
        {"meleeAttack": 4, "meleeDefence": 8, "weaponDamage": 3, "piercing": 1, "armor": 6, "chargeBonus": 1, "health": 17, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 75},
        "Korekta: PL O=10 → MD 8; niski AP (formacja defensywna).",
    ),
    "Miecznik galijski": (
        "Gaesatae",
        "Galijski miecznik (Żelazo); długi miecz, agresywna szarża.",
        {"meleeAttack": 8, "meleeDefence": 2, "weaponDamage": 6, "piercing": 2, "armor": 1, "chargeBonus": 7, "health": 12, "missileAttack": 0, "wallAttack": 0, "Morale bazowe": 90},
        "Korekta: PL A=9 → AP 8; więcej szarży niż Gaesatae.",
    ),
}

ORDER = [
    "Soldurii", "Konnica lancowa asyryjska", "Konnica łucznicza asyryjska", "Łucznik asyryjski",
    "Drużynnik", "Jeździec z szczepnikami", "Strażnik bram Harappy", "Piechota induska",
    "Garnizon Harappy", "Rydwan Kapadokijski", "Piechota hetycka", "Gwardia hetycka",
    "Gwardia Ishtar", "Wojownik babiloński", "Piechota neobabilońska", "Tyrski miecznik",
    "Wojownik fenicki", "Gwardia Tyr", "Thorakites", "Legionarius", "Evocati",
    "iButho z iklwa", "Gwardzista z champi", "Wojownik z żelaznym khopesh",
    "Mur tarcz (Sargonid)", "Miecznik galijski",
]

H_FILL = PatternFill("solid", fgColor="2E3440")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
Y_FILL = PatternFill("solid", fgColor="FFF3CD")
G_FILL = PatternFill("solid", fgColor="D4EDDA")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def field_power(u):
    ma = float(u.get("meleeAttack") or 0)
    wd = float(u.get("weaponDamage") or 0)
    pier = float(u.get("piercing") or 0)
    ch = float(u.get("chargeBonus") or 0)
    ms = float(u.get("missileAttack") or 0)
    md = float(u.get("meleeDefence") or 0)
    arm = float(u.get("armor") or 0)
    hp = float(u.get("health") or 0)
    a = ma + wd + pier + ch / 2 + ms / 2
    o = md + arm + hp / 2
    return round(a, 1), round(o, 1), round(a + o, 1)


def ref_stats(units, name):
    u = next((x for x in units if x["Jednostka"] == name), None)
    if not u:
        return {}
    return {k: u.get(k, "") for k in TW_COLS}


def style_header(ws, row, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row, c, h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    with open(UNITS_PATH, encoding="utf-8") as f:
        units = json.load(f)
    by_name = {u["Jednostka"]: u for u in units}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propozycje-statów"

    headers = [
        "#", "Jednostka", "Nacja", "Epoka", "Tech", "Rola", "Typ",
        "Atak PL", "Obrona PL", "AD PL", "Health PL (×2)",
        "Wzorzec w grze",
        "Historia / charakter",
        "Uzasadnienie propozycji",
        "Ma EN dziś?",
    ]
    headers += [f"obecne.{c}" for c in TW_COLS]
    headers += [f"PROPONUJ.{c}" for c in TW_COLS]
    headers += ["M_pole (prop)", "Akceptacja Macieja (TAK/NIE/korekta)"]

    style_header(ws, 1, headers)

    for i, name in enumerate(ORDER, 1):
        u = by_name.get(name)
        if not u:
            print("BRAK w units.json:", name, file=sys.stderr)
            continue
        ref_name, hist, prop, reason = SPECS[name]
        ref = ref_stats(units, ref_name)
        has_en = all(u.get(c) not in (None, "", "—") for c in ("meleeAttack", "meleeDefence"))
        row = i + 1
        hp_display = (u.get("Health") or 0) * 2
        ad_pl = u.get("Atak dystansowy") or 0
        base = [
            i, name, u.get("Nacja", ""), u.get("Epoka", ""), u.get("Tech", ""),
            u.get("Rola (linia)", ""), u.get("Typ", ""),
            u.get("Atak"), u.get("Obrona"), ad_pl, hp_display,
            ref_name,
            hist,
            reason,
            "TAK" if has_en else "NIE — silnik = 0",
        ]
        for c in TW_COLS:
            base.append(u.get(c, "—"))
        for c in TW_COLS:
            base.append(prop.get(c, 0))
        pu = dict(u)
        pu.update(prop)
        _, _, m = field_power(pu)
        base.append(m)
        base.append("")

        for col, val in enumerate(base, 1):
            cell = ws.cell(row, col, val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if not has_en:
                cell.fill = Y_FILL
            h = headers[col - 1]
            if h.startswith("PROPONUJ."):
                cell.fill = G_FILL

    # Arkusz wzorców
    ws2 = wb.create_sheet("Wzorce-referencyjne")
    ref_names = sorted({SPECS[n][0] for n in ORDER})
    rh = ["Jednostka", "Epoka", "Rola"] + TW_COLS + ["M_pole"]
    style_header(ws2, 1, rh)
    r = 2
    for rn in ref_names:
        u = by_name.get(rn)
        if not u:
            continue
        row = [rn, u.get("Epoka"), u.get("Rola (linia)")]
        for c in TW_COLS:
            row.append(u.get(c, ""))
        _, _, m = field_power(u)
        row.append(m)
        for col, val in enumerate(row, 1):
            c = ws2.cell(r, col, val)
            c.border = BORDER
        r += 1

    ws3 = wb.create_sheet("_INFO")
    ws3["A1"] = "Jednostki-staty-MACIEJ-20260706"
    ws3["A2"] = f"Wygenerowano: {date.today().isoformat()}"
    ws3["A3"] = "Źródło listy: gra-robocza/JEDNOSTKI-DO-POPRAWY-staty.md"
    ws3["A4"] = "Po akceptacji: wpisz TAK w kolumnie Akceptacja → eksport do units.json / Panel-C"
    ws3["A5"] = "Silnik TW v3 czyta pola EN (meleeAttack, meleeDefence, missileAttack, health…)"

    for wsx in (ws, ws2):
        for col in wsx.columns:
            wsx.column_dimensions[col[0].column_letter].width = 14
        wsx.column_dimensions["M"].width = 42
        wsx.column_dimensions["N"].width = 48
        wsx.column_dimensions["O"].width = 36

    wb.save(OUT)
    missing = sum(1 for n in ORDER if not all(by_name.get(n, {}).get(c) not in (None, "", "—") for c in ("meleeAttack", "meleeDefence")))
    print(f"OK: {OUT}")
    print(f"Jednostek: {len(ORDER)} · bez EN: {missing}")


if __name__ == "__main__":
    main()
