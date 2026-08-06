#!/usr/bin/env python3
# gen-panel-c.py — Panel-C.xlsx komplet (audyt PANEL-AUDYT opcja A)
import json
import os
import re
import shutil
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise SystemExit("Brak openpyxl: pip install openpyxl")

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
OUT = os.path.join(os.path.dirname(__file__), "Panel-C.xlsx")

COLS = ["ID", "Parametr", "Opis", "Wartość", "Zakres/dozwolone", "Jednostka", "Wpływ na grę", "Plik źródłowy"]
# TW v3 — klucze EN w units.json · etykieta PL w kolumnie Parametr
STAT_FIELDS = [
    ("meleeAttack", "Atak"),
    ("meleeDefence", "Obrona"),
    ("weaponDamage", "Obrażenia"),
    ("piercing", "Przebicie"),
    ("armor", "Pancerz"),
    ("chargeBonus", "Szarża"),
    ("health", "Zdrowie"),
    ("missileAttack", "Atak dystansowy"),
    ("wallAttack", "Atak vs Mur"),
    ("Morale bazowe", "Morale bazowe"),
]
COST_FIELDS = [
    ("Pieniądz (koszt)", "pkt", "Koszt produkcji w złocie"),
    ("Ludność", "os.", "Populacja miasta na jednostkę"),
    ("Utrzymanie (Pieniądz/turę)", "zł/t", "Utrzymanie co turę"),
    ("żywność/turę", "🌾/t", "Koszt żywności co turę"),
    ("Surowiec", "—", "Surowiec strategiczny wymagany do produkcji (Brąz/Żelazo/-)"),
    ("Surowiec (ilość)", "szt.", "Ilość surowca strategicznego na sztukę (rekrutacja ×5)"),
    ("Utrzymanie surowiec", "—", "Surowiec strategiczny utrzymania co turę (Drewno/Brąz/Żelazo/-)"),
    ("Utrzymanie surowiec (ilość)", "szt./t", "Ilość surowca utrzymania co turę (baza F250)"),
]
# Macierz-jednostek — wszystkie jednostki (pełna lista z units.json / dawne Jednostki.xlsx)
MACIERZ_FIELDS = [
    ("Widok pola", "heksy", "Mgła — visibility.ts (A-FOG-Q1=B)"),
    ("Ruch", "pkt/t", "Ruch na mapie strategicznej"),
    ("Ruch w bitwie (heksy)", "heksy/t", "Ruch w bitwie taktycznej"),
    ("Epoka", "—", "Epoka odblokowania"),
    ("Tech", "—", "Technologia wymagana"),
    ("Rola (linia)", "—", "Wręcz / Dystans / …"),
    ("Próg dezercji (% health)", "0–1", "Health % ucieczki z bitwy"),
]

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H_FILL = PatternFill("solid", fgColor="2E3440")
H_FONT = Font(bold=True, color="FFFFFF", size=11)


def load_json(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)


def slug(name):
    s = re.sub(r"[^a-zA-Z0-9]+", "-", name.upper()).strip("-")
    return s[:40]


def write_header(ws):
    for c, h in enumerate(COLS, 1):
        cell = ws.cell(1, c, h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_row(ws, r, row_data):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(r, c, val)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def is_combat_unit(u):
    """Jednostki z parametrami walki w arkuszu Jednostki-staty."""
    rola = u.get("Rola (linia)", "")
    if rola in ("Wręcz", "Flanka", "Dystans", "Morska", "Oblężnicza"):
        return True
    if u.get("Jednostka") == "Zwiadowca":
        return True
    if u.get("Super-jednostka") == "TAK":
        return True
    return False


def parse_bonus_pct(s):
    if not s:
        return 0
    m = re.search(r"(-?\d+)", str(s))
    return int(m.group(1)) if m else 0


def sheet_info(wb):
    ws = wb.active
    ws.title = "_INFO"
    ws["A1"] = "Panel-C — Grupa C (Walka)"
    ws["A3"] = "Jak używać:"
    ws["A4"] = "1. Edytuj kolumnę Wartość w arkuszach poniżej."
    ws["A5"] = "2. Zapisz plik."
    ws["A6"] = "3. Poproś agenta: „eksportuj panel C” (bez terminala)."
    ws["A8"] = "Macierz-jednostek = meta (Ruch, Widok…); Jednostki-staty = TW v3 (AP, Obraż, AD, wallAttack…)."
    ws["A9"] = "Po zmianie JSON: python gen-panel-c.py · Po edycji Excel: python export-c.py → units.json"
    ws["A10"] = "wallAttack = Katapulta, Taran, Wieża · Auto-walka → auto-battle-params.json (2 mnożniki strat)"
    ws["A11"] = "fieldPower/siegePower w units.json = auto przy gen-panel-c (derived, nie edytuj ręcznie)"
    ws["A12"] = f"Wygenerowano: {date.today().isoformat()}"


def sheet_auto_walka(wb):
    """Balans auto-walki na mocy M — Maciej kręci głównie 2 mnożniki."""
    ab_path = os.path.join(DATA, "auto-battle-params.json")
    if os.path.isfile(ab_path):
        with open(ab_path, encoding="utf-8") as f:
            ab = json.load(f).get("straty", {})
    else:
        ab = {}
    ws = wb.create_sheet("Auto-walka")
    write_header(ws)
    r = 2
    rows = [
        (
            "C-AUTO-COEF-WIN",
            "coef_zwyciezca",
            "Auto-walka: ile % HP traci ZWYCIĘZCA po wygranej (× mnożnik krzywej)",
            ab.get("coef_zwyciezca", 1.0),
            "0.5–2.0",
            "×",
            "↑ więcej krwi u zwycięzcy · ↓ zwycięzca zostaje silniejszy",
            "auto-battle-params.json",
        ),
        (
            "C-AUTO-COEF-LOSE",
            "coef_przegrany",
            "Auto-walka: ile % HP traci PRZEGRANY na polu (× mnożnik lustra)",
            ab.get("coef_przegrany", 1.0),
            "0.5–2.0",
            "×",
            "↑ słabsza armia ginie szybciej · ↓ więcej ocalańców",
            "auto-battle-params.json",
        ),
        (
            "C-AUTO-P-ATK",
            "p_atk",
            "Stromość krzywej ATK vs R (↑ = przy większej przewadze ATK szybciej mniej/więcej strat)",
            ab.get("p_atk", ab.get("p", 0.58)),
            "0.35–1.0",
            "wykładnik",
            "↑ szybciej reaguje na R (mocniejszy crush / krwawszy remis ATK)",
            "auto-battle-params.json",
        ),
        (
            "C-AUTO-P-DEF",
            "p_def",
            "Stromość krzywej DEF vs R (↑ = przy przewadze DEF szybciej zmiana strat obrońcy)",
            ab.get("p_def", ab.get("p", 0.58)),
            "0.35–1.0",
            "wykładnik",
            "↑ obrońca szybciej 'twardnieje' przy dużej przewadze · ↓ łagodniej",
            "auto-battle-params.json",
        ),
        (
            "C-AUTO-L-MAX",
            "L_MAX",
            "Skala krzywej przy R≈1 (wspólna baza ATK/DEF)",
            ab.get("L_MAX", 0.42),
            "0.25–0.55",
            "ulamek",
            "Rzadko — głównie coef i p_atk/p_def",
            "auto-battle-params.json",
        ),
        (
            "C-AUTO-REMIS",
            "remis_pct",
            "Remis M_ATK=M_DEF: strata obu stron",
            ab.get("remis_pct", 0.4),
            "0.3–0.5",
            "ulamek",
            "↑ bardziej krwawy remis",
            "auto-battle-params.json",
        ),
    ]
    for row in rows:
        write_row(ws, r, row)
        r += 1


def sheet_stale(wb, p):
    ws = wb.create_sheet("Stale-walki")
    write_header(ws)
    r = 2
    tw = p.get("tw_v3", {"hit_base": 40, "hit_min": 15, "hit_max": 75, "max_rounds": 200})
    rows = [
        ("C-TW-HIT-BASE", "tw_v3.hit_base", "TW Rome 2: baza szansy trafienia", tw["hit_base"], "35-45", "pkt", "Wiecej trafien globalnie", "combat-params.json"),
        ("C-TW-HIT-MIN", "tw_v3.hit_min", "TW: min hit%", tw["hit_min"], "10-20", "%", "Floor szansy", "combat-params.json"),
        ("C-TW-HIT-MAX", "tw_v3.hit_max", "TW: max hit%", tw["hit_max"], "70-80", "%", "Cap szansy", "combat-params.json"),
        ("C-TW-MAX-RND", "tw_v3.max_rounds", "Limit rund walki 1v1", tw.get("max_rounds", 200), "50-300", "rundy", "Dluzsze bitwy", "combat-params.json"),
        ("C-COUNTER-MULT", "counter_multiplier", "Mnoznik counter +50%", p["counter_multiplier"], "1.0-2.0", "x", "Kontry typow", "combat-params.json"),
        ("C-RIVER-ATK", "river_attack_mult", "Kara ataku przez rzeke", p["river_attack_mult"], "0.5-1.0", "x", "Nizszy = wieksza kara", "combat-params.json"),
    ]
    for row in rows:
        write_row(ws, r, row)
        r += 1


def sheet_oblezenie(wb, p):
    ws = wb.create_sheet("Oblezenie")
    write_header(ws)
    r = 2
    o = p["oblężenie"]
    s = p.get("siege_ai", {})
    rows = [
        ("C-SIEGE-WALL-BASE", "wall_base_obrona", "Mur poziom 1 bonus Obrona", o["wall_base_obrona"], "0-20", "pkt", "Mur bazowy", "combat-params.json"),
        ("C-SIEGE-WALL-LVL", "wall_per_level_obrona", "Mur +Obrona / poziom", o["wall_per_level_obrona"], "1-10", "pkt", "Skala muru", "combat-params.json"),
        ("C-SIEGE-WALL-MAX", "wall_max_level", "Max poziom muru", o["wall_max_level"], "1-15", "poz", "Limit muru", "combat-params.json"),
        ("C-SIEGE-WALL-PANC", "wall_pancerz_fraction", "Frakcja muru -> Pancerz", o["wall_pancerz_fraction"], "0.2-0.8", "ulamek", "Cover z muru", "combat-params.json"),
        ("C-SIEGE-HILL", "hill_defense_mult", "Mnoznik obrony wzgorza", o["hill_defense_mult"], "1.0-2.0", "x", "Miasto na wzgorzu", "combat-params.json"),
        ("C-SIEGE-MTN", "mountain_defense_mult", "Mnoznik obrony gory", o["mountain_defense_mult"], "1.0-2.5", "x", "Gory", "combat-params.json"),
        ("C-SIEGE-FORT", "fortify_obrona_bonus", "Bonus fortifikacji garnizonu", o["fortify_obrona_bonus"], "0-5", "pkt", "Stoi w murze", "combat-params.json"),
        ("C-SIEGE-MIL-POP", "militia_pop_fraction", "Milicja z % populacji", o["militia_pop_fraction"], "0.1-0.4", "ulamek", "Obrona miasta", "combat-params.json"),
        ("C-SIEGE-MIL-STR", "militia_strength_fraction", "Sila milicji vs Wojownik", o["militia_strength_fraction"], "0.3-0.7", "ulamek", "Slaba milicja", "combat-params.json"),
        ("C-SIEGE-MAX-RND", "siege_max_rounds", "Max rund szturmu", o["siege_max_rounds"], "15-60", "rundy", "Limit walki", "combat-params.json"),
        ("C-SAI-T1", "siege_ai.t1_assault_ratio", "AI T1: stosunek armii do szturmu", s.get("t1_assault_ratio", 1.8), "1.5-2.5", "x", "Od razu szturm", "combat-params.json"),
        ("C-SAI-T2", "siege_ai.t2_build_min_ratio", "AI T2: min ratio budowa machin", s.get("t2_build_min_ratio", 1.4), "1.2-1.8", "x", "Oblezenie+machiny", "combat-params.json"),
        ("C-SAI-T3", "siege_ai.t3_starve_min_ratio", "AI T3: min ratio glodzenie", s.get("t3_starve_min_ratio", 1.1), "1.0-1.3", "x", "Tylko glod", "combat-params.json"),
        ("C-SAI-WAIT", "siege_ai.t2_max_wait_turns", "AI T2: max tur czekania", s.get("t2_max_wait_turns", 5), "3-10", "tury", "Timeout szturmu", "combat-params.json"),
        ("C-SAI-MACH", "siege_ai.units_per_extra_machine", "Jednostek na dodatkowa maszyna", s.get("units_per_extra_machine", 10), "5-20", "szt", "Skala machin AI", "combat-params.json"),
    ]
    for row in rows:
        write_row(ws, r, row)
        r += 1


def sheet_macierz_jednostek(wb, units):
    ws = wb.create_sheet("Macierz-jednostek")
    write_header(ws)
    r = 2
    for u in units:
        name = u["Jednostka"]
        for field, jed, opis in MACIERZ_FIELDS:
            if field not in u:
                continue
            val = u.get(field)
            if val is None:
                continue
            fid = f"C-M-{slug(name)}-{slug(field)}"
            write_row(ws, r, [
                fid,
                f"{name} · {field}",
                opis,
                val,
                "liczba" if field not in ("Epoka", "Tech", "Rola (linia)") else "tekst",
                jed,
                "Pełna macierz jednostek — merge Jednostki.xlsx",
                "units.json",
            ])
            r += 1


def sheet_jednostki_stats(wb, units):
    ws = wb.create_sheet("Jednostki-staty")
    write_header(ws)
    r = 2
    for u in units:
        if not is_combat_unit(u):
            continue
        name = u["Jednostka"]
        for field, pl_label in STAT_FIELDS:
            if field not in u and field != "weaponDamage":
                continue
            val = u.get(field, 0 if field != "Morale bazowe" else 100)
            fid = f"C-U-{slug(name)}-{slug(field)}"
            write_row(ws, r, [
                fid,
                f"{name} · {pl_label}",
                f"TW v3 hard input: {field}",
                val,
                "hard input 1:1" if field != "Morale bazowe" else "0-120",
                "pkt" if field != "Morale bazowe" else "morale",
                "Balans 1v1 / bitwa (bez dzielnikow)",
                "units.json",
            ])
            r += 1


def sheet_stale_moc(wb, p):
    ws = wb.create_sheet("Stale-moc")
    write_header(ws)
    r = 2
    up = p.get("unit_power", {})
    rows = [
        ("C-MOC-CHARGE-DIV", "unit_power.charge_divisor", "A += Szarża / n", up.get("charge_divisor", 2), "1-4", "n", "M pole — szarża", "combat-params.json"),
        ("C-MOC-MS-DIV", "unit_power.missile_divisor", "A += AD / n", up.get("missile_divisor", 2), "1-4", "n", "M pole — dystans", "combat-params.json"),
        ("C-MOC-HP-FIELD", "unit_power.hp_field_divisor", "O += HP / n (pole)", up.get("hp_field_divisor", 2), "1-20", "n", "M pole — wytrzymałość", "combat-params.json"),
        ("C-MOC-HP-SIEGE", "unit_power.hp_siege_divisor", "O += HP / n (oblężenie)", up.get("hp_siege_divisor", 10), "1-20", "n", "M_siege", "combat-params.json"),
    ]
    for row in rows:
        write_row(ws, r, row)
        r += 1


MOC_HDR = [
    "Jednostka", "Rola", "AP", "Obraż", "Przeb", "Szarża", "AD",
    "OBR", "Panc", "HP", "A", "O", "M_pole", "wallAttack", "A_siege", "O_siege", "M_siege",
]
MOC_FORMULA_NOTE = (
    "A=C+D+E+F/2+G/2 · O=H+I+J/2 · M_pole=K+L · "
    "A_siege=N+G · O_siege=H+I+J/10 · M_siege=O+P · Oblężnicza: M_pole liczone, ale nie w sumie armii"
)


def sheet_moc_jednostek(wb, units):
    ws = wb.create_sheet("Moc-jednostek")
    ws.cell(1, 1, "Moc jednostek TW v3 — formuły Excel (read-only, nie export-c)")
    ws.cell(2, 1, MOC_FORMULA_NOTE)
    hr = 4
    for c, h in enumerate(MOC_HDR, 1):
        cell = ws.cell(hr, c, h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.border = BORDER
    combat = [u for u in units if is_combat_unit(u)]
    combat.sort(key=lambda u: -(u.get("fieldPower") or 0))
    r = hr + 1
    for u in combat:
        name = u["Jednostka"]
        ws.cell(r, 1, name)
        ws.cell(r, 2, u.get("Rola (linia)", ""))
        ws.cell(r, 3, u.get("meleeAttack", 0))
        ws.cell(r, 4, u.get("weaponDamage", 0))
        ws.cell(r, 5, u.get("piercing", 0))
        ws.cell(r, 6, u.get("chargeBonus", 0))
        ws.cell(r, 7, u.get("missileAttack", 0))
        ws.cell(r, 8, u.get("meleeDefence", 0))
        ws.cell(r, 9, u.get("armor", 0))
        ws.cell(r, 10, u.get("health", 0))
        ws.cell(r, 11, f"=C{r}+D{r}+E{r}+F{r}/2+G{r}/2")
        ws.cell(r, 12, f"=H{r}+I{r}+J{r}/2")
        ws.cell(r, 13, f"=K{r}+L{r}")
        wall = u.get("wallAttack", "")
        ws.cell(r, 14, wall if wall != "" else "")
        if u.get("Rola (linia)") == "Oblężnicza":
            ws.cell(r, 15, f"=N{r}+G{r}")
            ws.cell(r, 16, f"=H{r}+I{r}+J{r}/10")
            ws.cell(r, 17, f"=O{r}+P{r}")
        for c in range(1, len(MOC_HDR) + 1):
            ws.cell(r, c).border = BORDER
        r += 1
    widths = [28, 12, 6, 7, 6, 7, 6, 6, 6, 6, 8, 8, 9, 10, 9, 9, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def sheet_koszty(wb, units):
    ws = wb.create_sheet("Koszty-jednostek")
    write_header(ws)
    r = 2
    for u in units:
        if not is_combat_unit(u):
            continue
        name = u["Jednostka"]
        for field, jed, opis in COST_FIELDS:
            if field not in u:
                continue
            fid = f"C-K-{slug(name)}-{slug(field)}"
            default = "-" if field in ("Surowiec", "Utrzymanie surowiec") else 0
            write_row(ws, r, [
                fid,
                f"{name} · {field}",
                opis,
                u.get(field, default),
                "0-999",
                jed,
                "Ekonomia produkcji",
                "units.json",
            ])
            r += 1


def sheet_countery(wb, counters):
    ws = wb.create_sheet("Countery")
    write_header(ws)
    r = 2
    for i, c in enumerate(counters, 1):
        write_row(ws, r, [
            f"C-CNT-{i:02d}",
            f"{c.get('Typ atakujący', '')} vs {c.get('Cel (typ)', '')}",
            f"Rodzaj: {c.get('Rodzaj (Atak/Obrona)', '')} · Status: {c.get('Status', '')}",
            c.get("Bonus", "+50%"),
            "+/-50%",
            "%",
            "Kontra typow w walce",
            "counters.json",
        ])
        r += 1


def sheet_teren(wb, terrain):
    ws = wb.create_sheet("Teren-walka")
    write_header(ws)
    r = 2
    for i, t in enumerate(terrain, 1):
        bonus = parse_bonus_pct(t.get("Bonus Obrona", "0"))
        write_row(ws, r, [
            f"C-TER-{i:02d}",
            t.get("Teren", ""),
            t.get("Efekt specjalny", ""),
            bonus,
            "-100 do +100",
            "%",
            "Bonus obrony obroncy na terenie",
            "terrain-combat.json",
        ])
        r += 1


def sync_units_power_cache(units):
    tools = os.path.join(ROOT, "gra", "tools")
    if tools not in sys.path:
        sys.path.insert(0, tools)
    from unit_power import apply_power_cache
    return apply_power_cache(units)


def main():
    units_path = os.path.join(DATA, "units.json")
    units = load_json("units.json")
    n_moc = sync_units_power_cache(units)
    if n_moc:
        bak = units_path + f".bak-MOC-{date.today().isoformat()}"
        if not os.path.isfile(bak):
            shutil.copy2(units_path, bak)
        with open(units_path, "w", encoding="utf-8") as f:
            json.dump(units, f, ensure_ascii=False, indent=2)
            f.write("\n")
    params = load_json("combat-params.json")
    counters = load_json("counters.json")
    terrain = load_json("terrain-combat.json")
    wb = openpyxl.Workbook()
    sheet_info(wb)
    sheet_stale(wb, params)
    sheet_auto_walka(wb)
    sheet_stale_moc(wb, params)
    sheet_oblezenie(wb, params)
    sheet_macierz_jednostek(wb, units)
    sheet_jednostki_stats(wb, units)
    sheet_moc_jednostek(wb, units)
    sheet_koszty(wb, units)
    sheet_countery(wb, counters)
    sheet_teren(wb, terrain)
    wb.save(OUT)
    n_all = len(units)
    n = sum(1 for u in units if is_combat_unit(u))
    print(f"OK: {OUT} ({n_all} jednostek, {n} bojowych, M-cache={n_moc} pol)")


if __name__ == "__main__":
    main()
