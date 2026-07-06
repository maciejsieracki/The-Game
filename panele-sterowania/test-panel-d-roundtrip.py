#!/usr/bin/env python3
# test-panel-d-roundtrip.py — zmiana w Excel → export-d → JSON (bez trwałej modyfikacji danych)
import json
import os
import shutil
import subprocess
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    print("SKIP: brak openpyxl")
    sys.exit(0)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
PANEL = os.path.join(os.path.dirname(__file__), "Panel-D.xlsx")
EXPORT = os.path.join(os.path.dirname(__file__), "export-d.py")
DATA = os.path.join(ROOT, "gra", "data")

DIP_KEY = "handelZawarcie_zaufanie"
BAR_KEY = "barbarzyncy_start_tura"


def fail(msg):
    print(f"FAIL: {msg}")
    sys.exit(1)


def find_row(ws, param_col, key):
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, param_col).value == key:
            return r
    return None


def main():
    if not os.path.isfile(PANEL):
        fail(f"brak {PANEL} — uruchom gen-panel-d.py")

    with tempfile.TemporaryDirectory() as tmp:
        data_dir = os.path.join(tmp, "data")
        shutil.copytree(DATA, data_dir)
        xlsx = os.path.join(tmp, "Panel-D.xlsx")
        shutil.copy(PANEL, xlsx)

        dip_path = os.path.join(data_dir, "diplomacy.json")
        ai_path = os.path.join(data_dir, "ai-params.json")
        dip_before = json.load(open(dip_path, encoding="utf-8"))["params"][DIP_KEY]
        bar_before = json.load(open(ai_path, encoding="utf-8"))[BAR_KEY]["wartosc"]

        dip_probe = dip_before + 7 if isinstance(dip_before, (int, float)) else 99
        bar_probe = bar_before + 3 if isinstance(bar_before, (int, float)) else 8

        wb = openpyxl.load_workbook(xlsx)
        ws_d = wb["Dyplomacja"]
        r_d = find_row(ws_d, 2, DIP_KEY)
        if not r_d:
            fail(f"nie znaleziono {DIP_KEY} w Dyplomacja")
        ws_d.cell(r_d, 4, dip_probe)

        ws_b = wb["Barbarzyńcy"]
        r_b = find_row(ws_b, 2, BAR_KEY)
        if not r_b:
            fail(f"nie znaleziono {BAR_KEY} w Barbarzyńcy")
        ws_b.cell(r_b, 4, bar_probe)
        wb.save(xlsx)

        cmd = [sys.executable, EXPORT, "--xlsx", xlsx, "--data-dir", data_dir]
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0:
            fail(r.stderr or r.stdout)

        dip_after = json.load(open(dip_path, encoding="utf-8"))["params"][DIP_KEY]
        bar_after = json.load(open(ai_path, encoding="utf-8"))[BAR_KEY]["wartosc"]
        if dip_after != dip_probe:
            fail(f"diplomacy {DIP_KEY}: oczekiwano {dip_probe}, jest {dip_after}")
        if bar_after != bar_probe:
            fail(f"ai-params {BAR_KEY}: oczekiwano {bar_probe}, jest {bar_after}")

    print("OK: Panel-D round-trip (dyplomacja + barbarzyńcy)")


if __name__ == "__main__":
    main()
