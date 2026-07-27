#!/usr/bin/env python3
# generate-terrain-yields-xlsx.py — terrain-yields.json → Panel-A-Plony-Terenu.xlsx
# Uruchom z root: python panele-sterowania/generate-terrain-yields-xlsx.py
import json
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
OUT = os.path.join(os.path.dirname(__file__), "Panel-A-Plony-Terenu.xlsx")

INFO_TEXT = "Edycja → w czacie: eksportuj plony terenu"
HDR_BASE = ["Teren", "Żywność", "Praca", "Podatek", "Uwagi"]
HDR_MOD = ["Modyfikator", "Żywność", "Praca", "Podatek", "Uwagi"]
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
INFO_FONT = Font(bold=True, size=11, color="1F4E79")


def load_yields():
    path = os.path.join(DATA, "terrain-yields.json")
    if not os.path.isfile(path):
        print(f"Brak pliku: {path}")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_table_sheet(ws, title, headers, rows, key_field):
    ws.title = title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    info = ws.cell(1, 1, INFO_TEXT)
    info.font = INFO_FONT
    info.alignment = Alignment(horizontal="left")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    widths = [22, 10, 10, 10, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    for idx, row in enumerate(rows, 3):
        ws.cell(idx, 1, row.get(key_field, ""))
        ws.cell(idx, 2, row.get("Żywność", 0))
        ws.cell(idx, 3, row.get("Praca", 0))
        ws.cell(idx, 4, row.get("Podatek", row.get("Handel", 0)))
        uwagi = row.get("Uwagi")
        ws.cell(idx, 5, uwagi if uwagi is not None else "")
        ws.cell(idx, 5).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A3"


def main():
    data = load_yields()
    wb = openpyxl.Workbook()
    ws_base = wb.active
    write_table_sheet(
        ws_base,
        "Teren-bazowy",
        HDR_BASE,
        data.get("terrain_types", []),
        "Teren",
    )
    write_table_sheet(
        wb.create_sheet("Bonusy-nakladki"),
        "Bonusy-nakladki",
        HDR_MOD,
        data.get("terrain_modifiers", []),
        "Modyfikator",
    )

    wb.properties.title = "Panel A — Plony terenu"
    wb.properties.subject = f"gen {date.today().isoformat()}"
    wb.save(OUT)
    print(f"Zapisano: {OUT}")
    print(f"Arkusze: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
