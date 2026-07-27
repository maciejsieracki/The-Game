#!/usr/bin/env python3
# export-tereny-maciej.py — Tereny-i-ulepszenia-MACIEJ.xlsx → terrain-yields.json + terrain-improvements.json
# Uruchom: python panele-sterowania/export-tereny-maciej.py
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "Tereny-i-ulepszenia-MACIEJ.xlsx")
YIELDS_PATH = os.path.join(DATA, "terrain-yields.json")
TERRAIN_PATH = os.path.join(DATA, "terrain-improvements.json")

YIELD_COLS = ("Żywność", "Praca", "Podatek", "Drewno", "Kamień")
BONUS_KEYS = ("zywnosc", "praca", "handel", "pieniadz", "kamien", "drewno", "glina")
DATA_ROW = 3


def _num_equal(a, b):
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return False


def coerce_num(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return old, False
    try:
        nv = int(round(float(val)))
        return nv, not _num_equal(nv, old)
    except (TypeError, ValueError):
        return old, False


def coerce_optional_num(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        if old is None:
            return None, False
        return None, True
    try:
        nv = int(round(float(val)))
        if old is None:
            return nv, True
        return nv, not _num_equal(nv, old)
    except (TypeError, ValueError):
        return old, False


def recalc_suma(row):
    podatek = row.get("Podatek", row.get("Handel", 0))
    total = sum(int(row.get(f, 0) or 0) for f in ("Żywność", "Praca", "Drewno", "Kamień")) + int(podatek or 0)
    if row.get("Suma") == total:
        return False
    row["Suma"] = total
    return True


def read_yields_sheet(ws, key_header):
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    key_col = col_idx.get(key_header, 1)
    out = {}
    for r in range(DATA_ROW, ws.max_row + 1):
        key = ws.cell(r, key_col).value
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        row = {}
        for field in YIELD_COLS:
            col_name = field
            if field == "Podatek" and col_name not in col_idx and "Handel" in col_idx:
                col_name = "Handel"
            if col_name in col_idx:
                row[field] = ws.cell(r, col_idx[col_name]).value
        if "Uwagi" in col_idx:
            row["Uwagi"] = ws.cell(r, col_idx["Uwagi"]).value
        out[key] = row
    return out


def overlay_yields(json_rows, excel_rows, key_field):
    changed = 0
    for row in json_rows:
        key = row.get(key_field)
        if not key or key not in excel_rows:
            continue
        src = excel_rows[key]
        row_changed = False
        for field in YIELD_COLS:
            old = row.get(field, row.get("Handel", 0) if field == "Podatek" else 0)
            src_val = src.get(field, src.get("Handel") if field == "Podatek" else None)
            nv, did = coerce_num(old, src_val)
            if did:
                row[field] = nv
                row_changed = True
                changed += 1
        if "Uwagi" in src and src["Uwagi"] is not None:
            uw = str(src["Uwagi"]).strip()
            old_uw = row.get("Uwagi")
            if uw != ("" if old_uw is None else str(old_uw)):
                row["Uwagi"] = uw if uw else None
                row_changed = True
                changed += 1
        if row_changed and recalc_suma(row):
            changed += 1
    return changed


def read_improvements_sheet(ws):
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    out = {}
    for r in range(DATA_ROW, ws.max_row + 1):
        key = ws.cell(r, col_idx.get("Klucz (nie zmieniaj)", 1)).value
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        bonus = {}
        bonus_labels = {
            "Żywność (bonus)": "zywnosc",
            "Praca (bonus)": "praca",
            "Podatek (bonus)": "handel",
            "Handel (bonus)": "handel",
            "Pieniądz (bonus)": "pieniadz",
            "Kamień (bonus)": "kamien",
            "Drewno (bonus)": "drewno",
            "Glina (bonus)": "glina",
        }
        for label, bk in bonus_labels.items():
            if label not in col_idx:
                continue
            val = ws.cell(r, col_idx[label]).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            bonus[bk] = int(round(float(val)))
        row = {"bonus": bonus}
        if "Koszt (Praca)" in col_idx:
            row["koszt_praca"] = ws.cell(r, col_idx["Koszt (Praca)"]).value
        if "Ilość surowca/turę" in col_idx:
            row["surowiec_ilosc_tura"] = ws.cell(r, col_idx["Ilość surowca/turę"]).value
        out[key] = row
    return out


def overlay_improvements(terrain, excel_rows):
    changed = 0
    for imp_key, src in excel_rows.items():
        if imp_key not in terrain or not isinstance(terrain[imp_key], dict):
            continue
        node = terrain[imp_key]
        new_bonus = src.get("bonus") or {}
        old_bonus = node.get("bonus") or {}
        merged = dict(old_bonus)
        bonus_changed = False
        for bk in BONUS_KEYS:
            if bk in new_bonus:
                if merged.get(bk) != new_bonus[bk]:
                    merged[bk] = new_bonus[bk]
                    bonus_changed = True
                    changed += 1
            elif bk in merged:
                del merged[bk]
                bonus_changed = True
                changed += 1
        if bonus_changed:
            node["bonus"] = merged if merged else {}
        if "koszt_praca" in src:
            nv, did = coerce_num(node.get("koszt_praca", 0), src["koszt_praca"])
            if did:
                node["koszt_praca"] = nv
                changed += 1
        if "surowiec_ilosc_tura" in src:
            nv, did = coerce_optional_num(node.get("surowiec_ilosc_tura"), src["surowiec_ilosc_tura"])
            if did:
                if nv is None:
                    node.pop("surowiec_ilosc_tura", None)
                else:
                    node["surowiec_ilosc_tura"] = nv
                changed += 1
    return changed


def save_json(obj, path):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
        f.write("\n")
    json.load(open(tmp, encoding="utf-8"))
    os.replace(tmp, path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"Brak pliku: {args.xlsx}")
        print("Uruchom: python panele-sterowania/gen-tereny-maciej-xlsx.py")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    for sheet in ("Plony-bazowe", "Nakladki", "Ulepszenia"):
        if sheet not in wb.sheetnames:
            print(f"Brak arkusza: {sheet}")
            sys.exit(1)

    yields_data = json.load(open(YIELDS_PATH, encoding="utf-8"))
    terrain = json.load(open(TERRAIN_PATH, encoding="utf-8"))

    base_rows = read_yields_sheet(wb["Plony-bazowe"], "Teren")
    mod_rows = read_yields_sheet(wb["Nakladki"], "Modyfikator")
    imp_rows = read_improvements_sheet(wb["Ulepszenia"])

    ch_base = overlay_yields(yields_data.get("terrain_types", []), base_rows, "Teren")
    ch_mod = overlay_yields(yields_data.get("terrain_modifiers", []), mod_rows, "Modyfikator")
    ch_imp = overlay_improvements(terrain, imp_rows)
    total = ch_base + ch_mod + ch_imp

    print(f"terrain_types: {ch_base} zmian")
    print(f"terrain_modifiers: {ch_mod} zmian")
    print(f"terrain-improvements: {ch_imp} zmian")
    print(f"RAZEM: {total} zmian")

    if total and not args.dry_run:
        save_json(yields_data, YIELDS_PATH)
        save_json(terrain, TERRAIN_PATH)
        print(f"Zapisano: {YIELDS_PATH}")
        print(f"Zapisano: {TERRAIN_PATH}")
    elif args.dry_run:
        print("(dry-run — JSON nie zapisany)")
    else:
        print("(brak zmian)")


if __name__ == "__main__":
    main()
