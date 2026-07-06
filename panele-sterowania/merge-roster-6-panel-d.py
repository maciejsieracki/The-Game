#!/usr/bin/env python3
"""
merge-roster-6-panel-d.py — dopisuje 6 nacji (draft REZERWA) do Panel-D.xlsx.

Uruchom z root projektu (po D-ROSTER ABC, przed edycją Macieja):
  python panele-sterowania/merge-roster-6-panel-d.py
  python panele-sterowania/merge-roster-6-panel-d.py --dry-run

Nie dotyka gra/data/*.json.
"""
import json
import os
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, ".."))
PANEL = os.path.join(HERE, "Panel-D.xlsx")
DRAFT = os.path.join(ROOT, "Civ-CYWILIZACJE", "draft", "roster-6-REZERWA.json")
GEN = os.path.join(HERE, "gen-panel-d.py")

SHEETS = {
    "bonusy": ("Bonusy-cywilizacji", ["Nacja", "typCywilizacji", "Typ efektu", "Cel", "Wartosc", "Opis", "Realizuje"]),
    "roster": ("Cywilizacje-roster", None),
    "params": ("Parametry-cyw", ["Cywilizacja", "preferowaneBudynki", "preferowaneJednostki", "modWzrostu", "modEkonomii", "uwagi"]),
    "ai": ("AI-per-nacja", ["Cywilizacja", "agresywnosc", "ekspansywnosc", "priorytetMilitarny", "priorytetEkonomia", "priorytetNauka", "tolerancjaRyzyka", "sklonnoscDoPodboju", "profilMapy", "uwagi"]),
    "dip": ("Dyplomacja-per-nacja", ["Cywilizacja", "sklonnoscSojusze", "lojalnosc", "progWojny", "pamietliwosc", "otwartoscHandel", "nastawienieBazowe", "uwagi"]),
}

PARAMS_DEFAULT = {
    "Harappa": ("Rynek, Mur", "piechota, łucznicy", 1.0, 1.04),
    "Hetyci": ("Mur, Koszary", "rydwany, piechota", 1.0, 1.0),
    "Słowianie": ("Koszary, Tartak", "piechota", 1.03, 1.0),
    "Babilonia": ("Biblioteka, Rynek", "piechota, łucznicy", 1.0, 1.03),
    "Asyria": ("Koszary, Kuźnia", "łucznicy, piechota", 1.0, 0.98),
    "Fenicjanie": ("Rynek", "piechota, łucznicy", 1.0, 1.05),
}


def load_draft():
    with open(DRAFT, encoding="utf-8") as f:
        return json.load(f)["cywilizacje"]


def ensure_panel():
    if not os.path.isfile(PANEL):
        print(f"Brak {PANEL} — uruchamiam gen-panel-d.py …")
        subprocess.check_call([sys.executable, GEN], cwd=ROOT)


def sheet_col_map(ws):
    return {str(c.value).strip(): j for j, c in enumerate(ws[1], 1) if c.value}


def existing_names(ws, col_name="Cywilizacja"):
    cmap = sheet_col_map(ws)
    col = cmap.get(col_name) or cmap.get("Nacja") or 1
    names = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v:
            names.add(str(v).strip())
    return names


def append_row(ws, rec, col_order=None):
    cmap = sheet_col_map(ws)
    row_idx = ws.max_row + 1
    if col_order:
        for j, key in enumerate(col_order, 1):
            ws.cell(row_idx, j, rec.get(key, ""))
    else:
        for key, val in rec.items():
            if key in cmap:
                ws.cell(row_idx, cmap[key], val)


def build_rows(cyw):
    name = cyw["Cywilizacja"]
    klastry = cyw.get("nazwyKlastra") or []
    roster = {
        "Cywilizacja": name,
        "mnoznikHandelPieniadz": cyw.get("mnoznikHandelPieniadz", 2.0),
        "ikonaId": cyw.get("ikonaId", ""),
    }
    for i in range(10):
        roster[f"Klaster_{i+1:02d}"] = klastry[i] if i < len(klastry) else ""

    bonusy = []
    for b in cyw.get("bonusy") or []:
        bonusy.append({
            "Nacja": name,
            "typCywilizacji": cyw.get("typCywilizacji", ""),
            "Typ efektu": b.get("typ", ""),
            "Cel": b.get("cel", ""),
            "Wartosc": b.get("wartosc", ""),
            "Opis": b.get("opis", ""),
            "Realizuje": b.get("realizuje", ""),
        })

    p = PARAMS_DEFAULT.get(name, ("", "piechota", 1.0, 1.0))
    params = {
        "Cywilizacja": name,
        "preferowaneBudynki": p[0],
        "preferowaneJednostki": p[1],
        "modWzrostu": p[2],
        "modEkonomii": p[3],
        "uwagi": f"draft roster-6 REZERWA tier={cyw.get('tier', '?')}",
    }

    ai = cyw.get("civAi") or {}
    ai_row = {
        "Cywilizacja": name,
        "agresywnosc": ai.get("agresywnosc", 4),
        "ekspansywnosc": 0,
        "priorytetMilitarny": ai.get("priorytetMilitarny", 5),
        "priorytetEkonomia": ai.get("priorytetEkonomia", 5),
        "priorytetNauka": ai.get("priorytetNauka", 5),
        "tolerancjaRyzyka": ai.get("tolerancjaRyzyka", 4),
        "sklonnoscDoPodboju": ai.get("sklonnoscDoPodboju", 2),
        "profilMapy": ai.get("profilMapy", "kopia_typu_obronna"),
        "uwagi": "draft roster-6",
    }

    pn = cyw.get("perNacja") or {}
    dip = {
        "Cywilizacja": name,
        "sklonnoscSojusze": pn.get("sklonnoscSojusze", 5),
        "lojalnosc": pn.get("lojalnosc", 5),
        "progWojny": pn.get("progWojny", 5),
        "pamietliwosc": pn.get("pamietliwosc", 5),
        "otwartoscHandel": pn.get("otwartoscHandel", 5),
        "nastawienieBazowe": pn.get("nastawienieBazowe", 50),
        "uwagi": "draft roster-6",
    }

    return roster, bonusy, params, ai_row, dip


def patch_sumer_roster(ws):
    """Q1A: Sumerowie ikonaId -> sumer (jesli wiersz istnieje)."""
    cmap = sheet_col_map(ws)
    col_cyw = cmap.get("Cywilizacja", 1)
    col_ik = cmap.get("ikonaId", 3)
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, col_cyw).value or "").strip() == "Sumerowie":
            ws.cell(r, col_ik, "sumer")
            print("  Sumerowie: ikonaId -> sumer")
            return True
    return False


def patch_sumer_bonusy(ws):
    """Q1A: wiersze bonusow Sumerowie typCywilizacji babilon -> sumer."""
    cmap = sheet_col_map(ws)
    col_n = cmap.get("Nacja", 1)
    col_t = cmap.get("typCywilizacji", 2)
    n = 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, col_n).value or "").strip() == "Sumerowie":
            if str(ws.cell(r, col_t).value or "").strip() == "babilon":
                ws.cell(r, col_t, "sumer")
                n += 1
    if n:
        print(f"  Sumerowie: {n} wierszy bonusow typCywilizacji -> sumer")
    return n > 0


def main():
    dry = "--dry-run" in sys.argv
    ensure_panel()
    civs = load_draft()
    wb = openpyxl.load_workbook(PANEL)

    added = 0
    for cyw in civs:
        name = cyw["Cywilizacja"]
        roster, bonusy, params, ai_row, dip = build_rows(cyw)

        ws_r = wb[SHEETS["roster"][0]]
        if name not in existing_names(ws_r):
            append_row(ws_r, roster)
            print(f"+ roster: {name}")
            added += 1

        ws_b = wb[SHEETS["bonusy"][0]]
        if name not in existing_names(ws_b, "Nacja"):
            for b in bonusy:
                append_row(ws_b, b, SHEETS["bonusy"][1])
            print(f"+ bonusy: {name} ({len(bonusy)} wierszy)")

        ws_p = wb[SHEETS["params"][0]]
        if name not in existing_names(ws_p):
            append_row(ws_p, params, SHEETS["params"][1])
            print(f"+ parametry-cyw: {name}")

        ws_a = wb[SHEETS["ai"][0]]
        if name not in existing_names(ws_a):
            append_row(ws_a, ai_row, SHEETS["ai"][1])
            print(f"+ AI-per-nacja: {name}")

        ws_d = wb[SHEETS["dip"][0]]
        if name not in existing_names(ws_d):
            append_row(ws_d, dip, SHEETS["dip"][1])
            print(f"+ Dyplomacja-per-nacja: {name}")

    patch_sumer_roster(wb[SHEETS["roster"][0]])
    patch_sumer_bonusy(wb[SHEETS["bonusy"][0]])

    if dry:
        print(f"DRY-RUN: dopisano/byłoby {added} nacji (bez zapisu)")
        return

    wb.save(PANEL)
    print(f"OK -> {PANEL} (nowe nacje w arkuszach: {added})")
    print("Maciej: otworz Panel-D.xlsx -> edytuj Wartosci -> w czacie: eksportuj panel")


if __name__ == "__main__":
    main()
