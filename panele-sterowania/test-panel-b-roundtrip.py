#!/usr/bin/env python3
# test-panel-b-roundtrip.py — zmiana w Excel → export-b → JSON (bez trwałej modyfikacji danych)
import json
import os
import shutil
import subprocess
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    print("SKIP: brak openpyxl")
    sys.exit(0)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
PANEL = os.path.join(os.path.dirname(__file__), "Panel-B.xlsx")
EXPORT = os.path.join(os.path.dirname(__file__), "export-b.py")
DATA = os.path.join(ROOT, "gra", "data")


def fail(msg):
    print(f"FAIL: {msg}")
    sys.exit(1)


def main():
    if not os.path.isfile(PANEL):
        fail(f"brak {PANEL} — uruchom gen-panel-b.py")

    with tempfile.TemporaryDirectory() as tmp:
        data_dir = os.path.join(tmp, "data")
        shutil.copytree(DATA, data_dir)
        xlsx = os.path.join(tmp, "Panel-B.xlsx")
        shutil.copy(PANEL, xlsx)

        miasto_path = os.path.join(data_dir, "miasto-params.json")
        before = json.load(open(miasto_path, encoding="utf-8"))["min_dystans_miast"]["wartosc"]

        wb = openpyxl.load_workbook(xlsx)
        ws = wb["Miasto"]
        probe = before + 1 if isinstance(before, int) else 6
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value == "min_dystans_miast":
                ws.cell(r, 4, probe)
                break
        else:
            fail("nie znaleziono wiersza min_dystans_miast")
        wb.save(xlsx)

        cmd = [sys.executable, EXPORT, "--xlsx", xlsx, "--data-dir", data_dir]
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0:
            fail(r.stderr or r.stdout)

        after = json.load(open(miasto_path, encoding="utf-8"))["min_dystans_miast"]["wartosc"]
        if after != probe:
            fail(f"round-trip: oczekiwano {probe}, jest {after}")

        # Budynki: kosztBudowy stolarnia
        if "Budynki" in openpyxl.load_workbook(xlsx).sheetnames:
            bld_path = os.path.join(data_dir, "buildings.json")
            wb3 = openpyxl.load_workbook(xlsx)
            ws3 = wb3["Budynki"]
            hdr = [c.value for c in ws3[1]]
            koszt_col = hdr.index("kosztBudowy") + 1 if "kosztBudowy" in hdr else None
            if koszt_col:
                for r in range(2, ws3.max_row + 1):
                    if ws3.cell(r, hdr.index("id") + 1).value == "stolarnia":
                        ws3.cell(r, koszt_col, 777)
                        break
                wb3.save(xlsx)
                subprocess.run(cmd, check=True, capture_output=True)
                kb = next(b["kosztBudowy"] for b in json.load(open(bld_path, encoding="utf-8")) if b["id"] == "stolarnia")
                if kb != 777:
                    fail(f"buildings stolarnia.kosztBudowy: oczekiwano 777, jest {kb}")

            # koszt_surowce.drewno (R-PANEL-SYNC)
            drewno_col = hdr.index("koszt_surowce.drewno") + 1 if "koszt_surowce.drewno" in hdr else None
            if drewno_col:
                wb4 = openpyxl.load_workbook(xlsx)
                ws4 = wb4["Budynki"]
                hdr4 = [c.value for c in ws4[1]]
                id_col4 = hdr4.index("id") + 1
                drewno_col4 = hdr4.index("koszt_surowce.drewno") + 1
                for r in range(2, ws4.max_row + 1):
                    if ws4.cell(r, id_col4).value == "stolarnia":
                        ws4.cell(r, drewno_col4, 88)
                        break
                wb4.save(xlsx)
                subprocess.run(cmd, check=True, capture_output=True)
                ks = next(
                    b["koszt_surowce"]["drewno"]
                    for b in json.load(open(bld_path, encoding="utf-8"))
                    if b["id"] == "stolarnia"
                )
                if ks != 88:
                    fail(f"buildings stolarnia.koszt_surowce.drewno: oczekiwano 88, jest {ks}")

    print("OK: Panel-B round-trip (miasto + budynki + koszt_surowce)")

    # Power P-A: jednostka_wojskowa pkt
    if "Potega-P-A" in openpyxl.load_workbook(PANEL).sheetnames:
        with tempfile.TemporaryDirectory() as tmp2:
            data_dir2 = os.path.join(tmp2, "data")
            shutil.copytree(DATA, data_dir2)
            xlsx2 = os.path.join(tmp2, "Panel-B.xlsx")
            shutil.copy(PANEL, xlsx2)
            power_path = os.path.join(data_dir2, "power-params.json")
            before_pkt = json.load(open(power_path, encoding="utf-8"))["skladniki"]["jednostka_wojskowa"]["pkt"]
            wb2 = openpyxl.load_workbook(xlsx2)
            ws2 = wb2["Potega-P-A"]
            probe = before_pkt + 1 if isinstance(before_pkt, (int, float)) else 26
            for r in range(2, ws2.max_row + 1):
                if ws2.cell(r, 2).value == "jednostka_wojskowa":
                    ws2.cell(r, 4, probe)
                    break
            else:
                fail("nie znaleziono wiersza jednostka_wojskowa w Potega-P-A")
            wb2.save(xlsx2)
            cmd2 = [sys.executable, EXPORT, "--xlsx", xlsx2, "--data-dir", data_dir2]
            subprocess.run(cmd2, check=True, capture_output=True)
            after_pkt = json.load(open(power_path, encoding="utf-8"))["skladniki"]["jednostka_wojskowa"]["pkt"]
            if after_pkt != probe:
                fail(f"power round-trip: oczekiwano {probe}, jest {after_pkt}")
        print("OK: Panel-B round-trip (Potega-P-A)")


if __name__ == "__main__":
    main()
