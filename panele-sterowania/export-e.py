#!/usr/bin/env python3
# export-e.py — Panel-E.xlsx → ui-params.json + e-start-params.json
# Uruchom z root: python panele-sterowania/export-e.py [--dry-run]
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "Panel-E.xlsx")

COL_PARAM = 2
COL_VALUE = 4

MAP_LABELS = ["Malenki", "Mały", "Standardowy", "Duży", "Ogromny"]
MAP_FIELDS = ["rywale_ai", "miasta_panstwa", "typy_cywilizacji", "hex_w", "hex_h"]


def coerce(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return old, False
    if isinstance(old, bool):
        if isinstance(val, bool):
            return val, val != old
        nv = str(val).strip().lower() in ("tak", "true", "1", "prawda", "yes")
        return nv, nv != old
    if isinstance(old, int) and not isinstance(old, bool):
        try:
            nv = int(round(float(val)))
            return nv, nv != old
        except (TypeError, ValueError):
            return old, False
    if isinstance(old, float):
        try:
            nv = float(val)
            return nv, abs(nv - old) > 1e-9
        except (TypeError, ValueError):
            return old, False
    if old is None:
        return val, True
    nv = str(val).strip()
    return nv, nv != str(old)


def read_sheet_params(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, COL_PARAM).value
        if key is None or str(key).strip() == "":
            continue
        out[str(key).strip()] = ws.cell(r, COL_VALUE).value
    return out


def apply_nowa_gra(ui, params):
    changed = 0
    for s in ui.get("nowa_gra", {}).get("ustawienia", []):
        key = s.get("key")
        if key not in params:
            continue
        nv, did = coerce(s.get("domyslny", 0), params[key])
        if did:
            s["domyslny"] = nv
            changed += 1
    return changed


def apply_menu(ui, params):
    changed = 0
    if "menu_wersja" in params:
        nv, did = coerce(ui.get("menu", {}).get("wersja", ""), params["menu_wersja"])
        if did:
            ui.setdefault("menu", {})["wersja"] = nv
            changed += 1
    for s in ui.get("menu", {}).get("ustawienia", []):
        key = s.get("key")
        if key not in params:
            continue
        nv, did = coerce(s.get("domyslny", 0), params[key])
        if did:
            s["domyslny"] = nv
            changed += 1
    return changed


def apply_e_start(ep, sheets):
    changed = 0

    def set_path(parts, val):
        nonlocal changed
        node = ep
        for p in parts[:-1]:
            node = node.setdefault(p, {})
        old = node.get(parts[-1])
        nv, did = coerce(old if old is not None else val, val)
        if did or parts[-1] not in node:
            node[parts[-1]] = nv
            if did:
                changed += 1

    if "Defaulty-startu" in sheets:
        for k, v in sheets["Defaulty-startu"].items():
            if k in ("player_civ_id", "start_epoch_id", "map_quality_default", "render_quality_bundled"):
                set_path(["defaulty", k], v)

    if "Skala-mapy" in sheets:
        for label in MAP_LABELS:
            for field in MAP_FIELDS:
                pk = f"{label}.{field}"
                if pk not in sheets["Skala-mapy"]:
                    continue
                set_path(["skala_mapy", label, field], sheets["Skala-mapy"][pk])

    if "Generator-E2" in sheets:
        for k, v in sheets["Generator-E2"].items():
            set_path(["generator_e2", k], v)

    if "Tempo-gry" in sheets:
        for k, v in sheets["Tempo-gry"].items():
            set_path(["tempo_gry", k], v)

    if "Zwyciestwo" in sheets:
        for k, v in sheets["Zwyciestwo"].items():
            set_path(["zwyciestwo", k], v)

    if "Kreator-zaaw" in sheets:
        for k, v in sheets["Kreator-zaaw"].items():
            set_path(["kreator_zaawansowane", k], v)

    if "Decyzje-kanon" in sheets:
        for k, v in sheets["Decyzje-kanon"].items():
            set_path(["decyzje_kanon", k], v)

    return changed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"BRAK: {args.xlsx} — najpierw: python panele-sterowania/gen-panel-e.py")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        if name.startswith("_"):
            continue
        sheets[name] = read_sheet_params(wb[name])

    ui_path = os.path.join(DATA, "ui-params.json")
    ep_path = os.path.join(DATA, "e-start-params.json")

    with open(ui_path, encoding="utf-8") as f:
        ui = json.load(f)
    if os.path.isfile(ep_path):
        with open(ep_path, encoding="utf-8") as f:
            ep = json.load(f)
    else:
        ep = {"_opis": "Panel-E export"}

    c_ui = 0
    if "Nowa-gra" in sheets:
        c_ui += apply_nowa_gra(ui, sheets["Nowa-gra"])
    if "Menu" in sheets:
        c_ui += apply_menu(ui, sheets["Menu"])
    c_ep = apply_e_start(ep, sheets)

    print(f"Panel-E: ui-params {c_ui} zmian, e-start-params {c_ep} zmian")

    if args.dry_run:
        print("(dry-run — pliki nie zapisane)")
        return

    with open(ui_path, "w", encoding="utf-8") as f:
        json.dump(ui, f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(ep_path, "w", encoding="utf-8") as f:
        json.dump(ep, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"Zapisano: {ui_path}")
    print(f"Zapisano: {ep_path}")


if __name__ == "__main__":
    main()
