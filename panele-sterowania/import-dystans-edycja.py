#!/usr/bin/env python3
# import-dystans-edycja.py — TW-dystans-edycja.xlsx → units.json
# Pola: Zasięg, Ilość pocisków, missileAttack, wallAttack (oblężnicze)
from __future__ import annotations

import json
import shutil
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Brak openpyxl: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
UNITS = ROOT / "gra" / "data" / "units.json"
XLSX = Path(__file__).resolve().parent / "TW-dystans-edycja.xlsx"

SIEGE_WALL = {"Katapulta", "Taran", "Wieża oblężnicza"}

FIELD_MAP = {
    "Zasięg (hex)": "Zasięg ataku (hex)",
    "Zasięg ataku (hex)": "Zasięg ataku (hex)",
    "Ilość pocisków": "Ilość pocisków",
    "missileAttack": "missileAttack",
    "wallAttack": "wallAttack",
}


def num(v):
    if v is None or v == "" or v == "—":
        return None
    if isinstance(v, (int, float)):
        return int(round(float(v)))
    s = str(v).strip().replace(",", ".")
    if not s or s == "—":
        return None
    return int(round(float(s)))


def find_header(ws) -> tuple[int, dict[str, int]]:
    for r in range(1, 25):
        if ws.cell(r, 1).value != "Jednostka":
            continue
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(r, c).value
            if h:
                headers[str(h).strip()] = c
        return r, headers
    raise SystemExit("Nie znaleziono nagłówka (Jednostka)")


def col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        if n in headers:
            return headers[n]
    return None


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Brak pliku: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Dystans-TW-v3"]
    header_row, headers = find_header(ws)

    c_rng = col(headers, "Zasięg (hex)", "Zasięg ataku (hex)")
    c_sh = col(headers, "Ilość pocisków")
    c_ms = col(headers, "missileAttack")
    c_wall = col(headers, "wallAttack", "missileAttackVsWall (Katapulta)", "missileAttackVsWall")

    edits: dict[str, dict] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        ms = num(ws.cell(r, c_ms).value) if c_ms else None
        if name == "Katapulta" and ms is None and c_ms:
            legacy = col(headers, "missileAttackVsUnit (Katapulta)", "missileAttackVsUnit")
            if legacy:
                ms = num(ws.cell(r, legacy).value)
        edits[name] = {
            "Zasięg ataku (hex)": num(ws.cell(r, c_rng).value) if c_rng else None,
            "Ilość pocisków": num(ws.cell(r, c_sh).value) if c_sh else None,
            "missileAttack": ms,
            "wallAttack": num(ws.cell(r, c_wall).value) if c_wall else None,
        }

    backup = UNITS.with_suffix(f".json.bak-dystans-{date.today().isoformat()}")
    shutil.copy2(UNITS, backup)
    units = json.loads(UNITS.read_text(encoding="utf-8"))
    changed: list[str] = []

    for u in units:
        name = u.get("Jednostka")
        if name not in edits:
            continue
        e = edits[name]
        for json_key in ("Zasięg ataku (hex)", "Ilość pocisków", "missileAttack", "wallAttack"):
            val = e.get(json_key)
            if val is None:
                continue
            if u.get(json_key) != val:
                u[json_key] = val
                changed.append(f"{name}.{json_key}={val}")
        if name not in SIEGE_WALL:
            u.pop("wallAttack", None)
        u.pop("missileAttackVsUnit", None)
        u.pop("missileAttackVsWall", None)

    UNITS.write_text(json.dumps(units, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {UNITS}")
    print(f"Backup: {backup}")
    print(f"Zaktualizowano pól: {len(changed)}")
    for line in changed:
        print(f"  {line}")
    if not changed:
        print("  (JSON już zgodny z Excelem)")


if __name__ == "__main__":
    main()
