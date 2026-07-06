#!/usr/bin/env python3
"""
gen-panel-d.py — buduje Panel-D.xlsx (Grupa D) z bieżących JSON + stałych lane.

Uruchom z root projektu:
  python panele-sterowania/gen-panel-d.py

Seed z gra/data/*.json (gdy brak legacy xlsx w repo).
"""
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("pip install openpyxl")

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
OUT = os.path.join(os.path.dirname(__file__), "Panel-D.xlsx")
DATA = os.path.join(ROOT, "gra", "data")

COLS = ["ID", "Parametr", "Opis", "Wartość", "Zakres/dozwolone", "Jednostka", "Wpływ na grę", "Plik źródłowy"]
H_FILL = PatternFill("solid", fgColor="3D5A80")
H_FONT = Font(bold=True, color="FFFFFF", size=10)
VAL_FONT = Font(color="0000FF")
THIN = Side(style="thin", color="CCCCCC")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BONUSY_COLS = ["Nacja", "typCywilizacji", "Typ efektu", "Cel", "Wartosc", "Opis", "Realizuje"]
ROSTER_COLS = ["Cywilizacja", "mnoznikHandelPieniadz", "ikonaId"] + [f"Klaster_{i:02d}" for i in range(1, 11)]
PARAMETRY_CYW_COLS = [
    "Cywilizacja", "preferowaneBudynki", "preferowaneJednostki",
    "modWzrostu", "modEkonomii", "uwagi",
]
AI_PER_NACJA_COLS = [
    "Cywilizacja", "agresywnosc", "ekspansywnosc", "priorytetMilitarny",
    "priorytetEkonomia", "priorytetNauka", "tolerancjaRyzyka",
    "sklonnoscDoPodboju", "profilMapy", "uwagi",
]
DIP_PER_NACJA_COLS = [
    "Cywilizacja", "sklonnoscSojusze", "lojalnosc", "progWojny",
    "pamietliwosc", "otwartoscHandel", "nastawienieBazowe", "uwagi",
]
DIP_AKCJE_COLS = [
    "Akcja", "Opis", "Dostępne: Główni rywale", "Dostępne: Poboczni", "Koszt", "Efekt",
]


def load_json(name):
    p = os.path.join(DATA, name)
    if not os.path.isfile(p):
        return {}
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def row(sheet_id, param, opis, wartosc, zakres, jedn, wplyw, plik):
    return [f"D-{sheet_id}-{param}", param, opis, wartosc, zakres, jedn, wplyw, plik]


def write_sheet(wb, title, rows):
    ws = wb.create_sheet(title[:31])
    for j, c in enumerate(COLS, 1):
        cell = ws.cell(1, j, c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.border = BD
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            cell = ws.cell(i, j, v)
            cell.border = BD
            if j == 4:
                cell.font = VAL_FONT
    ws.freeze_panes = "A2"
    return ws


def write_table_sheet(wb, title, columns, rows):
    """Arkusz tabelaryczny — nagłówek w wierszu 1, dane od wiersza 2."""
    ws = wb.create_sheet(title[:31])
    for j, h in enumerate(columns, 1):
        cell = ws.cell(1, j, h)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.border = BD
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, rec in enumerate(rows, 2):
        for j, col in enumerate(columns, 1):
            val = rec.get(col) if isinstance(rec, dict) else None
            cell = ws.cell(i, j, val)
            cell.border = BD
    ws.freeze_panes = "A2"
    return ws


def bonusy_cywilizacji_rows():
    civs = load_json("civs.json").get("cywilizacje", [])
    rows = []
    for cyw in civs:
        name = cyw.get("Cywilizacja", "")
        typ_cyw = cyw.get("typCywilizacji", "")
        for b in cyw.get("bonusy") or []:
            rows.append({
                "Nacja": name,
                "typCywilizacji": typ_cyw,
                "Typ efektu": b.get("typ", ""),
                "Cel": b.get("cel", ""),
                "Wartosc": b.get("wartosc", ""),
                "Opis": b.get("opis", ""),
                "Realizuje": b.get("realizuje", ""),
            })
    return rows


def cywilizacje_roster_rows():
    civs = load_json("civs.json").get("cywilizacje", [])
    rows = []
    for cyw in civs:
        klastry = cyw.get("nazwyKlastra") or [""] * 10
        if len(klastry) < 10:
            klastry = klastry + [""] * (10 - len(klastry))
        rec = {
            "Cywilizacja": cyw.get("Cywilizacja", ""),
            "mnoznikHandelPieniadz": cyw.get("mnoznikHandelPieniadz", 2.0),
            "ikonaId": cyw.get("ikonaId", ""),
        }
        for i in range(10):
            rec[f"Klaster_{i + 1:02d}"] = klastry[i]
        rows.append(rec)
    return rows


def parametry_cyw_rows():
    return load_json("civ-params.json").get("cywilizacje", [])


def ai_per_nacja_rows():
    return load_json("civ-ai.json").get("cywilizacje", [])


def dyplomacja_akcje_rows():
    return load_json("diplomacy.json").get("akcje_dyplomatyczne") or []


def dyplomacja_per_nacja_rows():
    return load_json("diplomacy.json").get("perNacja") or []


def dyplomacja_rows():
    params = load_json("diplomacy.json").get("params", {})
    rows = []
    for k, v in sorted(params.items()):
        rows.append(row(
            "DIP", k, f"Parametr dyplomacji ({k})", v,
            "liczba (ujemne OK)", "pkt/tura lub próg",
            "Wyżej = łatwiej sojusze / szybszy wzrost Zaufania",
            "gra/data/diplomacy.json → diplomacy.ts",
        ))
    return rows


def ai_params_rows(prefix_filter=None, sheet_id="AI"):
    ai = load_json("ai-params.json")
    rows = []
    for k, obj in sorted(ai.items()):
        if not isinstance(obj, dict):
            continue
        if prefix_filter and not k.startswith(prefix_filter):
            continue
        w = obj.get("wartosc", obj.get("wartość", 0))
        sec = obj.get("sekcja", "")
        opis = obj.get("opis", k)
        rows.append(row(
            sheet_id, k, f"{sec} — {opis}" if sec else opis, w,
            "zależy od parametru", "× lub szt.",
            "Patrz opis w ai-params.json",
            "gra/data/ai-params.json → ai.ts / barbarians.ts",
        ))
    return rows


def ai_zachowanie_rows():
    ai = load_json("ai-params.json")
    prefixes = ("ekspansja_", "dyplomacja_", "ai_wycofanie")
    rows = []
    for k, obj in sorted(ai.items()):
        if not isinstance(obj, dict):
            continue
        if not any(k.startswith(p) for p in prefixes):
            continue
        w = obj.get("wartosc", obj.get("wartość", 0))
        sec = obj.get("sekcja", "")
        opis = obj.get("opis", k)
        rows.append(row(
            "AIZ", k, f"{sec} — {opis}" if sec else opis, w,
            "zależy od parametru", "×, pkt lub próg",
            "Patrz opis w ai-params.json",
            "gra/data/ai-params.json → ai.ts",
        ))
    return rows


def barbarzyncy_rows():
    fallbacks = {
        "barbarzyncy_start_tura": (5, "tura", "Od której tury spawn obozów"),
        "barbarzyncy_max_obozy": (6, "szt.", "Max obozów na mapie"),
        "barbarzyncy_min_dystans_miasto": (5, "heksy", "Min. odległość obozu od miasta"),
        "barbarzyncy_odstep_obozow": (6, "heksy", "Min. odległość między obozami"),
        "barbarzyncy_interwal_spawnu": (6, "tury", "Co ile tur nowy obóz"),
        "barbarzyncy_jednostek_na_oboz": (2, "szt.", "Spawn jednostek z obozu"),
        "barbarzyncy_zasieg_kontroli": (3, "heksy", "Zasięg kontroli obozu"),
        "barbarzyncy_zasieg_agresji": (6, "heksy", "Zasięg agresji barbarzyńców"),
        "barbarzyncy_prog_odwrotu_hp": (0.3, "0–1", "Poniżej tego HP % — odwrót"),
    }
    ai = load_json("ai-params.json")
    rows = []
    for k, (fb, jedn, wplyw) in fallbacks.items():
        w = fb
        if k in ai and isinstance(ai[k], dict):
            w = ai[k].get("wartosc", ai[k].get("wartość", fb))
        rows.append(row("BAR", k, wplyw, w, "≥0", jedn, wplyw, "ai-params.json → barbarians.ts"))
    rows.append(row(
        "BAR", "epoka_cutoff_barbarzy", "Barbarzyńcy OFF od epoki Średniowiecze (11=C*)", 4,
        "1–4", "epoka", "≥4 = buntownicy zamiast obozów", "barbarians.ts (kod; v1.1 → JSON)",
    ))
    return rows


def zwyciestwo_rows():
    return [
        row("VIC", "prog_dominacji_power", "Próg dominacji Power gracza (10=A*)", 0.5, "0.01–0.99", "udział",
            ">50% sumy Power w ostatniej epoce = wygrana", "victory.ts"),
        row("VIC", "ostatnia_epoka_gry_v1", "Ostatnia epoka v1.0 (dominacja)", 3, "1–3", "epoka",
            "Dominacja liczy się tylko w tej epoce", "victory.ts"),
        row("VIC", "nauka_wymaga_rakiety", "Nauka: wymaga flagi rakiety", "TAK", "TAK/NIE", "—",
            "Oprócz wszystkich tech w scope", "victory.ts + SILNIK"),
    ]


def eksporty_rows():
    return [
        ["—", "Skrypt", "Co eksportuje", "Arkusze Panel-D", "JSON docelowy", "—", "—", "—"],
        ["—", "export-d.py", "Jeden orchestrator Panel-D", "wszystkie poniżej", "wiele", "—", "python panele-sterowania/export-d.py", "—"],
        ["—", "—", "Parametry dyplomacji", "Dyplomacja", "diplomacy.json → params", "—", "—", "—"],
        ["—", "—", "Akcje dyplomatyczne", "Dyplomacja-akcje", "diplomacy.json → akcje_dyplomatyczne", "—", "—", "—"],
        ["—", "—", "Dyplomacja per nacja", "Dyplomacja-per-nacja", "diplomacy.json → perNacja", "—", "—", "—"],
        ["—", "—", "Bonusy cywilizacji", "Bonusy-cywilizacji", "civs.json → bonusy[]", "—", "—", "—"],
        ["—", "—", "Roster cywilizacji", "Cywilizacje-roster", "civs.json → klastry/mnoznik/ikona", "—", "—", "—"],
        ["—", "—", "Parametry cyw", "Parametry-cyw", "civ-params.json", "—", "—", "—"],
        ["—", "—", "AI per nacja", "AI-per-nacja", "civ-ai.json", "—", "—", "—"],
        ["—", "—", "AI global + barbarzyńcy", "AI-trudnosc, AI-archetyp, AI-zachowanie, Barbarzyńcy", "ai-params.json", "—", "—", "—"],
    ]


def write_info(wb):
    ws = wb.active
    ws.title = "_INFO"
    lines = [
        ("Panel sterowania — Grupa D (Cywilizacje, Dyplomacja, AI, Barbarzyńcy)", ""),
        (f"Wygenerowano: {datetime.now():%Y-%m-%d %H:%M}", ""),
        ("", ""),
        ("JAK UŻYWAĆ", ""),
        ("1. Edytuj kolumnę Wartość (parametry) lub komórki w arkuszach tabelarycznych.", ""),
        ("2. Zapisz plik Panel-D.xlsx.", ""),
        ("3. Napisz w czacie grupy D: eksportuj panel — agent odpali export-d.py.", ""),
        ("", ""),
        ("Arkusze tabelaryczne: Bonusy-cywilizacji, Cywilizacje-roster, Parametry-cyw, AI-per-nacja, Dyplomacja-akcje, Dyplomacja-per-nacja.", ""),
        ("Spec: docs/obieg/PANEL-STEROWANIA-SPEC.md", ""),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ws.cell(i, 1, a).font = Font(bold=("JAK" in a or "Panel sterowania" in a))
        ws.cell(i, 2, b)
    ws.column_dimensions["A"].width = 72


def main():
    wb = openpyxl.Workbook()
    write_info(wb)
    write_sheet(wb, "Dyplomacja", dyplomacja_rows())
    write_table_sheet(wb, "Bonusy-cywilizacji", BONUSY_COLS, bonusy_cywilizacji_rows())
    write_table_sheet(wb, "Cywilizacje-roster", ROSTER_COLS, cywilizacje_roster_rows())
    write_table_sheet(wb, "Parametry-cyw", PARAMETRY_CYW_COLS, parametry_cyw_rows())
    write_table_sheet(wb, "AI-per-nacja", AI_PER_NACJA_COLS, ai_per_nacja_rows())
    akcje = dyplomacja_akcje_rows()
    if akcje:
        write_table_sheet(wb, "Dyplomacja-akcje", DIP_AKCJE_COLS, akcje)
    per = dyplomacja_per_nacja_rows()
    if per:
        write_table_sheet(wb, "Dyplomacja-per-nacja", DIP_PER_NACJA_COLS, per)
    write_sheet(wb, "AI-trudnosc", ai_params_rows("trudnosc_", "AIT"))
    write_sheet(wb, "AI-archetyp", ai_params_rows("archetype_", "AIA"))
    write_sheet(wb, "AI-zachowanie", ai_zachowanie_rows())
    write_sheet(wb, "Barbarzyńcy", barbarzyncy_rows())
    write_sheet(wb, "Zwycięstwo", zwyciestwo_rows())
    write_sheet(wb, "_Eksporty", eksporty_rows())
    wb.save(OUT)
    print(f"OK -> {OUT}")
    print(f"arkusze: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
