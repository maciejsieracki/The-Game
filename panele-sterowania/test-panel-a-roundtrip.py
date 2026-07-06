#!/usr/bin/env python3
# test-panel-a-roundtrip.py — zmiana w Panel-A → export-a → JSON (temp dir)
import json
import os
import shutil
import subprocess
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    print("SKIP: brak openpyxl")
    sys.exit(0)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
PANEL = os.path.join(os.path.dirname(__file__), "Panel-A.xlsx")
EXPORT = os.path.join(os.path.dirname(__file__), "export-a.py")
DATA = os.path.join(ROOT, "gra", "data")


def fail(msg):
    print(f"FAIL: {msg}")
    sys.exit(1)


def main():
    if not os.path.isfile(PANEL):
        fail(f"brak {PANEL} — uruchom gen-panel-a.py")

    with tempfile.TemporaryDirectory() as tmp:
        data_dir = os.path.join(tmp, "data")
        shutil.copytree(DATA, data_dir)
        xlsx = os.path.join(tmp, "Panel-A.xlsx")
        shutil.copy(PANEL, xlsx)

        terrain_path = os.path.join(data_dir, "terrain-improvements.json")
        before = json.load(open(terrain_path, encoding="utf-8"))["farma"]["bonus"]["zywnosc"]

        wb = openpyxl.load_workbook(xlsx)
        if "Ulepszenia-FOOD" not in wb.sheetnames:
            fail("brak arkusza Ulepszenia-FOOD")
        ws = wb["Ulepszenia-FOOD"]
        probe = 88
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value == "farma.bonus.zywnosc":
                ws.cell(r, 4, probe)
                break
        else:
            fail("nie znaleziono farma.bonus.zywnosc")

        if "Plony-terenow" in wb.sheetnames:
            ws2 = wb["Plony-terenow"]
            for r in range(2, ws2.max_row + 1):
                if ws2.cell(r, 2).value == "plony.types.laka.zywnosc":
                    ws2.cell(r, 4, 77)
                    break

        wb.save(xlsx)
        cmd = [sys.executable, EXPORT, "--xlsx", xlsx, "--data-dir", data_dir]
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0:
            fail(r.stderr or r.stdout)

        after = json.load(open(terrain_path, encoding="utf-8"))["farma"]["bonus"]["zywnosc"]
        if after != probe:
            fail(f"farma.bonus.zywnosc: oczekiwano {probe}, jest {after}")

        yields_path = os.path.join(data_dir, "terrain-yields.json")
        if os.path.isfile(yields_path):
            for row in json.load(open(yields_path, encoding="utf-8"))["terrain_types"]:
                if row.get("Teren") == "Łąka":
                    if row.get("Żywność") != 77:
                        fail(f"Łąka Żywność: oczekiwano 77, jest {row.get('Żywność')}")
                    break

    print("OK: Panel-A round-trip (ulepszenia + plony)")


if __name__ == "__main__":
    main()
