#!/usr/bin/env python3
# export-dystans-edycja.py — Excel do ręcznej korekty statów dystansowych (TW v3)
# Uruchom: python panele-sterowania/export-dystans-edycja.py
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:
    raise SystemExit("Brak openpyxl: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
UNITS = ROOT / "gra" / "data" / "units.json"
BACKUP = UNITS.with_suffix(".json.bak-TW-v3-2026-06-29")
OUT = Path(__file__).resolve().parent / "TW-dystans-edycja.xlsx"

HEADERS = [
    "Jednostka",
    "Nazwa EN",
    "Epoka",
    "Rola (linia)",
    "Zasięg (hex)",
    "Ilość pocisków",
    "Stary Atak dystansowy (ref)",
    "missileAttack",
    "wallAttack",
    "Uwagi",
]

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT = Font(bold=True, color="FFFFFF")
EDIT_FILL = PatternFill("solid", fgColor="FFF8E1")
REF_FILL = PatternFill("solid", fgColor="F0F0F0")
KAT_FILL = PatternFill("solid", fgColor="E8F4FC")


def load_old_missile() -> dict[str, int | float]:
    if not BACKUP.is_file():
        return {}
    data = json.loads(BACKUP.read_text(encoding="utf-8"))
    return {
        u["Jednostka"]: u.get("Atak dystansowy", 0) or 0
        for u in data
    }


def is_ranged(u: dict) -> bool:
    if u.get("missileAttack") or u.get("wallAttack"):
        return True
    shots = u.get("Ilość pocisków", 0)
    rng = u.get("Zasięg ataku (hex)", 0)
    if isinstance(shots, (int, float)) and shots > 0:
        return True
    if isinstance(rng, (int, float)) and rng > 0:
        return True
    rola = (u.get("Rola (linia)") or "").lower()
    return "dystans" in rola or "obleż" in rola or "oblezen" in rola


def main() -> None:
    units = json.loads(UNITS.read_text(encoding="utf-8"))
    old_ms = load_old_missile()
    ranged = [u for u in units if is_ranged(u)]
    ranged.sort(key=lambda u: -(u.get("missileAttack") or u.get("wallAttack") or 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dystans-TW-v3"
    ws.append(["TW v3 — edycja statów dystansowych"])
    ws.append([f"Wygenerowano: {date.today().isoformat()}"])
    ws.append(["Katapulta: missileAttack + wallAttack. Taran/Wieża: wallAttack (oblężenie). Reszta dystansu: missileAttack."])
    ws.append([])

    hr = ws.max_row + 1
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for u in ranged:
        name = u["Jednostka"]
        is_siege = (u.get("Rola (linia)") == "Oblężnicza")
        is_kat = name == "Katapulta"
        row = [
            name,
            u.get("Nazwa EN", ""),
            u.get("Epoka", ""),
            u.get("Rola (linia)", ""),
            u.get("Zasięg ataku (hex)", "—"),
            u.get("Ilość pocisków", "—"),
            old_ms.get(name, "—"),
            u.get("missileAttack", 0),
            u.get("wallAttack", "") if is_siege else "",
            (u.get("Uwagi") or "")[:120],
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c == 7:
                cell.fill = REF_FILL
            elif c in (5, 6, 8, 9):
                cell.fill = KAT_FILL if (is_siege and c == 9) or (is_kat and c in (8, 9)) else EDIT_FILL

    widths = [28, 14, 10, 14, 12, 14, 18, 14, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    instr = wb.create_sheet("Instrukcja")
    for line in [
        ["Import: python panele-sterowania/import-dystans-edycja.py"],
        [""],
        ["Edytuj żółte kolumny — wszystkie trafiają do units.json:"],
        ["• Zasięg (hex) → Zasięg ataku (hex)"],
        ["• Ilość pocisków"],
        ["• missileAttack (atak dystansowy; Katapulta vs jednostki)"],
        ["• wallAttack — Katapulta, Taran, Wieża oblężnicza (vs mur/brama)"],
    ]:
        instr.append(line)

    wb.save(OUT)
    print(f"OK: {OUT} ({len(ranged)} jednostek)")


if __name__ == "__main__":
    main()
