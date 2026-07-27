#!/usr/bin/env python3
# gen-panel-a.py — generuje Panel-A.xlsx (Grupa A: mapa, ulepszenia terenu, E2, zasięgi)
# Uruchom z root projektu: python panele-sterowania/gen-panel-a.py
import json
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "gra", "data")
OUT = os.path.join(os.path.dirname(__file__), "Panel-A.xlsx")

HDR = ["ID", "Parametr", "Opis", "Wartość", "Zakres/dozwolone", "Jednostka", "Wpływ na grę", "Plik źródłowy"]
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

BONUS_FIELDS = ("zywnosc", "praca", "handel", "pieniadz", "kamien", "drewno")
FOOD_SHEET_KEYS = {
    "farma", "irygacja", "tarasy", "lodzie_rybackie", "oboz_lowiecki", "warzelnia_soli",
    "bydlo", "owce", "lama",
}

KANON_FOOD = {
    "farma.bonus.zywnosc": 3,
    "irygacja.bonus.zywnosc": 5,
    "tarasy.bonus.zywnosc": 3,
    "bydlo.bonus.zywnosc": 2,
    "bydlo.bonus.praca": 3,
    "owce.bonus.zywnosc": 1,
    "owce.bonus.praca": 2,
    "lama.bonus.zywnosc": 1,
    "lama.bonus.praca": 3,
    "lodzie_rybackie.bonus.zywnosc": 2,
    "lodzie_rybackie.bonus.praca": 3,
    "oboz_lowiecki.bonus.zywnosc": 1,
    "warzelnia_soli.bonus.zywnosc": 1,
}

DEFAULT_MAP_GEN = {
    "mgla": {
        "default_sight_jednostki": {"wartosc": 3, "opis": "Domyślny promień wzroku jednostki (hexDistance)"},
    },
    "gestosc": {
        "surowce_mult": {"low": 0.6, "medium": 1.0, "high": 1.4},
        "baseline_rarity_mult": 1.35,
        "rzeki_max_mala_mapa": {"low": 2, "medium": 5, "high": 8},
        "river_scale": {"mala": 1.0, "srednia": 1.35, "duza": 1.7, "ogromna": 2.1},
        "desert_noise_threshold": {"low": 0.68, "medium": 0.63, "high": 0.58},
        "forest_noise_threshold": {"low": 0.65, "medium": 0.58, "high": 0.50},
    },
    "mapa_skala": {
        "aktywne_typy": {"mala": 3, "srednia": 5, "duza": 7, "ogromna": 9},
        "domyslni_rywale": {"mala": 2, "srednia": 4, "duza": 6, "ogromna": 8},
    },
}

# generator.ts — rozmiary siatki (kod jeszcze hardcoded; JSON = Panel-A export)
ROZMIAR_DIMS = {
    "malenki": (38, 26),
    "maly": (54, 37),
    "standardowy": (84, 60),
    "duzy": (120, 84),
    "ogromny": (168, 119),
}
DEFAULT_GEN_WIDTH = 36
DEFAULT_GEN_HEIGHT = 28

# gen-helpers DEPOSIT_RULES — tylko rarity (allowedOn zostaje w kodzie)
DEPOSIT_RARITY_DEFAULT = {
    "miedz": 0.10, "zelazo": 0.08, "glina": 0.10, "konie": 0.10, "wegiel": 0.10,
    "owce": 0.08, "bydlo": 0.07, "lama": 0.06, "luksus": 0.06, "sol": 0.12,
}
METAL_DEPOSIT_MIN_ERA = {"miedz": 2, "zelazo": 3}

YIELD_FIELDS = (
    ("zywnosc", "Żywność"), ("praca", "Praca"), ("podatek", "Podatek"),
    ("drewno", "Drewno"), ("kamien", "Kamień"),
)
TEREN_SLUG = {
    "Łąka": "laka", "Równina": "rownina", "Wzgórza": "wzgorza", "Góry": "gory",
    "Wybrzeże": "wybrzeze", "Morze": "morze", "Pustynia": "pustynia",
}
MOD_SLUG = {"Rzeka": "rzeka", "Las (nakładka)": "las"}


def load_json(name):
    path = os.path.join(DATA, name)
    if not os.path.isfile(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_info(ws):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    rows = [
        ("Panel sterowania — GRUPA A (Mapa / ulepszenia / generator)", ""),
        ("", ""),
        ("Jak używać (Maciej)", ""),
        ("1.", "Otwórz ten plik, edytuj kolumnę Wartość, zapisz."),
        ("2.", "W czacie grupy A napisz: eksportuj panel"),
        ("3.", "Agent zaktualizuje JSON w grze — bez terminala po Twojej stronie."),
        ("", ""),
        ("Priorytety lane A", "P1 Panel ✅ · P2 FOOD ✅ · P3 E2 ✅ · A5 miasta ✅ · balans = ten plik"),
        ("Kanon FOOD", "docs/decyzje/KANON-ULEPSZENIA-ZYWNOSC-HODOWLA.md"),
        ("Usunięte z gry", "plantacja / luksus (D3) — brak wierszy w panelu"),
        ("Łodzie rybackie", "A-R7: tylko terytorium miasta — reguła w kodzie, nie parametr Excel"),
        ("E2 vs Panel-E", "Balans gęstości = Generator-E2 (ten plik); suwaki kreatora = Panel-E"),
        ("Merge PANEL", "A-M1…A-M3 ✅ · legacy w docs/archiwum/panele-legacy/"),
        ("Wygenerowano", f"gen-panel-a.py · {date.today().isoformat()}"),
    ]
    for i, (a, b) in enumerate(rows, 1):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=14)


def write_param_sheet(wb, title, items):
    ws = wb.create_sheet(title)
    for c, h in enumerate(HDR, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    widths = [14, 34, 46, 12, 16, 14, 38, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for idx, it in enumerate(items, 2):
        ws.cell(idx, 1, it.get("id", idx - 1))
        ws.cell(idx, 2, it["key"])
        ws.cell(idx, 3, it.get("opis", ""))
        ws.cell(idx, 4, it["value"])
        ws.cell(idx, 5, it.get("zakres", ""))
        ws.cell(idx, 6, it.get("jednostka", ""))
        ws.cell(idx, 7, it.get("wplyw", ""))
        ws.cell(idx, 8, it.get("plik", ""))
        for c in (3, 7):
            ws.cell(idx, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


def node_val(terrain, imp_key, kind, sub=None):
    if imp_key not in terrain or not isinstance(terrain[imp_key], dict):
        return None
    node = terrain[imp_key]
    if kind == "koszt_praca":
        return node.get("koszt_praca")
    if kind == "bonus" and sub:
        return node.get("bonus", {}).get(sub)
    return node.get(kind)


def flatten_zywnosc_kanon(terrain):
    specs = [
        ("A-FOOD-FARMA-E", "farma.epoka", "farma", "epoka", None, "Farma — epoka odblokowania"),
        ("A-FOOD-FARMA-Z", "farma.bonus.zywnosc", "farma", "bonus", "zywnosc", "Farma — bonus żywności na heksie"),
        ("A-FOOD-FARMA-K", "farma.koszt_praca", "farma", "koszt_praca", None, "Koszt budowy farmy (Praca ze skarbca)"),
        ("A-FOOD-IRYG-E", "irygacja.epoka", "irygacja", "epoka", None, "Irygacja — epoka"),
        ("A-FOOD-IRYG-Z", "irygacja.bonus.zywnosc", "irygacja", "bonus", "zywnosc", "Irygacja — bonus żywności"),
        ("A-FOOD-IRYG-K", "irygacja.koszt_praca", "irygacja", "koszt_praca", None, "Koszt irygacji"),
        ("A-FOOD-TAR-E", "tarasy.epoka", "tarasy", "epoka", None, "Tarasy — epoka (Brąz)"),
        ("A-FOOD-TAR-Z", "tarasy.bonus.zywnosc", "tarasy", "bonus", "zywnosc", "Tarasy (Wzgórza, Chińczycy+Inkowie)"),
        ("A-FOOD-TAR-K", "tarasy.koszt_praca", "tarasy", "koszt_praca", None, "Koszt tarasów"),
        ("A-FOOD-BYD-E", "bydlo.epoka", "bydlo", "epoka", None, "Bydło — epoka odblokowania"),
        ("A-FOOD-BYD-Z", "bydlo.bonus.zywnosc", "bydlo", "bonus", "zywnosc", "Bydło — bonus żywności (płaski, + farma lub solo)"),
        ("A-FOOD-BYD-P", "bydlo.bonus.praca", "bydlo", "bonus", "praca", "Bydło — produkcja"),
        ("A-FOOD-BYD-K", "bydlo.koszt_praca", "bydlo", "koszt_praca", None, "Koszt pastwiska bydła"),
        ("A-FOOD-OWC-E", "owce.epoka", "owce", "epoka", None, "Owce — epoka odblokowania"),
        ("A-FOOD-OWC-Z", "owce.bonus.zywnosc", "owce", "bonus", "zywnosc", "Owce — solo wzgórze"),
        ("A-FOOD-OWC-P", "owce.bonus.praca", "owce", "bonus", "praca", "Owce — produkcja"),
        ("A-FOOD-OWC-K", "owce.koszt_praca", "owce", "koszt_praca", None, "Koszt owiec"),
        ("A-FOOD-LAM-E", "lama.epoka", "lama", "epoka", None, "Lama — epoka (Inkowie)"),
        ("A-FOOD-LAM-Z", "lama.bonus.zywnosc", "lama", "bonus", "zywnosc", "Lama — solo (Inkowie)"),
        ("A-FOOD-LAM-P", "lama.bonus.praca", "lama", "bonus", "praca", "Lama — produkcja"),
        ("A-FOOD-LAM-K", "lama.koszt_praca", "lama", "koszt_praca", None, "Koszt lamy"),
        ("A-FOOD-LOD-E", "lodzie_rybackie.epoka", "lodzie_rybackie", "epoka", None, "Łodzie rybackie — epoka"),
        ("A-FOOD-LOD-Z", "lodzie_rybackie.bonus.zywnosc", "lodzie_rybackie", "bonus", "zywnosc", "Łodzie rybackie — żywność"),
        ("A-FOOD-LOD-P", "lodzie_rybackie.bonus.praca", "lodzie_rybackie", "bonus", "praca", "Łodzie — produkcja"),
        ("A-FOOD-LOD-K", "lodzie_rybackie.koszt_praca", "lodzie_rybackie", "koszt_praca", None, "Koszt łodzi rybackich"),
        ("A-FOOD-OBO-E", "oboz_lowiecki.epoka", "oboz_lowiecki", "epoka", None, "Obóz łowiecki — epoka"),
        ("A-FOOD-OBO-Z", "oboz_lowiecki.bonus.zywnosc", "oboz_lowiecki", "bonus", "zywnosc", "Obóz łowiecki — żywność"),
        ("A-FOOD-OBO-$", "oboz_lowiecki.bonus.pieniadz", "oboz_lowiecki", "bonus", "pieniadz", "Obóz łowiecki — pieniądze"),
        ("A-FOOD-OBO-K", "oboz_lowiecki.koszt_praca", "oboz_lowiecki", "koszt_praca", None, "Koszt obozu"),
        ("A-FOOD-SOL-E", "warzelnia_soli.epoka", "warzelnia_soli", "epoka", None, "Warzelnia soli — epoka"),
        ("A-FOOD-SOL-Z", "warzelnia_soli.bonus.zywnosc", "warzelnia_soli", "bonus", "zywnosc", "Warzelnia soli — żywność"),
        ("A-FOOD-SOL-$", "warzelnia_soli.bonus.pieniadz", "warzelnia_soli", "bonus", "pieniadz", "Warzelnia soli — pieniądze"),
        ("A-FOOD-SOL-K", "warzelnia_soli.koszt_praca", "warzelnia_soli", "koszt_praca", None, "Koszt warzelni soli"),
    ]
    out = []
    for pid, param_key, imp_key, kind, sub, opis in specs:
        val = KANON_FOOD.get(param_key)
        if imp_key:
            got = node_val(terrain, imp_key, kind, sub)
            if got is not None and param_key not in KANON_FOOD:
                val = got
            elif param_key in KANON_FOOD:
                val = KANON_FOOD[param_key]
            elif got is not None:
                val = got
        if kind == "koszt_praca":
            jednostka = "Praca"
        elif kind == "epoka":
            jednostka = "1=Kamień…3=Żelazo"
        elif sub == "zywnosc":
            jednostka = "żywność/turę"
        elif sub == "pieniadz":
            jednostka = "zł/turę"
        else:
            jednostka = "Praca/turę"
        zakres = "eksport → JSON" if imp_key else "info — bez eksportu"
        out.append({
            "id": pid,
            "key": param_key,
            "value": val,
            "opis": opis,
            "jednostka": jednostka,
            "zakres": zakres,
            "wplyw": "okolica tileYield · kod warstw = P2",
            "plik": "terrain-improvements.json",
        })
    return out


def flatten_ulepszenia_inne(terrain):
    skip = {"_meta", "_miasto_zasieg_ref"} | FOOD_SHEET_KEYS
    out = []
    seq = 0
    scalar_meta = {
        "koszt_praca": ("Koszt budowy", "Praca", "terrain-improvements.json"),
        "epoka": ("Epoka odblokowania", "1=Kamień…3=Żelazo", "terrain-improvements.json"),
        "surowiecOdblokowany": ("Odblokowany surowiec (ASCII)", "klucz lub puste", "terrain-improvements.json"),
        "bonus_obrona_proc": ("Bonus obrony (obozowanie)", "%", "terrain-improvements.json"),
        "bonus_wymaga_obozowania": ("Bonus obrony tylko przy obozowaniu", "tak/nie", "terrain-improvements.json"),
        "zasieg_terytorium": ("Rozszerzenie terytorium", "heksy", "terrain-improvements.json"),
        "zasieg_pol": ("Zasięg kontroli pola", "heksy", "terrain-improvements.json"),
        "zasieg_kontroli": ("Zasięg kontroli (fort)", "heksy", "terrain-improvements.json"),
    }
    for imp_key, node in terrain.items():
        if imp_key in skip or not isinstance(node, dict):
            continue
        nazwa = node.get("nazwa", imp_key)
        for field, (desc, unit, plik) in scalar_meta.items():
            if field not in node:
                continue
            seq += 1
            param = f"{imp_key}.{field}"
            out.append({
                "id": f"A-IMP-{seq:03d}",
                "key": param,
                "value": node[field],
                "opis": f"{nazwa} — {desc}",
                "jednostka": unit,
                "zakres": "",
                "wplyw": f"ulepszenie {imp_key} · placement/render MAPA",
                "plik": plik,
            })
        bonus = node.get("bonus") or {}
        for bf in BONUS_FIELDS:
            if bf not in bonus:
                continue
            seq += 1
            param = f"{imp_key}.bonus.{bf}"
            out.append({
                "id": f"A-IMP-{seq:03d}",
                "key": param,
                "value": bonus[bf],
                "opis": f"{nazwa} — bonus {bf}",
                "jednostka": "pkt/turę",
                "zakres": "",
                "wplyw": "tileYield / surowce",
                "plik": "terrain-improvements.json",
            })
        wyc = node.get("wycinka")
        if isinstance(wyc, dict):
            for wf, wdesc in (("praca_per_tura", "Praca na turę wycinki"), ("tury", "Liczba tur wycinki")):
                if wf not in wyc:
                    continue
                seq += 1
                param = f"{imp_key}.wycinka.{wf}"
                out.append({
                    "id": f"A-IMP-{seq:03d}",
                    "key": param,
                    "value": wyc[wf],
                    "opis": f"{nazwa} — {wdesc}",
                    "jednostka": "Praca" if wf == "praca_per_tura" else "tury",
                    "zakres": "",
                    "wplyw": "wyrąb lasu — 3 tury potem teren bazowy",
                    "plik": "terrain-improvements.json",
                })
    return out


def flatten_zasiegi_mgla(terrain, map_gen):
    mg = (map_gen or DEFAULT_MAP_GEN).get("mgla", DEFAULT_MAP_GEN["mgla"])
    items = [
        {
            "id": "A-FOG-SIGHT",
            "key": "mgla.default_sight_jednostki",
            "value": mg.get("default_sight_jednostki", {}).get("wartosc", 3),
            "opis": "Domyślny promień wzroku jednostki (hexDistance)",
            "jednostka": "heksy",
            "zakres": "2–5 typowo",
            "wplyw": "mgła — visibility.ts DEFAULT_SIGHT (wpięcie Integrator)",
            "plik": "map-gen-params.json",
        },
        {
            "id": "A-ZAS-MIASTO",
            "key": "nota.miasto_zasieg_terytorium",
            "value": 10,
            "opis": "NOTA: miasto bazowe zasięg terytorium (nie ulepszenie)",
            "jednostka": "heksy",
            "zakres": "stałe ref EKONOMIA",
            "wplyw": "zakładanie kolejnego miasta — Panel-B / cities",
            "plik": "(tylko nota — bez eksportu)",
        },
    ]
    for imp_key, label in (("posterunek", "Posterunek"), ("fort", "Fort")):
        node = terrain.get(imp_key, {})
        if not isinstance(node, dict):
            continue
        for field in ("zasieg_terytorium", "zasieg_pol", "bonus_obrona_proc"):
            if field not in node:
                continue
            items.append({
                "id": f"A-ZAS-{imp_key[:4].upper()}-{field[:3].upper()}",
                "key": f"{imp_key}.{field}",
                "value": node[field],
                "opis": f"{label} — {field.replace('_', ' ')}",
                "jednostka": "%" if "proc" in field else "heksy",
                "zakres": "",
                "wplyw": "terytorium + mgła posterunku",
                "plik": "terrain-improvements.json",
            })
    return items


def flatten_generator_e2(map_gen):
    g = (map_gen or DEFAULT_MAP_GEN).get("gestosc", DEFAULT_MAP_GEN["gestosc"])
    sm = g.get("surowce_mult", {})
    rm = g.get("rzeki_max_mala_mapa", {})
    dt = g.get("desert_noise_threshold", {})
    ft = g.get("forest_noise_threshold", {})
    rs = g.get("river_scale", {})
    rows = [
        ("A-E2-RES-LOW", "gestosc.surowce_mult.low", sm.get("low", 0.6), "Mnożnik rzadkości złóż — Mało", "0,4–0,8", "×", "P3 generator placeDeposits"),
        ("A-E2-RES-MED", "gestosc.surowce_mult.medium", sm.get("medium", 1.0), "Mnożnik złóż — Normalnie", "0,8–1,2", "×", "P3"),
        ("A-E2-RES-HI", "gestosc.surowce_mult.high", sm.get("high", 1.4), "Mnożnik złóż — Dużo", "1,2–1,8", "×", "P3"),
        ("A-E2-RES-BASE", "gestosc.baseline_rarity_mult", g.get("baseline_rarity_mult", 1.35), "Bazowy boost rarity (więcej złóż niż dziś)", "1,0–2,0", "×", "P3 E2-Q2"),
        ("A-E2-RIV-LOW", "gestosc.rzeki_max_mala_mapa.low", rm.get("low", 2), "Max rzek na małej mapie — Mało", "1–4", "szt", "P3 E2-Q3"),
        ("A-E2-RIV-MED", "gestosc.rzeki_max_mala_mapa.medium", rm.get("medium", 5), "Max rzek mała mapa — Normalnie", "3–8", "szt", "P3"),
        ("A-E2-RIV-HI", "gestosc.rzeki_max_mala_mapa.high", rm.get("high", 8), "Max rzek mała mapa — Dużo", "6–12", "szt", "P3"),
        ("A-E2-DES-LOW", "gestosc.desert_noise_threshold.low", dt.get("low", 0.68), "Próg szumu pustyni — Mało", "0,5–0,8", "0–1", "P3 E2-Q4"),
        ("A-E2-DES-MED", "gestosc.desert_noise_threshold.medium", dt.get("medium", 0.63), "Próg pustyni — Normalnie", "", "0–1", "P3"),
        ("A-E2-DES-HI", "gestosc.desert_noise_threshold.high", dt.get("high", 0.58), "Próg pustyni — Dużo", "", "0–1", "P3"),
        ("A-E2-FOR-LOW", "gestosc.forest_noise_threshold.low", ft.get("low", 0.65), "Próg lasu logicznego — Mało", "", "0–1", "P3"),
        ("A-E2-FOR-MED", "gestosc.forest_noise_threshold.medium", ft.get("medium", 0.58), "Próg lasu — Normalnie", "", "0–1", "P3"),
        ("A-E2-FOR-HI", "gestosc.forest_noise_threshold.high", ft.get("high", 0.50), "Próg lasu — Dużo", "", "0–1", "P3"),
    ]
    for size in ("mala", "srednia", "duza", "ogromna"):
        rows.append((
            f"A-E2-RIV-SCL-{size[:3].upper()}",
            f"gestosc.river_scale.{size}",
            rs.get(size, DEFAULT_MAP_GEN["gestosc"]["river_scale"][size]),
            f"Skala liczby rzek vs mała mapa — {size}",
            "0,8–2,5",
            "×",
            "P3 skala rozmiaru mapy",
        ))
    out = []
    for pid, key, val, opis, zakres, jed, wplyw in rows:
        out.append({
            "id": pid,
            "key": key,
            "value": val,
            "opis": opis,
            "zakres": zakres,
            "jednostka": jed,
            "wplyw": wplyw,
            "plik": "map-gen-params.json (kod P3)",
        })
    return out


def flatten_mapa_skala(map_gen):
    ms = (map_gen or DEFAULT_MAP_GEN).get("mapa_skala", DEFAULT_MAP_GEN["mapa_skala"])
    at = ms.get("aktywne_typy", {})
    dr = ms.get("domyslni_rywale", {})
    out = []
    for size, label in (("mala", "Mała"), ("srednia", "Średnia"), ("duza", "Duża"), ("ogromna", "Ogromna")):
        out.append({
            "id": f"A-MAP-TYP-{size[:3].upper()}",
            "key": f"mapa_skala.aktywne_typy.{size}",
            "value": at.get(size, DEFAULT_MAP_GEN["mapa_skala"]["aktywne_typy"][size]),
            "opis": f"Aktywne typy cywilizacji na mapie {label}",
            "jednostka": "typy",
            "zakres": "3–9",
            "wplyw": "cluster spawn · E2-Q1",
            "plik": "map-gen-params.json",
        })
        out.append({
            "id": f"A-MAP-RYW-{size[:3].upper()}",
            "key": f"mapa_skala.domyslni_rywale.{size}",
            "value": dr.get(size, DEFAULT_MAP_GEN["mapa_skala"]["domyslni_rywale"][size]),
            "opis": f"Domyślna liczba rywali AI — mapa {label}",
            "jednostka": "cyw",
            "zakres": "1–8",
            "wplyw": "kreator nowej gry",
            "plik": "map-gen-params.json",
        })
    return out


def flatten_plony_terenow(yields_data):
    out = []
    seq = 0
    for row in yields_data.get("terrain_types", []):
        teren = row.get("Teren", "")
        slug = TEREN_SLUG.get(teren, teren.lower())
        for fk, col in YIELD_FIELDS:
            val = row.get(col)
            if val is None and fk == "podatek":
                val = row.get("Handel")
            if val is None and col not in row:
                continue
            seq += 1
            out.append({
                "id": f"A-PLON-{seq:03d}",
                "key": f"plony.types.{slug}.{fk}",
                "value": val,
                "opis": f"{teren} — bazowy plon {fk}",
                "jednostka": "pkt/turę",
                "zakres": "0–8 typowo",
                "wplyw": "economy.ts tileYield · okolica miasta",
                "plik": "terrain-yields.json",
            })
    for row in yields_data.get("terrain_modifiers", []):
        mod = row.get("Modyfikator", "")
        mslug = MOD_SLUG.get(mod, mod.lower())
        for fk, col in YIELD_FIELDS:
            val = row.get(col)
            if val is None and fk == "podatek":
                val = row.get("Handel")
            if val is None and col not in row:
                continue
            seq += 1
            out.append({
                "id": f"A-PLON-{seq:03d}",
                "key": f"plony.mod.{mslug}.{fk}",
                "value": val,
                "opis": f"Modyfikator {mod} — {fk}",
                "jednostka": "pkt/turę",
                "zakres": "-3…+5",
                "wplyw": "bonus nakładki / rzeki na heks",
                "plik": "terrain-yields.json",
            })
    return out


def flatten_ruch_po_terenie(movement):
    out = []
    costs = movement.get("costs", {})
    cost_labels = {
        "Laka": "Łąka", "Rownina": "Równina", "Pustynia": "Pustynia",
        "Wybrzeze": "Wybrzeże", "Wzgorza": "Wzgórza", "Gory": "Góry", "Morze": "Morze",
    }
    for ck, label in cost_labels.items():
        if ck not in costs:
            continue
        out.append({
            "id": f"A-RUCH-{ck[:3].upper()}",
            "key": f"ruch.costs.{ck}",
            "value": costs[ck],
            "opis": f"Koszt ruchu jednostki — {label}",
            "jednostka": "punkty ruchu",
            "zakres": "1–99 (99=nieprzechodnie)",
            "wplyw": "units/setup.ts movementCost",
            "plik": "terrain-movement.json",
        })
    if "forestExtra" in movement:
        out.append({
            "id": "A-RUCH-LAS",
            "key": "ruch.forestExtra",
            "value": movement["forestExtra"],
            "opis": "Dopłata ruchu przez las (nakładka)",
            "jednostka": "punkty ruchu",
            "zakres": "0–3",
            "wplyw": "ruch po heksie z lasem",
            "plik": "terrain-movement.json",
        })
    return out


def flatten_generator_rozmiary(map_gen):
    g = (map_gen or {}).get("generator", {})
    out = [
        {
            "id": "A-GEN-DEF-W",
            "key": "generator.default_width",
            "value": g.get("default_width", DEFAULT_GEN_WIDTH),
            "opis": "Domyślna szerokość siatki (dev / fallback generator.ts)",
            "jednostka": "heksy",
            "zakres": "20–200",
            "wplyw": "generateMap() gdy brak rozmiaru z menu",
            "plik": "map-gen-params.json (wpięcie Integrator P3)",
        },
        {
            "id": "A-GEN-DEF-H",
            "key": "generator.default_height",
            "value": g.get("default_height", DEFAULT_GEN_HEIGHT),
            "opis": "Domyślna wysokość siatki (dev / fallback)",
            "jednostka": "heksy",
            "zakres": "15–150",
            "wplyw": "generateMap() fallback",
            "plik": "map-gen-params.json",
        },
    ]
    dims = g.get("rozmiar_dims", {})
    labels = {
        "malenki": "Malenki", "maly": "Mały", "standardowy": "Standardowy",
        "duzy": "Duży", "ogromny": "Ogromny",
    }
    for key, (dw, dh) in ROZMIAR_DIMS.items():
        stored = dims.get(key, [dw, dh])
        w = stored[0] if isinstance(stored, (list, tuple)) else dw
        h = stored[1] if isinstance(stored, (list, tuple)) and len(stored) > 1 else dh
        lbl = labels.get(key, key)
        out.append({
            "id": f"A-GEN-{key[:3].upper()}-W",
            "key": f"generator.rozmiar_dims.{key}.w",
            "value": w,
            "opis": f"Rozmiar mapy {lbl} — szerokość",
            "jednostka": "heksy",
            "zakres": "",
            "wplyw": "menuLabelToDims · generujSwiat",
            "plik": "map-gen-params.json",
        })
        out.append({
            "id": f"A-GEN-{key[:3].upper()}-H",
            "key": f"generator.rozmiar_dims.{key}.h",
            "value": h,
            "opis": f"Rozmiar mapy {lbl} — wysokość",
            "jednostka": "heksy",
            "zakres": "",
            "wplyw": "menuLabelToDims · generujSwiat",
            "plik": "map-gen-params.json",
        })
    return out


def flatten_zloza_generator(map_gen):
    deps = (map_gen or {}).get("deposit_rules", {})
    metal = (map_gen or {}).get("metal_deposit_min_era", METAL_DEPOSIT_MIN_ERA)
    out = []
    for zid, rarity in DEPOSIT_RARITY_DEFAULT.items():
        node = deps.get(zid, {})
        rv = node.get("rarity", rarity) if isinstance(node, dict) else rarity
        out.append({
            "id": f"A-ZLO-R-{zid[:4].upper()}",
            "key": f"zloza.{zid}.rarity",
            "value": rv,
            "opis": f"Rzadkość złoża „{zid}” (placeDeposits)",
            "jednostka": "0–1 ułamek",
            "zakres": "0,02–0,25",
            "wplyw": "generator — ile heksów dostaje złoże (teren w kodzie)",
            "plik": "map-gen-params.json",
        })
    for zid, era in METAL_DEPOSIT_MIN_ERA.items():
        out.append({
            "id": f"A-ZLO-E-{zid[:4].upper()}",
            "key": f"zloza.{zid}.min_epoka",
            "value": metal.get(zid, era),
            "opis": f"Minimalna epoka widoczności złoża „{zid}”",
            "jednostka": "1=Kamień…3=Żelazo",
            "zakres": "1–3",
            "wplyw": "deposit-era.ts · mgła surowców",
            "plik": "map-gen-params.json",
        })
    return out


def write_eksporty(wb):
    ws = wb.create_sheet("_Eksporty")
    ws.cell(1, 1, "Po edycji Panel-A.xlsx — napisz w czacie: eksportuj panel").font = Font(bold=True, size=12)
    cmds = [
        ("Panel A (ten plik)", "python panele-sterowania/export-a.py"),
        ("Regeneracja z JSON", "python panele-sterowania/gen-panel-a.py"),
        ("JSON-y", "terrain-improvements · terrain-yields · terrain-movement · map-gen-params"),
    ]
    for i, (label, cmd) in enumerate(cmds, 3):
        ws.cell(i, 1, label)
        ws.cell(i, 2, cmd)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55


def main():
    terrain = load_json("terrain-improvements.json")
    if not terrain:
        print("Brak terrain-improvements.json")
        sys.exit(1)
    map_gen = load_json("map-gen-params.json")
    yields_data = load_json("terrain-yields.json") or {"terrain_types": [], "terrain_modifiers": []}
    movement = load_json("terrain-movement.json") or {"costs": {}, "forestExtra": 1}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_info(wb.create_sheet("_INFO", 0))

    write_param_sheet(wb, "Ulepszenia-FOOD", flatten_zywnosc_kanon(terrain))
    write_param_sheet(wb, "Ulepszenia-inne", flatten_ulepszenia_inne(terrain))
    write_param_sheet(wb, "Plony-terenow", flatten_plony_terenow(yields_data))
    write_param_sheet(wb, "Ruch-po-terenie", flatten_ruch_po_terenie(movement))
    write_param_sheet(wb, "Zasięgi-mgła", flatten_zasiegi_mgla(terrain, map_gen))
    write_param_sheet(wb, "Generator-rozmiary", flatten_generator_rozmiary(map_gen))
    write_param_sheet(wb, "Zloza-generator", flatten_zloza_generator(map_gen))
    write_param_sheet(wb, "Generator-E2", flatten_generator_e2(map_gen))
    write_param_sheet(wb, "Mapa-skala", flatten_mapa_skala(map_gen))
    write_eksporty(wb)

    wb.save(OUT)
    print(f"Zapisano: {OUT}")
    print(f"Arkusze: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
