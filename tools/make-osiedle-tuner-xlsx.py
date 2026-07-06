#!/usr/bin/env python3
"""Generuje D-START-OSIEDLE-tuner.xlsx — edycja bonusów osiedla (Pop, Praw, Sz)."""
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    raise SystemExit("Brak openpyxl: pip install openpyxl")

OUT = Path(__file__).resolve().parent.parent / "docs" / "balans" / "D-START-OSIEDLE-tuner.xlsx"

# Kanon z society-params.json (Maciej final 2026-07-02)
DATA = {
    "Easy": {
        "praw": [9, 7, 5, 3],
        "sz": [4, 3, 2, 1],
    },
    "Normal": {
        "praw": [7, 5, 3, 1],
        "sz": [3, 2, 1, 0],
    },
    "Hard": {
        "praw": [5, 3, 2, 1],
        "sz": [1, 1, 0, 0],
    },
}

wb = Workbook()
ws = wb.active
ws.title = "Osiedle"

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF")
section_fill = PatternFill("solid", fgColor="D9E2F3")
section_font = Font(bold=True)

ws["A1"] = "D-START-OSIEDLE — tuner bonusów osiedla"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:D1")
ws["A2"] = "Edytuj żółte komórki (Bonus Praw / Bonus Sz). Pop = 1–4. Po grze wklej wartości do Mastera."
ws.merge_cells("A2:D2")

row = 4
for diff in ("Easy", "Normal", "Hard"):
    ws.cell(row, 1, diff).font = section_font
    ws.cell(row, 1).fill = section_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    for col, label in enumerate(["Pop", "Bonus Praw", "Bonus Sz"], start=1):
        c = ws.cell(row, col, label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    row += 1

    edit_fill = PatternFill("solid", fgColor="FFF2CC")
    for i, pop in enumerate(range(1, 5)):
        ws.cell(row, 1, pop).alignment = Alignment(horizontal="center")
        pr = ws.cell(row, 2, DATA[diff]["praw"][i])
        sz = ws.cell(row, 3, DATA[diff]["sz"][i])
        pr.fill = edit_fill
        sz.fill = edit_fill
        pr.alignment = Alignment(horizontal="center")
        sz.alignment = Alignment(horizontal="center")
        row += 1
    row += 1

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 4

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Zapisano: {OUT}")
