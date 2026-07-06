#!/usr/bin/env python3
"""Import docs/balans/D-START-OSIEDLE-tuner.xlsx → gra/data/society-params.json (Praw + Sz)."""
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "docs" / "balans" / "D-START-OSIEDLE-tuner.xlsx"
JSON = ROOT / "gra" / "data" / "society-params.json"


def read_xlsx() -> dict[str, dict[str, list[int]]]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Osiedle"]
    data: dict[str, dict[str, list[int]]] = {}
    current: str | None = None
    for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
        if not row or row[0] is None:
            continue
        if row[0] in ("Easy", "Normal", "Hard"):
            current = row[0].lower()
            data[current] = {"praw": [], "sz": []}
            continue
        if row[0] == "Pop" or current is None:
            continue
        if isinstance(row[0], (int, float)) and 1 <= int(row[0]) <= 4:
            data[current]["praw"].append(int(row[1] or 0))
            data[current]["sz"].append(int(row[2] or 0))
    for diff in ("easy", "normal", "hard"):
        if len(data.get(diff, {}).get("praw", [])) != 4:
            raise SystemExit(f"Brak 4 wierszy Praw dla {diff} w {XLSX}")
        if len(data[diff]["sz"]) != 4:
            raise SystemExit(f"Brak 4 wierszy Sz dla {diff} w {XLSX}")
    return data


def main() -> None:
    tables = read_xlsx()
    sp = json.loads(JSON.read_text(encoding="utf-8"))
    for diff in ("easy", "normal", "hard"):
        sp["prawo"]["prawo_bonus_osiedle_pop"][diff] = tables[diff]["praw"]
        sp["szczescie"]["szczescie_bonus_osiedle_pop"][diff] = tables[diff]["sz"]
    JSON.write_text(json.dumps(sp, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {XLSX.name} -> society-params.json")
    for diff in ("easy", "normal", "hard"):
        print(f"  {diff}: Praw={tables[diff]['praw']} Sz={tables[diff]['sz']}")


if __name__ == "__main__":
    main()
