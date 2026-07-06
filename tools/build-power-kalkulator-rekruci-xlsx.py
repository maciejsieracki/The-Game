#!/usr/bin/env python3
"""Excel pomocniczy: rekruci / werb / ekw. jednostek — epoki 1–5 (Maciej)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "docs" / "decyzje" / "POWER-pomoc-rekruci-epoki-1-5.xlsx"

EPOKI_5 = [
    # ep, ludekNaLudka, maxRekNaLudka, kosztWerbu
    (1, 10000, 1000, 100),
    (2, 20000, 2000, 200),
    (3, 40000, 4000, 400),
    (4, 80000, 8000, 800),
    (5, 160000, 16000, 1600),
]

HEADER_FILL = PatternFill("solid", fgColor="1E2430")
HEADER_FONT = Font(bold=True, color="E0B24A")
INPUT_FILL = PatternFill("solid", fgColor="FFF8E0")
COEFF_FILL = PatternFill("solid", fgColor="E8F4E8")
INFO_FILL = PatternFill("solid", fgColor="F0F0F0")
ACCENT_FILL = PatternFill("solid", fgColor="E8F8E8")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def build_zalozenia(wb):
    ws = wb.create_sheet("Założenia", 0)
    ws["A1"] = "POWER — pomocnik rekruci / werb (epoki 1–5)"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "Kanon manpower (The Game):",
        "• 1 ludek (slot populacji miasta 1–10) → ludność abs. i max rekrutów zależą od epoki imperium.",
        "• Co epokę: ludność abs. ×2, max rekrutów na ludka ×2, koszt werbu 1 jednostki ×2.",
        "• Koszt werbu = 10% max puli rekrutów na ludka (stała reguła we wszystkich epokach).",
        "",
        "Obiektywny parametr Power (rekruci) — NIE surowa pula, NIE % max (sufit 100):",
        "  ekw. jednostek = floor(rekruci_bieżący ÷ koszt_werbu[epoka])",
        "  → ile jednostek możesz werbować z obecnej puli (neutralne względem skali epoki).",
        "",
        "Pełna pula (100% max): ekw. = suma_ludków × 10  (zawsze, każda epoka).",
        "",
        "Power z rekrutów (do kalibracji P-A):",
        "  pkt_rekruci = ekw. jednostek × współczynnik  (domyślnie 10 — jak armia na mapie).",
        "",
        "Armia na mapie (osobny składnik Power):",
        "  pkt_armia = liczba_jednostek × współczynnik_armia  (domyślnie 10).",
        "",
        "Pułap 100% puli: przy tym samym % wypełnienia Power z rekrutów jest stały między epokami.",
        "Pułap „stara pula”: jeśli epoka skoczyła, a pula nie urosła — ekw. spada (realny stan gry).",
        "",
        "Źródło liczb: gra/data/epoka-ludnosc-manpower.json · epoki 1–5.",
        "Powiązany: POWER-kalkulator-Maciej.xlsx (pełny Power P-A).",
    ]
    for i, text in enumerate(lines, start=2):
        ws.cell(i, 1, text)
        if text.startswith("•") or text.startswith("  "):
            ws.cell(i, 1).font = Font(size=11)
    ws.column_dimensions["A"].width = 95


def build_tabela_epok(wb):
    ws = wb.create_sheet("Tabela epok 1-5")
    ws["A1"] = "Skala kanonu — mnożnik ×2 co epokę (ep.1 → ep.5)"
    ws["A1"].font = Font(bold=True)

    headers = [
        "Epoka",
        "Ludzi / 1 ludka",
        "Max rekr. / 1 ludka",
        "Koszt werbu 1 jedn.",
        "×2 vs poprzednia",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))

    for i, (ep, lud, mp, cost) in enumerate(EPOKI_5, start=4):
        ws.cell(i, 1, ep)
        ws.cell(i, 2, lud)
        ws.cell(i, 3, mp)
        ws.cell(i, 4, cost)
        if ep == 1:
            ws.cell(i, 5, "—")
        else:
            ws.cell(i, 5, 2)
        for c in range(2, 5):
            ws.cell(i, c).number_format = "#,##0"

    r = 10
    ws.cell(r, 1, "Przykład: 10 ludków, pełna pula (100%)")
    ws.cell(r, 1).font = Font(bold=True)
    sub = ["Max rekruci", "Ekw. jednostek", "→ stałe 100 ekw."]
    for j, label in enumerate(sub):
        ws.cell(r + 1 + j, 1, label)
    for i, ep_row in enumerate(range(4, 9)):
        col = i + 2
        L = get_column_letter(col)
        ws.cell(r, col, f"Ep. {i + 1}")
        ws.cell(r, col).font = Font(bold=True)
        ws.cell(r + 1, col, f"=10*C{ep_row}")
        ws.cell(r + 2, col, f"=IF(D{ep_row}=0,0,{L}{r+1}/D{ep_row})")
        ws.cell(r + 3, col, f"=IF({L}{r+2}=100,\"OK\",\"!\")")

    ws.column_dimensions["A"].width = 28
    for col in "BCDE":
        ws.column_dimensions[col].width = 18


def build_kalkulator(wb):
    ws = wb.create_sheet("Kalkulator epok")
    ws["A1"] = "Przelicznik — wpisz żółte · patrz ekw. i Power"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "Suma ludków (wszystkie miasta)"
    ws["B3"] = 10
    ws["B3"].fill = INPUT_FILL
    ws["A4"] = "Współcz. Power / ekw. jednostki"
    ws["B4"] = 10
    ws["B4"].fill = COEFF_FILL
    ws["A5"] = "Współcz. Power / jednostka na mapie"
    ws["B5"] = 10
    ws["B5"].fill = COEFF_FILL

    # Scenariusze % puli (kolumny)
    pct_row = 7
    ws.cell(pct_row, 1, "% wypełnienia puli (scenariusze → kolumny)")
    ws.cell(pct_row, 1).font = Font(bold=True)
    pcts = [("D", 25), ("E", 50), ("F", 75), ("G", 100)]
    for col_letter, pct in pcts:
        col = ord(col_letter) - ord("A") + 1
        ws.cell(pct_row, col, f"{pct}%")
        ws.cell(pct_row, col).fill = INPUT_FILL
        ws.cell(pct_row, col).border = BORDER

    # Jednostki na mapie per epoka
    ws.cell(pct_row + 1, 1, "Jednostki na mapie (per epoka → wiersz)")
    ws.cell(pct_row + 1, 1).font = Font(italic=True, color="555555")

    hdr = 9
    headers = [
        "Epoka",
        "Ludzi/ludek",
        "Max rek/ludek",
        "Koszt werbu",
        "Max rekruci",
        "Jedn. mapa",
        "25% pula",
        "50% pula",
        "75% pula",
        "100% pula",
        "Ekw. @25%",
        "Ekw. @50%",
        "Ekw. @75%",
        "Ekw. @100%",
        "Pkt rekr.@100%",
        "Pkt armia",
        "Suma P(rek+arm)",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(hdr, i, h)
    style_header(ws, hdr, len(headers))

    data_start = hdr + 1
    for i, (ep, lud, mp, cost) in enumerate(EPOKI_5):
        r = data_start + i
        ws.cell(r, 1, ep)
        ws.cell(r, 2, lud)
        ws.cell(r, 3, mp)
        ws.cell(r, 4, cost)
        ws.cell(r, 5, f"=$B$3*C{r}")
        c_arm = ws.cell(r, 6, [3, 5, 8, 12, 18][i])  # rosnąca armia demo
        c_arm.fill = INPUT_FILL
        c_arm.border = BORDER
        # pule per %
        for j, pct in enumerate([25, 50, 75, 100]):
            col = 7 + j
            ws.cell(r, col, f"=ROUND(E{r}*{pct}/100,0)")
            ws.cell(r, col).number_format = "#,##0"
        # ekw per %
        for j in range(4):
            pool_col = get_column_letter(7 + j)
            ekw_col = 11 + j
            ws.cell(r, ekw_col, f"=IF(D{r}=0,0,FLOOR({pool_col}{r}/D{r},1))")
        ws.cell(r, 15, f"=IF(D{r}=0,0,FLOOR(J{r}/D{r},1))")
        ws.cell(r, 16, f"=F{r}*$B$5")
        ws.cell(r, 17, f"=O{r}*$B$4+F{r}*$B$5")
        ws.cell(r, 15).fill = ACCENT_FILL
        ws.cell(r, 17).font = Font(bold=True)

    ws.column_dimensions["A"].width = 8
    for c in range(2, 18):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.column_dimensions["Q"].width = 14

    note = data_start + len(EPOKI_5) + 2
    ws.cell(note, 1, "Kolumny G–J: rekruci_bieżący przy danym % max · K–N: ekw. jednostek · O: pkt z pełnej puli")
    ws.cell(note + 1, 1, "P: armia · Q: pkt rekr (100%) + pkt armia — kręć B3–B5 i jednostki w kol. F")


def build_case_stuck(wb):
    """Kejs: ta sama surowa pula 8000 — epoka rośnie, koszt ×2."""
    ws = wb.create_sheet("Kejs stara pula")
    ws["A1"] = "Kejs: 10 ludków · pula ZAMROŻONA 8000 rekrutów · epoka rośnie"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:I1")

    ws["A3"] = "Stała pula (rekruci)"
    ws["B3"] = 8000
    ws["B3"].fill = INPUT_FILL
    ws["A4"] = "Ludki"
    ws["B4"] = 10
    ws["B4"].fill = INPUT_FILL
    ws["A5"] = "Współcz. pkt"
    ws["B5"] = 10
    ws["B5"].fill = COEFF_FILL

    headers = [
        "Epoka",
        "Koszt werbu",
        "Max rekruci",
        "% max",
        "Ekw. jednostek",
        "Pkt rekr",
        "Jedn. mapa",
        "Pkt armia",
        "Uwaga",
    ]
    hdr = 7
    for i, h in enumerate(headers, 1):
        ws.cell(hdr, i, h)
    style_header(ws, hdr, len(headers))

    armies = [3, 5, 8, 12, 18]
    for i, (ep, _, mp, cost) in enumerate(EPOKI_5):
        r = hdr + 1 + i
        ws.cell(r, 1, ep)
        ws.cell(r, 2, cost)
        ws.cell(r, 3, f"=$B$4*{mp}")
        ws.cell(r, 4, f"=IF(C{r}=0,0,ROUND($B$3/C{r}*100,1))")
        ws.cell(r, 5, f"=IF(B{r}=0,0,FLOOR($B$3/B{r},1))")
        ws.cell(r, 6, f"=E{r}*$B$5")
        ws.cell(r, 7, armies[i])
        ws.cell(r, 7).fill = INPUT_FILL
        ws.cell(r, 8, f"=G{r}*$B$5")
        uwagi = [
            "baseline",
            "epoka ↑ koszt ×2 — ekw. spada o połowę",
            "pula coraz mniejszy % max",
            "…",
            "…",
        ]
        ws.cell(r, 9, uwagi[i])

    r2 = hdr + len(EPOKI_5) + 3
    ws.cell(r2, 1, "Porównanie: ta sama mobilizacja (80% max) — pula rośnie z epoką")
    ws.cell(r2, 1).font = Font(bold=True, color="336633")
    headers2 = ["Epoka", "Pula (80% max)", "Koszt", "Ekw.", "Pkt rekr", "vs ep.1"]
    for i, h in enumerate(headers2, 1):
        ws.cell(r2 + 1, i, h)
    style_header(ws, r2 + 1, len(headers2))

    base_row = r2 + 2
    for i, (ep, _, mp, cost) in enumerate(EPOKI_5):
        row = base_row + i
        ws.cell(row, 1, ep)
        ws.cell(row, 2, f"=ROUND($B$4*{mp}*0.8,0)")
        ws.cell(row, 3, cost)
        ws.cell(row, 4, f"=IF(C{row}=0,0,FLOOR(B{row}/C{row},1))")
        ws.cell(row, 5, f"=D{row}*$B$5")
        if i == 0:
            ws.cell(row, 6, "—")
        else:
            ws.cell(row, 6, f"=E{row}-E{base_row}")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["I"].width = 36
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 14


def build_armia_vs_rekr(wb):
    ws = wb.create_sheet("Armia vs rekruci")
    ws["A1"] = "Ten sam ekw. mobilizacji — różna armia już na mapie (ep.3 · 10 ludków · 50% pula)"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Epoka"
    ws["B3"] = 3
    ws["B3"].fill = INPUT_FILL
    ws["A4"] = "Ludki"
    ws["B4"] = 10
    ws["A5"] = "% puli"
    ws["B5"] = 50
    ws["B5"].fill = INPUT_FILL
    ws["A6"] = "Współcz."
    ws["B6"] = 10
    ws["B6"].fill = COEFF_FILL

    ep3 = EPOKI_5[2]
    ws["A7"] = "Koszt werbu (auto ep.3)"
    ws["B7"] = ep3[3]
    ws["A8"] = "Max rekruci (auto)"
    ws["B8"] = f"=B4*{ep3[2]}"

    hdr = 10
    for i, h in enumerate(["Jedn. na mapie", "Ekw. z puli", "Pkt armia", "Pkt rekr", "Suma", "Armia %"], 1):
        ws.cell(hdr, i, h)
    style_header(ws, hdr, 6)

    armies = [0, 2, 5, 10, 20, 40, 80]
    for i, army in enumerate(armies):
        r = hdr + 1 + i
        ws.cell(r, 1, army)
        ws.cell(r, 1).fill = INPUT_FILL
        ws.cell(r, 2, f"=IF(B7=0,0,FLOOR(ROUND(B8*B5/100,0)/B7,1))")
        ws.cell(r, 3, f"=A{r}*$B$6")
        ws.cell(r, 4, f"=B{r}*$B$6")
        ws.cell(r, 5, f"=C{r}+D{r}")
        ws.cell(r, 6, f"=IF(E{r}=0,0,C{r}/E{r})")
        ws.cell(r, 6).number_format = "0%"

    ws.column_dimensions["A"].width = 16
    for c in "BCDEF":
        ws.column_dimensions[c].width = 14


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_zalozenia(wb)
    build_tabela_epok(wb)
    build_kalkulator(wb)
    build_case_stuck(wb)
    build_armia_vs_rekr(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Zapisano: {OUT}")


if __name__ == "__main__":
    main()
