#!/usr/bin/env python3
"""Generuje docs/decyzje/POWER-kalkulator-Maciej.xlsx — kalibracja P-A (Maciej)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

OUT = Path(__file__).resolve().parents[1] / "docs" / "decyzje" / "POWER-kalkulator-Maciej.xlsx"

EPOKI = [
    (1, 10000, 1000, 100),
    (2, 20000, 2000, 200),
    (3, 40000, 4000, 400),
    (4, 80000, 8000, 800),
    (5, 160000, 16000, 1600),
    (6, 320000, 32000, 3200),
    (7, 640000, 64000, 6400),
    (8, 1200000, 120000, 12000),
    (9, 2400000, 240000, 24000),
    (10, 4800000, 480000, 48000),
]

HEADER_FILL = PatternFill("solid", fgColor="1E2430")
HEADER_FONT = Font(bold=True, color="E0B24A")
INPUT_FILL = PatternFill("solid", fgColor="FFF8E0")
COEFF_FILL = PatternFill("solid", fgColor="E8F4E8")
NEW_FILL = PatternFill("solid", fgColor="E8EEF8")
WARN_FILL = PatternFill("solid", fgColor="FFE8E8")
INFO_FILL = PatternFill("solid", fgColor="F0F0F0")
REJECT_FILL = PatternFill("solid", fgColor="E8E8E8")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Pierwszy wiersz tabeli składników (Excel)
TBL = 19
TBL_END = 27  # 9 wierszy


def style_header(ws, row, cols, fill=HEADER_FILL):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def build_epoki(wb):
    ws = wb.create_sheet("Epoki", 0)
    ws["A1"] = "Epoka imperium"
    ws["B1"] = "Ludzi na 1 ludka"
    ws["C1"] = "Max rekrutów na 1 ludka"
    ws["D1"] = "Koszt werbu 1 jednostki"
    style_header(ws, 1, 4)
    for i, (e, ludek, mp, cost) in enumerate(EPOKI, start=2):
        ws.cell(i, 1, e)
        ws.cell(i, 2, ludek)
        ws.cell(i, 3, mp)
        ws.cell(i, 4, cost)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws["F1"] = "Ekw. jednostek z puli = rekruci_bieżący ÷ koszt werbu (neutralne względem epoki przy tym samym % puli)."
    ws["F1"].font = Font(italic=True, color="336633")


def build_kalkulator(wb):
    ws = wb.create_sheet("Kalkulator")
    ws["A1"] = "POWER — kalkulator P-A (The Game · Maciej)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "BEZ × epoki. Ludność = LUDKI. Rekruci = EKW. JEDNOSTEK (pula ÷ koszt werbu w epoce) — skaluje z grą, bez sufitu 100. "
        "Żółte wpisy · zielone współczynniki · niebieskie NOWE"
    )
    ws["A2"].font = Font(italic=True, color="553333")
    ws.merge_cells("A2:G2")

    labels = [
        (4, "Epoka imperium (1–10) — koszt werbu + INFO", 1),
        (5, "Suma ludków (sloty populacji, wszystkie miasta)", 100),
        (6, "Jednostki na mapie", 10),
        (7, "Wygrane bitwy (łącznie)", 10),
        (8, "Rekruci — bieżąca pula", 8000),
        (9, "Miasta", 10),
        (10, "Heksy terytorium", 500),
        (11, "Budynki wybudowane", 60),
        (12, "Odkrycia / tech zbadane", 16),
        (13, "Ulepszenia w terenie", 50),
    ]
    ws["A3"] = "WEJŚCIA (scenariusz)"
    ws["A3"].font = Font(bold=True)
    for row, label, default in labels:
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, default)
        c.fill = INPUT_FILL
        c.border = BORDER
        if row in (12, 13):
            c.fill = NEW_FILL

    # Pola pochodne (INFO)
    info = [
        (14, "Ludność absolutna (INFO — nie w Power)", "=B5*VLOOKUP(B4,Epoki!$A$2:$B$11,2,FALSE)"),
        (15, "Max rekrutów imperium (INFO)", "=B5*VLOOKUP(B4,Epoki!$A$2:$C$11,3,FALSE)"),
        (16, "Koszt werbu 1 jednostki (INFO)", "=VLOOKUP(B4,Epoki!$A$2:$D$11,4,FALSE)"),
        (
            17,
            "Ekwiwalent jednostek z puli (→ Power)",
            "=IF(B16=0,0,FLOOR(B8/B16,1))",
        ),
        (18, "% max puli (INFO — odrzucone jako Power)", "=IF(B15=0,0,ROUND(B8/B15*100,0))"),
    ]
    for row, label, formula in info:
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, formula)
        c.fill = INFO_FILL if row != 17 else PatternFill("solid", fgColor="E8F8E8")
        c.border = BORDER
        if row in (14, 15, 16):
            c.number_format = "#,##0"
        if row in (17, 18):
            c.number_format = "0"

    ws["D4"] = "Power (= suma, bez × epoka)"
    ws["E4"] = f"=E5"
    ws["E4"].font = Font(bold=True, size=13)
    ws["D5"] = "powerBase (suma składników)"
    ws["E5"] = f"=SUM(D{TBL}:D{TBL_END})"
    ws["D6"] = "Mnożnik epoki (ODRZUCONE)"
    ws["E6"] = 1
    ws["E6"].fill = REJECT_FILL
    ws["F6"] = "Było ×E — zaburzało balans (skok epoki = ×2 Power)"

    start = 18
    ws.cell(start, 1, "Składnik")
    ws.cell(start, 2, "Surowa liczba")
    ws.cell(start, 3, "Współczynnik")
    ws.cell(start, 4, "Suma punktów")
    ws.cell(start, 5, "Waga stary model %")
    ws.cell(start, 6, "Równoważnik / uwagi")
    ws.cell(start, 7, "% bazy")
    style_header(ws, start, 7)

    rows = [
        ("Armia (jednostki)", "=B6", 25, 0, "1 jedn. = 0,5 miasta", False),
        ("Wygrane bitwy", "=B7", 25, 0, "1 bitwa = 1 jednostka (pkt)", False),
        (
            "Ludki (sloty populacji)",
            "=B5",
            5,
            0,
            "NIE ludność abs · ~10 ludków/miasto @ 10 miast = ~100",
            False,
        ),
        (
            "Rekruci (ekw. jednostek z puli)",
            "=B17",
            5,
            0,
            "pula ÷ koszt werbu · pełna pula ≈ ludki×10 szt.",
            False,
        ),
        ("Miasta", "=B9", 50, 0, "1 miasto = 2 jednostki (pkt)", False),
        ("Terytorium (heksy)", "=B10", 0.5, 0, "100 heks = 1 miasto (pkt)", False),
        ("Budynki", "=B11", 5, 0, "zastępuje stary slot gospodarka", False),
        (
            "Odkrycia / tech zbadane",
            "=B12",
            20,
            0,
            "Progres zamiast mnożnika epoki",
            True,
        ),
        (
            "Ulepszenia w terenie",
            "=B13",
            5,
            0,
            "farmy, drogi, hodowla…",
            True,
        ),
    ]

    r = TBL
    for name, raw_f, coeff, weight, note, is_new in rows:
        ws.cell(r, 1, name)
        ws.cell(r, 2, raw_f)
        c_coeff = ws.cell(r, 3, coeff)
        c_coeff.fill = NEW_FILL if is_new else COEFF_FILL
        c_coeff.border = BORDER
        ws.cell(r, 4, f"=B{r}*C{r}")
        ws.cell(r, 5, weight if weight else "—")
        ws.cell(r, 6, note)
        ws.cell(r, 7, f"=IF($E$5=0,0,D{r}/$E$5)")
        ws.cell(r, 7).number_format = "0.0%"
        if is_new:
            ws.cell(r, 1).fill = NEW_FILL
        r += 1

    # Porównanie starej ludności abs (tylko edukacja — nie w sumie)
    cmp = r + 2
    ws.cell(cmp, 1, "PORÓWNANIE (nie w Power) — stary sposób ludności abs.")
    ws.cell(cmp, 1).font = Font(bold=True, color="884444")
    ws.cell(cmp + 1, 1, "Gdyby: ludność abs /1000")
    ws.cell(cmp + 1, 2, "=FLOOR(B14/1000,1)")
    ws.cell(cmp + 1, 3, 1)
    ws.cell(cmp + 1, 3).fill = REJECT_FILL
    ws.cell(cmp + 1, 4, f"=B{cmp+1}*C{cmp+1}")
    ws.cell(cmp + 1, 6, "Podwojenie przy skoku epoki — dlaczego odrzuciliśmy")
    ws.cell(cmp + 2, 1, "Gdyby: Power × epoka")
    ws.cell(cmp + 2, 2, "=B4")
    ws.cell(cmp + 2, 3, 1)
    ws.cell(cmp + 2, 3).fill = REJECT_FILL
    ws.cell(cmp + 2, 4, f"=E5*B{cmp+2}")
    ws.cell(cmp + 2, 6, "Stary model v2 — odrzucony")
    ws.cell(cmp + 3, 1, "Gdyby: rekruci % max × 1 (sufit 100)")
    ws.cell(cmp + 3, 2, "=B18")
    ws.cell(cmp + 3, 3, 1)
    ws.cell(cmp + 3, 3).fill = REJECT_FILL
    ws.cell(cmp + 3, 4, f"=B{cmp+3}*C{cmp+3}")
    ws.cell(cmp + 3, 6, "Traci znaczenie gdy inne składniki rosną — odrzucone")
    ws.cell(cmp + 4, 1, "Gdyby: rekruci /100 × 1")
    ws.cell(cmp + 4, 2, "=FLOOR(B8/100,1)")
    ws.cell(cmp + 4, 3, 1)
    ws.cell(cmp + 4, 3).fill = REJECT_FILL
    ws.cell(cmp + 4, 4, f"=B{cmp+4}*C{cmp+4}")
    ws.cell(cmp + 4, 6, "Surowa pula rośnie z epoką — odrzucone")

    for col, w in zip("ABCDEFG", (28, 16, 14, 14, 16, 48, 10)):
        ws.column_dimensions[col].width = w


def build_porownanie_epok(wb):
    """Dwie równe cywilizacje — skok epoki bez wzrostu ludków/tech."""
    ws = wb.create_sheet("Test skok epoki")
    ws["A1"] = "Test: 10 ludków, 6 jedn., 2 bitwy, reszta jak spec — tylko epoka się zmienia"
    ws["A1"].font = Font(bold=True)
    ws.merge_cells("A1:F1")

    headers = ["", "Kamień ep.1", "Brąz ep.2", "Różnica"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 4)

    params = [
        ("Epoka", 1, 2, "=C4-B4"),
        ("Suma ludków", 10, 10, 0),
        ("Jednostki", 6, 6, 0),
        ("Bitwy", 2, 2, 0),
        ("Rekruci pula", 8000, 16000, "=C8-B8"),
        ("Miasta", 1, 1, 0),
        ("Heksy", 45, 45, 0),
        ("Budynki", 6, 6, 0),
        ("Tech zbadane", 8, 9, "=C12-B12"),
        ("Ulepszenia", 12, 12, 0),
    ]
    base_row = 4
    for i, (name, v1, v2, diff) in enumerate(params):
        r = base_row + i
        ws.cell(r, 1, name)
        ws.cell(r, 2, v1)
        ws.cell(r, 3, v2)
        if isinstance(diff, str):
            ws.cell(r, 4, diff)
        else:
            ws.cell(r, 4, diff)

    r0 = base_row + len(params) + 1
    ws.cell(r0, 1, "Ludność abs (INFO)")
    ws.cell(r0, 2, "=B5*VLOOKUP(B4,Epoki!$A$2:$B$11,2,FALSE)")
    ws.cell(r0, 3, "=C5*VLOOKUP(C4,Epoki!$A$2:$B$11,2,FALSE)")
    ws.cell(r0, 4, "=C{r0}-B{r0}".format(r0=r0))

    ws.cell(r0 + 1, 1, "Pkt gdyby abs/1000 (stary)")
    ws.cell(r0 + 1, 2, "=FLOOR(B{r}/1000,1)".format(r=r0))
    ws.cell(r0 + 1, 3, "=FLOOR(C{r}/1000,1)".format(r=r0))
    ws.cell(r0 + 1, 4, "=C{r}-B{r}".format(r=r0 + 1))

    ws.cell(r0 + 2, 1, "Pkt ludki×12 (nowy)")
    ws.cell(r0 + 2, 2, "=B5*12")
    ws.cell(r0 + 2, 3, "=C5*12")
    ws.cell(r0 + 2, 4, 0)

    ws.cell(r0 + 3, 1, "Power × epoka (odrzucone)")
    ws.cell(r0 + 3, 2, 478)
    ws.cell(r0 + 3, 3, "=956")
    ws.cell(r0 + 3, 4, "=C{r}-B{r}".format(r=r0 + 3))

    ws.cell(r0 + 4, 1, "Koszt werbu 1 jedn.")
    ws.cell(r0 + 4, 2, "=VLOOKUP(B4,Epoki!$A$2:$D$11,4,FALSE)")
    ws.cell(r0 + 4, 3, "=VLOOKUP(C4,Epoki!$A$2:$D$11,4,FALSE)")
    ws.cell(r0 + 4, 4, "=C{r}-B{r}".format(r=r0 + 4))

    ws.cell(r0 + 5, 1, "Ekw. jednostek (pula÷koszt)")
    ws.cell(r0 + 5, 2, "=IF(B{r}=0,0,FLOOR(B8/B{r},1))".format(r=r0 + 4))
    ws.cell(r0 + 5, 3, "=IF(C{r}=0,0,FLOOR(C8/C{r},1))".format(r=r0 + 4))
    ws.cell(r0 + 5, 4, "=C{r}-B{r}".format(r=r0 + 5))

    ws.cell(r0 + 7, 1, "Przy 80% puli: ep.1 8000/100=80 · ep.2 16000/200=80 — ten sam mobilization Power.")
    ws.cell(r0 + 7, 1).font = Font(italic=True, color="336633")

    ws.cell(r0 + 8, 1, "Wniosek: skok epoki nie podbija Power bez realnego wzrostu imperium.")
    ws.cell(r0 + 8, 1).font = Font(italic=True, color="336633")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14


def build_scenariusze(wb):
    ws = wb.create_sheet("Scenariusze")
    ws["A1"] = "Scenariusze — skopiuj do Kalkulator!B4:B13"
    ws["A1"].font = Font(bold=True)

    headers = ["Parametr", "Start ep.1", "Mid ep.3", "Late ep.7", "Równe cyw · ep1", "Równe cyw · ep2 +1 tech"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))

    data = [
        ("Epoka", 1, 3, 7, 1, 2),
        ("Suma ludków", 10, 20, 30, 10, 10),
        ("Jednostki", 3, 10, 25, 6, 6),
        ("Wygrane bitwy", 0, 3, 12, 2, 2),
        ("Rekruci (pula)", 5000, 12000, 60000, 8000, 16000),
        ("Miasta", 1, 2, 4, 1, 1),
        ("Heksy", 30, 55, 120, 45, 45),
        ("Budynki", 2, 8, 24, 6, 6),
        ("Tech zbadane", 3, 12, 35, 8, 9),
        ("Ulepszenia terenu", 4, 18, 60, 12, 12),
    ]
    for i, row in enumerate(data, start=4):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)

    ws.cell(16, 1, "Kolumny E–F: ten sam % puli (80%) — ep.2 ma 2× rekrutów i 2× koszt werbu → ten sam ekw. jednostek")
    ws.cell(16, 1).font = Font(italic=True, color="336633")

    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16


def build_stary_model(wb):
    ws = wb.create_sheet("Stary model wag")
    ws["A1"] = "Historia wag % + odrzucone elementy v2"
    ws["A1"].font = Font(bold=True)
    headers = ["Klucz", "Waga %", "Status", "W modelu P-A (Excel)"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 4)
    old = [
        ("wielkoscArmii", 24, "stary 0–100", "Armia ×10"),
        ("wygraneBitwy", 17, "stary 0–100", "Bitwy ×25"),
        ("ludnosc", 15, "stary 0–100", "→ Ludki ×12 (NIE abs/1000)"),
        ("rekruci", 15, "stary 0–100", "→ ekw. jednostek (pula÷koszt) ×10"),
        ("miasta", 12, "stary 0–100", "Miasta ×50 + heksy ×2"),
        ("gospodarka", 10, "stary 0–100", "→ Budynki ×8"),
        ("epoka", 7, "stary 0–100", "ODRZUCONE jako mnożnik"),
        ("—", "—", "NOWE", "Tech zbadane ×15 · ulepszenia ×5"),
    ]
    for i, row in enumerate(old, start=4):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["D"].width = 40


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_epoki(wb)
    build_kalkulator(wb)
    build_porownanie_epok(wb)
    build_scenariusze(wb)
    build_stary_model(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Zapisano: {OUT}")


if __name__ == "__main__":
    main()
