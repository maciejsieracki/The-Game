#!/usr/bin/env python3
"""
import-city-names-from-xlsx.py — Excel → city-names-pools.json (+ sync civs.json).

Źródło: panele-sterowania/Nazwy-miast-cywilizacji.xlsx
Cel:    gra/data/city-names-pools.json → export-city-names.py → gra/data/civs.json

Arkusze (pierwszy pasujący):
  Miasto / Macierz_100 — kolumny Miasto_01..100 → miasta_cywilizacji
  Panstwo / Macierz_10 — kolumny Panstwo_01..10 → miasta_panstwa

Alternatywnie: arkusz Nazwy (długi format) gdy brak macierzy.

Użycie (z katalogu gra/):
  python tools/import-city-names-from-xlsx.py
  python tools/import-city-names-from-xlsx.py --xlsx ../panele-sterowania/Nazwy-miast-cywilizacji.xlsx
  python tools/import-city-names-from-xlsx.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_XLSX = os.path.normpath(
    os.path.join(HERE, "..", "..", "panele-sterowania", "Nazwy-miast-cywilizacji.xlsx")
)
DEF_POOLS = os.path.normpath(os.path.join(HERE, "..", "data", "city-names-pools.json"))
DEF_CIVS = os.path.normpath(os.path.join(HERE, "..", "data", "civs.json"))
DEF_EXPORT = os.path.join(HERE, "export-city-names.py")

MIASTA_CYW = 100
MIASTA_PAN = 10

SHEET_MIASTO = ("Miasto", "Macierz_100")
SHEET_PANSTWO = ("Panstwo", "Macierz_10")


def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: str, data: dict) -> None:
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def norm_name(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return s


def find_sheet(wb, names: tuple[str, ...]):
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    return None


def parse_matrix_headers(ws, prefix: str, count: int) -> tuple[dict[str, int], list[str]]:
    """Map column index → slot 1..count; return errors."""
    errs: list[str] = []
    col_map: dict[str, int] = {}
    header_row = 1
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, col).value
        if raw is None:
            continue
        h = str(raw).strip()
        m = re.fullmatch(rf"{prefix}_(\d{{1,3}})", h, re.IGNORECASE)
        if m:
            idx = int(m.group(1))
            if 1 <= idx <= count:
                col_map[f"{prefix}_{idx:02d}"] = col
    expected = {f"{prefix}_{i:02d}" for i in range(1, count + 1)}
    found = {k for k in col_map if k in expected or True}
    # Rebuild by numeric index
    by_idx: dict[int, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, col).value
        if raw is None:
            continue
        h = str(raw).strip()
        m = re.fullmatch(rf"{prefix}_(\d{{1,3}})", h, re.IGNORECASE)
        if m:
            idx = int(m.group(1))
            if 1 <= idx <= count:
                by_idx[idx] = col
    missing = [i for i in range(1, count + 1) if i not in by_idx]
    if missing:
        errs.append(f"Arkusz {ws.title}: brak kolumn {prefix}_XX dla: {missing[:5]}{'…' if len(missing) > 5 else ''}")
    return by_idx, errs


def read_matrix(ws, prefix: str, count: int, civ_ids: set[str]) -> tuple[dict[str, list[str]], list[str]]:
    errs: list[str] = []
    by_idx, hdr_errs = parse_matrix_headers(ws, prefix, count)
    errs.extend(hdr_errs)
    if hdr_errs:
        return {}, errs

    # IkonaId column
    ikona_col = None
    for col in range(1, min(ws.max_column, 6) + 1):
        h = ws.cell(1, col).value
        if h and str(h).strip().lower() == "ikonaid":
            ikona_col = col
            break
    if ikona_col is None:
        errs.append(f"Arkusz {ws.title}: brak kolumny IkonaId")
        return {}, errs

    pools: dict[str, list[str]] = {}
    for row in range(2, ws.max_row + 1):
        cid = norm_name(ws.cell(row, ikona_col).value)
        if not cid:
            continue
        names: list[str] = []
        for i in range(1, count + 1):
            col = by_idx.get(i)
            val = norm_name(ws.cell(row, col).value) if col else ""
            names.append(val)
        pools[cid] = names

    for cid in civ_ids:
        if cid not in pools:
            errs.append(f"Arkusz {ws.title}: brak wiersza dla {cid}")
    return pools, errs


def read_long_sheet(ws, civ_ids: set[str]) -> tuple[dict[str, dict[str, list[str]]], list[str]]:
    """Arkusz Nazwy: Cywilizacja | IkonaId | Kategoria | Lp | Nazwa | …"""
    errs: list[str] = []
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h:
            headers[str(h).strip().lower()] = col

    required = ["ikonaid", "kategoria", "lp", "nazwa"]
    for r in required:
        if r not in headers:
            errs.append(f"Arkusz Nazwy: brak kolumny {r}")
    if errs:
        return {}, errs

    data: dict[str, dict[str, list[str]]] = {}
    for row in range(2, ws.max_row + 1):
        cid = norm_name(ws.cell(row, headers["ikonaid"]).value)
        if not cid:
            continue
        kat = norm_name(ws.cell(row, headers["kategoria"]).value)
        lp_raw = ws.cell(row, headers["lp"]).value
        name = norm_name(ws.cell(row, headers["nazwa"]).value)
        try:
            lp = int(lp_raw)
        except (TypeError, ValueError):
            continue
        if kat == "miasto_cywilizacji":
            key, max_lp = "miasta_cywilizacji", MIASTA_CYW
        elif kat == "miasto_panstwa":
            key, max_lp = "miasta_panstwa", MIASTA_PAN
        else:
            continue
        if lp < 1 or lp > max_lp:
            continue
        entry = data.setdefault(cid, {"miasta_cywilizacji": [""] * MIASTA_CYW, "miasta_panstwa": [""] * MIASTA_PAN})
        entry[key][lp - 1] = name

    for cid in civ_ids:
        if cid not in data:
            errs.append(f"Arkusz Nazwy: brak wpisów dla {cid}")
    return data, errs


def validate_pools(pools: dict, civ_ids: list[str]) -> list[str]:
    errs: list[str] = []
    for cid in civ_ids:
        if cid not in pools:
            errs.append(f"Brak wpisu: {cid}")
            continue
        e = pools[cid]
        cyw = e.get("miasta_cywilizacji", [])
        pan = e.get("miasta_panstwa", [])
        if len(cyw) != MIASTA_CYW:
            errs.append(f"{cid}: miasta_cywilizacji {len(cyw)}/{MIASTA_CYW}")
        if len(pan) != MIASTA_PAN:
            errs.append(f"{cid}: miasta_panstwa {len(pan)}/{MIASTA_PAN}")
        for i, n in enumerate(cyw, 1):
            if not n:
                errs.append(f"{cid}: puste Miasto_{i:02d}")
        for i, n in enumerate(pan, 1):
            if not n:
                errs.append(f"{cid}: puste Panstwo_{i:02d}")
        if len(set(cyw)) != len(cyw):
            dupes = [x for x in cyw if cyw.count(x) > 1]
            errs.append(f"{cid}: duplikaty miasta_cywilizacji ({len(set(dupes))} unikalnych)")
        if len(set(pan)) != len(pan):
            dupes = [x for x in pan if pan.count(x) > 1]
            errs.append(f"{cid}: duplikaty miasta_panstwa ({len(set(dupes))} unikalnych)")
    return errs


def build_nazwa_pl_map(civs: dict, existing: dict) -> dict[str, str]:
    out: dict[str, str] = {}
    for c in civs.get("cywilizacje", []):
        cid = c.get("ikonaId")
        if not cid:
            continue
        out[cid] = (
            existing.get(cid, {}).get("nazwa_pl")
            or c.get("Cywilizacja")
            or cid
        )
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEF_XLSX)
    ap.add_argument("--pools", default=DEF_POOLS)
    ap.add_argument("--civs", default=DEF_CIVS)
    ap.add_argument("--dry-run", action="store_true", help="tylko walidacja, bez zapisu")
    ap.add_argument("--skip-export", action="store_true", help="nie uruchamiaj export-city-names.py")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        sys.exit(f"[import-city-names] brak pliku: {args.xlsx}")

    civs = load_json(args.civs)
    existing = load_json(args.pools) if os.path.isfile(args.pools) else {}
    civ_ids = [c["ikonaId"] for c in civs["cywilizacje"] if c.get("ikonaId")]
    civ_id_set = set(civ_ids)
    nazwa_pl = build_nazwa_pl_map(civs, existing)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    errs: list[str] = []

    ws_miasto = find_sheet(wb, SHEET_MIASTO)
    ws_panstwo = find_sheet(wb, SHEET_PANSTWO)

    cyw_by_civ: dict[str, list[str]] = {}
    pan_by_civ: dict[str, list[str]] = {}

    if ws_miasto and ws_panstwo:
        cyw_by_civ, e1 = read_matrix(ws_miasto, "Miasto", MIASTA_CYW, civ_id_set)
        pan_by_civ, e2 = read_matrix(ws_panstwo, "Panstwo", MIASTA_PAN, civ_id_set)
        errs.extend(e1)
        errs.extend(e2)
    elif "Nazwy" in wb.sheetnames:
        long_data, e3 = read_long_sheet(wb["Nazwy"], civ_id_set)
        errs.extend(e3)
        for cid, entry in long_data.items():
            cyw_by_civ[cid] = entry["miasta_cywilizacji"]
            pan_by_civ[cid] = entry["miasta_panstwa"]
    else:
        sys.exit(
            "[import-city-names] brak arkuszy Miasto+Panstwo ani Nazwy w "
            f"{args.xlsx} (dostępne: {wb.sheetnames})"
        )

    pools: dict = {}
    for cid in civ_ids:
        pools[cid] = {
            "nazwa_pl": nazwa_pl.get(cid, cid),
            "miasta_cywilizacji": list(cyw_by_civ.get(cid, [])),
            "miasta_panstwa": list(pan_by_civ.get(cid, [])),
        }

    errs.extend(validate_pools(pools, civ_ids))
    if errs:
        print("[import-city-names] WALIDACJA:")
        for e in errs:
            print(f"  ! {e}")
        sys.exit(1)

    if args.dry_run:
        print(f"[import-city-names] dry-run OK: {len(civ_ids)} cywilizacji × (100+10)")
        return

    save_json(args.pools, pools)
    print(f"[import-city-names] zapisano {args.pools} ({len(civ_ids)} cywilizacji)")

    if not args.skip_export:
        cmd = [sys.executable, DEF_EXPORT, "--pools", args.pools, "--civs", args.civs]
        print("[import-city-names] -> export-city-names.py")
        rc = subprocess.call(cmd)
        if rc != 0:
            sys.exit(rc)
        print("[import-city-names] civs.json zsynchronizowany")


if __name__ == "__main__":
    main()
