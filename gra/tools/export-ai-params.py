#!/usr/bin/env python3
# DEPRECATED (2026-06-30 — PANEL-MERGE): Excel w docs/archiwum/panele-legacy/.
# Kanon: panele-sterowania/Panel-{A..E}.xlsx + export-{a..e}.py — nie używać do nowych zmian.
# export-ai-params.py -- BEZPIECZNY eksport panelu AI-parametry.xlsx -> gra/data/ai-params.json.
# Dzial: Civ-AI. Wzorowane na export-panel.py (Civ-MIASTO).
#
# Zasada: nakladamy wartosci z Excela na ORYGINALNY ai-params.json
#   - zachowujemy strukture i kolejnosc istniejacych kluczy,
#   - kolumna "Wartosc" z Excela -> pole "wartosc" w JSON (ASCII, zgodne z danymi),
#   - pusta komorka Wartosc = brak zmiany,
#   - klucz z panelu spoza JSON = DOPISANY (wartosc + sekcja + opis),
#   - kolumna "Status" jest IGNOROWANA (to adnotacja panelu LIVE/PLANOWANE, nie schema).
# NIE rusza innych JSON-ow. NIE uruchamiac export-data.py ani npm run build.
#
# Uzycie:
#   python3 gra/tools/export-ai-params.py
#   python3 gra/tools/export-ai-params.py --xlsx X.xlsx --json J.json   # do testow
import openpyxl, json, os, sys, argparse

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(HERE, "..", "data"))
CIV_ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
DEFAULT_XLSX = os.path.join(CIV_ROOT, "docs", "archiwum", "panele-legacy", "Civ-AI", "AI-parametry.xlsx")  # panel przeniesiony do Civ-AI/
DEFAULT_JSON = os.path.join(DATA_DIR, "ai-params.json")
SHEET = "AI-parametry"


def coerce_number(val, old=None):
    """Zwroc liczbe (int gdy calkowita, inaczej float). Pusta/zla -> old."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return old
    try:
        f = float(val)
    except Exception:
        return old
    if f == int(f):
        return int(f)
    return f


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--json", default=DEFAULT_JSON)
    ap.add_argument("--out", default=None, help="domyslnie nadpisuje --json")
    args = ap.parse_args()
    out_path = args.out or args.json

    if not os.path.exists(args.xlsx):
        print("BRAK pliku panelu:", args.xlsx); sys.exit(1)
    if not os.path.exists(args.json):
        print("BRAK pliku JSON:", args.json); sys.exit(1)

    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if SHEET not in wb.sheetnames:
        print("BRAK arkusza:", SHEET, "w", args.xlsx); sys.exit(1)
    ws = wb[SHEET]

    # Znajdz wiersz naglowka (Parametr | Wartosc | Sekcja | Opis | ...).
    header_row = None
    for r in range(1, min(ws.max_row, 6) + 1):
        vals = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[r]]
        if "parametr" in vals and any(v.startswith("warto") for v in vals):
            header_row = r
            cols = {v: i for i, v in enumerate(vals)}
            break
    if header_row is None:
        print("Nie znaleziono naglowka (Parametr/Wartosc)."); sys.exit(1)

    c_key = cols.get("parametr")
    c_val = next((i for v, i in cols.items() if v.startswith("warto")), None)
    c_sek = cols.get("sekcja")
    c_opis = cols.get("opis")

    changed, added = 0, 0
    for r in range(header_row + 1, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if c_key >= len(row):
            continue
        key = row[c_key]
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        new_val = coerce_number(row[c_val] if c_val is not None and c_val < len(row) else None)
        if key in data:
            entry = data[key]
            if new_val is not None and entry.get("wartosc") != new_val:
                entry["wartosc"] = new_val
                changed += 1
        else:
            sek = row[c_sek] if c_sek is not None and c_sek < len(row) else None
            opis = row[c_opis] if c_opis is not None and c_opis < len(row) else None
            data[key] = {
                "wartosc": new_val if new_val is not None else 0,
                "sekcja": str(sek).strip() if sek else "",
                "opis": str(opis).strip() if opis else "",
            }
            added += 1

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"OK: {out_path}  (zmienione: {changed}, dopisane: {added}, kluczy total: {len(data)})")


if __name__ == "__main__":
    main()
