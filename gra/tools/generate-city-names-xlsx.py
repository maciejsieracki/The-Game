#!/usr/bin/env python3
"""
generate-city-names-xlsx.py — Excel do przeglądu nazw miast (Maciej).

Źródło: gra/data/city-names-pools.json + gra/data/civs.json (ikonaId, nazwyKlastra).
Wynik:  panele-sterowania/Nazwy-miast-cywilizacji.xlsx

Arkusze:
  README      — instrukcja + status audytu państw-miast
  Podsumowanie — liczba nazw per cywilizacja
  Nazwy       — długi format: Cywilizacja | IkonaId | Kategoria | Lp | Nazwa | Uwagi | Status
  Macierz_100 — wiersze=cywilizacje, kolumny Miasto_01..100
  Macierz_10  — wiersze=cywilizacje, kolumny Panstwo_01..10

Użycie (z katalogu gra/):
  python tools/generate-city-names-xlsx.py
  python tools/generate-city-names-xlsx.py --json ../data/city-names-pools.json
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_POOLS = os.path.normpath(os.path.join(HERE, "..", "data", "city-names-pools.json"))
DEF_CIVS = os.path.normpath(os.path.join(HERE, "..", "data", "civs.json"))
DEF_OUT = os.path.normpath(os.path.join(HERE, "..", "..", "panele-sterowania", "Nazwy-miast-cywilizacji.xlsx"))

HEADER_FILL = PatternFill("solid", fgColor="2A2416")
HEADER_FONT = Font(bold=True, color="F4E6A8")
CATEGORY_FILL = {
    "miasto_cywilizacji": PatternFill("solid", fgColor="1A2230"),
    "miasto_panstwa": PatternFill("solid", fgColor="1E2A1A"),
}


def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def validate_pools(pools: dict, civs: dict) -> list[str]:
    errs: list[str] = []
    civ_ids = {c.get("ikonaId") for c in civs.get("cywilizacje", [])}
    for cid in civ_ids:
        if cid not in pools:
            errs.append(f"Brak wpisu w city-names-pools.json: {cid}")
            continue
        entry = pools[cid]
        cyw = entry.get("miasta_cywilizacji", [])
        pan = entry.get("miasta_panstwa", [])
        if len(cyw) != 100:
            errs.append(f"{cid}: miasta_cywilizacji ma {len(cyw)} pozycji (oczekiwano 100)")
        if len(pan) != 10:
            errs.append(f"{cid}: miasta_panstwa ma {len(pan)} pozycji (oczekiwano 10)")
        if len(set(cyw)) != len(cyw):
            errs.append(f"{cid}: duplikaty w miasta_cywilizacji")
        if len(set(pan)) != len(pan):
            errs.append(f"{cid}: duplikaty w miasta_panstwa")
    return errs


def write_readme(ws) -> None:
    ws.title = "README"
    lines = [
        ("Nazwy miast — The Game (Civ)", ""),
        ("Data wygenerowania", str(date.today())),
        ("", ""),
        ("Cel", "Przegląd Macieja przed importem do civs.json / silnika."),
        ("", ""),
        ("Kategorie", ""),
        ("miasto_cywilizacji", "100 nazw historycznych — kolejne miasta gracza/AI imperium (founding)."),
        ("miasto_panstwa", "10 nazw — miasta-państwa w klastrze gracza (Sparta, Kapua…). Obecnie nazwyKlastra[1..10] w civs.json."),
        ("", ""),
        ("Kolumna Status", "do_przeglądu = propozycja AI/research; zaakceptowane / do_poprawy — edytuj w Excelu."),
        ("", ""),
        ("AUDYT: czy państwa-miasta zakładają nowe miasta?", ""),
        ("Stan kodu 2026-07-07", "ZABLOKOWANE — AI miast-państw ma profil defensiveCopy (ai.ts decideDefensiveCopyTurn)."),
        ("", "Brak produkcji osadników, brak foundCity, tylko garnizon + riposta gdy zagrożenie."),
        ("", "Flaga startCityState na mieście dotyczy tylko dystansu founding (3 hex vs 5), nie pozwala AI zakładać."),
        ("", "Uwaga: defensiveCopy dostają też obce stolice klastrów (typCityCopyOwners) — osobna decyzja produktowa."),
        ("", ""),
        ("Następny krok po akceptacji", "export-city-names.py → civs.json pola nazwyMiast[100] + nazwyKlastra[10]."),
    ]
    for r, (a, b) in enumerate(lines, 1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if a.startswith("AUDYT") or a == "Nazwy miast":
            ws.cell(row=r, column=1).font = Font(bold=True, color="E8D88A")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 88


def write_summary(ws, pools: dict, civ_order: list[str]) -> None:
    ws.title = "Podsumowanie"
    headers = ["Cywilizacja", "IkonaId", "Miasta cyw (100)", "Miasta państw (10)", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    for i, cid in enumerate(civ_order, 2):
        e = pools[cid]
        ws.cell(row=i, column=1, value=e.get("nazwa_pl", cid))
        ws.cell(row=i, column=2, value=cid)
        ws.cell(row=i, column=3, value=len(e.get("miasta_cywilizacji", [])))
        ws.cell(row=i, column=4, value=len(e.get("miasta_panstwa", [])))
        ws.cell(row=i, column=5, value="do_przeglądu")
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22


def write_long(ws, pools: dict, civ_order: list[str]) -> None:
    ws.title = "Nazwy"
    headers = ["Cywilizacja", "IkonaId", "Kategoria", "Lp", "Nazwa", "Uwagi", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    row = 2
    for cid in civ_order:
        e = pools[cid]
        nazwa_pl = e.get("nazwa_pl", cid)
        for kat, key in [("miasto_cywilizacji", "miasta_cywilizacji"), ("miasto_panstwa", "miasta_panstwa")]:
            names = e.get(key, [])
            for lp, name in enumerate(names, 1):
                ws.cell(row=row, column=1, value=nazwa_pl)
                ws.cell(row=row, column=2, value=cid)
                ws.cell(row=row, column=3, value=kat)
                ws.cell(row=row, column=4, value=lp)
                ws.cell(row=row, column=5, value=name)
                ws.cell(row=row, column=6, value="historyczna / research 2026-07")
                ws.cell(row=row, column=7, value="do_przeglądu")
                row += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A2"


def write_matrix(ws, pools: dict, civ_order: list[str], key: str, prefix: str, count: int) -> None:
    ws.title = prefix
    headers = ["Cywilizacja", "IkonaId"] + [f"{prefix}_{i:02d}" for i in range(1, count + 1)]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    for r, cid in enumerate(civ_order, 2):
        e = pools[cid]
        ws.cell(row=r, column=1, value=e.get("nazwa_pl", cid))
        ws.cell(row=r, column=2, value=cid)
        names = e.get(key, [])
        for i in range(count):
            val = names[i] if i < len(names) else ""
            ws.cell(row=r, column=3 + i, value=val)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    for i in range(count):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16
    ws.freeze_panes = "C2"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pools", default=DEF_POOLS)
    ap.add_argument("--civs", default=DEF_CIVS)
    ap.add_argument("--out", default=DEF_OUT)
    ap.add_argument("--strict", action="store_true", help="przerwij gdy liczba nazw != 100/10")
    args = ap.parse_args()

    pools = load_json(args.pools)
    civs = load_json(args.civs)
    civ_order = [c["ikonaId"] for c in civs["cywilizacje"] if c.get("ikonaId")]

    errs = validate_pools(pools, civs)
    if errs:
        print("[generate-city-names-xlsx] WALIDACJA:")
        for e in errs:
            print(f"  ! {e}")
        if args.strict:
            sys.exit(1)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb = openpyxl.Workbook()
    write_readme(wb.active)
    write_summary(wb.create_sheet(), pools, civ_order)
    write_long(wb.create_sheet(), pools, civ_order)
    write_matrix(wb.create_sheet(), pools, civ_order, "miasta_cywilizacji", "Miasto", 100)
    write_matrix(wb.create_sheet(), pools, civ_order, "miasta_panstwa", "Panstwo", 10)
    wb.save(args.out)
    print(f"[generate-city-names-xlsx] zapisano: {args.out}")
    print(f"[generate-city-names-xlsx] cywilizacji: {len(civ_order)}")


if __name__ == "__main__":
    main()
