#!/usr/bin/env python3
# append-e1-status-xlsx.py — dopisuje wiersze E1 do Status-projektu-The-Game.xlsx
# Uruchom z root Civ: python gra/tools/append-e1-status-xlsx.py

import datetime
import os
import shutil
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Brak openpyxl: pip install openpyxl")
    sys.exit(1)

CIV_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(CIV_ROOT, "Status-projektu-The-Game.xlsx")
SHEET = "Grupa-E"
TODAY = datetime.date.today().isoformat()

HEADERS = ("ID", "Temat", "Status", "Decyzja", "Data", "Uwagi")

ROWS = [
    ("E1-D13", "Defaulty startu (framework)", "CZĘŚCIOWO", "D13=A", "2026-06-26", "Szczegóły E1-nowa-gra.md"),
    ("E1-kod", "UI+MAPA+SILNIK kreator", "WPIĘTE", "—", TODAY, "generujSwiat, seed, typSwiata, rzymianie"),
    ("E1-Q9", "Reset gracza Nowa gra", "OTWARTE ABC", "—", TODAY, "Provisional A w kodzie"),
    ("E1-Q10", "Start Brąz + tech", "OTWARTE ABC", "—", TODAY, "Provisional A w kodzie"),
    ("E1-Q11", "Kształt mapy Ziemia", "OTWARTE ABC", "—", TODAY, "ZIEMIA_LAND_CENTERS"),
    ("E1-Q12", "Zakres rywali menu", "OTWARTE ABC", "—", TODAY, "newGameMapDefaults ±1"),
    ("E1-Q6-8", "Menu główne Kampania/Multi", "OTWARTE", "—", TODAY, "Nie wysłane ABC"),
    ("E3-D14", "Surowce żelazo/stal", "CZĘŚCIOWO", "D14=A", "2026-06-26", "Brak złoża zelazo MAPA"),
    ("E2", "AI / zwycięstwo meta", "OTWARTE", "—", TODAY, "Brak pytań ABC"),
    ("E1-bramka", "Gra-podglad-TEST.html", "CZEKA", "—", TODAY, "F bramka Node lokalnie"),
]


def main():
    if not os.path.isfile(XLSX):
        print(f"Brak pliku: {XLSX}")
        sys.exit(1)
    bak = XLSX + f".bak-E1-{TODAY}"
    shutil.copy2(XLSX, bak)
    print(f"Backup: {bak}")
    wb = openpyxl.load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
        start_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(SHEET)
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="E8D88A")
        start_row = 2
    for i, row in enumerate(ROWS):
        r = start_row + i
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)
    wb.save(XLSX)
    print(f"Dopisano {len(ROWS)} wierszy do arkusza '{SHEET}'")


if __name__ == "__main__":
    main()
