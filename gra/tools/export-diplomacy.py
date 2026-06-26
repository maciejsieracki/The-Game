#!/usr/bin/env python3
"""
export-diplomacy.py -- TARGETED export: Dyplomacja.xlsx[arkusz "params"] -> diplomacy.json["params"].

Aktualizuje WYLACZNIE blok "params" w gra/data/diplomacy.json. NIE uruchamia globalnego
tools/export-data.py i NIE regeneruje innych plikow JSON (civs.json itd. pozostaja nietkniete).

Przeplyw sterowania parametrami:
  Dyplomacja.xlsx (arkusz "params", kolumna B = edytowalna) --> [ten skrypt] --> diplomacy.json["params"]
  --> loadDiplomacyParams(json) w src/game/diplomacy.ts --> { ...DIPLOMACY_PARAMS, ...override } w SILNIK.

Uzycie (z katalogu gra/):
  python3 tools/export-diplomacy.py             # zapis do diplomacy.json
  python3 tools/export-diplomacy.py --dry-run   # tylko wypisz parametry, bez zapisu
  python3 tools/export-diplomacy.py --xlsx SCIEZKA --json SCIEZKA

Domyslne sciezki liczone wzgledem tego pliku (xlsx: hub Civ/Dyplomacja/, fallback root Civ/):
  xlsx = ../../Dyplomacja/Dyplomacja.xlsx  (po reorganizacji) | ../../Dyplomacja.xlsx (przed)
  json = ../data/diplomacy.json
"""
import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl:  pip install openpyxl --break-system-packages")

HERE = os.path.dirname(os.path.abspath(__file__))
# xlsx: hub Civ/Dyplomacja/ (po reorganizacji) lub root Civ/ (przed) -- bierzemy ten, ktory istnieje.
_XLSX_CANDIDATES = [
    os.path.normpath(os.path.join(HERE, "..", "..", "Dyplomacja", "Dyplomacja.xlsx")),
    os.path.normpath(os.path.join(HERE, "..", "..", "Dyplomacja.xlsx")),
]


def default_xlsx():
    for p in _XLSX_CANDIDATES:
        if os.path.exists(p):
            return p
    return _XLSX_CANDIDATES[0]


DEF_JSON = os.path.normpath(os.path.join(HERE, "..", "data", "diplomacy.json"))
SHEET = "params"


def num(v):
    """Zwroc int gdy liczba calkowita, inaczej float. Akceptuje '60', '0,5', '0.5'."""
    if isinstance(v, bool):
        raise ValueError("wartosc logiczna nie jest dozwolona")
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else v
    s = str(v).strip().replace(",", ".")
    f = float(s)
    return int(f) if f.is_integer() else f


def read_params(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Brak arkusza '{SHEET}' w {xlsx_path}")
    ws = wb[SHEET]
    params = {}
    for r in range(2, ws.max_row + 1):          # wiersz 1 = naglowek
        key = ws.cell(row=r, column=1).value     # kolumna A = klucz
        val = ws.cell(row=r, column=2).value     # kolumna B = wartosc
        if key is None or str(key).strip() == "":
            continue
        try:
            params[str(key).strip()] = num(val)
        except (TypeError, ValueError):
            sys.exit(f"Wiersz {r}: klucz '{key}' ma niepoprawna liczbe: {val!r}")
    return params


def main():
    ap = argparse.ArgumentParser(description="Export params: Dyplomacja.xlsx -> diplomacy.json")
    ap.add_argument("--xlsx", default=None)
    ap.add_argument("--json", default=DEF_JSON)
    ap.add_argument("--dry-run", action="store_true", help="tylko wypisz, nie zapisuj")
    a = ap.parse_args()
    if a.xlsx is None:
        a.xlsx = default_xlsx()

    params = read_params(a.xlsx)
    print(f"[export-diplomacy] odczytano {len(params)} parametrow z arkusza '{SHEET}' ({a.xlsx})")
    for k, v in params.items():
        print(f"    {k} = {v}")

    if a.dry_run:
        print("[dry-run] JSON nietkniety.")
        return

    with open(a.json, "r", encoding="utf-8") as f:
        data = json.load(f)

    old = data.get("params", {})
    changed = {k: (old.get(k), v) for k, v in params.items() if old.get(k) != v}

    # blok "params" na poczatku, reszta kluczy w niezmienionej kolejnosci
    newdata = {"params": params}
    for k, v in data.items():
        if k != "params":
            newdata[k] = v

    with open(a.json, "w", encoding="utf-8") as f:
        json.dump(newdata, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"[export-diplomacy] zapisano blok 'params' -> {a.json}")
    if changed:
        print("Zmienione wartosci:")
        for k, (o, n) in changed.items():
            print(f"    {k}: {o} -> {n}")
    else:
        print("Brak zmian wartosci.")


if __name__ == "__main__":
    main()
