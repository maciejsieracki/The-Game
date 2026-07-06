#!/usr/bin/env python3
# gen-ulepszenia-xlsx.py -- buduje EDYTOWALNY MIASTO/Ulepszenia-terenu.xlsx z data/terrain-improvements.json.
# Liczby (niebieskie) = strojysz; kolumna "Komentarz Naster" (zolta). Edycja -> JSON: export-ulepszenia.py.
import openpyxl, json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data", "terrain-improvements.json"))
OUT  = os.path.normpath(os.path.join(HERE, "..", "..", "MIASTO", "Ulepszenia-terenu.xlsx"))

d = json.load(open(DATA, encoding="utf-8"))
rows = [(k, v) for k, v in d.items() if k != "_meta" and not str(k).startswith("_")]

# kolumny: id + nazwa + epoka + bonus_* (zebrane) + reszta skalarow
bonus_keys = []
for _, v in rows:
    for bk in (v.get("bonus") or {}):
        if bk not in bonus_keys: bonus_keys.append(bk)
scalar_keys = []
for _, v in rows:
    for sk in v:
        if sk not in ("bonus",) and sk not in scalar_keys: scalar_keys.append(sk)
# kolejnosc skalarow: nazwa, epoka, potem reszta
order = [s for s in ["nazwa", "epoka"] if s in scalar_keys] + [s for s in scalar_keys if s not in ("nazwa", "epoka")]
cols = ["id"] + ["nazwa", "epoka"] + [f"bonus.{b}" for b in bonus_keys] + [s for s in order if s not in ("nazwa", "epoka")] + ["Komentarz Naster"]

FONT = "Arial"
H = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="1F6F63")
NF = Font(name=FONT, color="0000FF")          # liczby = input
TF = Font(name=FONT, color="1A1A1A")
KF = Font(name=FONT, bold=True, color="1A1A1A")
NOTE = PatternFill("solid", fgColor="FFF6CC")
thin = Side(style="thin", color="D9D9D9"); B = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical="top"); TOP = Alignment(vertical="top")

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Ulepszenia terenu"
for j, c in enumerate(cols, 1):
    cell = ws.cell(1, j, c); cell.font = H; cell.fill = HF; cell.border = B; cell.alignment = WRAP
for i, (k, v) in enumerate(rows):
    r = i + 2
    out = {"id": k, "nazwa": v.get("nazwa", ""), "epoka": v.get("epoka", "")}
    for b in bonus_keys: out[f"bonus.{b}"] = (v.get("bonus") or {}).get(b, "")
    for s in order:
        if s not in ("nazwa", "epoka"): out[s] = v.get(s, "")
    for j, c in enumerate(cols, 1):
        if c == "Komentarz Naster":
            cell = ws.cell(r, j, ""); cell.fill = NOTE; cell.border = B; continue
        raw = out.get(c, "")
        cell = ws.cell(r, j, raw if not isinstance(raw, (dict, list)) else json.dumps(raw, ensure_ascii=False))
        cell.border = B
        if c == "id": cell.font = KF; cell.alignment = TOP
        elif isinstance(raw, (int, float)) and not isinstance(raw, bool): cell.font = NF; cell.alignment = TOP
        else: cell.font = TF; cell.alignment = WRAP
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
for col in range(1, len(cols) + 1):
    w = 10
    for row in range(1, len(rows) + 2):
        val = ws.cell(row, col).value
        if val is not None: w = max(w, min(len(str(val)) + 2, 55))
    ws.column_dimensions[get_column_letter(col)].width = w
ws.sheet_properties.tabColor = "1F6F63"
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("OK ->", OUT, "| wierszy:", len(rows), "| kolumn:", len(cols))
