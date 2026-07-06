#!/usr/bin/env python3
"""DEPRECATED (2026-06-30 — PANEL-MERGE):
  Excel: docs/archiwum/panele-legacy/
  Kanon: panele-sterowania/Panel-{A..E}.xlsx + export-{a..e}.py
  Nie używać do nowych zmian balansu.


export-civ-ai.py — TARGETED: Cywilizacje.xlsx[AI-zachowanie] -> gra/data/civ-ai.json

Kolumny (wiersz 2): Cywilizacja | Agresywnosc | Ekspansywnosc | Priorytet militarny |
  Priorytet ekonomia | Priorytet nauka | Tolerancja ryzyka | Sklonnosc do podboju |
  Profil mapy | Uwagi

  python tools/export-civ-ai.py [--dry-run]
"""
import argparse
import json
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_XLSX = os.path.normpath(os.path.join(HERE, "..", "..", "docs", "archiwum", "panele-legacy", "Cywilizacje.xlsx"))
DEF_JSON = os.path.normpath(os.path.join(HERE, "..", "data", "civ-ai.json"))
SHEET = "AI-zachowanie"


def num(v, default=None):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return default


def str_val(v, default=""):
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Brak arkusza '{SHEET}' w {xlsx_path}")
    ws = wb[SHEET]
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[2]]
    col = {}
    for i, h in enumerate(header):
        if "cywilizacja" in h:
            col["nacja"] = i
        elif "agresywn" in h:
            col["agresywnosc"] = i
        elif "ekspansywn" in h:
            col["ekspansywnosc"] = i
        elif "militarn" in h:
            col["priorytetMilitarny"] = i
        elif "ekonom" in h and "priorytet" in h:
            col["priorytetEkonomia"] = i
        elif "nauka" in h and "priorytet" in h:
            col["priorytetNauka"] = i
        elif "tolerancja" in h:
            col["tolerancjaRyzyka"] = i
        elif "podboj" in h:
            col["sklonnoscDoPodboju"] = i
        elif "profil" in h:
            col["profilMapy"] = i
        elif h == "uwagi":
            col["uwagi"] = i

    rows = []
    for r in range(3, ws.max_row + 1):
        cells = [ws.cell(r, c + 1).value for c in range(len(header))]
        if col.get("nacja") is None or not cells[col["nacja"]]:
            continue
        name = str(cells[col["nacja"]]).strip()
        entry = {"Cywilizacja": name}
        for key, idx in col.items():
            if key == "nacja":
                continue
            if key == "profilMapy" or key == "uwagi":
                entry[key] = str_val(cells[idx] if idx < len(cells) else None)
            else:
                entry[key] = num(cells[idx] if idx < len(cells) else None, 0)
        if entry.get("agresywnosc") is None and entry.get("profilMapy") == "":
            continue
        rows.append(entry)
    wb.close()
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEF_XLSX)
    ap.add_argument("--json", default=DEF_JSON)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    rows = read_rows(args.xlsx)
    print(f"[export-civ-ai] nacje: {len(rows)}")
    for e in rows:
        print(f"  {e['Cywilizacja']}: agres={e.get('agresywnosc')} profil={e.get('profilMapy')}")

    out = {"cywilizacje": rows, "_meta": {"zrodlo": "Cywilizacje.xlsx/AI-zachowanie"}}

    if args.dry_run:
        print("[dry-run] bez zapisu")
        return

    bak = args.json + ".bak-CYWILIZACJE-export-ai"
    if os.path.exists(args.json):
        shutil.copy2(args.json, bak)
        print(f"backup: {bak}")

    tmp = args.json + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")
    json.load(open(tmp, encoding="utf-8"))
    os.replace(tmp, args.json)
    print(f"[export-civ-ai] GOTOWE: {args.json}")


if __name__ == "__main__":
    main()
