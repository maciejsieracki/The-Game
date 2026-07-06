#!/usr/bin/env python3
# audit-units-tw-v3.py — weryfikacja statów TW v3 (melee + dystans)
"""Reguły melee (z backupu units.json.bak-TW-v3):
  ÷10: Atak, Obrona, Obrażenia, Pancerz, Przebicie, Uderzenie
  health × 0,25
  Hastati: meleeAttack=8, weaponDamage=8

Reguły dystans (z panele-sterowania/TW-dystans-edycja.xlsx):
  Zasięg, Ilość pocisków, missileAttack, wallAttack (Katapulta)

Użycie:
  python gra/tools/audit-units-tw-v3.py
  python gra/tools/audit-units-tw-v3.py --fix
"""
from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

ROOT = Path(__file__).resolve().parents[2]
UNITS = ROOT / "gra" / "data" / "units.json"
BACKUP = UNITS.with_suffix(".json.bak-TW-v3-2026-06-29")
XLSX = ROOT / "panele-sterowania" / "TW-dystans-edycja.xlsx"
REPORT = Path(__file__).resolve().parent / "TW-v3-audyt-jednostek.md"

OLD_TO_NEW = {
    "Atak": "meleeAttack",
    "Obrona": "meleeDefence",
    "Obrażenia": "weaponDamage",
    "Pancerz": "armor",
    "Przebicie": "piercing",
    "Uderzenie": "chargeBonus",
    "Health": "health",
}
DIV10 = {"Atak", "Obrona", "Obrażenia", "Pancerz", "Przebicie", "Uderzenie"}
OVERRIDES = {"Hastati": {"meleeAttack": 8, "weaponDamage": 8}, "Konnica": {"meleeAttack": 30}}
MELEE_KEYS = list(OLD_TO_NEW.values())


def to_int(v) -> int:
    return int(round(float(v)))


def expected_melee(old_u: dict) -> dict[str, int]:
    name = old_u["Jednostka"]
    out: dict[str, int] = {}
    for ok, nk in OLD_TO_NEW.items():
        raw = old_u.get(ok, 0)
        if raw in ("—", "", None):
            raw = 0
        if ok == "Health":
            out[nk] = to_int(float(raw) * 0.25)
        elif ok in DIV10:
            out[nk] = to_int(float(raw) / 10)
        else:
            out[nk] = to_int(float(raw))
    if name in OVERRIDES:
        out.update(OVERRIDES[name])
    return out


def load_excel_dystans() -> dict[str, dict]:
    if not XLSX.is_file() or openpyxl is None:
        return {}
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Dystans-TW-v3"]
    header_row = None
    for r in range(1, 25):
        if ws.cell(r, 1).value == "Jednostka":
            header_row = r
            break
    if not header_row:
        return {}
    out: dict[str, dict] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        wall = ws.cell(r, 9).value
        out[name] = {
            "Zasięg ataku (hex)": to_int(ws.cell(r, 5).value) if ws.cell(r, 5).value not in (None, "", "—") else 0,
            "Ilość pocisków": to_int(ws.cell(r, 6).value) if ws.cell(r, 6).value not in (None, "", "—") else 0,
            "missileAttack": to_int(ws.cell(r, 8).value) if ws.cell(r, 8).value not in (None, "", "—") else 0,
            "wallAttack": to_int(wall) if wall not in (None, "", "—") else None,
        }
    return out


def audit(fix: bool) -> tuple[list, list, int]:
    old = {u["Jednostka"]: u for u in json.loads(BACKUP.read_text(encoding="utf-8"))}
    units = json.loads(UNITS.read_text(encoding="utf-8"))
    excel_d = load_excel_dystans()
    melee_err: list[tuple] = []
    dyst_err: list[tuple] = []
    fixes = 0

    for u in units:
        name = u["Jednostka"]
        ou = old.get(name)
        if not ou:
            melee_err.append((name, "brak w backupie"))
            continue

        exp = expected_melee(ou)
        for k, v in exp.items():
            if u.get(k) != v:
                melee_err.append((name, k, v, u.get(k)))
                if fix:
                    u[k] = v
                    fixes += 1

        # dystans: tylko jednostki z Excela
        if name in excel_d:
            ed = excel_d[name]
            for k, v in ed.items():
                if k == "wallAttack":
                    if name != "Katapulta":
                        if u.get(k) is not None:
                            dyst_err.append((name, k, None, u.get(k)))
                            if fix:
                                u.pop(k, None)
                                fixes += 1
                        continue
                    if v is None:
                        continue
                if u.get(k) != v:
                    dyst_err.append((name, k, v, u.get(k)))
                    if fix:
                        u[k] = v
                        fixes += 1
            if name != "Katapulta":
                u.pop("wallAttack", None)
        u.pop("missileAttackVsUnit", None)
        u.pop("missileAttackVsWall", None)

        # pozostałe: missileAttack=0 jeśli brak strzału
        if name not in excel_d and u.get("missileAttack", 0) != 0:
            dyst_err.append((name, "missileAttack", 0, u.get("missileAttack")))
            if fix:
                u["missileAttack"] = 0
                fixes += 1

    if fix and fixes:
        UNITS.write_text(json.dumps(units, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return melee_err, dyst_err, fixes


def write_report(melee_err, dyst_err) -> None:
    lines = [
        "# TW v3 — audyt jednostek",
        "",
        f"Data: {date.today().isoformat()}",
        "",
        "## Reguły",
        "",
        "**Melee (50 jednostek):** ÷10 ze skali 0–100 + health×0,25 + Hastati Atak/Obrażenia=8",
        "",
        "**Dystans (15 jednostek):** wartości z `TW-dystans-edycja.xlsx` (zasięg, pociski, missileAttack, wallAttack)",
        "",
        f"## Wynik",
        "",
        f"- Błędy melee: **{len(melee_err)}**",
        f"- Błędy dystans: **{len(dyst_err)}**",
        "",
    ]
    if melee_err:
        lines += ["### Melee", "", "| Jednostka | Pole | Powinno | Jest |", "|---|---|---:|---:|"]
        for e in melee_err:
            if len(e) == 4:
                lines.append(f"| {e[0]} | {e[1]} | {e[2]} | {e[3]} |")
            else:
                lines.append(f"| {e[0]} | — | — | {e[1]} |")
        lines.append("")
    if dyst_err:
        lines += ["### Dystans", "", "| Jednostka | Pole | Powinno | Jest |", "|---|---|---:|---:|"]
        for e in dyst_err:
            lines.append(f"| {e[0]} | {e[1]} | {e[2]} | {e[3]} |")
        lines.append("")
    if not melee_err and not dyst_err:
        lines += ["**OK** — wszystkie jednostki zgodne z wytycznymi.", ""]

    # próbki
    units = json.loads(UNITS.read_text(encoding="utf-8"))
    samples = [
        "Wojownik", "Wojownik z mieczem i tarczą", "Falanga", "Hastati",
        "Łucznik", "Kusznik", "Katapulta", "Zwiadowca", "Taran",
    ]
    lines += ["## Próbki (melee + dystans)", "", "| Jednostka | A | O | Ob | P | Pr | Sz | HP | ms | mur |", "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|"]
    for u in units:
        if u["Jednostka"] not in samples:
            continue
        n = u["Jednostka"]
        lines.append(
            f"| {n} | {u.get('meleeAttack','')} | {u.get('meleeDefence','')} | {u.get('weaponDamage','')} "
            f"| {u.get('armor','')} | {u.get('piercing','')} | {u.get('chargeBonus','')} | {u.get('health','')} "
            f"| {u.get('missileAttack','')} | {u.get('wallAttack','—')} |"
        )
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--fix", action="store_true")
    args = ap.parse_args()
    melee_err, dyst_err, fixes = audit(args.fix)
    write_report(melee_err, dyst_err)
    print(f"Audyt: melee={len(melee_err)} dystans={len(dyst_err)} fix={fixes}")
    print(f"Raport: {REPORT}")
    if melee_err or dyst_err:
        for e in melee_err[:10]:
            print(" MELEE", e)
        for e in dyst_err[:10]:
            print(" DYST", e)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
