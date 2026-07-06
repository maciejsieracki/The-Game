#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DEPRECATED (2026-06-30 — PANEL-MERGE):
  NIGDY nie uruchamiaj — nadpisuje wiele JSON naraz.
  Kanon: panele-sterowania/Panel-{A..E}.xlsx + export-{a..e}.py

export-data.py
Converts Excel design files from Civ/ into UTF-8 JSON in gra/data/.
"""

import json
import re
import openpyxl
from pathlib import Path

BASE_DIR = Path("/sessions/epic-jolly-heisenberg/mnt/Civ")
DATA_DIR = BASE_DIR / "gra" / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)


# ─── helpers ──────────────────────────────────────────────────────────────────

def clean_value(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    if isinstance(v, float):
        return int(v) if v == int(v) else v
    return v


def is_empty_row(row):
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in row)


def is_section_label(row, n_cols):
    """True when text only in first cell — epoch/section separator."""
    if row[0] is None:
        return False
    non_none = [v for v in row[1:] if v is not None and str(v).strip() != ""]
    return len(non_none) == 0 and isinstance(row[0], str)


def rows_to_records(ws, header_row_index=1, skip_section_labels=True):
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) <= header_row_index:
        return []
    headers = [clean_value(h) for h in all_rows[header_row_index]]
    headers = [h if h is not None else f"col_{i}" for i, h in enumerate(headers)]
    n_cols = len(headers)
    records = []
    for row in all_rows[header_row_index + 1:]:
        padded = list(row) + [None] * n_cols
        padded = padded[:n_cols]
        if is_empty_row(padded):
            continue
        if skip_section_labels and is_section_label(padded, n_cols):
            continue
        obj = {k: clean_value(v) for k, v in zip(headers, padded)}
        records.append(obj)
    return records


def save_json(data, filename):
    path = DATA_DIR / filename
    with open(str(path), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  Saved {filename} — {record_count(data)} record(s)")


def record_count(data):
    if isinstance(data, list):
        return len(data)
    if isinstance(data, dict):
        total = 0
        for v in data.values():
            if isinstance(v, list):
                total += len(v)
            elif isinstance(v, dict):
                for vv in v.values():
                    if isinstance(vv, list):
                        total += len(vv)
        return total
    return 1


# ─── Jednostki.xlsx ───────────────────────────────────────────────────────────

def export_jednostki():
    print("Jednostki.xlsx")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Jednostki.xlsx"), data_only=True)
    save_json(rows_to_records(wb["Jednostki"], header_row_index=1), "units.json")
    save_json(rows_to_records(wb["Countery"], header_row_index=1), "counters.json")
    save_json(rows_to_records(wb["Teren"], header_row_index=1), "terrain-combat.json")


# ─── Budynki.xlsx ─────────────────────────────────────────────────────────────

def export_budynki():
    print("Budynki.xlsx")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Budynki.xlsx"), data_only=True)
    ws = wb["Budynki"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Row 0 = formula note (skip), Row 1 = headers, Row 2+ = data
    headers = [clean_value(h) for h in all_rows[1]]

    def idx(name):
        for i, h in enumerate(headers):
            if h and str(h).strip() == name:
                return i
        raise KeyError(f"Column not found: {name!r} in {headers}")

    buildings = []
    for row in all_rows[2:]:
        if is_empty_row(row):
            continue
        if is_section_label(row, len(headers)):
            continue

        def get(name):
            try:
                return clean_value(row[idx(name)])
            except (KeyError, IndexError):
                return None

        def get_num(name):
            v = get(name)
            if v is None:
                return 0
            try:
                return int(v) if float(v) == int(float(v)) else float(v)
            except (TypeError, ValueError):
                return 0

        nazwy_raw = get("Nazwy poziomow") or ""
        nazwy_list = [n.strip() for n in str(nazwy_raw).split("|") if n.strip()] if nazwy_raw else []

        building = {
            "id":              get("ID"),
            "nazwa":           get("Nazwa bazowa"),
            "kategoria":       get("Kategoria"),
            "epokaWejscia":    get_num("Epoka wejscia"),
            "maksPoziom":      get_num("Maks poziom"),
            "nazwyPoziomow":   nazwy_list,
            "baza": {
                "praca":       get_num("Baza Praca"),
                "pieniadz":    get_num("Baza Pieniadz"),
                "zywnosc":     get_num("Baza Zywnosc"),
                "nauka":       get_num("Baza Nauka"),
                "kultura":     get_num("Baza Kultura"),
                "zadowolenie": get_num("Baza Zadowolenie"),
                "obrona":      get_num("Baza Obrona"),
                "mnoznik":     get_num("Baza Mnoznik %"),
            },
            "przyrost": {
                "praca":       get_num("Przyrost Praca"),
                "pieniadz":    get_num("Przyrost Pieniadz"),
                "zywnosc":     get_num("Przyrost Zywnosc"),
                "nauka":       get_num("Przyrost Nauka"),
                "kultura":     get_num("Przyrost Kultura"),
                "zadowolenie": get_num("Przyrost Zadowolenie"),
                "obrona":      get_num("Przyrost Obrona"),
                "mnoznik":     get_num("Przyrost Mnoznik %"),
            },
            "kosztBudowy":       get_num("Koszt budowy (Praca)"),
            "przyrostKosztu":    get_num("Przyrost kosztu/poz"),
            "utrzymanie":        get_num("Utrzymanie (Pieniadz/ture)"),
            "przyrostUtrzymania": get_num("Przyrost utrzymania/poz"),
            "wymagania":         get("Wymagania") or "",
            "uwagi":             get("Uwagi") or "",
            "techUnlock":        get("Technologia odblokowujaca") or "",
        }
        if building["id"] is not None:
            buildings.append(building)

    save_json(buildings, "buildings.json")




# ─── Surowce.xlsx ─────────────────────────────────────────────────────────────

def export_surowce():
    print("Surowce.xlsx")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Surowce.xlsx"), data_only=True)
    save_json(rows_to_records(wb["Surowce"], header_row_index=1), "resources.json")


# ─── Technologie-drzewko.xlsx ─────────────────────────────────────────────────

def export_technologie():
    print("Technologie-drzewko.xlsx")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Technologie-drzewko.xlsx"), data_only=True)
    save_json(rows_to_records(wb["Technologie"], header_row_index=1), "tech.json")


# ─── Cywilizacje.xlsx ─────────────────────────────────────────────────────────

def export_cywilizacje():
    print("Cywilizacje.xlsx")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Cywilizacje.xlsx"), data_only=True)
    result = {}
    for sname in wb.sheetnames:
        key = sname.lower().replace(" ", "_")
        result[key] = rows_to_records(wb[sname], header_row_index=1)
    save_json(result, "civs.json")


# ─── Plony-terenow.xlsx ───────────────────────────────────────────────────────
# Two logical tables:
#   Table 1 (terrain types):  header=row2 (idx1), data=rows3-9 (idx2-8)
#   Table 2 (modifiers):      header=row12 (idx11), data=rows13-14 (idx12-13)
#   Rows 15+ are legend text — skip.

def export_plony():
    print("Plony-terenow.xlsx")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Plony-terenow.xlsx"), data_only=True)
    ws = wb["Tereny"]
    all_rows = list(ws.iter_rows(values_only=True))

    def make_records(header_idx, data_indices):
        headers = [clean_value(h) for h in all_rows[header_idx]]
        headers = [h if h is not None else f"col_{i}" for i, h in enumerate(headers)]
        n = len(headers)
        records = []
        for idx in data_indices:
            if idx >= len(all_rows):
                continue
            row = list(all_rows[idx]) + [None] * n
            row = row[:n]
            if is_empty_row(row):
                continue
            records.append({k: clean_value(v) for k, v in zip(headers, row)})
        return records

    result = {
        "terrain_types": make_records(1, range(2, 9)),
        "terrain_modifiers": make_records(11, range(12, 14))
    }
    save_json(result, "terrain-yields.json")


# ─── Dyplomacja.xlsx ──────────────────────────────────────────────────────────

def parse_panel_sterowania(ws):
    """
    Panel sterowania is a config panel with labeled sections A-F.
    Exported as dict: section_letter -> {name, records}.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    section_pattern = re.compile(r"^([A-F])\.\s+(.+)")

    sections = {}
    current_key = None
    current_name = None
    current_headers = None
    current_records = []

    def flush():
        if current_key and current_records:
            sections[current_key] = {
                "name": current_name,
                "records": current_records[:]
            }

    for row in all_rows:
        if is_empty_row(row):
            continue
        first = clean_value(row[0])
        if first is None:
            continue
        if isinstance(first, str):
            m = section_pattern.match(first)
            if m:
                flush()
                current_key = m.group(1)
                current_name = m.group(2).strip()
                current_headers = None
                current_records = []
                continue
        if current_key and current_headers is None:
            hdrs = [clean_value(v) for v in row]
            non_none = [h for h in hdrs if h is not None]
            if len(non_none) >= 2:
                current_headers = hdrs
                continue
        if current_key and current_headers is not None:
            n = len(current_headers)
            padded = list(row) + [None] * n
            padded = padded[:n]
            obj = {}
            for k, v in zip(current_headers, padded):
                if k is not None:
                    obj[k] = clean_value(v)
            if obj:
                current_records.append(obj)

    flush()
    return sections


def export_dyplomacja():
    print("Dyplomacja.xlsx")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Dyplomacja.xlsx"), data_only=True)
    result = {}
    for sname in wb.sheetnames:
        key = sname.lower().replace(" ", "_").replace("—", "-").replace("–", "-")
        ws = wb[sname]
        if sname == "Panel sterowania":
            result[key] = parse_panel_sterowania(ws)
        else:
            result[key] = rows_to_records(ws, header_row_index=1)
    save_json(result, "diplomacy.json")


# ─── Ekonomia-parametry.xlsx ──────────────────────────────────────────────────
# 4 economic sheets only (society sheets moved to Spoleczenstwo-parametry.xlsx):
#   Ekonomia miasta | Ekonomia budynkow | Ekonomia terenu i mapy | Globalne
# Output: econ-params.json  { sheet_key: { param_key: {easy,normal,hard,jednostka,opis} } }

ECON_SHEET_KEY_MAP = {
    "Ekonomia miasta":         "ekonomia_miasta",
    "Ekonomia budynków": "budynki",
    "Ekonomia terenu i mapy":  "teren_mapa",
    "Globalne":                "globalne",
}


def _parse_econ_sheet(ws):
    """Return {param_key: {easy, normal, hard, jednostka, opis}} for one thematic sheet.
    Row 1 = sheet title (merged), Row 2 = column headers (Parametr|Easy|Normal|Hard|Jednostka|Opis),
    Row 3+ = data.  Section separator rows (▸ …) are skipped.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    params = {}
    for row in all_rows[2:]:            # skip title row and header row
        if is_empty_row(row):
            continue
        key = clean_value(row[0])
        if key is None:
            continue
        if isinstance(key, str) and key.startswith("▸"):
            continue                    # section separator (▸)
        easy   = clean_value(row[1])
        normal = clean_value(row[2])
        hard   = clean_value(row[3])
        unit   = clean_value(row[4])
        opis   = clean_value(row[5])
        params[key] = {
            "easy":      easy,
            "normal":    normal,
            "hard":      hard,
            "jednostka": unit,
            "opis":      opis,
        }
    return params


def export_econ_params():
    print("Ekonomia-parametry.xlsx")
    xlsx = BASE_DIR / "Ekonomia-parametry.xlsx"
    if not xlsx.exists():
        print("  SKIP — file not found")
        return
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)

    result = {}
    for sheet_name, json_key in ECON_SHEET_KEY_MAP.items():
        if sheet_name in wb.sheetnames:
            sheet_params = _parse_econ_sheet(wb[sheet_name])
            result[json_key] = sheet_params
            print(f"  {json_key}: {len(sheet_params)} parametrow")
        else:
            print(f"  WARNING: sheet '{sheet_name}' not found — skipped")

    total = sum(len(v) for v in result.values())
    save_json(result, "econ-params.json")
    print(f"  [{total} parametrow lacznie]")


# ─── Spoleczenstwo-parametry.xlsx ─────────────────────────────────────────────
# 5 society sheets:
#   Zdrowie | Szczescie | Kultura | Religia — numeric params {easy,normal,hard,...}
#   Religie cywilizacji — list of records
# Output: society-params.json

SOCIETY_SHEET_KEY_MAP = {
    "Zdrowie":         "zdrowie",
    "Szczęście": "szczescie",
    "Kultura":         "kultura",
    "Religia":         "religia",
}

SOCIETY_SHEET_KEY_TABLE = {
    "Religie cywilizacji": "religie_cywilizacji",
}


def export_society_params():
    print("Spoleczenstwo-parametry.xlsx")
    xlsx = BASE_DIR / "Spoleczenstwo-parametry.xlsx"
    if not xlsx.exists():
        print("  SKIP — file not found")
        return
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)

    result = {}

    # Numeric param sheets (same layout as econ sheets)
    for sheet_name, json_key in SOCIETY_SHEET_KEY_MAP.items():
        if sheet_name in wb.sheetnames:
            sheet_params = _parse_econ_sheet(wb[sheet_name])
            result[json_key] = sheet_params
            print(f"  {json_key}: {len(sheet_params)} parametrow")
        else:
            print(f"  WARNING: sheet '{sheet_name}' not found — skipped")

    # Table-layout sheet (list of records, header in row 2)
    for sheet_name, json_key in SOCIETY_SHEET_KEY_TABLE.items():
        if sheet_name in wb.sheetnames:
            records = rows_to_records(wb[sheet_name], header_row_index=1)
            result[json_key] = records
            print(f"  {json_key}: {len(records)} rekordow")
        else:
            print(f"  WARNING: sheet '{sheet_name}' not found — skipped")

    total = sum(len(v) for v in result.values())
    save_json(result, "society-params.json")
    print(f"  [{total} rekordow/parametrow lacznie]")


# ─── AI-parametry.xlsx ────────────────────────────────────────────────────────

def export_ai_params():
    print("AI-parametry.xlsx")
    xlsx = BASE_DIR / "AI-parametry.xlsx"
    if not xlsx.exists():
        print("  SKIP — file not found")
        return
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    # Row 0 = title merge, Row 1 = headers, Row 2+ = data
    params = {}
    for row in all_rows[2:]:
        if is_empty_row(row):
            continue
        key = clean_value(row[0])
        if key is None:
            continue
        params[key] = {
            "wartosc": clean_value(row[1]),
            "sekcja": clean_value(row[2]),
            "opis": clean_value(row[3]),
        }
    save_json(params, "ai-params.json")
    print(f"  [{len(params)} parametrow]")


# ─── Plony-terenow.xlsx / Ruch terenu ────────────────────────────────────────
# Sheet 'Ruch terenu':
#   Row 0  = title (skip)
#   Row 1  = column headers: Teren bazowy | Punkty ruchu (koszt wejscia) | Uwagi
#   Rows 2-8 = 7 base terrains  (key=col0, cost=col1)
#   Row 9 = empty
#   Row 10 = empty
#   Row 11 = modifier header: Modyfikator nakladki | Dodatkowy koszt ruchu | Uwagi
#   Row 12 = Las (nakladka) modifier data
#   Row 13+ = legend / notes (skip)
#
# Output: terrain-movement.json
#   { "costs": { "Laka":1, ... }, "forestExtra": 1 }

def export_ruch_terenu():
    print("Plony-terenow.xlsx / Ruch terenu")
    wb = openpyxl.load_workbook(str(BASE_DIR / "Plony-terenow.xlsx"), data_only=True)
    if "Ruch terenu" not in wb.sheetnames:
        print("  WARNING: sheet 'Ruch terenu' not found — skipped")
        return

    ws = wb["Ruch terenu"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Parse base terrain costs (rows 2-8, i.e. index 2 through 8)
    costs = {}
    DATA_START = 2
    DATA_END   = 9              # exclusive

    for idx in range(DATA_START, DATA_END):
        if idx >= len(all_rows):
            break
        row = all_rows[idx]
        key  = clean_value(row[0])
        cost = clean_value(row[1])
        if key is None or cost is None:
            continue
        costs[key] = cost

    # Parse forest modifier (row 12, i.e. index 12)
    FOREST_IDX = 12
    forest_extra = 1  # default fallback
    if FOREST_IDX < len(all_rows):
        row = all_rows[FOREST_IDX]
        val = clean_value(row[1])
        if val is not None:
            forest_extra = val

    result = {
        "costs": costs,
        "forestExtra": forest_extra,
    }
    save_json(result, "terrain-movement.json")
    print(f"  costs: {costs}")
    print(f"  forestExtra: {forest_extra}")

# ─── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    print("=== export-data.py ===")
    _exports = [
        ("export_jednostki",      export_jednostki),
        ("export_budynki",        export_budynki),
        ("export_surowce",        export_surowce),
        ("export_technologie",    export_technologie),
        ("export_cywilizacje",    export_cywilizacje),
        ("export_plony",          export_plony),
        ("export_dyplomacja",     export_dyplomacja),
        ("export_econ_params",    export_econ_params),
        ("export_society_params", export_society_params),
        ("export_ai_params",      export_ai_params),
        ("export_ruch_terenu",    export_ruch_terenu),
    ]
    for _name, _fn in _exports:
        try:
            _fn()
        except Exception as _e:
            print(f"WARNING: {_name} failed: {_e} (zachowuje istniejacy JSON)", file=sys.stderr)
    print("\nDone. Files in:", DATA_DIR)
