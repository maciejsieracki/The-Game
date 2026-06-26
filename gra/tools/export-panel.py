#!/usr/bin/env python3
# export-panel.py -- BEZPIECZNY eksport z Panel-przeglad-danych.xlsx z powrotem do JSON-ow.
# Na razie TYLKO moje (Civ-MIASTO) zakladki: Budynki -> buildings.json, Spoleczenstwo -> society-params.json.
# Zasada: nakladamy wartosci z Excela na ORYGINALNY JSON (zachowujemy strukture, typy, kolejnosc kluczy).
# Puste komorki = brak zmiany. Kolumna "Komentarz Naster" jest IGNOROWANA (to ludzkie notatki, nie schema).
# NIE rusza cudzych JSON-ow. Uzycie:
#   python3 gra/tools/export-panel.py                 # czyta Panel-przeglad-danych.xlsx, pisze do gra/data
#   python3 gra/tools/export-panel.py --xlsx X --out-dir D   # do testow
import openpyxl, json, os, sys, argparse

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(HERE, "..", "data"))
CIV_ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
XLSX = os.path.join(CIV_ROOT, "MIASTO", "Panel-przeglad-danych.xlsx")
TOP = {"zdrowie", "szczescie", "kultura", "religia", "religie_cywilizacji"}

def coerce(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""): return old
    if isinstance(old, bool):
        if isinstance(val, bool): return val
        return str(val).strip().lower() in ("tak", "true", "1", "prawda")
    if isinstance(old, int):
        try: return int(round(float(val)))
        except Exception: return old
    if isinstance(old, float):
        try: return float(val)
        except Exception: return old
    if old is None: return val
    return str(val)

def overlay_buildings(ws, orig):
    header = [c.value for c in ws[1]]
    by_id = {b["id"]: b for b in orig}
    changed = 0
    for r in range(2, ws.max_row + 1):
        row = {header[j]: ws.cell(r, j + 1).value for j in range(len(header))}
        rid = row.get("id")
        if rid is None or rid not in by_id: continue
        b = by_id[rid]
        for col, val in row.items():
            if not col or col in ("Komentarz Naster", "id"): continue
            if col == "nazwyPoziomow":
                cur = b.get("nazwyPoziomow", [])
                if val is None or str(val).strip() == "": continue
                if ", ".join(map(str, cur)).strip() == str(val).strip(): continue
                b["nazwyPoziomow"] = [s.strip() for s in str(val).split(",")]; changed += 1
                continue
            if "." in col:
                p, c = col.split(".", 1)
                if isinstance(b.get(p), dict) and c in b[p]:
                    nv = coerce(b[p][c], val)
                    if nv != b[p][c]: b[p][c] = nv; changed += 1
            elif col in b:
                nv = coerce(b[col], val)
                if nv != b[col]: b[col] = nv; changed += 1
    return changed

def overlay_society(ws, orig):
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]
    header, mode = None, None
    map_over, list_over = {}, {}
    for row in rows:
        c1 = row[0]
        s = str(c1).strip() if c1 is not None else ""
        if s in TOP:
            header, mode = None, "await"; continue
        if not any(v not in (None, "") for v in row): continue
        if mode == "await":
            header = [(str(v).strip() if v is not None else "") for v in row]; mode = "data"; continue
        if mode == "data" and header:
            if header[0] == "klucz":
                for j in range(1, len(header)):
                    col = header[j]
                    if col in ("", "Komentarz Naster"): continue
                    map_over.setdefault(s, {})[col] = row[j]
            else:
                d = {}
                for j in range(len(header)):
                    col = header[j]
                    if col in ("", "Komentarz Naster"): continue
                    d[col] = row[j]
                list_over[s] = d
    changed = 0
    for k, v in orig.items():
        if isinstance(v, dict):
            for param, obj in v.items():
                ov = map_over.get(param)
                if not ov: continue
                for subk in list(obj.keys()):
                    if subk in ov:
                        nv = coerce(obj[subk], ov[subk])
                        if nv != obj[subk]: obj[subk] = nv; changed += 1
        elif isinstance(v, list):
            for elem in v:
                if not isinstance(elem, dict) or not elem: continue
                fk = next(iter(elem))
                key = str(elem.get(fk, "")).strip()
                ov = list_over.get(key)
                if not ov: continue
                for col, val in ov.items():
                    if col in elem:
                        nv = coerce(elem[col], val)
                        if nv != elem[col]: elem[col] = nv; changed += 1
    return changed

def overlay_miasto(ws, orig):
    # miasto-params.json = mapa param -> {wartosc, jednostka, opis}; arkusz = jedna tabela (klucz + subkolumny)
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    changed = 0
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if key is None: continue
        key = str(key).strip()
        if key not in orig or not isinstance(orig[key], dict): continue
        for c in range(2, len(header) + 1):
            col = header[c - 1]
            if col in (None, "", "Komentarz Naster", "klucz"): continue
            if col in orig[key]:
                nv = coerce(orig[key][col], ws.cell(r, c).value)
                if nv != orig[key][col]: orig[key][col] = nv; changed += 1
    return changed

def overlay_terrain(ws, orig):
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    changed = 0
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if rid is None: continue
        rid = str(rid).strip()
        if rid not in orig or not isinstance(orig[rid], dict): continue
        node = orig[rid]
        for c in range(2, len(header) + 1):
            col = header[c - 1]
            if col in (None, "", "Komentarz Naster", "klucz"): continue
            val = ws.cell(r, c).value
            if isinstance(col, str) and col.startswith("bonus."):
                bk = col.split(".", 1)[1]
                bonus = node.setdefault("bonus", {})
                if bk in bonus:
                    nv = coerce(bonus[bk], val)
                    if nv != bonus[bk]: bonus[bk] = nv; changed += 1
            elif col in node:
                nv = coerce(node[col], val)
                if nv != node[col]: node[col] = nv; changed += 1
    return changed

def save_json(obj, path):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2); f.write("\n")
    json.load(open(tmp, encoding="utf-8"))  # walidacja
    os.replace(tmp, path)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX)
    ap.add_argument("--out-dir", default=DATA_DIR)
    ap.add_argument("--src-dir", default=DATA_DIR, help="skad czytac oryginalne JSON-y (struktura)")
    args = ap.parse_args()
    wb = openpyxl.load_workbook(args.xlsx)
    total = 0
    # Budynki
    if "Budynki" in wb.sheetnames:
        orig = json.load(open(os.path.join(args.src_dir, "buildings.json"), encoding="utf-8"))
        ch = overlay_buildings(wb["Budynki"], orig)
        save_json(orig, os.path.join(args.out_dir, "buildings.json"))
        print(f"buildings.json: zmienionych wartosci = {ch}"); total += ch
    # Spoleczenstwo
    if "Spoleczenstwo" in wb.sheetnames:
        orig = json.load(open(os.path.join(args.src_dir, "society-params.json"), encoding="utf-8"))
        ch = overlay_society(wb["Spoleczenstwo"], orig)
        save_json(orig, os.path.join(args.out_dir, "society-params.json"))
        print(f"society-params.json: zmienionych wartosci = {ch}"); total += ch
    # Miasto-parametry
    if "Miasto-parametry" in wb.sheetnames:
        orig = json.load(open(os.path.join(args.src_dir, "miasto-params.json"), encoding="utf-8"))
        ch = overlay_miasto(wb["Miasto-parametry"], orig)
        save_json(orig, os.path.join(args.out_dir, "miasto-params.json"))
        print(f"miasto-params.json: zmienionych wartosci = {ch}"); total += ch
    # Ulepszenia-terenu
    if "Ulepszenia-terenu" in wb.sheetnames:
        orig = json.load(open(os.path.join(args.src_dir, "terrain-improvements.json"), encoding="utf-8"))
        ch = overlay_terrain(wb["Ulepszenia-terenu"], orig)
        save_json(orig, os.path.join(args.out_dir, "terrain-improvements.json"))
        print(f"terrain-improvements.json: zmienionych wartosci = {ch}"); total += ch
    print(f"RAZEM zmian: {total}")

if __name__ == "__main__":
    main()
