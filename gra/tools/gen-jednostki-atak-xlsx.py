#!/usr/bin/env python3
"""Eksport parametrów ataku jednostek z units.json → Excel (Maciej review)."""
from __future__ import annotations

import json
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("BRAK openpyxl — uruchom: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
UNITS_PATH = os.path.normpath(os.path.join(HERE, "..", "data", "units.json"))
OUT = os.path.normpath(os.path.join(HERE, "Jednostki-atak-TW-v3.xlsx"))

HEADERS = [
    "Jednostka",
    "Epoka",
    "Rola (linia)",
    "AP (meleeAttack)",
    "Obraż (weaponDamage)",
    "Przeb (piercing)",
    "Szarża (chargeBonus)",
    "AD (missileAttack)",
    "wallAttack",
    "Zasięg (hex)",
    "Ilość pocisków",
    "A = AP+Obraż+Przeb+Szarża/2+AD/2",
]

FONT = "Arial"
H = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="4A5568")
TF = Font(name=FONT, color="1A1A1A")
NF = Font(name=FONT, color="0000FF")
KF = Font(name=FONT, bold=True, color="1A1A1A")
NOTE_FILL = PatternFill("solid", fgColor="FFF6CC")
thin = Side(style="thin", color="D9D9D9")
B = Border(thin, thin, thin, thin)


def n(v, default=0):
    if v is None or v == "" or v == "—":
        return default
    if isinstance(v, (int, float)):
        return v
    return default


def attack_row(u: dict) -> list:
    ap = n(u.get("meleeAttack"))
    ob = n(u.get("weaponDamage"))
    pr = n(u.get("piercing"))
    sz = n(u.get("chargeBonus"))
    ad = n(u.get("missileAttack"))
    wall = u.get("wallAttack")
    wall_out = "" if wall is None else wall
    zas = u.get("Zasięg ataku (hex)", "—")
    poc = u.get("Ilość pocisków", "—")
    a = ap + ob + pr + sz / 2 + ad / 2
    return [
        u["Jednostka"],
        u.get("Epoka", ""),
        u.get("Rola (linia)", ""),
        ap,
        ob,
        pr,
        sz,
        ad,
        wall_out,
        zas if zas is not None else "—",
        poc if poc is not None else "—",
        round(a, 1),
    ]


def autosize(ws, nrows: int, ncols: int, max_w: int = 44):
    for col in range(1, ncols + 1):
        w = 10
        for row in range(1, nrows + 1):
            val = ws.cell(row, col).value
            if val is not None:
                w = max(w, min(len(str(val)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col)].width = w


def main() -> None:
    units = json.load(open(UNITS_PATH, encoding="utf-8"))
    rows = sorted((attack_row(u) for u in units), key=lambda r: r[0].lower())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atak"

    ws["A1"] = "Jednostki — parametry ATAKU (TW v3)"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = f"Źródło: gra/data/units.json · wygenerowano: {date.today().isoformat()}"
    ws["A2"].font = Font(name=FONT, italic=True, size=10, color="666666")
    ws["A3"] = "Obrona (OBR, Panc, HP) — osobny arkusz / następny eksport."
    ws["A3"].font = Font(name=FONT, size=10, color="666666")

    hdr_row = 5
    for j, h in enumerate(HEADERS, 1):
        c = ws.cell(hdr_row, j, h)
        c.font = H
        c.fill = HF
        c.border = B
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for i, row in enumerate(rows, hdr_row + 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j, val)
            cell.border = B
            cell.alignment = Alignment(vertical="top", wrap_text=j <= 3)
            if j == 1:
                cell.font = KF
            elif 4 <= j <= 12 and isinstance(val, (int, float)):
                cell.font = NF
            else:
                cell.font = TF

    ws.freeze_panes = "D6"
    autosize(ws, hdr_row + len(rows), len(HEADERS))

    # Legenda
    leg = wb.create_sheet("Legenda")
    leg["A1"] = "Pole JSON"
    leg["B1"] = "Skrót / opis"
    leg["A1"].font = H
    leg["B1"].font = H
    items = [
        ("meleeAttack", "AP — szansa trafienia wręcz"),
        ("weaponDamage", "Obraż — obrażenia broni"),
        ("piercing", "Przeb — przebicie pancerza"),
        ("chargeBonus", "Szarża — bonus szarży"),
        ("missileAttack", "AD — atak dystansowy"),
        ("wallAttack", "Atak vs mur/brama (oblężenie)"),
        ("A", "Suma składników ataku na polu (bez wallAttack)"),
    ]
    for i, (k, v) in enumerate(items, 2):
        leg.cell(i, 1, k).font = KF
        leg.cell(i, 2, v).font = TF
    leg.column_dimensions["A"].width = 18
    leg.column_dimensions["B"].width = 48

    wb.save(OUT)
    print(OUT)
    print(f"Wierszy jednostek: {len(rows)}")


if __name__ == "__main__":
    main()
