#!/usr/bin/env python3
"""
sync-status-tracker-xlsx.py — kanoniczny tracker zadań grup A–F + Master.

Plik: Status-projektu-The-Game.xlsx (root Civ)

Arkusze operacyjne (grupy uzupełniają na bieżąco):
  Dashboard, Grupa-A … Grupa-F, Master-Silnik, Otwarte-ABC

Statusy: NIE ROZPOCZĘTE | W TOKU | GOTOWE | → SILNIK | ZAMKNIĘTE | BLOK | CZEKA ABC

Uruchomienie (Master po playtest / czaty):
  python gra/tools/sync-status-tracker-xlsx.py
  python gra/tools/sync-status-tracker-xlsx.py --dry-run

Grupy: edytuj SWÓJ arkusz; Master odświeża Dashboard skryptem.
"""
from __future__ import annotations

import argparse
import datetime
import os
import shutil
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Brak openpyxl: pip install openpyxl")
    sys.exit(1)

_TOOLS = os.path.dirname(os.path.abspath(__file__))
if _TOOLS not in sys.path:
    sys.path.insert(0, _TOOLS)
from game_scope_data import (  # noqa: E402
    CIV_MAPA_STEPS,
    CIV_SILNIK_STEPS,
    POSTEP_ROWS,
    SCOPE_HEADERS,
    SCOPE_STATUS_COLORS,
    STATUS_WG_GRUP,
)

CIV_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(CIV_ROOT, "Status-projektu-The-Game.xlsx")

HEADERS = (
    "ID",
    "Temat",
    "Status",
    "Priorytet",
    "Owner",
    "Decyzja",
    "Data",
    "Uwagi",
)

# Kanon na 2026-06-27 — Master Silnik (playtest + weryfikacja bramki)
TODAY = datetime.date.today().isoformat()

GROUP_SHEETS: dict[str, list[tuple]] = {
    "Grupa-A": [
        ("A-START-01", "Playtest: auto tryb budowy + Załóż miasto przy starcie", "NIE ROZPOCZĘTE", "P0", "Grupa A", "—", TODAY, "Maciej playtest ROBOCZA"),
        ("A-START-02", "Playtest: kamera max zoom (nie w chmurach)", "NIE ROZPOCZĘTE", "P0", "Grupa A", "—", TODAY, ""),
        ("A-START-03", "Playtest: rzeki respektują fog", "NIE ROZPOCZĘTE", "P0", "Grupa A", "—", TODAY, "scene.ts setFog"),
        ("A-START-04", "Playtest: minimap + fog jak mapa 3D", "NIE ROZPOCZĘTE", "P0", "Grupa A", "—", TODAY, "minimap.ts + minimapHud"),
        ("A-START-05", "Playtest: panel 🔨 opcja Załóż miasto", "NIE ROZPOCZĘTE", "P0", "Grupa A", "—", TODAY, "buildModeHud"),
        ("A2-Q4", "Panel jednostki [H]", "→ SILNIK", "P0", "Grupa A", "A", "2026-06-27", "UI gotowe; F wpięte"),
        ("A4-D4", "Budowa ulepszeń z mapy", "→ SILNIK", "P1", "Grupa A", "A", "2026-06-27", "lane GOTOWE"),
        ("B2-Q5-hex", "Ikona buntu 🔥 na heksie", "→ SILNIK", "P1", "Grupa A", "C", "2026-06-27", "getRevolt w kodzie F"),
        ("A1-Q15", "Power HUD — tylko wyświetlanie", "W TOKU", "P1", "Grupa A", "A", TODAY, "lane MAPA wylicza"),
        ("A3-Q1", "Panel armii", "W TOKU", "P2", "Grupa A", "B", TODAY, "handoff UNITS"),
        ("A5-Q1", "Wygląd miast 10 poziomów", "W TOKU", "P2", "Grupa A", "custom", TODAY, "MAPA mockup"),
    ],
    "Grupa-B": [
        ("B2-Q1", "Mieszkańcy 3 koszyki", "ZAMKNIĘTE", "—", "Grupa B", "A", "2026-06-26", ""),
        ("B2-Q2", "Porządek sekcja", "ZAMKNIĘTE", "—", "Grupa B", "B", "2026-06-26", ""),
        ("B2-Q3", "Zdrowie sekcja", "ZAMKNIĘTE", "—", "Grupa B", "A", "2026-06-26", ""),
        ("B2-Q4", "Specjaliści", "ZAMKNIĘTE", "—", "Grupa B", "C", "2026-06-26", ""),
        ("B2-Q5", "Alert buntu chip+hex", "ZAMKNIĘTE", "—", "Grupa B", "C", TODAY, "F wpięte"),
        ("B2-Q6", "Kary buntu migracja", "ZAMKNIĘTE", "—", "Grupa B", "C", TODAY, "F-B2-porzadek"),
        ("B2-Q7-9", "Szczęście / Porządek ABC", "CZEKA ABC", "P1", "Grupa B", "—", TODAY, "czat B"),
        ("B1.2-4", "Budowa / rush / auto / pola", "CZEKA ABC", "P1", "Grupa B", "—", TODAY, ""),
        ("B5-stub", "Żywność imperium advanceEmpireFood", "BLOK", "P2", "Grupa B", "—", TODAY, "stub — nie wpinaj"),
    ],
    "Grupa-C": [
        ("C1-Q1-5", "Wejście w walkę preBattle", "ZAMKNIĘTE", "—", "Grupa C", "mix", "2026-06-27", "F-C1 w kodzie"),
        ("C2-Q2-7", "UX bitwa TW", "ZAMKNIĘTE", "—", "Grupa C", "D5=B", "2026-06-27", "F-C2 w kodzie"),
        ("C3", "Oblężenie / szturm miasta", "CZEKA ABC", "P1", "Grupa C", "—", TODAY, "paczki ABC w czacie C"),
    ],
    "Grupa-D": [
        ("D1-D4", "Decyzje karta D1–D15 szczegóły", "ZAMKNIĘTE", "—", "Grupa D", "mix", "2026-06-26", ""),
        ("D4-bonusy", "Bonusy cywilizacji w grze", "W TOKU", "P1", "Grupa D", "3A", TODAY, "civ-bonusy 4 FAIL test"),
        ("D-testy", "diplomacy+ai testy bramka", "GOTOWE", "—", "Grupa D", "7B", TODAY, "133+188 PASS"),
        ("D-UI-handoff", "UI bonusy + modal wojny D3", "NIE ROZPOCZĘTE", "P2", "Grupa D", "—", TODAY, "po Opus"),
    ],
    "Grupa-E": [
        ("E1-UX-01", "Kreator: nawigacja blisko treści krok 2–5", "NIE ROZPOCZĘTE", "P1", "Grupa E", "—", TODAY, "newGameFlow.ts"),
        ("E1-kod", "Defaulty startu + generujSwiat", "→ SILNIK", "—", "Grupa E", "D13", TODAY, "wpięte F-A2"),
        ("E1-Q9", "Reset gracza Nowa gra", "CZEKA ABC", "P1", "Grupa E", "—", TODAY, ""),
        ("E1-Q10", "Start Brąz + tech", "CZEKA ABC", "P1", "Grupa E", "—", TODAY, ""),
        ("E1-Q11", "Kształt mapy Ziemia", "CZEKA ABC", "P1", "Grupa E", "—", TODAY, ""),
        ("E1-Q12", "Zakres rywali menu", "CZEKA ABC", "P1", "Grupa E", "—", TODAY, ""),
        ("E1-Q6-8", "Menu Kampania/Multi/media", "CZEKA ABC", "P2", "Grupa E", "—", TODAY, ""),
    ],
    "Grupa-F": [
        ("F-BRAMKA", "ROBOCZA opublikowana", "GOTOWE", "P0", "Grupa F", "—", TODAY, "md5 d813159b; Master zweryfikował"),
        ("F-batchy", "F1 F-A2 F-B2 F-C1 F-HUD F-HUD-2", "→ SILNIK", "P0", "Grupa F", "—", TODAY, "kod w main.ts"),
        ("F-PLAYTEST-A", "Batch poprawek A-START po GOTOWE od A", "CZEKA", "P0", "Grupa F", "—", TODAY, "nie zaczynaj przed Grupa A"),
        ("F-PROD-SPAWN", "Spawn jednostek z produkcji", "NIE ROZPOCZĘTE", "P1", "Grupa F", "—", TODAY, ""),
        ("F-Opus", "Czeka APPROVE → kanon", "W TOKU", "P0", "Master", "—", TODAY, "OPUS-REVIEW-QUEUE"),
    ],
}

DASHBOARD = [
    ("P0-playtest-A", "Grupa A: A-START-01…05 (5 zadań)", "NIE ROZPOCZĘTE", "P0", "Grupa A", "—", TODAY, "Maciej playtest — PRIORYTET"),
    ("P0-Opus", "Opus review → Gra-podglad.html", "W TOKU", "P0", "Master", "—", TODAY, "ROBOCZA gotowa"),
    ("P0-bramka", "ROBOCZA na dysku", "GOTOWE", "P0", "Grupa F", "—", TODAY, "testy PASS (4 fail civ-bonusy P2)"),
    ("P1-E-UX", "Kreator nawigacja E1-UX-01", "NIE ROZPOCZĘTE", "P1", "Grupa E", "—", TODAY, ""),
    ("P1-ABC-B", "Paczka B2-Q7… B1.x", "CZEKA ABC", "P1", "Grupa B", "—", TODAY, "czat B"),
    ("P1-ABC-E", "E1-Q9…Q12", "CZEKA ABC", "P1", "Grupa E", "—", TODAY, "czat E"),
    ("P1-ABC-C", "C3 oblężenie", "CZEKA ABC", "P1", "Grupa C", "—", TODAY, "czat C"),
]

MASTER_SHEET = [
    ("M-OPUS", "Review ROBOCZA → APPROVE/BLOCK", "W TOKU", "P0", "Master", "—", TODAY, "Opus Ask ręcznie"),
    ("M-KANON", "Promocja Gra-podglad.html po APPROVE", "NIE ROZPOCZĘTE", "P0", "Master", "—", TODAY, ""),
    ("M-PLAYTEST", "Checklista playtest Maciej (finalna)", "NIE ROZPOCZĘTE", "P1", "Master", "—", TODAY, "po kanonie"),
    ("M-ROUTING", "Dyspozycje OD/DO-MASTERA", "W TOKU", "—", "Master", "—", TODAY, ""),
    ("M-EXCEL", "Sync tracker (ten plik)", "W TOKU", "—", "Master", "—", TODAY, "sync-status-tracker-xlsx.py"),
]

OTWARTE_ABC = [
    ("E1-Q9", "Reset gracza", "OTWARTE", "P1", "Grupa E", "—", TODAY, ""),
    ("E1-Q10", "Start Brąz", "OTWARTE", "P1", "Grupa E", "—", TODAY, ""),
    ("E1-Q11", "Mapa Ziemia", "OTWARTE", "P1", "Grupa E", "—", TODAY, ""),
    ("E1-Q12", "Liczba rywali", "OTWARTE", "P1", "Grupa E", "—", TODAY, ""),
    ("B2-Q7", "Model szczęścia", "OTWARTE", "P1", "Grupa B", "—", TODAY, ""),
    ("B2-Q8", "Porządek szczęście", "OTWARTE", "P1", "Grupa B", "—", TODAY, ""),
    ("B2-Q9", "Kary porządek", "OTWARTE", "P1", "Grupa B", "—", TODAY, ""),
    ("C3", "Oblężenie", "OTWARTE", "P1", "Grupa C", "—", TODAY, ""),
]

STATUS_COLORS = {
    "GOTOWE": "C6EFCE",
    "ZAMKNIĘTE": "C6EFCE",
    "→ SILNIK": "FFEB9C",
    "W TOKU": "BDD7EE",
    "NIE ROZPOCZĘTE": "FFC7CE",
    "CZEKA ABC": "FCE4D6",
    "CZEKA": "FCE4D6",
    "BLOK": "FF9999",
    "OTWARTE": "FCE4D6",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header(ws) -> None:
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def write_data_rows(ws, rows: list[tuple], start_row: int = 2) -> None:
    status_col = HEADERS.index("Status") + 1
    for i, row in enumerate(rows):
        r = start_row + i
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == status_col:
                color = STATUS_COLORS.get(str(val), None)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)


def ensure_sheet(wb, name: str, rows: list[tuple]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet(name)
    style_header(ws)
    write_data_rows(ws, rows)
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14 if col != 2 else 42


def write_meta(wb) -> None:
    if "Meta" not in wb.sheetnames:
        ws = wb.create_sheet("Meta", 0)
    else:
        ws = wb["Meta"]
    lines = [
        ("INSTRUKCJA — Status-projektu-The-Game.xlsx", ""),
        ("Ostatnia synchronizacja Master", TODAY),
        ("", ""),
        ("WARSTWA 1 — CALOSC GRY (postep projektu):", ""),
        ("Status wg grup", "WSZYSTKIE elementy gry: Zrobione / Niezrobione / Czesciowo"),
        ("POSTEP-%", "Procenty per dzial + grywalnosc v0.1 / v1.0"),
        ("Podsumowanie", "Statystyki liczbowe z Status wg grup"),
        ("Civ-SILNIK, Civ-MAPA", "Kroki integracji / mapy (historia lane)"),
        ("", ""),
        ("WARSTWA 2 — OPERACYJNA (grupy A-F, sprint):", ""),
        ("Grupa-A … Grupa-F", "Zadania biezace — grupa uzupełnia Status + Data"),
        ("Dashboard", "Skrot P0/P1 — Master skryptem"),
        ("Master-Silnik", "Opus, kanon, routing"),
        ("Otwarte-ABC", "Pytania czekajace na Macieja"),
        ("", ""),
        ("Skrypt Master:", "python gra/tools/sync-status-tracker-xlsx.py"),
        ("ROBOCZA md5", "d813159b0726b94f8e360c53dadf72a8"),
        ("Kanon md5", "2276ec0f (stary — czeka Opus)"),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)


def write_scope_rows(ws, headers: tuple, rows: list[tuple], status_col_name: str = "Status") -> None:
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    status_col = headers.index(status_col_name) + 1 if status_col_name in headers else 3
    for i, row in enumerate(rows):
        r = i + 2
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == status_col:
                color = SCOPE_STATUS_COLORS.get(str(val))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)


def ensure_scope_sheet(wb, name: str, headers: tuple, rows: list[tuple]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    write_scope_rows(ws, headers, rows)
    widths = {1: 28, 2: 44, 3: 14, 4: 36, 5: 36, 6: 16, 7: 12, 8: 10}
    for col, w in widths.items():
        if col <= len(headers):
            ws.column_dimensions[get_column_letter(col)].width = w


def write_postep_sheet(wb) -> None:
    name = "POSTEP-%"
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    title_font = Font(bold=True, size=12)
    for i, row in enumerate(POSTEP_ROWS, 1):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if i == 1:
                cell.font = title_font
            if col == 2 and i > 2 and val and str(val).endswith("%"):
                try:
                    pct = float(str(val).replace("%", "").replace("~", ""))
                    if pct >= 75:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    elif pct >= 50:
                        cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FFC7CE")
                except ValueError:
                    pass
    for col, w in enumerate([32, 12, 40, 40, 24], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def write_civ_lane_sheet(wb, name: str, title: str, domain: str, steps: list[tuple]) -> None:
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    ws["A1"] = f"ZADANIE: {title}"
    ws["A2"] = f"Domena: {domain}"
    ws["A3"] = "Po wykonaniu kroku ustaw STATUS = Zrobione. Master odswieza skryptem."
    headers = ("#", "Krok", "Status", "Data / uwaga")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="C5D9F1")
    for i, step in enumerate(steps):
        r = 6 + i
        for col, val in enumerate(step, 1):
            cell = ws.cell(row=r, column=col, value=val)
            if col == 3:
                color = SCOPE_STATUS_COLORS.get(str(val))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["D"].width = 40


def _count_scope_stats() -> dict:
    from collections import Counter

    st = Counter()
    gr_done: dict[str, list[int]] = {}
    for row in STATUS_WG_GRUP:
        grupo = str(row[0]).split(".")[0].strip()
        status = str(row[2])
        st[status] += 1
        if grupo not in gr_done:
            gr_done[grupo] = [0, 0]
        gr_done[grupo][1] += 1
        if status == "Zrobione":
            gr_done[grupo][0] += 1
    return {"status": st, "groups": gr_done, "total": len(STATUS_WG_GRUP)}


def write_podsumowanie(wb) -> None:
    stats = _count_scope_stats()
    st = stats["status"]
    total = stats["total"]
    name = "Podsumowanie"
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    bold = Font(bold=True)
    rows = [
        ("THE GAME (Civ) — postep CALOSCI gry", None, None, None),
        (f"Stan: {TODAY} (auto-sync Master Silnik)", None, None, None),
        ("", None, None, None),
        ("STATYSTYKA wg STATUSU", "Liczba", "Udzial", None),
        ("Zrobione", st.get("Zrobione", 0), f"{100 * st.get('Zrobione', 0) / total:.0f}%", None),
        ("Czesciowo", st.get("Czesciowo", 0), f"{100 * st.get('Czesciowo', 0) / total:.0f}%", None),
        ("Gotowe, niewpiete", st.get("Gotowe, niewpiete", 0), f"{100 * st.get('Gotowe, niewpiete', 0) / total:.0f}%", None),
        ("Niezrobione", st.get("Niezrobione", 0), f"{100 * st.get('Niezrobione', 0) / total:.0f}%", None),
        ("RAZEM", total, "100%", None),
        ("", None, None, None),
        ("GRUPY (zakres funkcjonalny)", "Razem", "Zrobione", "% done"),
    ]
    group_labels = {
        "A": "A. Jednostki i walka",
        "B": "B. Mapa i teren",
        "C": "C. Miasta i ekonomia",
        "D": "D. AI i rozgrywka",
        "E": "E. Cywilizacje",
        "F": "F. Interfejs UI/HUD",
        "G": "G. Zapis i meta",
        "H": "H. Infrastruktura",
        "I": "I. Przyszlosc M7",
    }
    for key in "ABCDEFGHI":
        if key not in stats["groups"]:
            continue
        done, cnt = stats["groups"][key]
        pct = f"{100 * done / cnt:.0f}%" if cnt else "—"
        rows.append((group_labels.get(key, key), cnt, done, pct))
    rows.extend([
        ("", None, None, None),
        ("OPERACYJNIE (arkusze Grupa-A..F)", "Patrz Dashboard", None, None),
        ("Pelna lista elementow", "arkusz: Status wg grup", None, None),
    ])
    for i, row in enumerate(rows, 1):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if i in (1, 4, 11) and val:
                cell.font = bold
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12


def sync_full_game_scope(wb) -> None:
    ensure_scope_sheet(wb, "Status wg grup", SCOPE_HEADERS, STATUS_WG_GRUP)
    write_postep_sheet(wb)
    write_podsumowanie(wb)
    write_civ_lane_sheet(
        wb,
        "Civ-SILNIK",
        "Civ-SILNIK",
        "main.ts + wpinanie game/* + publisher ROBOCZA/kanonu",
        CIV_SILNIK_STEPS,
    )
    write_civ_lane_sheet(
        wb,
        "Civ-MAPA",
        "Civ-MAPA",
        "render/scene.ts + map/* + render/cities.ts (Grupa A)",
        CIV_MAPA_STEPS,
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.isfile(XLSX):
        print(f"BLOK: brak {XLSX}")
        return 1

    bak = XLSX + f".bak-tracker-{TODAY}"
    if not args.dry_run:
        shutil.copy2(XLSX, bak)
        print(f"Backup: {bak}")

    wb = openpyxl.load_workbook(XLSX)
    write_meta(wb)
    sync_full_game_scope(wb)
    ensure_sheet(wb, "Dashboard", DASHBOARD)
    for name, rows in GROUP_SHEETS.items():
        ensure_sheet(wb, name, rows)
    ensure_sheet(wb, "Master-Silnik", MASTER_SHEET)
    ensure_sheet(wb, "Otwarte-ABC", OTWARTE_ABC)

    if args.dry_run:
        print(
            "DRY-RUN — arkusze:",
            "Status wg grup, POSTEP-%, Podsumowanie, Civ-SILNIK, Civ-MAPA,",
            "Dashboard,", ", ".join(GROUP_SHEETS), "Master-Silnik, Otwarte-ABC",
        )
        return 0

    wb.save(XLSX)
    print(f"OK: {XLSX}")
    stats = _count_scope_stats()
    z = stats["status"].get("Zrobione", 0)
    print(f"CALOSC GRY: {z}/{stats['total']} Zrobione ({100*z//stats['total']}%) — arkusz Status wg grup")
    print("Zaktualizowano operacyjnie:", ", ".join(GROUP_SHEETS.keys()))
    return 0


if __name__ == "__main__":
    sys.exit(main())
