#!/usr/bin/env python3
"""
export-tech.py -- TARGETED export: Technologie-drzewko.xlsx[arkusz "Technologie"] -> gra/data/tech.json

Czyta arkusz technologii i regeneruje tech.json. Od 2026-06-25 tech.json ma format:
  { "tempo_gry": {...}, "technologie": [...] }
Skrypt zachowuje blok tempo_gry z istniejacego pliku (lub wstawia domyslny) i
aktualizuje tylko tablice technologie z danych xlsx.

Tryb --dry-run: wypisuje wynik i porownuje z obecnym tech.json -- NIE zapisuje.
Bez --dry-run: backup -> zapis -> walidacja przez json.tool (python3 -m json.tool).

Wazne ograniczenie: jesli --dry-run wykryje roznice miedzy xlsx a tech.json,
skrypt konczy sie kodem 1 i raportuje roznice -- nie nalezy wtedy uruchamiac bez --dry-run
bez swiadomej decyzji Macieja.

Uzycie (z dowolnego katalogu):
  python3 gra/tools/export-tech.py --dry-run
  python3 gra/tools/export-tech.py

Sciezki domyslne (wzgledne do tego pliku):
  xlsx = ../../Technologie-drzewko.xlsx
  json = ../data/tech.json
"""
import argparse
import json
import os
import shutil
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_XLSX = os.path.normpath(os.path.join(HERE, "..", "..", "Technologie-drzewko.xlsx"))
DEF_JSON = os.path.normpath(os.path.join(HERE, "..", "data", "tech.json"))
SHEET = "Technologie"

# Dokladne nazwy pol w tech.json (w tej kolejnosci)
JSON_FIELDS = [
    "Technologia",
    "Epoka",
    "Poziom",
    "Dostep do surowca.",
    "wymagany budynek",
    "Wymaga (prereq)",
    "Odblokowuje surowiec.",
    "Odblokowuje budynek",
    "Koszt nauki",
    "Uwagi",
]

# Mapowanie: kluczowe pole JSON -> naglowek xlsx (po strip)
# Potrzebne dla pol z diakrytykami, by porownac naglowki xlsx z polami JSON
FIELD_TO_HDR = {
    "Dostep do surowca.": "Dostep do surowca.",
    "Odblokowuje surowiec.": "Odblokowuje surowiec.",
}

# Rzeczywiste nazwy pol w JSON (z diakrytykami) - te sa zapisywane do JSON
JSON_FIELD_NAMES = [
    "Technologia",
    "Epoka",
    "Poziom",
    "Dostep do surowca.",
    "wymagany budynek",
    "Wymaga (prereq)",
    "Odblokowuje surowiec.",
    "Odblokowuje budynek",
    "Koszt nauki",
    "Uwagi",
]

DEFAULT_TEMPO_GRY = {
    "szybka": 0.2,
    "standardowa": 1.0,
    "dluga": 5.0,
    "_opis": (
        "Globalny mnoznik kosztu badan wybierany przy starcie gry. "
        "Stosuje sie do KAZDEGO 'Koszt nauki' w tablicy 'technologie' "
        "przez funkcje applyTempoKoszt() z gra/src/game/tech-tempo.ts."
    ),
}


def load_techs_from_json(path):
    """Wczytuje tech.json i zwraca (lista_technologii, tempo_gry_dict_lub_None).
    Obsluguje stary format (plain array) i nowy format ({ tempo_gry, technologie }).
    """
    with open(path, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    if isinstance(data, list):
        return data, None  # stary format -- brak tempo_gry
    return data.get("technologie", []), data.get("tempo_gry", None)


def read_xlsx(xlsx_path):
    """Czyta arkusz 'Technologie' i zwraca liste slownikow zgodnych z formatem tech.json."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Brak arkusza '{SHEET}' w {xlsx_path}")
    ws = wb[SHEET]

    # Znajdz wiersz naglowka (zawiera 'Technologia')
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Technologia":
            header_row = r
            break
    if header_row is None:
        sys.exit(f"Nie znaleziono wiersza naglowka (z 'Technologia') w arkuszu '{SHEET}'")

    # Zbuduj mapowanie: nazwa_pola_json -> numer_kolumny (strip naglowkow xlsx)
    col_map = {}
    for c in range(1, ws.max_column + 1):
        hdr = ws.cell(header_row, c).value
        if hdr is not None:
            key = hdr.strip()
            col_map[key] = c

    # Sprawdz, czy wszystkie pola JSON sa pokryte (po strip)
    missing = [f for f in JSON_FIELD_NAMES if f not in col_map]
    if missing:
        sys.exit(f"Brak kolumn w arkuszu (po strip): {missing}\nDostepne: {list(col_map.keys())}")

    techs = []
    for r in range(header_row + 1, ws.max_row + 1):
        # Pomij puste wiersze i wiersze-komentarze (kol 1 None lub nie-string)
        first = ws.cell(r, 1).value
        if first is None or not isinstance(first, str) or not first.strip():
            continue
        # Pomij wiersze z tekstem-opisem (np. "Drzewko = kolumna...")
        if len(first.strip()) > 60 and "=" in first:
            continue

        obj = {}
        for field in JSON_FIELD_NAMES:
            c = col_map[field]
            val = ws.cell(r, c).value

            # Normalizacja typow
            if field == "Poziom":
                # Zawsze int
                if val is not None:
                    try:
                        val = int(val)
                    except (TypeError, ValueError):
                        pass
            elif field == "Koszt nauki":
                # Zawsze int (lub float jesli ktos wpisze ulamek)
                if val is not None:
                    try:
                        f = float(val)
                        val = int(f) if f.is_integer() else f
                    except (TypeError, ValueError):
                        pass
            elif isinstance(val, str):
                val = val.strip() if val.strip() else None
            # None pozostaje None (odpowiada null w JSON)

            obj[field] = val

        techs.append(obj)

    return techs


def diff_techs(from_xlsx, from_json):
    """Porownuje dwie listy technologii. Zwraca liste opisow roznic."""
    diffs = []

    # Indeksuj po nazwie technologii
    xlsx_by_name = {t["Technologia"]: t for t in from_xlsx}
    json_by_name = {t["Technologia"]: t for t in from_json}

    only_xlsx = set(xlsx_by_name) - set(json_by_name)
    only_json = set(json_by_name) - set(xlsx_by_name)

    for name in sorted(only_xlsx):
        diffs.append(f"  TYLKO W XLSX (brak w json): {name}")
    for name in sorted(only_json):
        diffs.append(f"  TYLKO W JSON (brak w xlsx): {name}")

    for name in sorted(set(xlsx_by_name) & set(json_by_name)):
        tx = xlsx_by_name[name]
        tj = json_by_name[name]
        for field in JSON_FIELD_NAMES:
            vx = tx.get(field)
            vj = tj.get(field)
            if vx != vj:
                diffs.append(f"  [{name}] pole '{field}': xlsx={vx!r}  vs  json={vj!r}")

    return diffs


def main():
    ap = argparse.ArgumentParser(
        description="Export: Technologie-drzewko.xlsx -> gra/data/tech.json"
    )
    ap.add_argument("--xlsx", default=DEF_XLSX, help="Sciezka do xlsx (domyslnie: auto)")
    ap.add_argument("--json", default=DEF_JSON, help="Sciezka do tech.json (domyslnie: auto)")
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Wypisz wynik i porownaj z obecnym tech.json. NIE zapisuj.",
    )
    a = ap.parse_args()

    if not os.path.exists(a.xlsx):
        sys.exit(f"Nie znaleziono xlsx: {a.xlsx}")

    print(f"[export-tech] xlsx: {a.xlsx}")
    print(f"[export-tech] json: {a.json}")

    techs = read_xlsx(a.xlsx)
    print(f"[export-tech] odczytano {len(techs)} technologii z arkusza '{SHEET}'")

    if a.dry_run:
        print("\n--- DRY-RUN: wynik z xlsx ---")
        print(json.dumps(techs, ensure_ascii=False, indent=2))

        if not os.path.exists(a.json):
            print(f"\n[dry-run] Brak pliku {a.json} -- nie mozna porownac.")
            return

        current_techs, _ = load_techs_from_json(a.json)

        print(f"\n--- DRY-RUN: porownanie xlsx vs obecny tech.json ({len(current_techs)} technologii) ---")
        diffs = diff_techs(techs, current_techs)
        if not diffs:
            print("  IDENTYCZNE -- xlsx i tech.json sa zgodne. Mozna uruchomic bez --dry-run.")
        else:
            print(f"  ZNALEZIONO {len(diffs)} ROZNIC:")
            for d in diffs:
                print(d)
            print(
                "\n[STOP] Roznice wykryte. NIE uruchamiaj bez --dry-run bez decyzji Macieja."
            )
            sys.exit(1)
        return

    # --- Tryb realnego zapisu ---
    if not os.path.exists(a.json):
        sys.exit(f"Nie znaleziono {a.json} -- nie mozna zrobic backupu. Abortuj.")

    # Wczytaj istniejace dane i zachowaj blok tempo_gry
    current_techs, existing_tempo_gry = load_techs_from_json(a.json)
    tempo_gry = existing_tempo_gry if existing_tempo_gry is not None else DEFAULT_TEMPO_GRY

    diffs = diff_techs(techs, current_techs)
    if diffs:
        print(f"[STOP] Wykryto {len(diffs)} roznic miedzy xlsx a tech.json:")
        for d in diffs:
            print(d)
        print("Nie zapisano. Uruchom --dry-run i zdecyduj.")
        sys.exit(1)

    # Backup
    bak = a.json + ".bak-CYWILIZACJE"
    shutil.copy2(a.json, bak)
    print(f"[export-tech] backup: {bak}")

    # Zapis w nowym formacie: { tempo_gry, technologie }
    output = {"tempo_gry": tempo_gry, "technologie": techs}
    with open(a.json, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"[export-tech] zapisano {len(techs)} technologii (+ blok tempo_gry) -> {a.json}")

    # Walidacja json.tool
    result = subprocess.run(
        [sys.executable, "-m", "json.tool", a.json],
        capture_output=True,
        text=True,
    )
    if result.returncode == 0:
        print("[export-tech] walidacja json.tool: OK")
    else:
        print(f"[export-tech] BLAD walidacji json.tool:\n{result.stderr}")
        sys.exit(1)


if __name__ == "__main__":
    main()
