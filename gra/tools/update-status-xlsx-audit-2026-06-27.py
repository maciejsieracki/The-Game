#!/usr/bin/env python3
"""Aktualizacja Status-projektu-The-Game.xlsx po audycie Master 2026-06-27."""
import datetime
import os
import shutil
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Brak openpyxl: pip install openpyxl")
    sys.exit(1)

CIV_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(CIV_ROOT, "Status-projektu-The-Game.xlsx")
TODAY = "2026-06-27"
HEADERS = ("ID", "Temat", "Status", "Decyzja", "Data", "Uwagi")

SHEETS = {
    "Dashboard": [
        ("P0-bramka", "F-BRAMKA → ROBOCZA", "BLOK", "—", TODAY, "Node u Macieja; fix newW/newH L3229"),
        ("P0-start", "Menu→mapa→B→miasto", "BLOK", "7B", TODAY, "ReferenceError newW w doStartGame"),
        ("P0-Opus", "Promocja finalna", "CZEKA", "—", TODAY, "Po ROBOCZA + Opus APPROVE"),
        ("ABC1", "Mockup HUD D1B", "ZAMKNIĘTE", "A", TODAY, "F-HUD częściowo w main"),
        ("B2-paczka", "Społeczeństwo Q1-Q6", "ZAMKNIĘTE", "mix", TODAY, "F-B2+F-B2-porzadek w kodzie"),
        ("C1-C2", "preBattle + UX bitwy", "ZAMKNIĘTE", "D5=B", TODAY, "F-C1 w kodzie; F-C2 kolejka"),
    ],
    "Grupa-A": [
        ("A1-Q11", "Kultura na pasku [A]", "ZAMKNIĘTE", "A", TODAY, "kulturaRate w hud"),
        ("A2-Q4", "Panel jednostki [H]", "ZAMKNIĘTE", "A", TODAY, "UI gotowe; F-HUD-2 czeka"),
        ("A1-Q9", "WYKONAJ + gate tury", "ZAMKNIĘTE", "A", TODAY, "UI gotowe; F-HUD-2 czeka"),
        ("B2-Q5-chip", "Chip buntu HUD", "ZAMKNIĘTE", "C", TODAY, "UI lane P1 Grupa A"),
        ("B2-Q5-hex", "Ikona buntu na heksie", "OTWARTE impl", "C", TODAY, "MAPA getRevolt — CZEKA"),
        ("A4-D4", "Budowa z mapy", "ZAMKNIĘTE", "A", TODAY, "P2 wpięcie po P1"),
    ],
    "Grupa-B": [
        ("B2-Q1", "Mieszkańcy 3 koszyki", "ZAMKNIĘTE", "A", "2026-06-26", ""),
        ("B2-Q2", "Porządek sekcja", "ZAMKNIĘTE", "B", "2026-06-26", ""),
        ("B2-Q3", "Zdrowie sekcja", "ZAMKNIĘTE", "A", "2026-06-26", ""),
        ("B2-Q4", "Specjaliści", "ZAMKNIĘTE", "C", "2026-06-26", ""),
        ("B2-Q5", "Alert buntu", "ZAMKNIĘTE", "C", TODAY, "chip+hex"),
        ("B2-Q6", "Kary buntu", "ZAMKNIĘTE", "C", TODAY, "F-B2-porzadek"),
        ("B2-Q7-9", "Szczęście/Porządek", "OTWARTE ABC", "—", TODAY, "B-OTWARTE-PYTANIA.md"),
        ("B1.2-4", "Budowa/rush/auto/pola", "OTWARTE ABC", "—", TODAY, "po B2"),
    ],
    "Grupa-F": [
        ("F-START-FIX", "newW/newH + hint B", "TODO", "—", TODAY, "P0 bloker"),
        ("F-BRAMKA", "ROBOCZA publish", "TODO", "—", TODAY, "Po fix + Node"),
        ("F-HUD", "D1B batch 1", "WPIĘTE kod", "ABC1=A", TODAY, "czeka bramka"),
        ("F-HUD-2", "WYKONAJ+[H]+budowa", "CZEKA A", "—", TODAY, "UI lane GOTOWE"),
        ("F-PROD-SPAWN", "Spawn z produkcji", "TODO", "—", TODAY, "L2360 brak units.push"),
        ("F-C2", "UX bitwy kanon", "TODO", "—", TODAY, "UNITS lane gotowy"),
    ],
    "Otwarte-ABC": [
        ("E1-Q9", "Reset gracza", "OTWARTE", "—", TODAY, "czat E"),
        ("E1-Q10", "Start Brąz", "OTWARTE", "—", TODAY, "czat E"),
        ("E1-Q11", "Mapa Ziemia", "OTWARTE", "—", TODAY, "czat E"),
        ("E1-Q12", "Liczba rywali", "OTWARTE", "—", TODAY, "czat E"),
        ("C3", "Oblężenie panel", "OTWARTE", "—", TODAY, "czat C"),
        ("B2-Q7", "Model szczęścia", "OTWARTE", "—", TODAY, "czat B"),
    ],
}


def write_sheet(wb, name, rows):
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(2, ws.max_row)
        start = 2
    else:
        ws = wb.create_sheet(name)
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="C5D9F1")
        start = 2
    for i, row in enumerate(rows):
        r = start + i
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)


def main():
    if not os.path.isfile(XLSX):
        print(f"Brak: {XLSX}")
        sys.exit(1)
    bak = XLSX + f".bak-audit-{TODAY}"
    shutil.copy2(XLSX, bak)
    print(f"Backup: {bak}")
    wb = openpyxl.load_workbook(XLSX)
    meta = wb["Meta"] if "Meta" in wb.sheetnames else None
    for sheet, rows in SHEETS.items():
        write_sheet(wb, sheet, rows)
    if meta is None:
        meta = wb.create_sheet("Meta", 0)
    meta["A1"] = "Ostatnia aktualizacja Master Silnik"
    meta["B1"] = TODAY
    meta["A2"] = "Kanon opublikowany md5"
    meta["B2"] = "2276ec0f (stary; ROBOCZA nie opublikowana)"
    meta["A3"] = "Workflow"
    meta["B3"] = "Grupa F=main.ts+ROBOCZA; Master=finalna po Opus"
    wb.save(XLSX)
    print("Zaktualizowano arkusze:", ", ".join(SHEETS.keys()))


if __name__ == "__main__":
    main()
