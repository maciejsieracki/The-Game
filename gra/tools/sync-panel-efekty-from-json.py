#!/usr/bin/env python3
"""
sync-panel-efekty-from-json.py — odswieza arkusz Bonusy w Panel-efekty-cyw-dyplomacja.xlsx z civs.json.

Kierunek: JSON -> Excel (do review Macieja). NIE nadpisuje recznych komentarzy poza arkuszem bonusow.
Uruchom gdy civs.json jest nowszy niz panel (np. po edycji recznej JSON).

Uzycie (z gra/):
  python3 tools/sync-panel-efekty-from-json.py
  python3 tools/sync-panel-efekty-from-json.py --dry-run

Wymaga: zamknietego Excela (brak lock .~lock).
"""
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Wymagany openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data", "civs.json"))
XLSX = os.path.normpath(os.path.join(
    HERE, "..", "..", "Civ-CYWILIZACJE", "Panel-efekty-cyw-dyplomacja.xlsx"))
SHEET = "Bonusy cywilizacji"

REALIZUJE_COLOR = {
    "walka": "FFCCCC",
    "miasto": "CCFFCC",
    "ekonomia": "FFFFCC",
    "mapa": "CCE5FF",
}
H = Font(bold=True, color="FFFFFF")
HF = PatternFill("solid", fgColor="4A5568")
B = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def main():
    dry = "--dry-run" in sys.argv
    civs = json.load(open(DATA, encoding="utf-8"))["cywilizacje"]
    if not os.path.exists(XLSX):
        sys.exit(f"Brak panelu: {XLSX}")
    if dry:
        n = sum(len(c.get("bonusy") or []) for c in civs)
        print(f"[dry-run] zapisalby {n} wierszy do {SHEET}")
        return
    wb = openpyxl.load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 1)
    headers = [
        "Nacja", "typCywilizacji", "Typ efektu", "Cel",
        "Wartosc", "Opis", "Realizuje (dzial)",
    ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.font = H
        c.fill = HF
        c.border = B
        c.alignment = WRAP
    row = 2
    for civ in civs:
        nacja = civ.get("Cywilizacja", "")
        tc = civ.get("typCywilizacji", civ.get("ikonaId", ""))
        for b in civ.get("bonusy") or []:
            vals = [
                nacja, tc,
                b.get("typ", ""), b.get("cel", ""),
                b.get("wartosc", ""), b.get("opis", ""),
                b.get("realizuje", ""),
            ]
            fill_hex = REALIZUJE_COLOR.get(str(b.get("realizuje", "")), "FFFFFF")
            fill = PatternFill("solid", fgColor=fill_hex)
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row, j, v)
                cell.border = B
                cell.alignment = WRAP
                if j == 5:
                    cell.font = Font(color="0000FF")
                cell.fill = fill
            row += 1
    wb.save(XLSX)
    print(f"[sync-panel-efekty] zapisano {row - 2} wierszy -> {XLSX}")


if __name__ == "__main__":
    main()
