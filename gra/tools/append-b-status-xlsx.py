#!/usr/bin/env python3
"""Aktualizacja arkusza Grupa-B w Status-projektu-The-Game.xlsx (audyt 2026-06-27)."""
import datetime
import os
import shutil
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Brak openpyxl: pip install openpyxl")
    sys.exit(1)

CIV_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(CIV_ROOT, "Status-projektu-The-Game.xlsx")
TODAY = "2026-06-27"
HEADERS = ("ID", "Temat", "Status", "Decyzja", "Data", "Uwagi")

ROWS = [
    ("B2-Q1", "Mieszkańcy 3 koszyki", "ZAMKNIĘTE", "A", "2026-06-26", "cityPanel + happinessBreakdown"),
    ("B2-Q2", "Porządek sekcja inline", "ZAMKNIĘTE", "B", "2026-06-26", "orderPanel buildOrderSectionHtml"),
    ("B2-Q3", "Zdrowie +/-", "ZAMKNIĘTE", "A", "2026-06-26", "computeCityHealthBreakdown"),
    ("B2-Q4", "Specjaliści panel", "ZAMKNIĘTE", "C", "2026-06-26", "usunięte v1.0"),
    ("B2-Q5", "Alert buntu", "ZAMKNIĘTE", "C", TODAY, "chip OK; hex CZEKA MAPA"),
    ("B2-Q6", "Kary buntu + migracja", "ZAMKNIĘTE", "C", TODAY, "F-B2-porzadek wpięte"),
    ("B2-Q7", "Model Szczęścia", "OTWARTE ABC", "—", TODAY, "propozycja D: model %"),
    ("B2-Q8", "Czynniki Szczęścia v1.0", "OTWARTE ABC", "—", TODAY, "Spec/JSON już istnieje"),
    ("B2-Q9", "Prawo w Porządku", "OTWARTE ABC", "—", TODAY, ""),
    ("B1.2", "Rush produkcji", "OTWARTE ABC", "—", TODAY, ""),
    ("B1.3", "Auto-zarządca UI", "OTWARTE ABC", "—", TODAY, ""),
    ("B1.4", "Worked tiles", "OTWARTE ABC", "—", TODAY, "okolica 👤 — priorytet"),
    ("B4.1", "Kultura w panelu", "OTWARTE ABC", "—", TODAY, "hak nie wpięty"),
    ("B4.2", "Religia w panelu", "OTWARTE ABC", "—", TODAY, ""),
    ("B5.1", "Suwak split żywności", "OTWARTE ABC", "—", TODAY, ""),
    ("B5.2", "Default split 70/30", "OTWARTE ABC", "—", TODAY, ""),
    ("B5-tick", "advanceEmpireFood", "TODO", "Q1 hybryda", TODAY, "stub empire-food.ts"),
    ("B1-ulepszenia", "terrain→tileYield", "TODO", "A4=A", TODAY, "luka P0 Civ5"),
    ("B-okolica-haki", "getWorkedTiles w main", "TODO", "—", TODAY, "panel ≠ ekonomia tury"),
    ("B3", "Suwaki 70/20/10", "ZAMKNIĘTE", "1A/3A", "2026-06-26", ""),
    ("B4-wealth", "Wealth D3=A", "ZAMKNIĘTE", "A", "2026-06-26", "wealth.ts 25/25"),
]


def main():
    if not os.path.isfile(XLSX):
        print(f"Brak pliku: {XLSX}")
        sys.exit(1)
    bak = XLSX + ".bak-grupa-b-" + TODAY
    shutil.copy2(XLSX, bak)
    wb = openpyxl.load_workbook(XLSX)
    name = "Grupa-B"
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet(name)
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="E2EFDA")
    for i, row in enumerate(ROWS):
        r = 2 + i
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)
    wb.save(XLSX)
    print(f"OK: {name} — {len(ROWS)} wierszy · backup: {bak}")


if __name__ == "__main__":
    main()
