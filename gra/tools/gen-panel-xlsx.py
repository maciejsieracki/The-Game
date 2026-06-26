#!/usr/bin/env python3
# gen-panel-xlsx.py -- buduje JEDEN edytowalny Excel z konsolidacja wszystkich data/*.json.
# To samo co Panel-przeglad-danych.html, ale EDYTOWALNE (liczby + kolumna "Komentarz Naster").
# Snapshot z biezacych JSON-ow. Uruchom:  python3 gra/tools/gen-panel-xlsx.py
import openpyxl, json, os, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(HERE, "..", "data"))
CIV_ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
OUT = os.path.join(CIV_ROOT, "MIASTO", "Panel-przeglad-danych.xlsx")

META = {
    "buildings.json":        ("Budynki",        "Civ-MIASTO",     "Budynki.xlsx",            True),
    "society-params.json":   ("Spoleczenstwo",  "Civ-MIASTO",     "Spoleczenstwo-parametry.xlsx", True),
    "miasto-params.json":    ("Miasto-parametry","Civ-MIASTO",    "miasto-params.json",      True),
    "terrain-improvements.json": ("Ulepszenia-terenu","Civ-MIASTO",  "terrain-improvements.json", True),
    "econ-params.json":      ("Ekonomia",       "Civ-EKONOMIA",   "Ekonomia-parametry.xlsx", False),
    "resources.json":        ("Surowce",        "Civ-EKONOMIA",   "Surowce.xlsx",            False),
    "terrain-yields.json":   ("Teren - plony",  "Civ-EKONOMIA",   "Plony-terenow.xlsx",      False),
    "terrain-movement.json": ("Teren - ruch",   "Civ-MAPA",       "Plony-terenow.xlsx",      False),
    "terrain-combat.json":   ("Teren - walka",  "Civ-MAPA",       "Plony-terenow.xlsx",      False),
    "units.json":            ("Jednostki",      "Civ-UNITS",      "Jednostki.xlsx",          False),
    "counters.json":         ("Countery walki", "Civ-UNITS",      "Macierz-walki.xlsx",      False),
    "diplomacy.json":        ("Dyplomacja",     "Civ-DYPLOMACJA", "Dyplomacja.xlsx",         False),
    "ai-params.json":        ("Parametry AI",   "Civ-AI",         "AI-parametry.xlsx",       False),
    "civs.json":             ("Cywilizacje",    "Civ-DANE",       "Cywilizacje.xlsx",        False),
    "tech.json":             ("Technologie",    "Civ-DANE",       "Technologie-drzewko.xlsx",False),
}
ORDER = ["buildings.json","society-params.json","miasto-params.json","terrain-improvements.json","econ-params.json","resources.json",
         "terrain-yields.json","terrain-movement.json","terrain-combat.json",
         "units.json","counters.json","diplomacy.json","ai-params.json","civs.json","tech.json"]

FONT = "Arial"
H_FILL      = PatternFill("solid", fgColor="2E3440")
H_FILL_MINE = PatternFill("solid", fgColor="1F6F63")
NOTE_FILL   = PatternFill("solid", fgColor="FFF6CC")
SECT_FILL   = PatternFill("solid", fgColor="EAF3F1")
H_FONT   = Font(name=FONT, bold=True, color="FFFFFF", size=11)
SECT_FONT= Font(name=FONT, bold=True, color="1F6F63", size=12)
KEY_FONT = Font(name=FONT, bold=True, color="1A1A1A")
NUM_FONT = Font(name=FONT, color="0000FF")   # niebieski = edytowalny input (konwencja)
TXT_FONT = Font(name=FONT, color="1A1A1A")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP  = Alignment(vertical="top")

def is_scalar(v): return not isinstance(v, (dict, list))
def is_uniform_map(d):
    vals = list(d.values())
    return len(vals) > 0 and all(isinstance(v, dict) for v in vals) and \
        all(all(is_scalar(x) for x in v.values()) for v in vals)
def norm(v):
    if v is None: return ""
    if isinstance(v, bool): return "TAK" if v else "NIE"
    if isinstance(v, list): return ", ".join(map(str, v)) if all(is_scalar(x) for x in v) else json.dumps(v, ensure_ascii=False)
    if isinstance(v, dict): return json.dumps(v, ensure_ascii=False)
    return v
def flatten(d):
    out = {}
    for k, v in d.items():
        if isinstance(v, dict):
            for k2, v2 in v.items(): out[f"{k}.{k2}"] = v2
        else:
            out[k] = v
    return out

def style_header(c, mine):
    c.font = H_FONT; c.fill = (H_FILL_MINE if mine else H_FILL); c.border = BORDER; c.alignment = WRAP
def style_value(c, raw):
    c.border = BORDER
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        c.font = NUM_FONT; c.alignment = TOP
    else:
        c.font = TXT_FONT; c.alignment = WRAP

def write_table(ws, r0, rows, mine):
    flat = [flatten(x) if isinstance(x, dict) else {"wartosc": x} for x in rows]
    cols, seen = [], set()
    for fr in flat:
        for k in fr:
            if k not in seen: seen.add(k); cols.append(k)
    disp = cols + ["Komentarz Naster"]
    for j, c in enumerate(disp, 1):
        style_header(ws.cell(r0, j, c), mine)
    for i, fr in enumerate(flat):
        for j, c in enumerate(cols, 1):
            raw = fr.get(c)
            cell = ws.cell(r0 + 1 + i, j, norm(raw)); style_value(cell, raw)
        nc = ws.cell(r0 + 1 + i, len(cols) + 1, ""); nc.fill = NOTE_FILL; nc.border = BORDER
    return r0 + 1 + len(flat), len(disp)

def write_map(ws, r0, d, mine):
    subcols, seen = [], set()
    for v in d.values():
        for k in v:
            if k not in seen: seen.add(k); subcols.append(k)
    disp = ["klucz"] + subcols + ["Komentarz Naster"]
    for j, c in enumerate(disp, 1):
        style_header(ws.cell(r0, j, c), mine)
    for i, (rk, rv) in enumerate(d.items()):
        kc = ws.cell(r0 + 1 + i, 1, rk); kc.font = KEY_FONT; kc.border = BORDER; kc.alignment = TOP
        for j, c in enumerate(subcols, 2):
            raw = rv.get(c)
            cell = ws.cell(r0 + 1 + i, j, norm(raw)); style_value(cell, raw)
        nc = ws.cell(r0 + 1 + i, len(subcols) + 2, ""); nc.fill = NOTE_FILL; nc.border = BORDER
    return r0 + 1 + len(d), len(disp)

def write_kv(ws, r0, d, mine):
    for j, c in enumerate(["Parametr", "Wartosc", "Komentarz Naster"], 1):
        style_header(ws.cell(r0, j, c), mine)
    i = 0
    for k, v in d.items():
        kc = ws.cell(r0 + 1 + i, 1, k); kc.font = KEY_FONT; kc.border = BORDER; kc.alignment = WRAP
        vc = ws.cell(r0 + 1 + i, 2, norm(v)); style_value(vc, v)
        nc = ws.cell(r0 + 1 + i, 3, ""); nc.fill = NOTE_FILL; nc.border = BORDER
        i += 1
    return r0 + 1 + i, 3

def render_block(ws, r, value, mine):
    if isinstance(value, list):
        if value and all(isinstance(x, dict) for x in value):
            nr, _ = write_table(ws, r, value, mine); return nr
        for i, x in enumerate(value):
            ws.cell(r + i, 1, norm(x)).font = TXT_FONT
        return r + max(1, len(value))
    if isinstance(value, dict):
        if all(is_scalar(v) for v in value.values()):
            nr, _ = write_kv(ws, r, value, mine); return nr
        if is_uniform_map(value):
            nr, _ = write_map(ws, r, value, mine); return nr
        for k, v in value.items():
            tc = ws.cell(r, 1, "  " + str(k)); tc.font = SECT_FONT; tc.fill = SECT_FILL
            r = render_block(ws, r + 1, v, mine) + 1
        return r
    ws.cell(r, 1, norm(value)).font = TXT_FONT
    return r + 1

def is_single_table(value):
    if isinstance(value, list): return bool(value) and all(isinstance(x, dict) for x in value)
    if isinstance(value, dict): return all(is_scalar(v) for v in value.values()) or is_uniform_map(value)
    return False

def flatten_improvements(d):
    out = {}
    for k, v in d.items():
        if k == "_meta" or not isinstance(v, dict): continue
        flat = {}
        for kk, vv in v.items():
            if kk == "bonus" and isinstance(vv, dict):
                for bk, bv in vv.items(): flat["bonus." + bk] = bv
            elif isinstance(vv, (dict, list)): flat[kk] = json.dumps(vv, ensure_ascii=False)
            else: flat[kk] = vv
        out[k] = flat
    return out

def autosize(ws):
    from openpyxl.utils import get_column_letter
    for col in range(1, ws.max_column + 1):
        w = 10
        for row in range(1, min(ws.max_row, 400) + 1):
            v = ws.cell(row, col).value
            if v is not None:
                w = max(w, min(len(str(v)) + 2, 60))
        ws.column_dimensions[get_column_letter(col)].width = w

# ---- load data ----
data_map, meta_map = {}, {}
for fn in ORDER:
    p = os.path.join(DATA_DIR, fn)
    if not os.path.exists(p): continue
    data_map[fn] = json.load(open(p, encoding="utf-8"))
    meta_map[fn] = META.get(fn, (fn, "?", "?", False))

if "terrain-improvements.json" in data_map:
    data_map["terrain-improvements.json"] = flatten_improvements(data_map["terrain-improvements.json"])

wb = openpyxl.Workbook()
# ---- Index sheet ----
idx = wb.active; idx.title = "Index"
idx["A1"] = "Panel przegladu danych - The Game (Civ)"; idx["A1"].font = Font(name=FONT, bold=True, size=14, color="1F6F63")
idx["A2"] = f"Edytowalny snapshot wszystkich data/*.json - wygenerowano {datetime.datetime.now():%Y-%m-%d %H:%M}"
idx["A2"].font = Font(name=FONT, italic=True, color="666666")
idx["A3"] = "Liczby (niebieskie) mozesz edytowac. Kolumna 'Komentarz Naster' (zolta) = Twoje uwagi. Zakladki MIASTO maja kolor morski."
idx["A3"].font = Font(name=FONT, color="666666")
hdr = ["Zakladka", "JSON", "Wlasciciel", "Zrodlowy Excel", "Pozycji"]
for j, c in enumerate(hdr, 1): style_header(idx.cell(5, j, c), False)
labels = {}
for i, fn in enumerate(data_map):
    label, owner, src, mine = meta_map[fn]
    title = label[:31]
    base, n = title, 2
    while title in labels.values(): title = f"{base[:28]}_{n}"; n += 1
    labels[fn] = title
    cnt = len(data_map[fn]) if isinstance(data_map[fn], (list, dict)) else 1
    row = [title, fn, owner, src, cnt]
    for j, v in enumerate(row, 1):
        cell = idx.cell(6 + i, j, v); cell.border = BORDER; cell.alignment = TOP
        cell.font = Font(name=FONT, bold=(j == 1), color=("1F6F63" if mine and j == 1 else "1A1A1A"))
idx.freeze_panes = "A6"
autosize(idx)

# ---- data sheets ----
for fn in data_map:
    label, owner, src, mine = meta_map[fn]
    ws = wb.create_sheet(labels[fn])
    if mine: ws.sheet_properties.tabColor = "1F6F63"
    value = data_map[fn]
    last = render_block(ws, 1, value, mine)
    if is_single_table(value):
        ws.freeze_panes = "A2"
        try: ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
        except Exception: pass
    autosize(ws)

wb.save(OUT)
print("OK ->", OUT)
print("zakladki:", len(data_map) + 1, "| rozmiar:", round(os.path.getsize(OUT) / 1024, 1), "KB")
