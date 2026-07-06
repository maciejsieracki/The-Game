#!/usr/bin/env python3
"""
build-tech-excel-mirror.py — generuje Technologie-drzewko.xlsx (wzór lustracyjny).

Wypełnia arkusz Technologie danymi z tech.json + budynki + jednostki + ulepszenia.
Arkusze pomocnicze: słownik, tempo, lookupy, bonusy cyw (szablon).

Użycie (z root projektu Civ):
  python gra/tools/build-tech-excel-mirror.py
  python gra/tools/build-tech-excel-mirror.py --out Technologie-drzewko.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import re
import unicodedata
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    raise SystemExit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
DATA = os.path.join(HERE, "..", "data")
DEF_OUT = os.path.join(ROOT, "Technologie-drzewko.xlsx")

TECH_ALIASES = {
    "Irygacja": "Gospodarka wodna",
    "Kalendarz": "Matematyka",
}
MULTI_TECH = {
    "posterunek": ["Obróbka drewna", "Murarstwo"],
}

# Kolumny eksportowane przez export-tech.py (kolejność = tech.json)
EXPORT_FIELDS = [
    "Technologia",
    "Epoka",
    "Poziom",
    "Dostęp do surowca.",
    "wymagany budynek",
    "Wymaga (prereq)",
    "Odblokowuje surowiec.",
    "Odblokowuje budynek",
    "Koszt nauki",
    "Uwagi",
]

# Kolumny rozszerzone (lustracja; część już w tech.json)
MIRROR_EXTRA = [
    "Odblokowuje ulepszenie terenu",
    "Odblokowuje jednostki",
    "Odblokowuje cud (lustracja)",
    "Slug (drzewko UI)",
    "Koniec epoki gracza (TAK/NIE)",
    "Odblokowuje epokę (1=Kamień 2=Brąz 3=Żelazo)",
    "Bramka specjalna (AND tech)",
    "Status (kanon/draft/placeholder)",
    "Ostatnia zmiana",
]

ALL_HEADERS = EXPORT_FIELDS + MIRROR_EXTRA


def slugify(name: str) -> str:
    s = unicodedata.normalize("NFD", name).lower()
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def load_json(path: str):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def norm_tech(raw: str | None) -> str | None:
    if raw is None or raw in ("-", "—", ""):
        return None
    return TECH_ALIASES.get(raw, raw)


def build_unlock_maps():
    bld = load_json(os.path.join(DATA, "buildings.json"))
    units = load_json(os.path.join(DATA, "units.json"))
    terr = load_json(os.path.join(DATA, "terrain-improvements.json"))

    by_bld: dict[str, list[str]] = {}
    for b in bld:
        t = (b.get("techUnlock") or "").strip()
        if not t:
            continue
        by_bld.setdefault(t, []).append(b.get("nazwa", b.get("id", "?")))

    by_unit: dict[str, list[str]] = {}
    for u in units:
        t = (u.get("Tech") or "").replace("—", "-").strip()
        if not t or t == "-":
            continue
        by_unit.setdefault(t, []).append(u.get("Jednostka", "?"))

    by_imp: dict[str, list[str]] = {}
    for key, row in terr.items():
        if key.startswith("_"):
            continue
        multi = MULTI_TECH.get(key)
        label = row.get("nazwa") or key
        if multi:
            for t in multi:
                by_imp.setdefault(t, []).append(label)
            continue
        t = norm_tech(row.get("tech"))
        if t:
            by_imp.setdefault(t, []).append(label)

    return by_bld, by_unit, by_imp


def build_wonder_unlock_map() -> dict[str, list[str]]:
    """Tech → lista etykiet cudów (z wonders.json; źródło prawdy techUnlock = cuda)."""
    wonders = load_json(os.path.join(DATA, "wonders.json"))
    by_wonder: dict[str, list[str]] = {}
    for w in wonders.get("cuda") or []:
        label = f"{w.get('nazwa', w.get('id', '?'))} ({w.get('dostep', '?')}, ep.{w.get('epokaWejscia', '?')})"
        for t in w.get("techUnlock") or []:
            by_wonder.setdefault(t, []).append(label)
    for w in wonders.get("parkowane_epoka4plus") or []:
        label = f"{w.get('nazwa', w.get('id', '?'))} [park.] ({w.get('dostep', '?')})"
        for t in w.get("techUnlock") or []:
            by_wonder.setdefault(t, []).append(label)
    return by_wonder


def civ_entry_label(ikona_id: str, civs: list) -> str:
    for c in civs:
        if c.get("ikonaId") == ikona_id:
            return c.get("epokaWejscia") or "kamien"
    return "?"


def end_era_flags(tech_name: str, uwagi: str | None) -> tuple[str, str]:
    u = (uwagi or "").lower()
    if "kończy epok" in u or "konczy epok" in u:
        if "epokę 3" in u or "epoke 3" in u or "żelazo" in u and "cap" in u:
            return "TAK", "3"
        if "epokę 2" in u or "epoke 2" in u or "brąz" in u:
            return "TAK", "2"
        if "epokę 1" in u or "kamień" in u:
            return "TAK", "1"
    if tech_name == "Brązownictwo":
        return "TAK", "2"
    if tech_name == "Waluta":
        return "TAK", "2"
    if tech_name == "Sztuka wojenna":
        return "TAK", "3"
    return "NIE", ""


def style_header(ws, row: int, ncol: int):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border


def autofit_columns(ws, min_width=10, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(min_width, max_len + 2)


def sheet_readme(wb):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("Technologie-drzewko.xlsx — wzór lustracyjny (The Game)", ""),
        ("", ""),
        ("Cel", "Jedno miejsce do edycji drzewka tech: prereq, koszty, odblokowania, uwagi."),
        ("", ""),
        ("Arkusze", ""),
        ("Technologie", "Główna tabela — 1 wiersz = 1 technologia. Eksport → gra/data/tech.json (export-tech.py)."),
        ("Tempo_gry", "Mnożniki tempa badań (szybka/standard/długa)."),
        ("Slownik_kolumn", "Opis każdej kolumny arkusza Technologie."),
        ("Lookup_budynki", "Lista budynków + techUnlock (referencja)."),
        ("Lookup_jednostki", "Jednostki + pole Tech (referencja)."),
        ("Lookup_ulepszenia", "Ulepszenia terenu + tech (referencja)."),
        ("Cuda_Antyk", "19 aktywnych + parkowane cuda — techUnlock, E/R, epoka (źródło: wonders.json)."),
        ("Bonusy_cyw_tech", "Szablon bonusów per cywilizacja powiązanych z tech (opcjonalnie)."),
        ("", ""),
        ("Eksport do gry", "python gra/tools/export-tech.py --dry-run  →  potem bez --dry-run"),
        ("Cuda", "Edycja w wonders.json / Panel — NIE przez export-tech.py."),
        ("Uwaga", "Kolumny po 'Uwagi' (Odblokowuje ulepszenie…) = lustracja; sync ręczny lub przyszły export."),
        ("Ostatnia generacja", str(date.today())),
    ]
    for r, (a, b) in enumerate(lines, 1):
        ws.cell(r, 1, a)
        ws.cell(r, 2, b)
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90


def sheet_slownik(wb):
    ws = wb.create_sheet("Slownik_kolumn")
    rows = [
        ("Technologia", "TAK", "Unikalna nazwa (id w grze). Musi być 1:1 z tech.json i bramkami ulepszeń."),
        ("Epoka", "TAK", "Kamień | Brąz | Żelazo — etykieta epoki w drzewku UI."),
        ("Poziom", "TAK", "Kolumna/poziom w drzewku (1–8). Wpływa na układ sciencePicker."),
        ("Dostęp do surowca.", "NIE", "Opis dostępu do surowca (boolean v0.1) — tekst dla gracza/designera."),
        ("wymagany budynek", "NIE", "Hint UI: budynek w imperium przed badaniem (twarda bramka opcjonalna)."),
        ("Wymaga (prereq)", "TAK", "— = brak. Pojedyncza tech lub A + B (wszystkie wymagane)."),
        ("Odblokowuje surowiec.", "TAK", "Lista surowców ASCII (drewno, ruda, bydlo…)."),
        ("Odblokowuje budynek", "TAK", "Budynki miasta + linia Jednostki: … jeśli brak budynku."),
        ("Koszt nauki", "TAK", "PN bazowe; tempo mnoży applyTempoKoszt() — patrz Tempo_gry."),
        ("Uwagi", "TAK", "Notatki design / koniec epoki / placeholder."),
        ("Odblokowuje ulepszenie terenu", "Lustracja", "Farma, Tartak, Bydło… — sync z terrain-improvements.json."),
        ("Odblokowuje jednostki", "Lustracja", "Lista jednostek z units.json (pole Tech)."),
        ("Odblokowuje cud (lustracja)", "Lustracja", "Cuda z wonders.json odblokowywane po tech (AND). Źródło prawdy: wonders.json, nie tech.json."),
        ("Slug (drzewko UI)", "Auto", "Generowany z nazwy — nie zmieniaj bez potrzeby."),
        ("Koniec epoki gracza", "Lustracja", "TAK gdy tech zamyka epokę (np. Brązownictwo, Waluta)."),
        ("Odblokowuje epokę", "Lustracja", "1/2/3 — numer epoki odblokowanej po zbadaniu."),
        ("Bramka specjalna (AND tech)", "Lustracja", "Np. Posterunek = Obróbka drewna + Murarstwo (kod, nie nowa tech)."),
        ("Status", "Edycja", "kanon | draft | placeholder — co jest w grze vs plan."),
        ("Ostatnia zmiana", "Edycja", "Data Twojej edycji wiersza."),
    ]
    ws.append(["Kolumna", "Eksport?", "Opis"])
    style_header(ws, 1, 3)
    for row in rows:
        ws.append(list(row))
    autofit_columns(ws)


def sheet_tempo(wb, tempo: dict):
    ws = wb.create_sheet("Tempo_gry")
    ws.append(["Parametr", "Wartość", "Opis"])
    style_header(ws, 1, 3)
    for key in ("szybka", "standardowa", "dluga"):
        ws.append([key, tempo.get(key), "Mnożnik kosztu badań"])
    ws.append(["_opis", tempo.get("_opis", ""), ""])
    ws.append([])
    ws.append(["Przykład", "Koszt bazowy 100", ""])
    ws.append(["szybka", round(100 * float(tempo.get("szybka", 0.2))), ""])
    ws.append(["standardowa", 100, ""])
    ws.append(["dluga", round(100 * float(tempo.get("dluga", 5))), ""])
    autofit_columns(ws)


def sheet_technologie(wb, tech_data: dict, by_bld, by_unit, by_imp, by_wonder):
    ws = wb.create_sheet("Technologie")
    techs = tech_data.get("technologie", [])

    # Wiersz instrukcji
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ALL_HEADERS))
    ws.cell(1, 1, "Drzewko technologii — edytuj od wiersza 3. Eksport: gra/tools/export-tech.py")
    ws.cell(1, 1).font = Font(italic=True, color="444444")

    for c, h in enumerate(ALL_HEADERS, 1):
        ws.cell(2, c, h)
    style_header(ws, 2, len(ALL_HEADERS))

    today = str(date.today())
    for r, t in enumerate(techs, 3):
        name = t.get("Technologia", "")
        uw = t.get("Uwagi")
        koniec, ep = end_era_flags(name, uw)

        bld_list = by_bld.get(name, [])
        json_bld = (t.get("Odblokowuje budynek") or "").strip()
        if json_bld and json_bld not in ("null", "None"):
            odbl_bud = json_bld
        elif bld_list:
            odbl_bud = "; ".join(bld_list)
        else:
            odbl_bud = t.get("Odblokowuje budynek")

        units = by_unit.get(name, [])
        odbl_units = ", ".join(units) if units else ""
        if odbl_units and odbl_bud and "Jednostki:" not in str(odbl_bud):
            odbl_bud = f"{odbl_bud}; Jednostki: {odbl_units}" if odbl_bud else f"Jednostki: {odbl_units}"

        imps = by_imp.get(name, [])
        odbl_imp = t.get("Odblokowuje ulepszenie terenu") or (", ".join(dict.fromkeys(imps)) if imps else "")

        wonders = by_wonder.get(name, [])
        odbl_wonder = "; ".join(dict.fromkeys(wonders)) if wonders else ""

        bramka = ""
        if name in ("Obróbka drewna", "Murarstwo"):
            bramka = "Posterunek wymaga OBIE (kod AND)"

        row_vals = {
            "Technologia": name,
            "Epoka": t.get("Epoka"),
            "Poziom": t.get("Poziom"),
            "Dostęp do surowca.": t.get("Dostęp do surowca."),
            "wymagany budynek": t.get("wymagany budynek"),
            "Wymaga (prereq)": t.get("Wymaga (prereq)"),
            "Odblokowuje surowiec.": t.get("Odblokowuje surowiec."),
            "Odblokowuje budynek": odbl_bud,
            "Koszt nauki": t.get("Koszt nauki"),
            "Uwagi": uw,
            "Odblokowuje ulepszenie terenu": odbl_imp,
            "Odblokowuje jednostki": odbl_units,
            "Odblokowuje cud (lustracja)": odbl_wonder,
            "Slug (drzewko UI)": slugify(name),
            "Koniec epoki gracza (TAK/NIE)": koniec,
            "Odblokowuje epokę (1=Kamień 2=Brąz 3=Żelazo)": ep,
            "Bramka specjalna (AND tech)": bramka,
            "Status (kanon/draft/placeholder)": "kanon",
            "Ostatnia zmiana": today,
        }
        for c, h in enumerate(ALL_HEADERS, 1):
            ws.cell(r, c, row_vals.get(h))

    # Walidacja listy Status
    dv = DataValidation(type="list", formula1='"kanon,draft,placeholder"', allow_blank=True)
    status_col = ALL_HEADERS.index("Status (kanon/draft/placeholder)") + 1
    dv.add(f"{get_column_letter(status_col)}3:{get_column_letter(status_col)}200")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A3"
    autofit_columns(ws)


def sheet_lookup_budynki(wb):
    ws = wb.create_sheet("Lookup_budynki")
    ws.append(["id", "nazwa", "techUnlock", "epokaWejscia", "kategoria", "uwagi"])
    style_header(ws, 1, 6)
    for b in load_json(os.path.join(DATA, "buildings.json")):
        ws.append([
            b.get("id"), b.get("nazwa"), b.get("techUnlock"),
            b.get("epokaWejscia"), b.get("kategoria"), b.get("uwagi"),
        ])
    autofit_columns(ws)


def sheet_lookup_jednostki(wb):
    ws = wb.create_sheet("Lookup_jednostki")
    ws.append(["Jednostka", "Epoka", "Tech", "Typ", "Uwagi"])
    style_header(ws, 1, 5)
    for u in load_json(os.path.join(DATA, "units.json")):
        ws.append([
            u.get("Jednostka"), u.get("Epoka"), u.get("Tech"),
            u.get("Typ"), (u.get("Uwagi") or "")[:120],
        ])
    autofit_columns(ws)


def sheet_lookup_ulepszenia(wb):
    ws = wb.create_sheet("Lookup_ulepszenia")
    ws.append(["klucz JSON", "nazwa", "tech (JSON)", "tech kanoniczna", "epoka", "koszt_praca", "surowiecOdblokowany"])
    style_header(ws, 1, 7)
    terr = load_json(os.path.join(DATA, "terrain-improvements.json"))
    for key, row in sorted(terr.items()):
        if key.startswith("_"):
            continue
        raw = row.get("tech")
        multi = MULTI_TECH.get(key)
        canon = " + ".join(multi) if multi else (norm_tech(raw) or "(brak — wyrąb/taras/posterunek spec.)")
        ws.append([
            key, row.get("nazwa"), raw, canon,
            row.get("epoka"), row.get("koszt_praca"), row.get("surowiecOdblokowany"),
        ])
    autofit_columns(ws)


def sheet_cuda_antyk(wb):
    ws = wb.create_sheet("Cuda_Antyk")
    headers = [
        "id", "nazwa", "dostep", "epokaWejscia_cudu", "techUnlock (AND)",
        "państwo E (ikonaId)", "epokaWejscia_państwa", "kosztBudowy",
        "utrzymanie", "wymagaTerenu", "maxNaSwiecie", "status",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    wonders = load_json(os.path.join(DATA, "wonders.json"))
    civs = load_json(os.path.join(DATA, "civs.json")).get("cywilizacje", [])

    note = (
        "Kanon cuda Antyku v1.0 — źródło prawdy: gra/data/wonders.json. "
        "Start budowy (D-CUD4): wszystkie tech z kolumny techUnlock odkryte. "
        "Reguła E (D-CUD-TECH): tech >= epokaWejscia państwa. R: wszyscy 15 graczy."
    )
    ws.cell(2, 1, note)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)

    def append_wonder(w, status: str):
        civ_e = w.get("cywilizacje") or []
        if w.get("dostep") == "E":
            pa = civ_e[0] if civ_e else ""
            ep_pa = civ_entry_label(pa, civs) if pa else ""
        else:
            pa = "wszyscy (15)"
            ep_pa = "—"
        ws.append([
            w.get("id"),
            w.get("nazwa"),
            w.get("dostep"),
            w.get("epokaWejscia"),
            " + ".join(w.get("techUnlock") or []),
            pa,
            ep_pa,
            w.get("kosztBudowy"),
            w.get("utrzymanie"),
            ", ".join(w.get("wymagaTerenu") or []) or "—",
            w.get("maxNaSwiecie", 1),
            status,
        ])

    for w in wonders.get("cuda") or []:
        append_wonder(w, "aktywny")
    for w in wonders.get("parkowane_epoka4plus") or []:
        append_wonder(w, "parkowany")

    ws.freeze_panes = "A3"
    autofit_columns(ws)


def sheet_bonusy_cyw(wb):
    ws = wb.create_sheet("Bonusy_cyw_tech")
    ws.append([
        "Cywilizacja", "ikonaId", "Technologia (opcjonalnie)", "Typ bonusu",
        "Cel", "Wartość", "Opis", "Realizuje (walka/ekonomia/nauka)", "Status",
    ])
    style_header(ws, 1, 9)
    note = (
        "Szablon: wpisz bonusy cywilizacji powiązane z tech (np. szybsza nauka, tańsza jednostka). "
        "Puste = brak specjalnego bonusu per tech. Ogólne bonusy cyw → Cywilizacje.xlsx / civs.json."
    )
    ws.cell(2, 1, note)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)
    civs = load_json(os.path.join(DATA, "civs.json")).get("cywilizacje", [])
    r = 4
    for c in civs:
        for b in c.get("bonusy") or []:
            if b.get("typ") in ("bonus_nauka",) or "tech" in (b.get("opis") or "").lower():
                ws.append([
                    c.get("Cywilizacja"), c.get("ikonaId"), "",
                    b.get("typ"), b.get("cel"), b.get("wartosc"),
                    b.get("opis"), b.get("realizuje"), "przykład z civs.json",
                ])
                r += 1
    # Puste wiersze do wypełnienia
    for _ in range(15):
        ws.append(["", "", "", "", "", "", "", "", ""])
    autofit_columns(ws)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=DEF_OUT, help="Ścieżka wyjściowa xlsx")
    args = ap.parse_args()

    tech_path = os.path.join(DATA, "tech.json")
    tech_data = load_json(tech_path)
    by_bld, by_unit, by_imp = build_unlock_maps()
    by_wonder = build_wonder_unlock_map()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    sheet_slownik(wb)
    sheet_tempo(wb, tech_data.get("tempo_gry", {}))
    sheet_technologie(wb, tech_data, by_bld, by_unit, by_imp, by_wonder)
    sheet_lookup_budynki(wb)
    sheet_lookup_jednostki(wb)
    sheet_lookup_ulepszenia(wb)
    sheet_cuda_antyk(wb)
    sheet_bonusy_cyw(wb)

    wb.save(args.out)
    print(f"[build-tech-excel-mirror] zapisano: {args.out}")
    print(f"  technologie: {len(tech_data.get('technologie', []))}")
    print(f"  cuda aktywne: {len(load_json(os.path.join(DATA, 'wonders.json')).get('cuda', []))}")
    print(f"  arkusze: {wb.sheetnames}")


if __name__ == "__main__":
    main()
