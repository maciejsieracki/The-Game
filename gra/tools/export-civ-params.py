#!/usr/bin/env python3
"""DEPRECATED (2026-06-30 — PANEL-MERGE):
  Excel: docs/archiwum/panele-legacy/
  Kanon: panele-sterowania/Panel-{A..E}.xlsx + export-{a..e}.py
  Nie używać do nowych zmian balansu.


export-civ-params.py — TARGETED: Cywilizacje.xlsx[Parametry-cyw] -> gra/data/civ-params.json

  python tools/export-civ-params.py [--dry-run]
"""
import argparse
import json
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_XLSX = os.path.normpath(os.path.join(HERE, "..", "..", "docs", "archiwum", "panele-legacy", "Cywilizacje.xlsx"))
DEF_JSON = os.path.normpath(os.path.join(HERE, "..", "data", "civ-params.json"))
SHEET = "Parametry-cyw"


def num(v, default=1.0):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def str_val(v, default=""):
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Brak arkusza '{SHEET}' w {xlsx_path}")
    ws = wb[SHEET]
    rows = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or str(name).strip() == "":
            continue
        rows.append({
            "Cywilizacja": str(name).strip(),
            "preferowaneBudynki": str_val(ws.cell(r, 2).value),
            "preferowaneJednostki": str_val(ws.cell(r, 3).value),
            "modWzrostu": num(ws.cell(r, 4).value, 1.0),
            "modEkonomii": num(ws.cell(r, 5).value, 1.0),
            "uwagi": str_val(ws.cell(r, 6).value),
        })
    wb.close()
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEF_XLSX)
    ap.add_argument("--json", default=DEF_JSON)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    rows = read_rows(args.xlsx)
    print(f"[export-civ-params] nacje: {len(rows)}")
    for e in rows:
        print(f"  {e['Cywilizacja']}: budynki={e['preferowaneBudynki'][:40]}")

    out = {"cywilizacje": rows, "_meta": {"zrodlo": "Cywilizacje.xlsx/Parametry-cyw"}}

    if args.dry_run:
        print("[dry-run] bez zapisu")
        return

    bak = args.json + ".bak-CYWILIZACJE-export-params"
    if os.path.exists(args.json):
        shutil.copy2(args.json, bak)

    tmp = args.json + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")
    json.load(open(tmp, encoding="utf-8"))
    os.replace(tmp, args.json)
    print(f"[export-civ-params] GOTOWE: {args.json}")


if __name__ == "__main__":
    main()
