#!/usr/bin/env python3
# gen-bonusy-cyw-xlsx.py — Excel 9×3 bonusów z gra/data/civs.json (D4-Q1: review Macieja przed kodem).
# Edycja wartości w Excelu -> potem targeted re-export civs.json (ręcznie / osobny skrypt).
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("BRAK openpyxl — uruchom: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data", "civs.json"))
OUT = os.path.normpath(os.path.join(
    HERE, "..", "..", "docs", "archiwum", "panele-legacy",
    "Civ-CYWILIZACJE", "Bonusy-cywilizacji-9x3.xlsx"))

FONT = "Arial"
H = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="4A5568")
NF = Font(name=FONT, color="0000FF")
TF = Font(name=FONT, color="1A1A1A")
KF = Font(name=FONT, bold=True, color="1A1A1A")
NOTE = PatternFill("solid", fgColor="FFF6CC")
HDR2 = PatternFill("solid", fgColor="718096")
thin = Side(style="thin", color="D9D9D9")
B = Border(thin, thin, thin, thin)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top", wrap_text=False)

BONUS_SLOTS = 3
BONUS_FIELDS = ["typ", "cel", "wartosc", "opis", "realizuje"]


def load_civs():
    with open(DATA, encoding="utf-8") as f:
        return json.load(f)["cywilizacje"]


def style_header(ws, row, cols, fill=HF):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row, j, c)
        cell.font = H
        cell.fill = fill
        cell.border = B
        cell.alignment = WRAP


def autosize(ws, nrows, ncols, max_w=50):
    for col in range(1, ncols + 1):
        w = 10
        for row in range(1, nrows + 1):
            val = ws.cell(row, col).value
            if val is not None:
                w = max(w, min(len(str(val)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col)].width = w


def write_cell(ws, r, c, val, editable_num=False, note=False):
    cell = ws.cell(r, c, val if not isinstance(val, (dict, list)) else json.dumps(val, ensure_ascii=False))
    cell.border = B
    if note:
        cell.fill = NOTE
        cell.font = TF
        cell.alignment = WRAP
    elif editable_num and isinstance(val, (int, float)) and not isinstance(val, bool):
        cell.font = NF
        cell.alignment = TOP
    elif c == 1:
        cell.font = KF
        cell.alignment = TOP
    else:
        cell.font = TF
        cell.alignment = WRAP


def build_matrix_sheet(ws, civs):
    ws.title = "Bonusy 9x3"
    base_cols = [
        "Cywilizacja",
        "Religia",
        "Jednostka specjalna",
        "mnoznikHandelPieniadz",
    ]
    bonus_cols = []
    for i in range(1, BONUS_SLOTS + 1):
        for f in BONUS_FIELDS:
            bonus_cols.append(f"B{i}_{f}")
    cols = base_cols + bonus_cols + ["Komentarz Maciej"]
    style_header(ws, 1, cols)

    for i, civ in enumerate(civs):
        r = i + 2
        bonusy = civ.get("bonusy") or []
        while len(bonusy) < BONUS_SLOTS:
            bonusy.append({})

        row_vals = {
            "Cywilizacja": civ.get("Cywilizacja", ""),
            "Religia": civ.get("Religia", ""),
            "Jednostka specjalna": civ.get("Jednostka specjalna", ""),
            "mnoznikHandelPieniadz": civ.get("mnoznikHandelPieniadz", ""),
        }
        for bi in range(BONUS_SLOTS):
            b = bonusy[bi] if bi < len(bonusy) else {}
            for f in BONUS_FIELDS:
                row_vals[f"B{bi + 1}_{f}"] = b.get(f, "")

        for j, c in enumerate(cols, 1):
            if c == "Komentarz Maciej":
                write_cell(ws, r, j, "", note=True)
                continue
            val = row_vals.get(c, "")
            write_cell(ws, r, j, val, editable_num=(c.endswith("_wartosc") or c == "mnoznikHandelPieniadz"))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(civs) + 1}"
    autosize(ws, len(civs) + 1, len(cols))
    ws.sheet_properties.tabColor = "4A5568"


def build_flat_sheet(ws, civs):
    ws.title = "Lista bonusów"
    cols = ["Cywilizacja", "Nr", "typ", "cel", "wartosc", "opis", "realizuje", "Komentarz Maciej"]
    style_header(ws, 1, cols)
    r = 2
    for civ in civs:
        name = civ.get("Cywilizacja", "")
        bonusy = civ.get("bonusy") or []
        for idx, b in enumerate(bonusy, 1):
            for j, c in enumerate(cols, 1):
                if c == "Cywilizacja":
                    write_cell(ws, r, j, name)
                elif c == "Nr":
                    write_cell(ws, r, j, idx)
                elif c == "Komentarz Maciej":
                    write_cell(ws, r, j, "", note=True)
                else:
                    write_cell(ws, r, j, b.get(c, ""), editable_num=(c == "wartosc"))
            r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r - 1}"
    autosize(ws, r - 1, len(cols))
    ws.sheet_properties.tabColor = "718096"


def build_readme_sheet(ws):
    ws.title = "Jak edytować"
    lines = [
        "Bonusy cywilizacji — review Macieja (D4-Q1)",
        "",
        "1. Edytuj NIEBIESKIE komórki (wartosc, mnoznikHandelPieniadz).",
        "2. Żółta kolumna 'Komentarz Maciej' — uwagi, nie idą do gry.",
        "3. Arkusz 'Bonusy 9x3' = 9 nacji × 3 sloty bonusów.",
        "4. Arkusz 'Lista bonusów' = ten sam content, 1 wiersz = 1 bonus.",
        "5. Po akceptacji: targeted re-export do gra/data/civs.json (CYWILIZACJE).",
        "6. Dopiero POTEM wdrożenie w lane (ekonomia / walka / miasto).",
        "",
        f"Źródło: gra/data/civs.json",
        f"Generator: gra/tools/gen-bonusy-cyw-xlsx.py",
    ]
    for i, line in enumerate(lines, 1):
        cell = ws.cell(i, 1, line)
        cell.font = TF if i > 1 else Font(name=FONT, bold=True, size=12)
    ws.column_dimensions["A"].width = 72


def main():
    civs = load_civs()
    wb = openpyxl.Workbook()
    build_readme_sheet(wb.active)
    build_matrix_sheet(wb.create_sheet(), civs)
    build_flat_sheet(wb.create_sheet(), civs)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    n_bonus = sum(len(c.get("bonusy") or []) for c in civs)
    print("OK ->", OUT)
    print("| nacji:", len(civs), "| bonusów:", n_bonus, "| sloty:", BONUS_SLOTS, "per nacja")


if __name__ == "__main__":
    main()
