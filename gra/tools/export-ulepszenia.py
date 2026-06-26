#!/usr/bin/env python3
# export-ulepszenia.py -- BEZPIECZNY eksport MIASTO/Ulepszenia-terenu.xlsx -> data/terrain-improvements.json.
# Overlay na ORYGINALNY JSON (zachowuje strukture/typy/kolejnosc), puste komorki = brak zmiany, "Komentarz Naster" ignorowany.
import openpyxl, json, os

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data", "terrain-improvements.json"))
XLSX = os.path.normpath(os.path.join(HERE, "..", "..", "MIASTO", "Ulepszenia-terenu.xlsx"))

def coerce(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""): return old
    if isinstance(old, bool): return str(val).strip().lower() in ("tak", "true", "1")
    if isinstance(old, int):
        try: return int(round(float(val)))
        except Exception: return old
    if isinstance(old, float):
        try: return float(val)
        except Exception: return old
    if old is None: return val
    return str(val)

orig = json.load(open(DATA, encoding="utf-8"))
wb = openpyxl.load_workbook(XLSX)
ws = wb["Ulepszenia terenu"]
header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
changed = 0
for r in range(2, ws.max_row + 1):
    rid = ws.cell(r, 1).value
    if rid is None: continue
    rid = str(rid).strip()
    if rid not in orig or not isinstance(orig[rid], dict): continue
    node = orig[rid]
    for c in range(2, len(header) + 1):
        col = header[c - 1]
        if col in (None, "", "Komentarz Naster", "id"): continue
        val = ws.cell(r, c).value
        if isinstance(col, str) and col.startswith("bonus."):
            bk = col.split(".", 1)[1]
            bonus = node.setdefault("bonus", {})
            if bk in bonus:
                nv = coerce(bonus[bk], val)
                if nv != bonus[bk]: bonus[bk] = nv; changed += 1
        elif col in node:
            nv = coerce(node[col], val)
            if nv != node[col]: node[col] = nv; changed += 1

tmp = DATA + ".tmp"
with open(tmp, "w", encoding="utf-8") as f:
    json.dump(orig, f, ensure_ascii=False, indent=2); f.write("\n")
json.load(open(tmp, encoding="utf-8"))
os.replace(tmp, DATA)
print(f"terrain-improvements.json: zmienionych wartosci = {changed}")
