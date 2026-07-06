#!/usr/bin/env python3
"""DEPRECATED (2026-06-30 — PANEL-MERGE):
  Excel: docs/archiwum/panele-legacy/
  Kanon: panele-sterowania/Panel-{A..E}.xlsx + export-{a..e}.py
  Nie używać do nowych zmian balansu.


export-civ-dyplomacy-nations.py — TARGETED: Cywilizacje.xlsx[Dyplomacja] -> diplomacy.json[perNacja]

NIE dotyka params, akcje_dyplomatyczne ani innych bloków diplomacy.json.

  python tools/export-civ-dyplomacy-nations.py [--dry-run]
"""
import argparse
import json
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_XLSX = os.path.normpath(os.path.join(HERE, "..", "..", "docs", "archiwum", "panele-legacy", "Cywilizacje.xlsx"))
DEF_JSON = os.path.normpath(os.path.join(HERE, "..", "data", "diplomacy.json"))
SHEET = "Dyplomacja"


def num(v, default=0):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return default


def str_val(v, default=""):
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def read_per_nacja(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Brak arkusza '{SHEET}' w {xlsx_path}")
    ws = wb[SHEET]
    rows = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or str(name).strip() == "":
            continue
        rows.append({
            "Cywilizacja": str(name).strip(),
            "sklonnoscSojusze": num(ws.cell(r, 2).value),
            "lojalnosc": num(ws.cell(r, 3).value),
            "progWojny": num(ws.cell(r, 4).value),
            "pamietliwosc": num(ws.cell(r, 5).value),
            "otwartoscHandel": num(ws.cell(r, 6).value),
            "nastawienieBazowe": num(ws.cell(r, 7).value, 50),
            "uwagi": str_val(ws.cell(r, 8).value),
        })
    wb.close()
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEF_XLSX)
    ap.add_argument("--json", default=DEF_JSON)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    per = read_per_nacja(args.xlsx)
    print(f"[export-civ-dyplomacy-nations] nacje: {len(per)}")

    if not os.path.exists(args.json):
        sys.exit(f"Brak {args.json}")

    with open(args.json, encoding="utf-8") as f:
        data = json.load(f)

    data["perNacja"] = per

    if args.dry_run:
        print("[dry-run] bez zapisu")
        return

    bak = args.json + ".bak-CYWILIZACJE-export-perNacja"
    shutil.copy2(args.json, bak)

    tmp = args.json + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    json.load(open(tmp, encoding="utf-8"))
    os.replace(tmp, args.json)
    print(f"[export-civ-dyplomacy-nations] GOTOWE: perNacja={len(per)} w {args.json}")


if __name__ == "__main__":
    main()
