#!/usr/bin/env python3
"""DEPRECATED (2026-06-30 — PANEL-MERGE):
  Excel: docs/archiwum/panele-legacy/
  Kanon: panele-sterowania/Panel-{A..E}.xlsx + export-{a..e}.py
  Nie używać do nowych zmian balansu.


export-civs.py -- TARGETED export: Cywilizacje.xlsx[arkusz "Cywilizacje"] -> civs.json["nazwyKlastra", "mnoznikHandelPieniadz", "ikonaId"].

Aktualizuje WYLACZNIE pola `nazwyKlastra`, `mnoznikHandelPieniadz` i `ikonaId` per cywilizacja w gra/data/civs.json.
Dopasowanie po kluczu `Cywilizacja`. Wszystkie inne pola (w tym `bonusy`, `ikonaId`, freetext) i kolejnosc cywilizacji
pozostaja NIETKNIETE — pola `bonusy`, `typCywilizacji`, `archetyp` sa zarzadzane recznie w civs.json (nie pochodza z xlsx).

Zrodlo danych:
  Cywilizacje.xlsx, arkusz "Cywilizacje", wiersz 2 = naglowki, wiersze 3+ = dane.
  Kolumny Klaster_01..Klaster_10 (col 9..18): [0]=stolica.
  Kolumna "Mnoznik Handel-Pieniadz" (col 19): mnoznik konwersji Handel->Pieniadz (baza 2.0).
  Kolumna "IkonaId" (col 20): stabilny identyfikator ikony/emblematu UI (lowercase ASCII).

WAZNE: pola `bonusy` (schemat T3=A), `typCywilizacji` (klucz enum TypCywilizacji) i `archetyp`
  (klucz CIV_TO_ARCH z ai.ts) sa zarzadzane recznie w civs.json.
  Ten skrypt ich NIE nadpisuje. Przy re-eksporcie z xlsx wszystkie te pola zostaja zachowane.
  Aby zaktualizowac bonusy/typCywilizacji/archetyp — edytuj civs.json bezposrednio.

Uzycie (z katalogu gra/):
  python3 tools/export-civs.py             # zapis do civs.json
  python3 tools/export-civs.py --dry-run   # tylko wypisz, bez zapisu
  python3 tools/export-civs.py --xlsx SCIEZKA --json SCIEZKA

Domyslne sciezki liczone wzgledem tego pliku:
  xlsx = ../../Cywilizacje.xlsx
  json = ../data/civs.json
"""
import argparse
import json
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl:  pip install openpyxl --break-system-packages")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_XLSX = os.path.normpath(os.path.join(HERE, "..", "..", "docs", "archiwum", "panele-legacy", "Cywilizacje.xlsx"))
DEF_JSON = os.path.normpath(os.path.join(HERE, "..", "data", "civs.json"))
SHEET = "Cywilizacje"
KLASTER_COLS = 10  # Klaster_01..Klaster_10
KLASTER_START_COL = 9  # kolumna I (1-indexed)
CYW_COL = 1  # kolumna A: klucz cywilizacji
MNOZNIK_COL_HEADER = "Mnoznik Handel-Pieniadz"
IKONA_COL_HEADER = "IkonaId"

# Pola zachowywane z civs.json (nie nadpisywane z xlsx):
PRESERVED_FIELDS = {"bonusy", "typCywilizacji", "archetyp", "Cywilizacja", "Styl / charakter", "Jednostka specjalna",
                    "Bonus startowy", "Bonusy/minusy (do dopracowania)", "Uwagi",
                    "Religia"}


def read_xlsx_data(xlsx_path):
    """Czyta nazwyKlastra, mnoznikHandelPieniadz i ikonaId z arkusza Cywilizacje.
    Zwraca dict {Cywilizacja: {"klastry": [str x 10], "mnoznikHandelPieniadz": float, "ikonaId": str}}."""
    if not os.path.exists(xlsx_path):
        sys.exit(f"Nie znaleziono pliku xlsx: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Brak arkusza '{SHEET}' w {xlsx_path}")
    ws = wb[SHEET]

    # Znajdz wiersz naglowkow (szukamy "Cywilizacja" w kol 1)
    header_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=CYW_COL).value
        if val == "Cywilizacja":
            header_row = r
            break
    if header_row is None:
        sys.exit(f"Nie znaleziono wiersza naglowkowego (kol A = 'Cywilizacja') w arkuszu '{SHEET}'")

    # Sprawdz naglowki Klaster_01..Klaster_10
    expected = [f"Klaster_{i:02d}" for i in range(1, KLASTER_COLS + 1)]
    actual = [ws.cell(row=header_row, column=KLASTER_START_COL + i).value for i in range(KLASTER_COLS)]
    if actual != expected:
        sys.exit(
            f"Nieoczekiwane naglowki klastra w wierszu {header_row}: "
            f"oczekiwano {expected}, znaleziono {actual}"
        )

    # Znajdz kolumne mnoznika (szukaj po naglowku)
    mnoznik_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == MNOZNIK_COL_HEADER:
            mnoznik_col = c
            break
    if mnoznik_col is None:
        sys.exit(
            f"Nie znaleziono kolumny '{MNOZNIK_COL_HEADER}' w wierszu naglowkowym {header_row}. "
            "Dodaj kolumne do arkusza 'Cywilizacje' w Cywilizacje.xlsx przed eksportem."
        )
    print(f"[export-civs] kolumna mnoznika: {mnoznik_col} ('{MNOZNIK_COL_HEADER}')")

    # Znajdz kolumne ikonaId (szukaj po naglowku)
    ikona_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == IKONA_COL_HEADER:
            ikona_col = c
            break
    if ikona_col is None:
        sys.exit(
            f"Nie znaleziono kolumny '{IKONA_COL_HEADER}' w wierszu naglowkowym {header_row}. "
            "Dodaj kolumne do arkusza 'Cywilizacje' w Cywilizacje.xlsx przed eksportem."
        )
    print(f"[export-civs] kolumna ikonaId: {ikona_col} ('{IKONA_COL_HEADER}')")

    result = {}
    for r in range(header_row + 1, ws.max_row + 1):
        cyw = ws.cell(row=r, column=CYW_COL).value
        if not cyw or str(cyw).strip() == "":
            continue
        cyw = str(cyw).strip()

        # Klastry
        nazwy = []
        for i in range(KLASTER_COLS):
            v = ws.cell(row=r, column=KLASTER_START_COL + i).value
            if v is None or str(v).strip() == "":
                sys.exit(
                    f"Wiersz {r} [{cyw}]: Klaster_{i+1:02d} jest pusty! "
                    "Uzupelnij xlsx przed eksportem."
                )
            nazwy.append(str(v).strip())

        # Mnoznik
        mnoznik_raw = ws.cell(row=r, column=mnoznik_col).value
        if mnoznik_raw is None:
            sys.exit(
                f"Wiersz {r} [{cyw}]: brak wartosci '{MNOZNIK_COL_HEADER}'! "
                "Uzupelnij xlsx przed eksportem."
            )
        try:
            mnoznik = float(mnoznik_raw)
        except (ValueError, TypeError):
            sys.exit(f"Wiersz {r} [{cyw}]: '{MNOZNIK_COL_HEADER}' nie jest liczba: {mnoznik_raw!r}")

        # IkonaId
        ikona_raw = ws.cell(row=r, column=ikona_col).value
        if ikona_raw is None or str(ikona_raw).strip() == "":
            sys.exit(
                f"Wiersz {r} [{cyw}]: brak wartosci '{IKONA_COL_HEADER}'! "
                "Uzupelnij xlsx przed eksportem."
            )
        ikona = str(ikona_raw).strip()

        result[cyw] = {"klastry": nazwy, "mnoznikHandelPieniadz": mnoznik, "ikonaId": ikona}

    return result


def main():
    ap = argparse.ArgumentParser(
        description="Export nazwyKlastra + mnoznikHandelPieniadz + ikonaId: Cywilizacje.xlsx -> civs.json"
    )
    ap.add_argument("--xlsx", default=DEF_XLSX)
    ap.add_argument("--json", default=DEF_JSON)
    ap.add_argument("--dry-run", action="store_true", help="tylko wypisz, nie zapisuj")
    a = ap.parse_args()

    xlsx_data = read_xlsx_data(a.xlsx)
    print(f"[export-civs] odczytano {len(xlsx_data)} cywilizacji z arkusza '{SHEET}' ({a.xlsx})")
    for cyw, d in xlsx_data.items():
        nazwy = d["klastry"]
        mnoznik = d["mnoznikHandelPieniadz"]
        ikona = d["ikonaId"]
        print(f"  {cyw:15s}: klastry={nazwy[0]}..{nazwy[-1]}  mnoznikHandelPieniadz={mnoznik}  ikonaId={ikona!r}")

    if a.dry_run:
        print("\n[dry-run] civs.json nietkniety.")
        return

    if not os.path.exists(a.json):
        sys.exit(f"Nie znaleziono civs.json: {a.json}")

    # Backup
    bak = a.json + ".bak-CYWILIZACJE"
    shutil.copy2(a.json, bak)
    print(f"\n[export-civs] backup: {bak}")

    with open(a.json, "r", encoding="utf-8") as f:
        data = json.load(f)

    if "cywilizacje" not in data:
        sys.exit("civs.json nie zawiera klucza 'cywilizacje'!")

    matched = 0
    unmatched_json = []
    unmatched_xlsx = set(xlsx_data.keys())

    for cyw_obj in data["cywilizacje"]:
        cyw_key = cyw_obj.get("Cywilizacja", "").strip()
        if cyw_key in xlsx_data:
            d = xlsx_data[cyw_key]

            # nazwyKlastra
            old_klastry = cyw_obj.get("nazwyKlastra", [])
            cyw_obj["nazwyKlastra"] = d["klastry"]

            # mnoznikHandelPieniadz
            old_mnoznik = cyw_obj.get("mnoznikHandelPieniadz")
            cyw_obj["mnoznikHandelPieniadz"] = d["mnoznikHandelPieniadz"]

            # ikonaId
            old_ikona = cyw_obj.get("ikonaId")
            cyw_obj["ikonaId"] = d["ikonaId"]

            # UWAGA: pola `bonusy`, `typCywilizacji`, `archetyp` sa celowo NIE aktualizowane z xlsx.
            # Schemat bonusow (T3=A), klucz enum i klucz archetypu zarzadzane sa recznie w civs.json.
            # Przy re-eksporcie wszystkie te pola zostaja zachowane bez zmian.

            unmatched_xlsx.discard(cyw_key)
            matched += 1

            klastry_changed = old_klastry != d["klastry"]
            mnoznik_changed = old_mnoznik != d["mnoznikHandelPieniadz"]
            ikona_changed = old_ikona != d["ikonaId"]
            bonusy_count = len(cyw_obj.get("bonusy", []))
            changes = []
            if klastry_changed:
                changes.append("klastry")
            if mnoznik_changed:
                changes.append(f"mnoznik {old_mnoznik}->{d['mnoznikHandelPieniadz']}")
            if ikona_changed:
                changes.append(f"ikona {old_ikona!r}->{d['ikonaId']!r}")
            status = f"[zmieniono: {', '.join(changes)}]" if changes else "[bez zmian]"
            print(f"  {cyw_key:15s}: {status}  bonusy={bonusy_count} (zachowane)")
        else:
            unmatched_json.append(cyw_key)

    print(f"\n[export-civs] dopasowano {matched}/{len(data['cywilizacje'])} cywilizacji.")
    if unmatched_json:
        print(f"  OSTRZEZENIE: cywilizacje w JSON bez odpowiednika w xlsx: {unmatched_json}")
    if unmatched_xlsx:
        print(f"  OSTRZEZENIE: cywilizacje w xlsx bez odpowiednika w JSON: {list(unmatched_xlsx)}")

    with open(a.json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[export-civs] zapisano: {a.json}")
    print("[export-civs] GOTOWE. Pola 'bonusy', 'typCywilizacji', 'archetyp', 'nazwyKlastra', 'mnoznikHandelPieniadz', 'ikonaId' zachowane.")


if __name__ == "__main__":
    main()
