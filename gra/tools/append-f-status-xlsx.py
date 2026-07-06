#!/usr/bin/env python3
# append-f-status-xlsx.py — dopisuje wiersze Grupa F / SILNIK do Status-projektu-The-Game.xlsx
# Uruchom z root Civ: python gra/tools/append-f-status-xlsx.py

import datetime
import os
import shutil
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Brak openpyxl: pip install openpyxl")
    sys.exit(1)

CIV_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(CIV_ROOT, "Status-projektu-The-Game.xlsx")
SHEET = "Grupa-F"
TODAY = datetime.date.today().isoformat()

HEADERS = ("ID", "Temat", "Status", "Decyzja", "Data", "Uwagi")

ROWS = [
    ("F-kod", "Integracja main.ts (8 batchy)", "WPIĘTE", "—", TODAY, "F1,F2,F-A2,F-B2,F-C1,F-HUD,F-B2-porzadek"),
    ("F-bramka", "Gra-podglad-ROBOCZA.html", "CZEKA", "—", TODAY, "Node lokalnie: bramka-test-publish.ps1"),
    ("F-HUD-1", "HUD D1B cz.1 (pasek, minimapa, wojny)", "WPIĘTE", "ABC1=A", TODAY, "hud.ts w main"),
    ("F-HUD-2", "WYKONAJ + panel [H] + tryb budowy", "TODO", "A1-Q9,A2-Q4", TODAY, "handoff D1B-A4-batch"),
    ("F-C2", "Bitwa TW pełna (deploy map)", "TODO", "D5=B", TODAY, "po ROBOCZA PASS"),
    ("F-B2-hex", "Ikona buntu 🔥 na heksie", "CZEKA MAPA", "B2-Q5=C", TODAY, "getRevolt w cities.ts"),
    ("F-save-B2", "Persist cityOrderState w save", "TODO", "—", TODAY, "luka po Ctrl+L"),
    ("F-D4", "Audyt bonusów cywilizacji", "TODO P2", "—", TODAY, "delegacja CYWILIZACJE"),
    ("F-A4-D4", "Ulepszenia terenu z mapy", "TODO P2", "A4-D4-Q1=A", TODAY, "BLK-04 po MAPA+UI"),
    ("F-B5", "advanceEmpireFood", "BLOK", "—", TODAY, "stub throws — nie wpinaj"),
]


def main():
    if not os.path.isfile(XLSX):
        print(f"Brak pliku: {XLSX}")
        sys.exit(1)
    bak = XLSX + f".bak-F-{TODAY}"
    shutil.copy2(XLSX, bak)
    print(f"Backup: {bak}")
    wb = openpyxl.load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
        start_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(SHEET)
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="A8D4E8")
        start_row = 2
    for i, row in enumerate(ROWS):
        r = start_row + i
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)
    wb.save(XLSX)
    print(f"Dopisano {len(ROWS)} wierszy do arkusza '{SHEET}'")


if __name__ == "__main__":
    main()
