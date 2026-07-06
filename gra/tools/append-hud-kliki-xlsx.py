#!/usr/bin/env python3
# append-hud-kliki-xlsx.py — dopisuje arkusz HUD-mapa-kliki do Status-projektu-The-Game.xlsx
# Uruchom z root Civ: python gra/tools/append-hud-kliki-xlsx.py

import datetime
import os
import shutil
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Brak openpyxl: pip install openpyxl")
    sys.exit(1)

CIV_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(CIV_ROOT, "Status-projektu-The-Game.xlsx")
SHEET = "HUD-mapa-kliki"
TODAY = "2026-06-26"

ROWS = [
    ("A1-K01", "[A]", "Żywność", "NIE", "—", "Tooltip zapasy B5 +X/t", "—", "ZAMKNIĘTE", TODAY, "A1-revA"),
    ("A1-K02", "[A]", "Złoto", "opcj.", "—", "Tooltip bilansu", "—", "ZAMKNIĘTE", TODAY, "Bez panelu v1.0"),
    ("A1-K03", "[A]", "Praca", "opcj.", "—", "Tooltip bilansu", "—", "ZAMKNIĘTE", TODAY, "Bez panelu v1.0"),
    ("A1-K04", "[A]", "Badania", "TAK", "sciencePicker", "Pełnoekranowe drzewko tech", "G3", "ZAMKNIĘTE", TODAY, "A1-revA"),
    ("A1-K05", "[A]", "Bogactwo", "opcj.", "—", "Tooltip; suwak w cityPanel", "—", "ZAMKNIĘTE", TODAY, "D3 Wealth PL"),
    ("A1-K06", "[A]", "Ludność", "opcj.", "—", "Tooltip suma miast + przyrost", "—", "ZAMKNIĘTE", TODAY, "A1-revA"),
    ("A1-K07", "[A]", "Epoka (pasek %)", "NIE", "—", "Podgląd postępu", "—", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K08", "[A]", "Nacja", "NIE", "—", "Etykieta frakcji", "—", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K09", "[A]", "Osiedla N/max", "NIE", "—", "Liczniki; lista → Miasta [I]", "—", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K10", "[A]", "Tura / data", "NIE", "—", "Informacja", "—", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K11", "[A]", "Przycisk Dyplomacja", "TAK", "diplomacyPanel", "Panel dyplomacji", "G3", "ZAMKNIĘTE", TODAY, "A1-Q5"),
    ("A1-K12", "(B)", "Chip wojny (nacja ⚔)", "TAK", "diplomacyPanel", "Dyplomacja z fokusem na nacji", "—", "ZAMKNIĘTE", TODAY, "A1-Q5"),
    ("A1-K13", "[C]", "🏛️ Cuda", "TAK", "overlay-cuda", "Lista cudów + postęp", "G3", "ZAMKNIĘTE", TODAY, "Lane UI"),
    ("A1-K14", "[C]", "🔨 Budowa", "TAK", "tryb-budowy", "Placement ulepszeń na [D] + panel", "G2", "ZAMKNIĘTE", TODAY, "D4"),
    ("A1-K15", "[D]", "Hex własnego miasta", "TAK", "cityPanel", "Panel miasta pełny ekran", "G3", "ZAMKNIĘTE", TODAY, "main.ts"),
    ("A1-K16", "[D]", "Hex obcego miasta", "TAK", "diplomacyPanel", "Dyplomacja / info miasta", "G3", "ZAMKNIĘTE", TODAY, "v1.0 propozycja"),
    ("A1-K17", "[D]", "Własna jednostka", "TAK", "panel [H]", "Zaznaczenie + panel jednostki", "—", "CZĘŚCIOWO", TODAY, "A2-Q4 OTWARTE"),
    ("A1-K18", "[D]", "Wroga jednostka (z wyborem)", "TAK", "preBattle", "Podsumowanie → bitwa", "G1 możliwe", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K19", "[D]", "Pusty hex (jednostka wybrana)", "TAK", "ruch", "Ruch jednostki", "—", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K20", "[D]", "Hex w trybie Budowa", "TAK", "placement", "Postaw ulepszenie terenu", "G2", "ZAMKNIĘTE", TODAY, "D4"),
    ("A1-K21", "[E]", "Chip informacyjny", "TAK", "onEventClick", "Akcja kontekstowa / skok", "—", "ZAMKNIĘTE", TODAY, "A1-Q8"),
    ("A1-K22", "[E]", "Chip informacyjny ✕", "TAK", "onEventDismiss", "Zamknij chip", "—", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K23", "[E]", "Chip blocking", "TAK", "onEventClick", "Rozstrzygnięcie wymagane", "G1", "ZAMKNIĘTE", TODAY, "A1-Q9"),
    ("A1-K24", "[F]", "Hex minimapy", "TAK", "kamera", "Skok kamery na hex [D]", "—", "ZAMKNIĘTE", TODAY, "D15=B"),
    ("A1-K25", "[F2]", "Toggle warstwy", "TAK", "warstwa mapy", "ON/OFF widoku 3D", "—", "ZAMKNIĘTE", TODAY, "A1-Q6 F2"),
    ("A1-K26", "[I]", "Miasta", "TAK", "cityPanel", "Panel miasta / lista", "G3", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K27", "[I]", "WYKONAJ", "TAK", "onExecutePending", "Pierwsze blocking z [E]", "G1", "ZAMKNIĘTE", TODAY, "A1-Q9"),
    ("A1-K28", "[I]", "Koniec tury", "TAK", "onEndTurn", "Następna tura", "G1 gate", "ZAMKNIĘTE", TODAY, "A1-Q10"),
    ("A1-K29", "[I]", "Menu", "TAK", "mainMenu", "Zapis / wczytaj / ustawienia", "G3", "ZAMKNIĘTE", TODAY, ""),
    ("A1-K30", "[I2]", "Okrąg Koniec tury", "TAK", "onEndTurn", "Jak Koniec tury [I]", "G1 gate", "ZAMKNIĘTE", TODAY, "A1-Q10 A+B"),
    ("A1-K31", "[F′]", "Ikona Kultura — toggle zasięgu", "TAK", "MAPA overlay", "ON/OFF zasięg kultury na [D]", "—", "ZAMKNIĘTE", TODAY, "MAPA-F2-Q1"),
    ("A1-K32", "[F′]", "Ikona Religia — toggle zasięgu", "TAK", "MAPA overlay", "ON/OFF zasięg naszej religii na [D]", "—", "ZAMKNIĘTE", TODAY, "MAPA-F2-Q1"),
    ("A1-K33", "[F″]", "Ikona Kultura — klik treść", "TAK", "overlay-kultura", "Panel imperium kultury (A1-Q12a OTW.)", "G3", "OTWARTE", TODAY, "Grupa A nie MAPA/D"),
    ("A1-K34", "[F″]", "Ikona Religia — klik treść", "TAK", "overlay-religia", "Panel imperium religii (A1-Q12b OTW.)", "G3", "OTWARTE", TODAY, "Grupa A nie MAPA/D"),
    ("MAPA-F2-Q1", "[F′]", "Routing toggle zasięgu", "—", "MAPA lane", "Wygląd ikon + render — nie Grupa D", "—", "ZAMKNIĘTE", TODAY, "Maciej korekta"),
    ("A1-Q6-rev", "[C]", "Toolbar bez Zasoby", "—", "—", "Tylko Cuda + Budowa", "—", "ZAMKNIĘTE", TODAY, "A1-revA"),
]

HEADERS = (
    "ID", "Strefa", "Element", "Klik?", "Moduł docelowy",
    "Efekt (PL)", "Bramka", "Status decyzji", "Data", "Uwagi",
)


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=row, column=col, value=HEADERS[col - 1])
        c.fill = fill
        c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def main():
    if not os.path.exists(XLSX):
        print("BRAK:", XLSX)
        sys.exit(1)

    bak = XLSX + f".bak-HUD-kliki-{TODAY}"
    if not os.path.exists(bak):
        shutil.copy2(XLSX, bak)
        print("Backup:", bak)

    wb = openpyxl.load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    ws["A1"] = "HUD MAPY STRATEGICZNEJ — mapa kliknięć (decyzja Maciej 2026-06-26)"
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"Źródło markdown: docs/A1-HUD-MAP-KLIKNIEC.md · Aktualizacja: {TODAY}"
    ws.merge_cells("A2:J2")

    start_row = 4
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row)

    for ri, row in enumerate(ROWS, start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == 8 and val == "ZAMKNIĘTE":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")

    widths = [10, 8, 28, 8, 18, 42, 10, 14, 12, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Civ-UI: dopisz wiersz zadania jeśli brak
    ui = wb["Civ-UI"]
    next_row = ui.max_row + 1
    already = False
    for row in ui.iter_rows(min_row=6, values_only=True):
        if row and len(row) > 1 and row[1] and "mapa kliknięć" in str(row[1]).lower():
            already = True
            break
    if not already:
        ui.cell(row=next_row, column=1, value=7)
        ui.cell(row=next_row, column=2, value="HUD mapy D1B: mockup + mapa kliknięć (arkusz HUD-mapa-kliki)")
        ui.cell(row=next_row, column=3, value="W toku")
        ui.cell(row=next_row, column=4, value=f"Spec {TODAY}; decyzja Maciej; docs/A1-HUD-MAP-KLIKNIEC.md")

    wb.save(XLSX)
    print("OK:", XLSX, "-> arkusz", SHEET, f"({len(ROWS)} wierszy)")


if __name__ == "__main__":
    main()
