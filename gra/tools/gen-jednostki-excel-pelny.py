#!/usr/bin/env python3
"""
gen-jednostki-excel-pelny.py

Dwa pliki Excel:
  1. Jednostki-parametry-TW-v3-STAN.xlsx     — stan z units.json (3 zakładki, bez zmian)
  2. Jednostki-parametry-TW-v3-PROPOZYCJA.xlsx — propozycja balansu (3 zakładki)

Uruchomienie (z gra/):
  python tools/gen-jednostki-excel-pelny.py
"""
from __future__ import annotations

import json
import os
import sys
from copy import deepcopy
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("BRAK openpyxl — pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
UNITS_PATH = os.path.normpath(os.path.join(HERE, "..", "data", "units.json"))
OUT_STAN = os.path.normpath(os.path.join(HERE, "Jednostki-parametry-TW-v3-STAN.xlsx"))
OUT_PROP = os.path.normpath(os.path.join(HERE, "Jednostki-parametry-TW-v3-PROPOZYCJA.xlsx"))

FONT = "Arial"
H = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="4A5568")
HF_ATK = PatternFill("solid", fgColor="9B2C2C")
HF_DEF = PatternFill("solid", fgColor="2C5282")
HF_MISC = PatternFill("solid", fgColor="276749")
HF_PROP = PatternFill("solid", fgColor="744210")
TF = Font(name=FONT, color="1A1A1A")
NF = Font(name=FONT, color="0000FF")
PF = Font(name=FONT, color="006400")
KF = Font(name=FONT, bold=True, color="1A1A1A")
thin = Side(style="thin", color="D9D9D9")
B = Border(thin, thin, thin, thin)

HEAD_ATK = [
    "Jednostka", "Epoka", "Rola (linia)",
    "AP", "Obraż", "Przeb", "Szarża", "AD", "wallAttack",
    "Zasięg (hex)", "Ilość pocisków", "A",
]
HEAD_DEF = [
    "Jednostka", "Epoka", "Rola (linia)",
    "OBR", "Panc", "HP", "O",
]
HEAD_MISC = [
    "Jednostka", "Epoka", "Rola (linia)", "Typ", "Klasa", "Nazwa EN",
    "Ruch", "Ruch w bitwie", "Widok pola",
    "Koszt ($)", "Ludność", "Utrzymanie/t", "Żywność/t",
    "Morale bazowe", "Morale ucieczki", "Próg dezercji",
    "Kara flanki %", "Kara tyłu %",
    "Bonus vs Miecz %", "Bonus vs Włócz %", "Bonus vs Falanga %",
    "Bonus vs Off %", "Bonus vs Dystans %", "Bonus vs Kon %",
    "Super", "W zamian za", "Tech", "Uwagi (skrót)",
]

# --- Propozycja balansu (subiektywna, skala TW v3) ---
# Kotwica: Wojownik Kamień AP/OBR/Obraż=4; supery AP 8–10 Obraż=10; dystans słabe AP, AD z Excela.

def C(**kw):
    return kw


PROPOZYCJA: dict[str, dict] = {
    # --- Linia standard ---
    "Wojownik": C(meleeAttack=4, meleeDefence=4, weaponDamage=4, armor=2, piercing=1, chargeBonus=2, health=11, missileAttack=0),
    "Wojownik z mieczem i tarczą": C(meleeAttack=6, meleeDefence=5, weaponDamage=6, armor=4, piercing=2, chargeBonus=4, health=14, missileAttack=0),
    "Włócznik": C(meleeAttack=5, meleeDefence=8, weaponDamage=5, armor=6, piercing=2, chargeBonus=4, health=16, missileAttack=0),
    "Falanga": C(meleeAttack=5, meleeDefence=10, weaponDamage=4, armor=8, piercing=2, chargeBonus=7, health=25, missileAttack=0),
    "Hastati": C(meleeAttack=7, meleeDefence=7, weaponDamage=7, armor=9, piercing=3, chargeBonus=7, health=19, missileAttack=8),
    "Triari": C(meleeAttack=8, meleeDefence=8, weaponDamage=10, armor=6, piercing=4, chargeBonus=10, health=21, missileAttack=0),
    "Konnica": C(meleeAttack=8, meleeDefence=5, weaponDamage=7, armor=4, piercing=6, chargeBonus=10, health=14, missileAttack=0),
    "Rydwan (woły)": C(meleeAttack=6, meleeDefence=4, weaponDamage=7, armor=3, piercing=3, chargeBonus=10, health=24, missileAttack=0),
    "Łucznik": C(meleeAttack=2, meleeDefence=2, weaponDamage=2, armor=1, piercing=1, chargeBonus=0, health=8, missileAttack=6),
    "Zwiadowca": C(meleeAttack=0, meleeDefence=1, weaponDamage=0, armor=0, piercing=0, chargeBonus=0, health=5, missileAttack=0),
    # --- Dystans standard / kamień ---
    "Procarz": C(meleeAttack=1, meleeDefence=1, weaponDamage=0, armor=0, piercing=0, chargeBonus=0, health=8, missileAttack=8),
    "Procarz (Huaracoc)": C(meleeAttack=1, meleeDefence=1, weaponDamage=0, armor=0, piercing=0, chargeBonus=0, health=8, missileAttack=8),
    "Oszczepnik": C(meleeAttack=1, meleeDefence=1, weaponDamage=0, armor=0, piercing=0, chargeBonus=0, health=8, missileAttack=8),
    "Oszczepnik (Estólica)": C(meleeAttack=1, meleeDefence=1, weaponDamage=0, armor=0, piercing=0, chargeBonus=0, health=8, missileAttack=8),
    "Oszczepnik Zulu (Izijula)": C(meleeAttack=1, meleeDefence=2, weaponDamage=0, armor=1, piercing=0, chargeBonus=0, health=10, missileAttack=9),
    "Kusznik": C(meleeAttack=2, meleeDefence=2, weaponDamage=0, armor=1, piercing=0, chargeBonus=0, health=8, missileAttack=15),
    # --- Kulturowe wręcz (skala TW, nie ÷10) ---
    "Wojownik mykeński": C(meleeAttack=6, meleeDefence=4, weaponDamage=5, armor=3, piercing=1, chargeBonus=3, health=14, missileAttack=0),
    "Wojownik tyrreński": C(meleeAttack=7, meleeDefence=4, weaponDamage=5, armor=2, piercing=2, chargeBonus=3, health=12, missileAttack=0),
    "Wojownik z khopesh": C(meleeAttack=6, meleeDefence=5, weaponDamage=5, armor=3, piercing=1, chargeBonus=3, health=12, missileAttack=0),
    "Wojownik z toporem": C(meleeAttack=6, meleeDefence=4, weaponDamage=5, armor=2, piercing=2, chargeBonus=3, health=12, missileAttack=0),
    "Wojownik z maczugą (Chaska)": C(meleeAttack=7, meleeDefence=4, weaponDamage=6, armor=2, piercing=2, chargeBonus=2, health=10, missileAttack=0),
    "Wojownik Sherden": C(meleeAttack=6, meleeDefence=5, weaponDamage=5, armor=3, piercing=1, chargeBonus=3, health=14, missileAttack=0),
    "Wojownik szekelesz": C(meleeAttack=5, meleeDefence=5, weaponDamage=4, armor=3, piercing=0, chargeBonus=2, health=14, missileAttack=0),
    "Wojownik celtycki": C(meleeAttack=7, meleeDefence=3, weaponDamage=5, armor=2, piercing=1, chargeBonus=5, health=14, missileAttack=0),
    "Wojownik germański": C(meleeAttack=6, meleeDefence=4, weaponDamage=5, armor=2, piercing=1, chargeBonus=4, health=15, missileAttack=9),
    "Berserker germański": C(meleeAttack=8, meleeDefence=2, weaponDamage=6, armor=1, piercing=1, chargeBonus=6, health=12, missileAttack=0),
    "Gaesatae": C(meleeAttack=7, meleeDefence=2, weaponDamage=5, armor=1, piercing=1, chargeBonus=5, health=11, missileAttack=0),
    "Impi": C(meleeAttack=5, meleeDefence=6, weaponDamage=4, armor=3, piercing=1, chargeBonus=3, health=18, missileAttack=0),
    "Halabardnik Shang": C(meleeAttack=6, meleeDefence=5, weaponDamage=5, armor=4, piercing=3, chargeBonus=2, health=12, missileAttack=0),
    "Włócznik sumeryjski": C(meleeAttack=5, meleeDefence=6, weaponDamage=4, armor=3, piercing=2, chargeBonus=2, health=20, missileAttack=0),
    # --- Kulturowe dystans ---
    "Łucznik sumeryjski": C(meleeAttack=1, meleeDefence=1, weaponDamage=0, armor=0, piercing=0, chargeBonus=0, health=8, missileAttack=8),
    "Łucznik egipski": C(meleeAttack=1, meleeDefence=2, weaponDamage=0, armor=1, piercing=0, chargeBonus=0, health=8, missileAttack=11),
    "Łucznik akadyjski": C(meleeAttack=1, meleeDefence=2, weaponDamage=0, armor=1, piercing=0, chargeBonus=0, health=9, missileAttack=10),
    # --- Kulturowe kon / rydwan (top Brąz — wysoka szarża + HP) ---
    "Rydwan sumeryjski": C(meleeAttack=7, meleeDefence=4, weaponDamage=6, armor=3, piercing=2, chargeBonus=9, health=28, missileAttack=0),
    "Rydwan mykeński": C(meleeAttack=7, meleeDefence=3, weaponDamage=7, armor=2, piercing=3, chargeBonus=9, health=26, missileAttack=0),
    "Rydwan konny": C(meleeAttack=7, meleeDefence=4, weaponDamage=6, armor=3, piercing=2, chargeBonus=8, health=28, missileAttack=0),
    "Rydwan Shang": C(meleeAttack=7, meleeDefence=5, weaponDamage=6, armor=3, piercing=3, chargeBonus=8, health=30, missileAttack=0),
    "Rydwan celtycki": C(meleeAttack=8, meleeDefence=3, weaponDamage=7, armor=2, piercing=2, chargeBonus=10, health=26, missileAttack=0),
    "Rydwan egipski": C(meleeAttack=6, meleeDefence=4, weaponDamage=5, armor=2, piercing=2, chargeBonus=7, health=28, missileAttack=10),
    "Jeździec chiński": C(meleeAttack=7, meleeDefence=4, weaponDamage=5, armor=3, piercing=2, chargeBonus=5, health=22, missileAttack=0),
    # --- Super ---
    "Gwardia Królewska Sumeru": C(meleeAttack=10, meleeDefence=8, weaponDamage=10, armor=6, piercing=4, chargeBonus=8, health=21, missileAttack=0),
    "Hieros Lochos (Święty Zastęp)": C(meleeAttack=8, meleeDefence=10, weaponDamage=10, armor=6, piercing=4, chargeBonus=8, health=21, missileAttack=0),
    "Hu Ben Wei (Gwardia Tygrysa)": C(meleeAttack=10, meleeDefence=7, weaponDamage=10, armor=6, piercing=4, chargeBonus=8, health=20, missileAttack=0),
    "Królewska Gwardia": C(meleeAttack=10, meleeDefence=8, weaponDamage=10, armor=6, piercing=4, chargeBonus=8, health=20, missileAttack=0),
    "Medżaj (Gwardia Faraona)": C(meleeAttack=10, meleeDefence=8, weaponDamage=10, armor=6, piercing=4, chargeBonus=8, health=21, missileAttack=6),
    "uThulwana (Białe Tarcze)": C(meleeAttack=8, meleeDefence=7, weaponDamage=10, armor=6, piercing=4, chargeBonus=10, health=21, missileAttack=0),
    # --- Oblężenie / spec ---
    "Katapulta": C(meleeAttack=0, meleeDefence=0, weaponDamage=0, armor=0, piercing=1, chargeBonus=0, health=125, missileAttack=8, wallAttack=16),
    "Taran": C(meleeAttack=0, meleeDefence=0, weaponDamage=0, armor=1, piercing=1, chargeBonus=1, health=350, missileAttack=0, wallAttack=14),
    "Wieża oblężnicza": C(meleeAttack=0, meleeDefence=0, weaponDamage=0, armor=1, piercing=0, chargeBonus=0, health=450, missileAttack=0, wallAttack=6),
    "Galera": C(meleeAttack=3, meleeDefence=4, weaponDamage=3, armor=2, piercing=0, chargeBonus=2, health=12, missileAttack=0),
}


def n(v, default=0.0):
    if v is None or v == "" or v == "—":
        return float(default)
    if isinstance(v, (int, float)):
        return float(v)
    return float(default)


def dash(v):
    if v is None or v == "":
        return "—"
    return v


def apply_proposal(u: dict) -> dict:
    out = deepcopy(u)
    p = PROPOZYCJA.get(u["Jednostka"], {})
    for k, v in p.items():
        out[k] = v
    return out


def apply_proposal_to_json() -> tuple[int, str]:
    """Zapisuje PROPOZYCJA → units.json (backup + marker)."""
    import shutil
    from pathlib import Path as P

    path = P(UNITS_PATH)
    backup = path.with_name(path.name + f".bak-BALANS-{date.today().isoformat()}")
    shutil.copy2(path, backup)
    units = json.load(open(path, encoding="utf-8"))
    field_changes = 0
    for u in units:
        name = u["Jednostka"]
        if name not in PROPOZYCJA:
            continue
        for k, v in PROPOZYCJA[name].items():
            if u.get(k) != v:
                u[k] = v
                field_changes += 1
        u["_tw_v3_balans"] = date.today().isoformat()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(units, f, ensure_ascii=False, indent=2)
    return field_changes, str(backup)


def attack_row(u: dict) -> list:
    ap, ob, pr, sz, ad = int(n(u.get("meleeAttack"))), int(n(u.get("weaponDamage"))), int(n(u.get("piercing"))), int(n(u.get("chargeBonus"))), int(n(u.get("missileAttack")))
    wall = u.get("wallAttack")
    a = ap + ob + pr + sz / 2 + ad / 2
    return [
        u["Jednostka"], u.get("Epoka", ""), u.get("Rola (linia)", ""),
        ap, ob, pr, sz, ad, "" if wall is None else int(wall),
        dash(u.get("Zasięg ataku (hex)")), dash(u.get("Ilość pocisków")), round(a, 1),
    ]


def defense_row(u: dict) -> list:
    obr, panc, hp = int(n(u.get("meleeDefence"))), int(n(u.get("armor"))), int(n(u.get("health")))
    o = obr + panc + hp / 2
    return [u["Jednostka"], u.get("Epoka", ""), u.get("Rola (linia)", ""), obr, panc, hp, round(o, 1)]


def misc_row(u: dict) -> list:
    uw = (u.get("Uwagi") or "")[:80]
    return [
        u["Jednostka"], u.get("Epoka", ""), u.get("Rola (linia)", ""),
        u.get("Typ", ""), u.get("Klasa", ""), u.get("Nazwa EN", ""),
        u.get("Ruch", ""), u.get("Ruch w bitwie (heksy)", ""), u.get("Widok pola", ""),
        u.get("Pieniądz (koszt)", ""), u.get("Ludność", ""), u.get("Utrzymanie (Pieniądz/turę)", ""), u.get("żywność/turę", ""),
        u.get("Morale bazowe", ""), dash(u.get("Morale ucieczki")), u.get("Próg dezercji (% health)", ""),
        u.get("Kara obrony z flanki (%)", ""), u.get("Kara obrony z tyłu (%)", ""),
        u.get("Bonus vs Swordsman %", 0), u.get("Bonus vs Spearman %", 0), u.get("Bonus vs Falangite %", 0),
        u.get("Bonus vs Offensive %", 0), u.get("Bonus vs Distance %", 0), u.get("Bonus vs Mount %", 0),
        dash(u.get("Super-jednostka")), dash(u.get("W zamian za")), u.get("Tech", ""), uw,
    ]


def autosize(ws, nrows: int, ncols: int, max_w: int = 42):
    for col in range(1, ncols + 1):
        w = 10
        for row in range(1, nrows + 1):
            val = ws.cell(row, col).value
            if val is not None:
                w = max(w, min(len(str(val)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col)].width = w


def write_sheet(wb, title: str, headers: list, rows: list, hdr_fill, subtitle: str, prop_mode: bool = False):
    if title in wb.sheetnames:
        ws = wb[title]
    else:
        ws = wb.create_sheet(title)
    ws.delete_rows(1, ws.max_row or 1)

    ws["A1"] = subtitle
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = f"Wygenerowano: {date.today().isoformat()} · źródło: gra/data/units.json"
    ws["A2"].font = Font(name=FONT, italic=True, size=10, color="666666")
    if prop_mode:
        ws["A3"] = "PROPOZYCJA agenta — do porównania ze STAN; nie wdrożone w JSON."
        ws["A3"].font = Font(name=FONT, size=10, color="8B4513")

    hdr_row = 5
    for j, h in enumerate(headers, 1):
        c = ws.cell(hdr_row, j, h)
        c.font = H
        c.fill = hdr_fill
        c.border = B
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for i, row in enumerate(rows, hdr_row + 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j, val)
            cell.border = B
            cell.alignment = Alignment(vertical="top", wrap_text=j <= 3)
            if j == 1:
                cell.font = KF
            elif isinstance(val, (int, float)) and j >= 4:
                cell.font = PF if prop_mode else NF
            else:
                cell.font = TF

    ws.freeze_panes = "D6"
    autosize(ws, hdr_row + len(rows), len(headers))
    return ws


def build_workbook(units: list[dict], prop_mode: bool) -> openpyxl.Workbook:
    if prop_mode:
        units = [apply_proposal(u) for u in units]
    units = sorted(units, key=lambda u: u["Jednostka"].lower())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sub = "PROPOZYCJA balansu TW v3" if prop_mode else "STAN bieżący (units.json)"
    write_sheet(wb, "Atak", HEAD_ATK, [attack_row(u) for u in units], HF_ATK, f"Jednostki — ATAK · {sub}", prop_mode)
    write_sheet(wb, "Obrona", HEAD_DEF, [defense_row(u) for u in units], HF_DEF, f"Jednostki — OBRONA · {sub}", prop_mode)
    write_sheet(wb, "Pozostałe", HEAD_MISC, [misc_row(u) for u in units], HF_MISC, f"Jednostki — POZOSTAŁE · {sub}", prop_mode)

    # Podsumowanie M (Atak+Obrona) tylko w propozycji
    if prop_mode:
        ws = wb.create_sheet("Moc M (pole)")
        ws["A1"] = "M = A + O (pole; oblężnicze liczone jak w JSON, bez wall w A)"
        ws["A1"].font = Font(name=FONT, bold=True, size=12)
        h = ["Jednostka", "Epoka", "Rola", "A", "O", "M", "Uwaga propozycji"]
        for j, x in enumerate(h, 1):
            c = ws.cell(3, j, x)
            c.font = H
            c.fill = HF_PROP
            c.border = B
        notes = {
            "Konnica": "AP 30→8; szarża elitarna",
            "Hastati": "M < Triari (pilum AD=8, słabszy melee niż triarii)",
            "Rydwan Shang": "top Brąz — HP 30, szarża 8",
            "Katapulta": "wallAttack=16 przy oblężeniu",
        }
        for i, u in enumerate(units, 4):
            if is_siege_skip_m(u):
                continue
            a = attack_row(u)[-1]
            o = defense_row(u)[-1]
            m = round(a + o, 1)
            row = [u["Jednostka"], u.get("Epoka"), u.get("Rola (linia)"), a, o, m, notes.get(u["Jednostka"], "")]
            for j, val in enumerate(row, 1):
                ws.cell(i, j, val).border = B
        ws.freeze_panes = "A4"
        autosize(ws, 3 + len(units), 7)

    return wb


def is_siege_skip_m(u: dict) -> bool:
    return (u.get("Rola (linia)") or "") == "Oblężnicza"


def main():
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--apply-json", action="store_true", help="Zapis PROPOZYCJA do units.json")
    args = ap.parse_args()

    if args.apply_json:
        n, bak = apply_proposal_to_json()
        print(f"Zapisano units.json · zmian pól: {n} · backup: {bak}")

    units = json.load(open(UNITS_PATH, encoding="utf-8"))
    names = {u["Jednostka"] for u in units}
    missing = names - set(PROPOZYCJA.keys())
    extra = set(PROPOZYCJA.keys()) - names
    if missing:
        print("UWAGA: brak propozycji dla:", ", ".join(sorted(missing)), file=sys.stderr)
        sys.exit(1)
    if extra:
        print("UWAGA: nadmiarowe klucze propozycji:", ", ".join(sorted(extra)), file=sys.stderr)

    wb_stan = build_workbook(units, prop_mode=False)
    wb_stan.save(OUT_STAN)

    wb_prop = build_workbook(units, prop_mode=True)
    wb_prop.save(OUT_PROP)

    print(OUT_STAN)
    print(OUT_PROP)
    print(f"Jednostek: {len(units)}")


if __name__ == "__main__":
    main()
