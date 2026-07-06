#!/usr/bin/env python3
"""DEPRECATED (2026-06-30 — PANEL-MERGE):
  Excel: docs/archiwum/panele-legacy/
  Kanon: panele-sterowania/Panel-{A..E}.xlsx + export-{a..e}.py
  Nie używać do nowych zmian balansu.


export-bonusy-cyw.py — TARGETED export bonusow cywilizacji z Excel -> civs.json.

Zrodlo (kanon review Macieja):
  Civ-CYWILIZACJE/Panel-efekty-cyw-dyplomacja.xlsx, arkusz "Bonusy cywilizacji"

Aktualizuje WYLACZNIE tablice bonusy[] per cywilizacja w gra/data/civs.json.
Dopasowanie po kolumnie Nacja (= Cywilizacja). Kolejnosc bonusow w JSON = kolejnosc wierszy Excel.
Pola typCywilizacji, archetyp, nazwyKlastra, mnoznikHandelPieniadz, ikonaId — NIETKNIETE.

Kolumny arkusza (wiersz 1):
  Nacja | typCywilizacji | Typ efektu | Cel | Wartosc | Opis | Realizuje (dzial)

Uzycie (z katalogu gra/):
  python3 tools/export-bonusy-cyw.py
  python3 tools/export-bonusy-cyw.py --dry-run
  python3 tools/export-bonusy-cyw.py --xlsx SCIEZKA --json SCIEZKA

Wzorzec: export-ulepszenia.py (overlay, puste komorki = brak zmiany).
"""
import argparse
import json
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_XLSX = os.path.normpath(os.path.join(
    HERE, "..", "..", "docs", "archiwum", "panele-legacy", "Civ-CYWILIZACJE", "Panel-efekty-cyw-dyplomacja.xlsx"))
DEF_JSON = os.path.normpath(os.path.join(HERE, "..", "data", "civs.json"))
SHEET = "Bonusy cywilizacji"

COL_NACJA = 1
COL_TYP = 3
COL_CEL = 4
COL_WARTOSC = 5
COL_OPIS = 6
COL_REALIZUJE = 7


def coerce_wartosc(old, val):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return old
    if isinstance(old, (int, float)) and not isinstance(old, bool):
        try:
            return float(val) if isinstance(old, float) else int(round(float(val)))
        except (TypeError, ValueError):
            return old
    if isinstance(old, str):
        return str(val).strip()
    if old is None:
        return val
    return val


def read_bonus_rows(xlsx_path):
    if not os.path.exists(xlsx_path):
        sys.exit(f"Brak pliku: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Brak arkusza '{SHEET}' w {xlsx_path}")
    ws = wb[SHEET]
    by_nation = {}
    for r in range(2, ws.max_row + 1):
        nacja = ws.cell(r, COL_NACJA).value
        if nacja is None or str(nacja).strip() == "":
            continue
        key = str(nacja).strip()
        typ = ws.cell(r, COL_TYP).value
        cel = ws.cell(r, COL_CEL).value
        wartosc = ws.cell(r, COL_WARTOSC).value
        opis = ws.cell(r, COL_OPIS).value
        realizuje = ws.cell(r, COL_REALIZUJE).value
        if typ is None and cel is None and wartosc is None:
            continue
        entry = {
            "typ": str(typ).strip() if typ is not None else "",
            "cel": str(cel).strip() if cel is not None else "",
            "wartosc": wartosc if wartosc is not None else "",
            "opis": str(opis).strip() if opis is not None else "",
            "realizuje": str(realizuje).strip() if realizuje is not None else "",
        }
        by_nation.setdefault(key, []).append(entry)
    wb.close()
    return by_nation


def apply_bonusy(json_path, by_nation, dry_run=False):
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    changed = 0
    for row in data.get("cywilizacje", []):
        name = row.get("Cywilizacja")
        if not name or name not in by_nation:
            continue
        new_bonusy = []
        old_bonusy = row.get("bonusy") or []
        for i, nb in enumerate(by_nation[name]):
            ob = old_bonusy[i] if i < len(old_bonusy) else {}
            merged = {
                "typ": nb["typ"] or ob.get("typ", ""),
                "cel": nb["cel"] or ob.get("cel", ""),
                "wartosc": coerce_wartosc(ob.get("wartosc"), nb["wartosc"]),
                "opis": nb["opis"] or ob.get("opis", ""),
                "realizuje": nb["realizuje"] or ob.get("realizuje", ""),
            }
            if merged != (ob if i < len(old_bonusy) else {}):
                changed += 1
            new_bonusy.append(merged)
        if new_bonusy != old_bonusy:
            row["bonusy"] = new_bonusy
            print(f"  {name}: {len(new_bonusy)} bonusow zaktualizowano")
    if dry_run:
        print(f"[dry-run] zmiany={changed}, zapis pominiety")
        return
    backup = json_path + ".bak-CYWILIZACJE-export-bonusy"
    shutil.copy2(json_path, backup)
    tmp = json_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    json.load(open(tmp, encoding="utf-8"))
    os.replace(tmp, json_path)
    print(f"[export-bonusy-cyw] GOTOWE. backup: {backup}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEF_XLSX)
    ap.add_argument("--json", default=DEF_JSON)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    print(f"[export-bonusy-cyw] xlsx={args.xlsx}")
    by_nation = read_bonus_rows(args.xlsx)
    print(f"[export-bonusy-cyw] nacje z bonusami: {len(by_nation)}")
    apply_bonusy(args.json, by_nation, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
