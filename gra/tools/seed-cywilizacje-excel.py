#!/usr/bin/env python3
"""
seed-cywilizacje-excel.py — uzupełnia arkusze AI-zachowanie, Parametry-cyw, Dyplomacja
w Cywilizacje.xlsx wartościami startowymi (Panel-CYWILIZACJE + archetypy kodu).

Uruchom PRZED pierwszym export-civ-ai/params/dyplomacy-nations gdy arkusze są puste.
Maciej może potem korygować komórki w Excelu i re-eksportować.

  python tools/seed-cywilizacje-excel.py
  python tools/seed-cywilizacje-excel.py --dry-run
"""
import argparse
import json
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Wymagany openpyxl: pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, "..", ".."))
XLSX = os.path.join(ROOT, "Cywilizacje.xlsx")
PANEL = os.path.join(ROOT, "Civ-CYWILIZACJE", "Panel-CYWILIZACJE.xlsx")
CIVS_JSON = os.path.join(HERE, "..", "data", "civs.json")

NATIONS = [
    "Grecy", "Rzymianie", "Chińczycy", "Inkowie", "Zulusi",
    "Egipt", "Sumerowie", "Celtowie", "Germanie",
]

# Panel-CYWILIZACJE „AI - archetypy” (2026-06-26) — wartości 0..1 / delty priorytetów
ARCH = {
    "Grecy":     {"agg": 0.40, "trade": 0.75, "woj": 0,  "nau": 1,  "eko": 0,  "obr": 0},
    "Rzymianie": {"agg": 0.75, "trade": 0.50, "woj": 1,  "nau": 0,  "eko": 0,  "obr": 0},
    "Chińczycy": {"agg": 0.20, "trade": 0.85, "woj": -1, "nau": 1,  "eko": 1,  "obr": 0},
    "Inkowie":   {"agg": 0.45, "trade": 0.25, "woj": 0,  "nau": 0,  "eko": 0,  "obr": 1},
    "Zulusi":    {"agg": 0.90, "trade": 0.20, "woj": 2,  "nau": -1, "eko": -1, "obr": 0},
    "Egipt":     {"agg": 0.35, "trade": 0.60, "woj": 0,  "nau": 0,  "eko": 1,  "obr": 0},
    "Sumerowie": {"agg": 0.30, "trade": 0.65, "woj": -1, "nau": 2,  "eko": 0,  "obr": 0},
    "Celtowie":  {"agg": 0.60, "trade": 0.35, "woj": 2,  "nau": -1, "eko": 0,  "obr": 1},
    "Germanie":  {"agg": 0.65, "trade": 0.30, "woj": 2,  "nau": -1, "eko": -1, "obr": 0},
}

PROFIL_MAPY = "kopia_typu_obronna"  # D-START: miasta na mapie = kopie typu (obrona)


def priorytet_0_10(delta: int) -> int:
    return max(0, min(10, round(5 + delta * 1.5)))


def budynki_z_delty(woj: int, nau: int, eko: int, obr: int) -> str:
    parts = []
    if woj >= 1:
        parts.append("Koszary")
    if nau >= 1:
        parts.append("Biblioteka")
    if eko >= 1:
        parts.append("Rynek")
    if obr >= 1:
        parts.append("Mur")
    if not parts:
        parts = ["Koszary", "Rynek"]
    return ", ".join(parts)


def jednostki_z_civ(civ_row: dict) -> str:
    spec = (civ_row.get("Jednostka specjalna") or "").lower()
    styl = (civ_row.get("Styl / charakter") or "").lower()
    if "łucz" in spec or "kusz" in spec or "dystans" in styl:
        return "łucznicy, kawaleria lekka"
    if "rydwan" in styl:
        return "rydwany, łucznicy"
    if "kawaler" in styl or "konnic" in styl:
        return "kawaleria, piechota"
    return "piechota, elita"


def load_civs_by_name():
    with open(CIVS_JSON, encoding="utf-8") as f:
        data = json.load(f)
    return {r["Cywilizacja"]: r for r in data.get("cywilizacje", [])}


def try_load_panel_archetypes():
    """Opcjonalnie nadpisz ARCH z Panel-CYWILIZACJE jeśli istnieje."""
    if not os.path.exists(PANEL):
        return
    wb = openpyxl.load_workbook(PANEL, data_only=True, read_only=True)
    if "AI - archetypy" not in wb.sheetnames:
        wb.close()
        return
    ws = wb["AI - archetypy"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        if name not in ARCH:
            continue
        if row[1] is not None:
            ARCH[name]["agg"] = float(row[1])
        if row[2] is not None:
            ARCH[name]["trade"] = float(row[2])
        if row[3] is not None:
            ARCH[name]["woj"] = int(row[3])
        if row[4] is not None:
            ARCH[name]["nau"] = int(row[4])
        if row[5] is not None:
            ARCH[name]["eko"] = int(row[5])
        if row[6] is not None:
            ARCH[name]["obr"] = int(row[6])
    wb.close()


def seed_ai_sheet(ws, civs):
    # Nagłówek: wstaw Profil mapy przed Uwagi (kolumna 9)
    ws.cell(2, 9, "Profil mapy")
    ws.cell(2, 10, "Uwagi")
    for i, name in enumerate(NATIONS, start=3):
        a = ARCH[name]
        ws.cell(i, 1, name)
        ws.cell(i, 2, round(a["agg"] * 10))
        ws.cell(i, 3, 0)  # ekspansywnosc=0 (D-START kopie typu)
        ws.cell(i, 4, priorytet_0_10(a["woj"]))
        ws.cell(i, 5, priorytet_0_10(a["eko"]))
        ws.cell(i, 6, priorytet_0_10(a["nau"]))
        ws.cell(i, 7, max(1, min(10, round(a["agg"] * 10))))
        ws.cell(i, 8, max(0, min(10, round(a["agg"] * 6))))
        ws.cell(i, 9, PROFIL_MAPY)
        ws.cell(i, 10, f"seed CYW 2026-06-27; handel={a['trade']}")


def seed_params_sheet(ws, civs):
    for i, name in enumerate(NATIONS, start=3):
        a = ARCH[name]
        row = civs.get(name, {})
        ws.cell(i, 1, name)
        ws.cell(i, 2, budynki_z_delty(a["woj"], a["nau"], a["eko"], a["obr"]))
        ws.cell(i, 3, jednostki_z_civ(row))
        ws.cell(i, 4, round(1.0 + a["eko"] * 0.05, 2))
        ws.cell(i, 5, round(1.0 + a["trade"] * 0.1 - 0.05, 2))
        ws.cell(i, 6, "seed CYW 2026-06-27")


def seed_diplo_sheet(ws):
    for i, name in enumerate(NATIONS, start=3):
        a = ARCH[name]
        ws.cell(i, 1, name)
        ws.cell(i, 2, max(0, min(10, round((1 - a["agg"]) * 10))))
        ws.cell(i, 3, max(3, min(10, round(5 + (1 - a["agg"]) * 3))))
        ws.cell(i, 4, max(1, min(10, round(a["agg"] * 10))))
        ws.cell(i, 5, max(3, min(10, round(4 + a["agg"] * 4))))
        ws.cell(i, 6, max(0, min(10, round(a["trade"] * 10))))
        base = 50 + round((a["trade"] - a["agg"]) * 25)
        ws.cell(i, 7, max(20, min(80, base)))
        ws.cell(i, 8, "seed CYW 2026-06-27")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if not os.path.exists(args.xlsx):
        sys.exit(f"Brak: {args.xlsx}")

    try_load_panel_archetypes()
    civs = load_civs_by_name()

    if args.dry_run:
        print("[dry-run] wartości dla", len(NATIONS), "nacji (bez zapisu Excel)")
        for name in NATIONS:
            a = ARCH[name]
            print(f"  {name}: agg={a['agg']} trade={a['trade']} profil={PROFIL_MAPY}")
        return

    bak = args.xlsx + ".bak-CYWILIZACJE-seed"
    shutil.copy2(args.xlsx, bak)
    print(f"backup: {bak}")

    wb = openpyxl.load_workbook(args.xlsx)
    for sn in ("AI-zachowanie", "Parametry-cyw", "Dyplomacja"):
        if sn not in wb.sheetnames:
            sys.exit(f"Brak arkusza '{sn}'")
    seed_ai_sheet(wb["AI-zachowanie"], civs)
    seed_params_sheet(wb["Parametry-cyw"], civs)
    seed_diplo_sheet(wb["Dyplomacja"])
    wb.save(args.xlsx)
    wb.close()
    print(f"[seed-cywilizacje-excel] GOTOWE: {args.xlsx}")


if __name__ == "__main__":
    main()
